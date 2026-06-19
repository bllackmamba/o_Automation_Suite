"""
masterapp.py — Syndicate System, Single File
All logic, scraping, matching, collation, and UI in one place.
Run: streamlit run masterapp.py
Requires: pip install streamlit pandas numpy openpyxl playwright requests beautifulsoup4
          playwright install chromium
"""

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════
import streamlit as st
import pandas as pd
import numpy as np
import re, json, time, random, asyncio, os, logging
from pathlib import Path
from datetime import datetime
from multiprocessing import Pool, cpu_count

# ── 1-based row numbering for ALL tables ─────────────────────────────────────
# Users count rows from 1, not 0.
# Streamlit's st.dataframe does NOT reliably render a custom pandas .index —
# it shows its own internal 1-to-N counter for the visible slice regardless of
# what the pandas index is set to.  The portable fix is to insert an explicit
# "#" column with the correct row numbers and hide the pandas index.
# Fully guarded — never breaks rendering.
_orig_st_dataframe = st.dataframe
def _dataframe_1based(data=None, *args, **kwargs):
    try:
        if isinstance(data, pd.DataFrame) and not kwargs.get("hide_index", False):
            data = data.copy()
            _col = "#"
            if _col in data.columns:
                _col = "_row_"
            data.insert(0, _col, range(1, len(data) + 1))
            kwargs["hide_index"] = True
    except Exception as _e:
        logging.warning("_dataframe_1based: failed to insert row-number column: %s", _e)
    return _orig_st_dataframe(data, *args, **kwargs)
st.dataframe = _dataframe_1based

# ── Paginated dataframe helper ────────────────────────────────────────────────
# Navigation bar above AND below every table:
#   [⏮ First] [◀ Back]  Rows X–Y of N | Page P/T  [goto#]  [rows/pg▼]  [▶ Next] [⏭ Last]
# key must be unique per call site so each table has independent page state.
_PAGE_SIZE         = 50
_PAGE_SIZE_OPTIONS = [25, 50, 100, 200]

def show_paginated_df(df, key, use_container_width=True, height=None, hide_index=False, **kwargs):
    # _next_page is set by button clicks and acted on OUTSIDE the try/except so
    # that Streamlit's internal RerunException is never swallowed by a broad catch.
    _next_page = None
    _sk        = f"_pg_{key}"      # exposed outside try so rerun can use it
    _n_pages   = 1                  # safe fallback

    try:
        if not isinstance(df, pd.DataFrame):
            return _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)

        n_rows = len(df)
        if n_rows == 0:
            return _orig_st_dataframe(df, use_container_width=use_container_width,
                                      hide_index=hide_index, **kwargs)

        sk       = _sk
        sk_ps    = f"_ps_{key}"
        goto_key = f"_goto_{key}"
        ps_key   = f"_pssel_{key}"

        # ── initialise session state ──────────────────────────────────────
        if sk    not in st.session_state: st.session_state[sk]    = 0
        if sk_ps not in st.session_state: st.session_state[sk_ps] = _PAGE_SIZE

        page_size = int(st.session_state[sk_ps])
        n_pages   = max(1, (n_rows + page_size - 1) // page_size)
        _n_pages  = n_pages   # expose to outer scope for rerun guard
        cur       = max(0, min(int(st.session_state[sk]), n_pages - 1))

        start_r = cur * page_size + 1
        end_r   = min((cur + 1) * page_size, n_rows)

        # Sync goto widget to current page so it always reflects the live page.
        st.session_state[goto_key] = cur + 1

        # on_change callbacks — Streamlit calls these before the next rerun;
        # they only update session state (no st.rerun() needed here).
        def _on_goto():
            val = st.session_state.get(goto_key, 1)
            st.session_state[sk] = max(0, min(int(val) - 1, n_pages - 1))

        def _on_ps():
            new_ps = st.session_state.get(ps_key, _PAGE_SIZE)
            st.session_state[sk_ps] = int(new_ps)
            st.session_state[sk]    = 0

        # ── layout: table on the left, vertical nav panel on the right ───
        col_table, col_nav = st.columns([6, 1])

        with col_table:
            page_df = df.iloc[cur * page_size : (cur + 1) * page_size].copy()
            page_df = page_df.reset_index(drop=True)
            if not hide_index:
                _row_col = "#"
                if _row_col in page_df.columns:
                    _row_col = "_row_"
                page_df.insert(0, _row_col,
                               range(cur * page_size + 1,
                                     cur * page_size + len(page_df) + 1))
            disp_kwargs = dict(use_container_width=True, hide_index=True, **kwargs)
            if height is not None:
                disp_kwargs["height"] = height
            _orig_st_dataframe(page_df, **disp_kwargs)

        with col_nav:
            # ── page info ─────────────────────────────────────────────
            st.markdown(
                f"<div style='font-size:.75rem;color:#aaa;text-align:center;"
                f"padding:4px 0 6px 0;line-height:1.5'>"
                f"<b>{start_r:,}</b>–<b>{end_r:,}</b><br>"
                f"of <b>{n_rows:,}</b><br>"
                f"pg <b>{cur+1}</b>/<b>{n_pages}</b>"
                f"</div>",
                unsafe_allow_html=True)

            # ── First ─────────────────────────────────────────────────
            if st.button("« First", key=f"{sk}_first",
                         use_container_width=True, disabled=(cur == 0),
                         help="Jump to first page"):
                _next_page = 0          # handled OUTSIDE try/except

            # ── Back ──────────────────────────────────────────────────
            if st.button("▲ Prev", key=f"{sk}_back",
                         use_container_width=True, disabled=(cur == 0),
                         help="Previous page"):
                _next_page = cur - 1    # handled OUTSIDE try/except

            # ── Go to page ────────────────────────────────────────────
            st.number_input(
                "pg", min_value=1, max_value=n_pages, step=1,
                key=goto_key,
                on_change=_on_goto,
                label_visibility="collapsed",
                help=f"Go to page (1–{n_pages})")

            # ── Next ──────────────────────────────────────────────────
            if st.button("Next ▼", key=f"{sk}_next",
                         use_container_width=True, disabled=(cur >= n_pages - 1),
                         help="Next page"):
                _next_page = cur + 1    # handled OUTSIDE try/except

            # ── Last ──────────────────────────────────────────────────
            if st.button("Last »", key=f"{sk}_last",
                         use_container_width=True, disabled=(cur >= n_pages - 1),
                         help="Jump to last page"):
                _next_page = n_pages - 1  # handled OUTSIDE try/except

            # ── Rows per page ─────────────────────────────────────────
            st.markdown(
                "<div style='font-size:.7rem;color:#888;text-align:center;"
                "padding:6px 0 2px 0'>rows/pg</div>",
                unsafe_allow_html=True)
            ps_idx = _PAGE_SIZE_OPTIONS.index(page_size) \
                     if page_size in _PAGE_SIZE_OPTIONS else 1
            st.selectbox(
                "rows", _PAGE_SIZE_OPTIONS,
                index=ps_idx,
                key=ps_key,
                on_change=_on_ps,
                label_visibility="collapsed",
                help="Rows per page")

    except Exception:
        # Fallback: render plain table if anything in the pagination UI fails.
        # _next_page is checked AFTER this block so a pending rerun is never lost.
        try:
            _fb = df.copy() if isinstance(df, pd.DataFrame) else df
            if isinstance(_fb, pd.DataFrame) and not kwargs.get("hide_index", False):
                _fc = "#"
                if _fc in _fb.columns:
                    _fc = "_row_"
                _fb.insert(0, _fc, range(1, len(_fb) + 1))
                _orig_st_dataframe(_fb, use_container_width=use_container_width,
                                   hide_index=True, **kwargs)
            else:
                _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)
        except Exception:
            _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)

    # ── SAFE rerun: executed outside try/except so RerunException is never
    # swallowed.  This is the fix for buttons that appeared to do nothing.
    if _next_page is not None:
        st.session_state[_sk] = max(0, min(int(_next_page), _n_pages - 1))
        st.rerun()

try:
    from playwright.async_api import async_playwright
    PW_OK = True
except ImportError:
    PW_OK = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQ_OK = True
except ImportError:
    REQ_OK = False

# ═══════════════════════════════════════════════════════════════════════════════
# 1. PATH AUTO-DETECTION  (find_root / ROOT / DIRS / migration → syndicate_core/config.py)
# ═══════════════════════════════════════════════════════════════════════════════

# ── Auto-create .streamlit/config.toml to raise upload limit ──────────────
_streamlit_dir = Path(__file__).parent / ".streamlit"
_streamlit_dir.mkdir(exist_ok=True)
_config_path = _streamlit_dir / "config.toml"
if not _config_path.exists():
    _config_path.write_text(
        "[server]\n"
        "maxUploadSize = 10000\n"
        "maxMessageSize = 10000\n\n"
        "[browser]\n"
        "gatherUsageStats = false\n\n"
        "[runner]\n"
        "fastReruns = true\n"
    )

# ═══════════════════════════════════════════════════════════════════════════════
# 2. CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
from syndicate_core.config import *
from syndicate_core.scraping import *
from syndicate_core.pipeline import *
from syndicate_core.matching import *
from syndicate_core.generators import *
from syndicate_core.collation import *
from syndicate_core.b_sync import *

SCRAPE_URL = "https://www.thelott.com/syndicates?postcode={pc}"
SCRAPE_URL_WA = "https://www.lotterywest.wa.gov.au/play-online/syndicate-games?postcode={pc}"

SCRAPE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-AU,en;q=0.9",
    "Referer": "https://www.thelott.com/syndicates",
}


# ── Number value-range colour helper ──────────────────────────────────────────
def _num_colour(n: int) -> tuple:
    """Return (bg_hex, fg_hex) for a lottery number based on its value, not its SL."""
    if n <= 9:   return ("#FFFF00", "#000")
    if n <= 19:  return ("#00B0F0", "#000")
    if n <= 29:  return ("#A0A0A0", "#000")
    if n <= 40:  return ("#92D050", "#000")
    return               ("#FF69B4", "#fff")


# ── Naming convention helpers ──────────────────────────────────────────────
def parse_main_filename(fname: str) -> dict:
    """
    Parse `{cluster_label}_{lotto_type}_D{draw_no}.csv`
    e.g. `1n_oz_D1567.csv` → {cluster:'1n', lotto:'oz', draw:'D1567'}
    """
    stem = Path(fname).stem           # strip .csv
    parts = stem.split("_")
    result = {"cluster": "", "lotto": "", "draw": "", "raw": fname}
    if len(parts) >= 3:
        result["cluster"] = parts[0]
        result["lotto"]   = parts[1]
        result["draw"]    = "_".join(parts[2:])
    elif len(parts) == 2:
        result["lotto"]  = parts[0]
        result["draw"]   = parts[1]
    return result


def parse_cvi_filename(fname: str) -> dict:
    """
    Parse `CVI_{lotto_type}_{formula}_{date}.csv`
    e.g. `CVI_oz_BRD_2026_05_28.csv`
    """
    stem  = Path(fname).stem
    parts = stem.split("_")
    result = {"lotto": "", "formula": "", "date": "", "raw": fname}
    if len(parts) >= 3 and parts[0] == "CVI":
        result["lotto"]   = parts[1]
        result["formula"] = parts[2]
        result["date"]    = "_".join(parts[3:])
    return result


def scan_main_data_files() -> list[dict]:
    """
    Scan the active game's main_data_{game}/ folder and return only files
    matching the naming convention: {cluster}_{lotto_type}_D{draw_no}.csv
    e.g. 1n_oz_D1567.csv
    Rejects CVI files, SC files, or any file without a draw number (_D prefix).
    """
    lotto_codes = set(LOTTO_TYPES.keys())
    files = []
    for fp in sorted(active_game_dirs()["Main_Data"].glob("*.csv")):
        name = fp.stem   # without .csv
        parts = name.split("_")

        # Must have at least 3 parts: cluster, lotto, draw
        if len(parts) < 3:
            continue

        # Reject files that start with known non-main-data prefixes
        if parts[0] in ("CVI", "SC", "CF", "Container", "D", "MAIN"):
            continue

        # Must contain a lotto type code
        lotto_found = next((p for p in parts if p in lotto_codes), None)
        if not lotto_found:
            continue

        # Must contain a draw part starting with D followed by digits
        draw_found = next((p for p in parts
                          if p.startswith("D") and p[1:].isdigit()), None)
        if not draw_found:
            continue

        # Quick column check: reject if columns look like w1,w2,w3 (CVI file)
        try:
            sample = pd.read_csv(fp, nrows=2)
            w_cols = [c for c in sample.columns if re.match(r'^w\d+$', c)]
            n_cols_found = [c for c in sample.columns if re.match(r'^n\d+$', c, re.I)]
            if w_cols and not n_cols_found:
                continue   # looks like a CVI file, not main data
        except Exception:
            continue

        info = parse_main_filename(fp.name)
        info["path"] = str(fp)
        info["rows"] = 0
        try:
            info["rows"] = sum(1 for _ in open(fp)) - 1
        except Exception:
            pass
        files.append(info)
    return files


def scan_cvi_files(lotto_type: str = "") -> list[dict]:
    """Scan the active game's container_variable_inputs_{game}/ for CVI files."""
    files = []
    for fp in sorted(active_game_dirs()["CVI"].glob("CVI_*.csv")):
        info = parse_cvi_filename(fp.name)
        if not lotto_type or info["lotto"] == lotto_type:
            info["path"] = str(fp)
            files.append(info)
    return files


# ── Cluster registry ──────────────────────────────────────────────────────
CLUSTER_REGISTRY = ROOT / "clusters.json"

def load_clusters() -> list[dict]:
    if CLUSTER_REGISTRY.exists():
        try:
            return json.loads(CLUSTER_REGISTRY.read_text())
        except Exception as _e:
            import logging as _lg
            _lg.warning("load_clusters: failed to parse %s — cluster history unavailable (%s)",
                        CLUSTER_REGISTRY, _e)
    return []

def save_clusters(clusters: list[dict]):
    import os as _os
    tmp = CLUSTER_REGISTRY.with_suffix(".tmp")
    try:
        tmp.write_text(json.dumps(clusters, indent=2))
        _os.replace(tmp, CLUSTER_REGISTRY)
    except Exception:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
        raise

def next_cluster_id(clusters: list[dict]) -> str:
    if not clusters:
        return "001"
    last = max(int(c["id"]) for c in clusters if str(c.get("id","0")).isdigit())
    return f"{last+1:03d}"


# ═══════════════════════════════════════════════════════════════════════════════
# 3. SESSION STATE — single shared dict S (unscoped) + game-scoped helpers
# ═══════════════════════════════════════════════════════════════════════════════

def gkey(name: str) -> str:
    """Return a session_state key scoped to the active game."""
    return f"{name}__{active_game()}"

def gs(name: str, default=None):
    """Get a game-scoped session_state value, with default."""
    return st.session_state.get(gkey(name), default)

def gs_set(name: str, value):
    """Set a game-scoped session_state value."""
    st.session_state[gkey(name)] = value
    return value

def _auto_load_b(game_key: str = "sat") -> pd.DataFrame:
    """Try to auto-load the Base file for game_key from the project folder tree.

    Search order:
      1. Game-specific file under ROOT/Games/<GAME>/…/Base_<game>.xlsx
      2. Anywhere under ROOT matching Base_<game>.xlsx
      3. Shared Base.xlsx anywhere under ROOT
      4. Legacy f_rules_Gclaude.xlsx anywhere under ROOT
    Returns a DataFrame of w-columns (column-oriented), or empty DataFrame.
    Data is NEVER reordered or altered — values are read exactly as stored.
    """
    gcfg = GAMES_CFG.get(game_key, {})
    b_file = gcfg.get("b_file", f"Base_{game_key}.xlsx")
    b_sheet = gcfg.get("b_sheet", f"B_{game_key}")
    b_sheet_legacy = gcfg.get("b_sheet_legacy", "")

    candidates = (list(ROOT.rglob(b_file))
                  or list(ROOT.rglob("Base.xlsx"))
                  or list(ROOT.rglob("f_rules_Gclaude.xlsx")))
    if not candidates:
        return pd.DataFrame()

    for path in candidates:
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
            sheet = None
            for cand in (b_sheet, game_key.upper(), b_sheet_legacy):
                if cand and cand in xl.sheet_names:
                    sheet = cand
                    break
            if sheet is None and len(xl.sheet_names) == 1:
                sheet = xl.sheet_names[0]
            if sheet is None:
                continue
            raw = xl.parse(sheet, header=None)

            def _safe_nums_from_row(series):
                out = []
                for v in series.dropna():
                    try:
                        fv = float(v)
                        if fv >= 1:
                            out.append(int(fv))
                    except (ValueError, TypeError):
                        pass
                return out

            # ── Strategy 0: col 0 has w-prefixed row labels (row-oriented) ────
            # B_sat layout: col A = "w1"/"w701"…; numbers run across each row.
            # Returns a row-oriented DataFrame — each row is one w-set.
            _col0 = [str(raw.iloc[r, 0]).strip() for r in range(raw.shape[0])]
            _w_row_pairs = [(r, _col0[r]) for r in range(len(_col0))
                            if _col0[r].lower().startswith("w")]
            if _w_row_pairs:
                _rows = []
                for _rr, _wlabel in _w_row_pairs:
                    _nums = _safe_nums_from_row(raw.iloc[_rr, 1:])
                    if _nums:
                        _row = {"w": _wlabel}
                        _row.update({f"pos_{i+1}": n for i, n in enumerate(_nums)})
                        _rows.append(_row)
                if _rows:
                    return pd.DataFrame(_rows)

            # ── Strategy 1: row 0 has w-prefixed column headers ───────────────
            w_col_pairs = [
                (c, str(raw.iloc[0, c]).strip())
                for c in range(raw.shape[1])
                if str(raw.iloc[0, c]).strip().lower().startswith("w")
            ]
            b_data = {}
            for actual_col, wc in w_col_pairs:
                col_vals = raw.iloc[1:, actual_col].dropna()
                nums = [int(float(v)) for v in col_vals
                        if str(v).replace(".", "").replace("-", "").isdigit()
                        and float(v) >= 1]
                if nums:
                    b_data[wc] = pd.Series(nums)
            if b_data:
                return pd.DataFrame(b_data)
        except Exception:
            continue
    return pd.DataFrame()


def _init_state():
    # Re-init if S exists but uses the old unscoped key format (pre-game-isolation refactor).
    # Guard: skip if S (unscoped) already initialised AND game-scoped keys exist.
    # gkey("B") == "B__sat" at startup (active_game() defaults to "sat").
    if "S" in st.session_state and gkey("B") in st.session_state:
        return
    # Auto-load B for the default game (sat) at startup.
    # The active game can change via the game selector; B will reload in the B tab.
    _b_init = _auto_load_b("sat")
    _sat_dirs = game_dirs("sat")
    if _b_init.empty:
        _b_init = _load_file(_sat_dirs["Base"] / "B.xlsx")   # legacy fallback

    # Unscoped infrastructure — lives in the S dict as before.
    st.session_state.S = {
        "cf_active":         {r[1]: True for r in CF_ROWS},
        "auto":              {},
        "confirmed_api_url": "",
        "cookie_str":        "",
        "scrape_log":        [],
        "container_status":  {},
    }
    # Game-scoped DataFrames — stored directly in st.session_state via gs_set().
    # Only the default game (sat) is pre-loaded; other games populate on demand.
    gs_set("B",              _b_init)
    gs_set("R",              pd.DataFrame())
    gs_set("D",              _load_file(_sat_dirs["Direct"]       / "D.xlsx"))
    gs_set("Sp",             _load_sets_file(_sat_dirs["Splits"]       / "data_1b.xlsx"))
    gs_set("So",             _load_sets_file(_sat_dirs["Splits_Combi"] / "data.xlsx"))
    gs_set("main_data",      pd.DataFrame())
    gs_set("main_data_path", "")
    gs_set("cvi",            {})
    gs_set("results",        {})
    # Init container status
    for db in DASHBOARDS:
        st.session_state.S["container_status"][db] = pd.DataFrame({
            "Name": [db], "Status": ["Stopped"],
            "Memory Usage": [""], "CPU Usage": [""],
            "Percent progress %": [0],
            "Time Guestimate Min": [""], "Time Guestimate Max": [""],
        })

S = None  # set after _init_state() call in main

# ═══════════════════════════════════════════════════════════════════════════════
# 4. FILE I/O — large-file safe, all formats
# ═══════════════════════════════════════════════════════════════════════════════
LARGE_FILE_THRESHOLD = 50 * 1024 * 1024   # 50 MB


def _load_file(path: Path, max_rows: int = 0,
               numeric: bool = True) -> pd.DataFrame:
    """
    Load any supported file format.
    Large CSVs (>50 MB) are read in 500K-row chunks to avoid freezing.
    accdb/mdb files must be placed directly in the project folder —
    Streamlit cannot upload them via the browser (browser restriction, not ours).
    Use mdb-tools on Mac: brew install mdb-tools
    """
    if not path or not path.exists():
        return pd.DataFrame()
    try:
        ext  = path.suffix.lower()
        size = path.stat().st_size

        if ext == ".csv":
            if max_rows > 0:
                df = pd.read_csv(path, nrows=max_rows)
            elif size > LARGE_FILE_THRESHOLD:
                chunks = pd.read_csv(path, chunksize=500_000)
                df = pd.concat(chunks, ignore_index=True)
            else:
                df = pd.read_csv(path)
        elif ext in [".xlsx", ".xlsm"]:
            df = pd.read_excel(path, engine="openpyxl",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext == ".xls":
            df = pd.read_excel(path, engine="xlrd",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext in [".accdb", ".mdb"]:
            try:
                import pyodbc
                conn = pyodbc.connect(
                    f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={path};")
                cursor = conn.cursor()
                tables = [t.table_name for t in cursor.tables(tableType="TABLE")]
                q = (f"SELECT TOP {max_rows} * FROM [{tables[0]}]"
                     if max_rows > 0 else f"SELECT * FROM [{tables[0]}]")
                df = pd.read_sql(q, conn)
                conn.close()
            except Exception:
                import subprocess, io
                try:
                    tables_raw = subprocess.check_output(
                        ["mdb-tables", "-1", str(path)], timeout=30).decode()
                    tables = [t for t in tables_raw.strip().split("\n") if t]
                    csv_data = subprocess.check_output(
                        ["mdb-export", str(path), tables[0]],
                        timeout=120).decode()
                    df = pd.read_csv(io.StringIO(csv_data))
                    if max_rows > 0:
                        df = df.head(max_rows)
                except Exception:
                    return pd.DataFrame()
        elif ext == ".txt":
            for sep in ["\t", ",", " ", "|"]:
                try:
                    df = pd.read_csv(path, sep=sep,
                                     nrows=max_rows if max_rows > 0 else None)
                    if len(df.columns) >= 3:
                        break
                except Exception:
                    pass
            else:
                return pd.DataFrame()
        elif ext == ".xml":
            df = pd.read_xml(path)
        else:
            return pd.DataFrame()

        if numeric:
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()

def _load_file_preview(path: Path, n: int = 20) -> pd.DataFrame:
    """Load first N rows only — never freezes."""
    return _load_file(path, max_rows=n, numeric=False)

def _load_sets_file(path: Path) -> pd.DataFrame:
    """Load a Sp / So / Ep CSV saved row-oriented (via _sets_df_to_rows) and
    restore it to column-oriented format expected by S['Sp'] / S['So'] / S['Ep'].

    If the file has a 'set' column (row-oriented), applies _rows_to_sets_df to
    transpose back to column-oriented.  Falls back to plain _load_file for legacy
    column-oriented files.  R files are never passed here — they stay column-oriented.
    """
    df = _load_file(path, numeric=False)
    if df.empty:
        return df
    if "set" in df.columns:
        # Row-oriented on disk — transpose back to column-oriented for in-memory use
        return _rows_to_sets_df(df, set_col="set")
    return df




def _save_df(df: pd.DataFrame, path: Path):
    """Save DataFrame to CSV, padding n-columns to consistent width.

    Operates on a copy so the caller's DataFrame is never mutated. (Previously
    this added missing n-columns in place, silently altering the object the
    caller still held.)
    """
    if df.empty:
        return
    out = df.copy()
    nc = sorted([c for c in out.columns if re.match(r'^n\d+$', c)],
                key=lambda x: int(x[1:]))
    if nc:
        max_n = int(nc[-1][1:])
        for i in range(1, max_n+1):
            if f"n{i}" not in out.columns:
                out[f"n{i}"] = None
    _write_csv_atomic(path, out, index=False)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_styled_excel(df: pd.DataFrame, cvi_set: set,
                    n_cols: list, sheet_name: str = "Data") -> bytes:
    """
    Export DataFrame to Excel with highlighted cells.
    Cells whose value is in cvi_set get a salmon/pink background.
    Count column gets a yellow background.
    Returns bytes suitable for st.download_button.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    PINK   = PatternFill("solid", fgColor="FFB3B3")
    YELLOW = PatternFill("solid", fgColor="FFFF99")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    EXCEL_ROW_LIMIT = 1_048_575  # Excel max rows minus header
    truncated = False
    if len(df) > EXCEL_ROW_LIMIT:
        df = df.iloc[:EXCEL_ROW_LIMIT].copy()
        truncated = True

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    cols = list(df.columns)
    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = GREY
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = 8

    if truncated:
        # Write a notice row beneath the header
        ws.insert_rows(2)
        notice_cell = ws.cell(row=2, column=1,
                              value=f"⚠ Truncated to {EXCEL_ROW_LIMIT:,} rows "
                                    f"(Excel limit). Full data available as CSV.")
        notice_cell.font = Font(bold=True, color="FF0000")

    # Data rows (start at row 3 if truncated notice inserted, else row 2)
    row_offset = 3 if truncated else 2
    for ri, (_, row) in enumerate(df.iterrows(), row_offset):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci,
                           value=None if pd.isna(val) else val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            if col == "Count":
                cell.fill = YELLOW
            elif col in n_cols:
                try:
                    if int(round(float(val))) in cvi_set:
                        cell.fill = PINK
                except (ValueError, TypeError):
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# 5. HIGH-PERFORMANCE MATCHING ENGINE (numpy vectorised, chunked, parallel)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 6. COLLATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
# _to_w_rows is imported from syndicate_core.collation (above) so it can be
# tested independently without importing the full Streamlit app.  The import
# is part of `from syndicate_core.collation import *` in the imports block.


# ── Universal file reader ──────────────────────────────────────────────────
_UPLOAD_TYPES = [
    "csv", "txt", "tsv",
    "xlsx", "xls", "xlsb", "xlsm", "ods",
    "html", "htm",
    "json", "xml",
    "accdb", "mdb",
]

def _read_uploaded(f, **kw) -> "pd.DataFrame":
    """Read any Streamlit UploadedFile into a DataFrame regardless of extension.

    Supports: CSV/TSV/TXT, Excel (xlsx/xls/xlsb/xlsm/ods),
              HTML tables, JSON, XML, and Access (accdb/mdb via mdbtools).
    Extra kwargs are forwarded to the underlying pandas reader.
    """
    import io as _io
    name = f.name.lower()
    raw = f.getvalue()

    if name.endswith((".xlsx", ".xlsm", ".xlsb")):
        return pd.read_excel(_io.BytesIO(raw), engine="openpyxl", **kw)
    if name.endswith(".xls"):
        try:
            return pd.read_excel(_io.BytesIO(raw), engine="xlrd", **kw)
        except Exception:
            return pd.read_excel(_io.BytesIO(raw), engine="openpyxl", **kw)
    if name.endswith(".ods"):
        return pd.read_excel(_io.BytesIO(raw), engine="odf", **kw)
    if name.endswith((".html", ".htm")):
        tables = pd.read_html(_io.BytesIO(raw), **kw)
        return tables[0] if tables else pd.DataFrame()
    if name.endswith(".json"):
        return pd.read_json(_io.BytesIO(raw), **kw)
    if name.endswith(".xml"):
        return pd.read_xml(_io.BytesIO(raw), **kw)
    if name.endswith(".tsv"):
        return pd.read_csv(_io.BytesIO(raw), sep="\t", **kw)
    if name.endswith((".accdb", ".mdb")):
        try:
            import tempfile, subprocess, os
            with tempfile.NamedTemporaryFile(suffix=os.path.splitext(name)[1],
                                            delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            tables_out = subprocess.run(
                ["mdb-tables", "-1", tmp_path],
                capture_output=True, text=True, timeout=30)
            tbl = tables_out.stdout.strip().splitlines()[0] if tables_out.returncode == 0 else ""
            if tbl:
                result = subprocess.run(
                    ["mdb-export", tmp_path, tbl],
                    capture_output=True, text=True, timeout=30)
                if result.returncode == 0:
                    return pd.read_csv(_io.StringIO(result.stdout), **kw)
            raise RuntimeError(f"mdb-export failed. stderr: {tables_out.stderr}")
        except FileNotFoundError:
            raise RuntimeError(
                "Access (.accdb/.mdb) files need mdbtools installed on the server.\n"
                "  macOS: brew install mdbtools\n"
                "  Linux: sudo apt install mdbtools")
    # Default → CSV (handles .txt and plain .csv)
    for sep in (",", ";", "\t", "|", None):
        try:
            kwargs = dict(**kw)
            if sep is not None:
                kwargs["sep"] = sep
            else:
                kwargs.update(sep=None, engine="python")
            return pd.read_csv(_io.BytesIO(raw), **kwargs)
        except Exception:
            continue
    raise ValueError(f"Could not parse uploaded file: {f.name}")


def execute_collation(components: list[str]) -> pd.DataFrame:
    """
    Build a formula's CVI by STACKING each variable's w-sets as ROWS (vertically)
    and numbering the position columns w1, w2, … across the widest combination.

    Data model (confirmed — ROW orientation, to respect the spreadsheet column
    ceiling of ~16k columns vs ~1M rows):
      • Each w-set (a syndicate pick, a base set, a rainbow set, …) is one ROW.
      • D (Direct) is already row-oriented (one syndicate per row) → kept as rows.
      • B, R, Ep, Sp, So are stored column-wise → TRANSPOSED so each becomes a row.
      • A numbered/pure formula (e.g. D1D2D3) = "all of that variable joined" → a
        SINGLE block (strip trailing digits + de-duplicate): D1D2D3 == one D block.
    Result is TALL. e.g. sat BRD = B rows + R rows + 354,682 D rows, columns
    w1…w(longest combination). Matching is row-vs-row against Main_Data.
    """
    # Resolve tokens → base variables: strip trailing digits (D1→D, Sp2→Sp,
    # B3→B) and de-duplicate while preserving left-to-right order.
    seen, base_vars = set(), []
    for comp in components:
        base = re.sub(r"\d+$", "", str(comp)).strip()
        if base and base not in seen:
            seen.add(base)
            base_vars.append(base)

    pieces = []
    for var in base_vars:
        df = gs(var)
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        block = _to_w_rows(df, is_direct=(var == "D"),
                           force_column_oriented=(var == "Sp"))
        if block is None or block.empty:
            continue
        # Rename value columns (all except Set_Label) to w1, w2, …
        val_cols = [c for c in block.columns if c != "Set_Label"]
        block = block.rename(columns={c: f"w{i+1}" for i, c in enumerate(val_cols)})
        block.insert(0, "Source", var)
        pieces.append(block)

    if not pieces:
        return pd.DataFrame()

    # Stack vertically; align to the widest combination (pad missing positions).
    combined = pd.concat(pieces, axis=0, ignore_index=True)
    wcols = sorted([c for c in combined.columns if str(c).startswith("w")],
                   key=lambda x: int(x[1:]))
    combined = combined[["Source", "Set_Label"] + wcols]
    combined.insert(0, "Row_ID", range(1, len(combined) + 1))
    return combined

# ═══════════════════════════════════════════════════════════════════════════════
# 7. SCRAPER ENGINE (Playwright + requests)
# ═══════════════════════════════════════════════════════════════════════════════
def _extract_nums(text) -> list[int]:
    """Extract positive integers from any text. No upper cap — supports any jurisdiction."""
    return sorted(set(int(n) for n in re.findall(r'\b(\d{1,3})\b', str(text))
                  if int(n) >= 1))

def _valid_combo(nums: list) -> bool:
    if not (6 <= len(nums) <= 25): return False
    if max(nums) - min(nums) < 10: return False
    return True

def _parse_json(body: str, pc: int, state: str) -> list[dict]:
    """Walk any JSON shape to extract syndicate number combinations."""
    try:
        data = json.loads(body)
    except Exception:
        return []
    records, seen = [], set()

    def walk(obj, depth=0):
        if depth > 8: return
        if isinstance(obj, list):
            for item in obj: walk(item, depth+1)
        elif isinstance(obj, dict):
            nums = []
            for f in ["numbers","combination","selections","picks","balls",
                      "gameNumbers","game_numbers","selectedNumbers","ticketNumbers"]:
                if f in obj and isinstance(obj[f], list):
                    nums = sorted(set(int(x) for x in obj[f]
                                  if str(x).isdigit() and int(x) >= 1))
                    if nums: break
            if _valid_combo(nums):
                key = tuple(nums)
                if key not in seen:
                    seen.add(key)
                    def g(*keys):
                        for k in keys:
                            v = obj.get(k)
                            if v is not None and str(v).strip(): return str(v).strip()
                        return ""
                    row = {
                        "Postcode":        g("postcode","post_code") or str(pc),
                        "State":           g("state","jurisdiction") or state,
                        "Syndicate Number": g("id","syndicateId","syndicateNumber"),
                        "Agent":           g("agentName","retailer","outlet"),
                        "Title":           g("title","name","description"),
                        "Price per Share": g("price","sharePrice"),
                        "Shares":          g("shares","sharesAvailable"),
                        "Draw Number":     g("drawNumber","draw_number","drawId","drawNo"),
                        "Draw Date":       g("drawDate","draw_date","scheduledDate"),
                        "Game Type":       g("gameType","type","product","gameName") or
                                          ("Normal" if len(nums)<=7 else f"System {len(nums)}"),
                        "Length":          len(nums),
                    }
                    for j,n in enumerate(nums,1): row[f"n{j}"] = n
                    records.append(row)
            for v in obj.values(): walk(v, depth+1)

    walk(data)
    return records


def _parse_html(html: str, pc: int, state: str) -> list[dict]:
    """HTML fallback — extracts number clusters from rendered page."""
    if not REQ_OK or not html: return []
    soup = BeautifulSoup(html, "html.parser")
    for t in soup(["script","style","noscript","head"]): t.decompose()
    records, seen = [], set()
    for el in soup.find_all(True):
        if len(list(el.children)) > 15: continue
        nums = _extract_nums(el.get_text(" ",strip=True))
        if _valid_combo(nums):
            key = tuple(nums)
            if key in seen: continue
            seen.add(key)
            row = {"Postcode":pc,"State":state,
                   "Syndicate Number":"","Agent":"","Title":"",
                   "Price per Share":"","Shares":"",
                   "Draw Number":"","Draw Date":"",
                   "Game Type":"Normal" if len(nums)<=7 else f"System {len(nums)}",
                   "Length":len(nums)}
            for j,n in enumerate(nums,1): row[f"n{j}"] = n
            records.append(row)
    return records


async def _pw_scrape_state(state: str, postcodes: list,
                            api_url: str, cookie_str: str,
                            pb, stx) -> list[dict]:
    """Playwright scrape — intercepts JSON API responses per postcode."""
    all_recs = []
    total = len(postcodes)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            user_agent=SCRAPE_HEADERS["User-Agent"],
            locale="en-AU", timezone_id="Australia/Sydney",
            ignore_https_errors=True,
        )
        if cookie_str:
            cookies = [{"name":k.strip(),"value":v.strip(),
                        "domain":".thelott.com","path":"/"}
                       for kv in cookie_str.split(";")
                       for k,_,v in [kv.partition("=")]]
            await ctx.add_cookies(cookies)
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        page = await ctx.new_page()

        for i, pc in enumerate(postcodes):
            pb.progress((i+1)/total,
                text=f"[{state}] {pc} ({i+1}/{total}) — {len(all_recs)} records")
            stx.caption(f"Scraping {state} · postcode {pc}")

            captured = []
            async def on_resp(resp):
                ct = resp.headers.get("content-type","")
                if resp.status == 200 and \
                   ("json" in ct or any(k in resp.url.lower()
                    for k in ["syndic","api","play","search"])):
                    try:
                        body = (await resp.body()).decode("utf-8","ignore")
                        if body.strip()[:1] in ["{","["]:
                            captured.append({"url":resp.url,"body":body})
                    except Exception as _e:
                        logging.warning("scraper: could not read response body from %s: %s", resp.url, _e)

            page.on("response", on_resp)
            try:
                url = api_url.format(pc=pc) if "{pc}" in api_url \
                      else f"{api_url}?postcode={pc}"
                await page.goto(url, timeout=25000, wait_until="load")
                await asyncio.sleep(3)
                for sel in ["input[placeholder*='postcode' i]",
                            "input[type='search']"]:
                    try:
                        el = page.locator(sel).first
                        if await el.is_visible(timeout=1000):
                            await el.fill(str(pc))
                            await el.press("Enter")
                            await asyncio.sleep(2)
                            break
                    except Exception as _e:
                        logging.warning("scraper: locator %r interaction failed: %s", sel, _e)
                await asyncio.sleep(2)

                recs = []
                for cap in captured:
                    recs.extend(_parse_json(cap["body"], pc, state))
                if not recs:
                    html = await page.content()
                    recs = _parse_html(html, pc, state)
                all_recs.extend(recs)

            except Exception as ex:
                st.session_state.S["scrape_log"].append(f"{state} {pc}: {ex}")
            finally:
                page.remove_listener("response", on_resp)

            await asyncio.sleep(random.uniform(1.0, 2.5))

        await browser.close()

    return all_recs


def _run_sync(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try: return loop.run_until_complete(coro)
    finally: loop.close()


def _data_status() -> list:
    """Return status info for each D_*.csv file in Main_Data/."""
    rows = []
    for state in ["NSW", "VIC", "QLD", "SA", "TAS"]:
        fp = DIRS["Global_Scraper"] / f"D_{state}.csv"
        if fp.exists():
            stat = fp.stat()
            age_h = (datetime.now().timestamp() - stat.st_mtime) / 3600
            try:
                n = sum(1 for _ in open(fp)) - 1
            except Exception:
                n = 0
            rows.append({
                "state": state, "file": fp.name, "rows": n,
                "size_kb": round(stat.st_size / 1024, 1),
                "age_h": round(age_h, 1),
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%d %b %H:%M"),
                "exists": True,
            })
        else:
            rows.append({"state": state, "file": f"D_{state}.csv",
                         "rows": 0, "size_kb": 0, "age_h": 9999,
                         "modified": "—", "exists": False})
    return rows




# ═══════════════════════════════════════════════════════════════════════════════
# 8. CVI MATRIX — transposition + variable slicing
# ═══════════════════════════════════════════════════════════════════════════════
def build_w_matrix(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Sort rows longest→shortest, transpose each into a w-column, pad with None.

    The source 'number' columns of a D file are the syndicate picks, labelled
    **w1…wN** (project convention: syndicate = w, main data = n). Older/main-data
    files may instead use n1…nN. Accept EITHER prefix as the input columns so D
    (w-columns) loads correctly. (Previously this only looked for n1…nN, so D —
    which is w-columns — produced an empty matrix and the "No n-columns found"
    error.) The OUTPUT is always w-columns regardless of input naming.
    """
    src_cols = sorted(
        [c for c in df_raw.columns if re.match(r'^[wn]\d+$', c, re.I)],
        key=lambda x: int(x[1:]))
    if not src_cols:
        return pd.DataFrame()
    rows_as_lists = []
    for _, row in df_raw.iterrows():
        nums = [int(row[c]) for c in src_cols
                if pd.notna(row[c]) and str(row[c]).strip().lstrip('-').isdigit()]
        if nums: rows_as_lists.append(sorted(nums))
    if not rows_as_lists: return pd.DataFrame()
    rows_as_lists.sort(key=len, reverse=True)
    max_len = max(len(r) for r in rows_as_lists)
    w_dict = {f"w{i+1}": (r + [None]*(max_len-len(r)))
              for i, r in enumerate(rows_as_lists)}
    return pd.DataFrame(w_dict)


def d_to_w_only(df: pd.DataFrame) -> pd.DataFrame:
    """Return a variable as CLEAN w-columns, transposing ONLY raw syndicate data.

    Two very different shapes arrive here:
      • RAW D  — one ROW per syndicate, picks in w1…wN, plus syndicate metadata
                 (Syndicate_ID, Game, Games, PB, Draw_Number, Outlet_ID, …).
                 This must be TRANSPOSED → one w-column per syndicate.
      • ALREADY-WIDE (B, R, Ep, Sp, So, or a built W-matrix) — already one COLUMN
                 per w-set. These must be returned AS-IS (never transposed); we only
                 drop any stray non-w label column.

    The previous heuristic transposed whenever ANY non-w column was present, which
    flattened B (720 w-columns) into one long column. The real signal that a frame
    is raw D is the presence of SYNDICATE METADATA — so we key on that instead.
    """
    if df is None or df.empty:
        return df
    wcols = [c for c in df.columns if re.match(r'^w\d+$', str(c), re.I)]
    ncols = [c for c in df.columns if re.match(r'^n\d+$', str(c), re.I)]
    D_META = {"syndicate_id", "syndicate_name", "game", "games", "pb",
              "draw_number", "draw_numbers", "outlet_id", "outlet_name",
              "postcode", "state", "share_cost", "available_shares",
              "total_shares", "address", "suburb"}
    has_d_meta = any(str(c).strip().lower() in D_META for c in df.columns)

    # RAW D (rows = syndicates): transpose to one w-column per syndicate.
    if has_d_meta:
        return build_w_matrix(df)
    # ALREADY-WIDE variable (B/R/Ep/Sp/So or built matrix): keep w-columns as-is.
    if wcols:
        return df[wcols]
    # No w-columns, no D metadata, but n-columns present (main-data style): transpose.
    if ncols:
        return build_w_matrix(df)
    # Nothing recognisable — return unchanged rather than risk a bad transpose.
    return df


def slice_variables(w_mat: pd.DataFrame) -> dict:
    """Auto-slice Ep (top 8), Sp (top 4 lanes a-d), So (union of Sp lanes)."""
    w_keys = sorted([c for c in w_mat.columns if re.match(r'^w\d+$',c)],
                    key=lambda x: int(x[1:]))
    def col_ints(k):
        result = []
        for v in w_mat[k].dropna():
            try:
                n = float(str(v))
                if not pd.isna(n) and n >= 1:
                    result.append(int(round(n)))
            except (ValueError, TypeError):
                pass
        return result

    ep_keys = w_keys[:8]
    sp_keys = w_keys[:4]
    ep_raw = {k: col_ints(k) for k in ep_keys}
    sp_raw = {"lane_a": col_ints(sp_keys[0]) if len(sp_keys)>0 else [],
              "lane_b": col_ints(sp_keys[1]) if len(sp_keys)>1 else [],
              "lane_c": col_ints(sp_keys[2]) if len(sp_keys)>2 else [],
              "lane_d": col_ints(sp_keys[3]) if len(sp_keys)>3 else []}

    def to_df(d):
        ml = max((len(v) for v in d.values()), default=0)
        return pd.DataFrame({k: v+[None]*(ml-len(v)) for k,v in d.items()})

    so_union = sorted(set(sum(sp_raw.values(),[])))
    return {"Ep": to_df(ep_raw), "Sp": to_df(sp_raw),
            "So": pd.DataFrame({"So_union": so_union})}





def append_draw_to_b(b_df: pd.DataFrame, numbers: list,
                     draw_label: str = "", game: str = None) -> pd.DataFrame:
    """Append a new draw (sorted list of ints) to a B DataFrame.

    Handles both row-oriented B (col 'w' = label, pos_1/pos_2/… = numbers)
    and legacy column-oriented B (each column = a w-set).
    Existing rows are NEVER modified.  Returns unchanged b_df if numbers is empty.

    Row-oriented insertion point
    -----------------------------
    B is reverse-chronological (newest-first): row index b_hist_start holds the
    newest draw, increasing row index = older draws (see b_sync.py). If `game`
    is given and GAMES_CFG[game]["b_hist_start"] is configured, the new draw is
    inserted at that row (shifting existing rows down) via b_sync._append_draws
    — the SAME insertion point used by the "🔄 Sync B to latest" button
    (b_sync.append_draws_to_b), so both paths stay consistent.

    If `game` is not given, or b_hist_start isn't configured for it yet,
    falls back to the legacy bottom-append.
    """
    if not numbers:
        return b_df
    nums = sorted([int(n) for n in numbers if n >= 1])

    # ── Row-oriented format (preferred) ───────────────────────────────────────
    is_row_oriented = (isinstance(b_df, pd.DataFrame)
                       and not b_df.empty
                       and "w" in b_df.columns)
    if b_df is None or b_df.empty:
        # Default to row-oriented for new B
        label = draw_label or "w1"
        row = {"w": label}
        row.update({f"pos_{i+1}": n for i, n in enumerate(nums)})
        return pd.DataFrame([row])

    if is_row_oriented:
        label = draw_label or f"w{len(b_df) + 1}"
        b_hist_start = GAMES_CFG.get(game, {}).get("b_hist_start") if game else None
        if b_hist_start is not None:
            # Same insertion point as b_sync.append_draws_to_b (newest-first).
            draw_dict = {"draw": label, "date": "", "numbers": nums}
            return _append_draws(b_df, [draw_dict], b_hist_start)
        # b_hist_start not configured for this game — legacy bottom-append
        # (old ordering bug, see TODO #5 / b_hist_start config item).
        row = {"w": label}
        row.update({f"pos_{i+1}": n for i, n in enumerate(nums)})
        new_row_df = pd.DataFrame([row])
        return pd.concat([b_df, new_row_df], ignore_index=True)

    # ── Legacy column-oriented format ─────────────────────────────────────────
    new_df = b_df.copy()
    n_existing = len(new_df)
    for i, n in enumerate(nums):
        wcol = f"w{i+1}"
        if wcol in new_df.columns:
            new_row = pd.Series([pd.NA] * n_existing + [n], dtype="Int64")
            new_df[wcol] = new_row.values
        else:
            col_data = [pd.NA] * n_existing + [n]
            new_df[wcol] = pd.array(col_data, dtype="Int64")
    return new_df


def save_since_last(sl_dict: dict, game_key: str, label: str, pool: int,
                    url: str, sl_file) -> dict:
    """Persist a since_last dict to since_last.json (shared by fetch + upload)."""
    all_wt = sorted(sl_dict.keys(), key=lambda n: (sl_dict[n], n))
    data = {
        "since_last_dict": {str(k): v for k, v in sl_dict.items()},
        "all_wt": all_wt,
        "to_keep": list(all_wt),
        "game": game_key,
        "game_name": label,
        "pool_size": pool,
        "scraped_at": datetime.now().isoformat(),
        "url": url,
    }
    sl_file.parent.mkdir(parents=True, exist_ok=True)
    sl_file.write_text(json.dumps(data, indent=2))
    return data


def _last_draw_date(game_key: str):
    """Return the most recent completed draw date for this game (date object or None).

    Uses the draw_day field from GAMES_CFG and today's date to find the last
    weekday matching the draw schedule.  Returns None for unknown schedules.
    """
    from datetime import date, timedelta
    gcfg = GAMES_CFG.get(game_key, {})
    draw_day = gcfg.get("draw_day", "")
    WEEKDAY = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
               "Friday": 4, "Saturday": 5, "Sunday": 6}
    today = date.today()
    if draw_day == "Daily":
        return today
    # Mon/Wed/Fri draws
    if "Mon" in draw_day and "Wed" in draw_day:
        target_days = {0, 2, 4}
    else:
        wd = WEEKDAY.get(draw_day)
        if wd is None:
            return None
        target_days = {wd}
    for delta in range(7):
        check = today - timedelta(days=delta)
        if check.weekday() in target_days:
            return check
    return None


def _data_freshness_banner(scraped_at_iso: str, game_key: str) -> tuple[str, str]:
    """Return (css_class, message) describing whether the cached data is current.

    'ok'   → scraped after the last scheduled draw
    'warn' → scraped before the last scheduled draw (data is stale)
    'note' → can't determine staleness
    """
    last_draw = _last_draw_date(game_key)
    if last_draw is None or not scraped_at_iso or scraped_at_iso == "unknown":
        return "note", "Cannot determine data currency."
    try:
        from datetime import datetime as _dt
        scraped_dt = _dt.fromisoformat(scraped_at_iso).date()
    except Exception:
        return "note", f"Scraped: {scraped_at_iso[:16]}"
    scraped_str = scraped_dt.strftime("%-d %b %Y")
    draw_str    = last_draw.strftime("%-d %b %Y")
    if scraped_dt >= last_draw:
        return "ok", (f"✅ Data is current — scraped {scraped_str}, "
                      f"last draw was {draw_str}.")
    else:
        return "warn", (f"⚠️ Data may be stale — scraped {scraped_str} but "
                        f"last draw was {draw_str}. "
                        f"Click <b>Fetch now</b> to refresh.")



# ═══════════════════════════════════════════════════════════════════════════════
# 9. UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def am_toggle(section: str) -> str:
    """Global Auto/Manual toggle — ONE state shared across ALL pages.
    Clicking AUTO or MANUAL anywhere sets it site-wide and reruns immediately,
    so no page can silently break the auto chain.
    """
    # ── Read / initialise global state ────────────────────────────────────
    mode = S["auto"].get("_global", "Manual")

    c1, c2, _ = st.columns([1, 1, 8])
    with c1:
        if st.button("▶ AUTO", key=f"_a_{section}",
                     type="primary" if mode == "Auto" else "secondary",
                     use_container_width=True):
            S["auto"]["_global"] = "Auto"
            st.rerun()
    with c2:
        if st.button("✋ MANUAL", key=f"_m_{section}",
                     type="primary" if mode == "Manual" else "secondary",
                     use_container_width=True):
            S["auto"]["_global"] = "Manual"
            st.rerun()
    return mode


def edit_var(var_key: str, folder: Path, filename: str, label: str, note=""):
    """Editable variable input table — reads from and saves to real folder."""
    st.markdown(f"**{label}**")
    if note: st.caption(note)
    p = folder / filename
    if var_key not in S or (isinstance(S.get(var_key), pd.DataFrame) and S[var_key].empty and p.exists()):
        S[var_key] = _load_file(p)
    df = S.get(var_key, pd.DataFrame())
    edited = st.data_editor(df, key=f"ed_{var_key}",
                            use_container_width=True, num_rows="dynamic", height=280)
    for col in edited.columns:
        edited[col] = pd.to_numeric(edited[col], errors="coerce")
    S[var_key] = edited
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {label} CSV", to_csv_bytes(edited),
                           f"{label}.csv","text/csv",key=f"dl_{var_key}")
    with c2:
        if st.button(f"💾 Save to {filename}", key=f"sv_{var_key}"):
            edited.to_csv(p, index=False); st.success(f"Saved to {p}")


def set_container_status(db: str, val: str):
    for i in range(len(S["container_status"][db])):
        S["container_status"][db].at[i,"Status"] = val

# ═══════════════════════════════════════════════════════════════════════════════
# 10. PAGE CONFIG & CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Syndicate System", layout="wide",
                   initial_sidebar_state="collapsed")
st.markdown("""
<style>
  .block-container{padding-top:.4rem}
  .sec-hdr{font-size:.9rem;font-weight:900;color:#fff;padding:5px 14px;
            border-radius:5px;margin:4px 0 6px 0;display:inline-block}
  .hdr-blue{background:#1d3557} .hdr-green{background:#2d6a4f}
  .hdr-orange{background:#e07c00} .hdr-purple{background:#6d3b8e}
  .hdr-red{background:#c1121f} .hdr-teal{background:#0077b6}
  .hdr-brown{background:#6b4226}
  /* Status boxes — fixed foreground colour so they show on dark AND light themes */
  .note{background:#fff3cd;border-left:4px solid #e07c00;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c3d00 !important}
  .info{background:#cce5ff;border-left:4px solid #0077b6;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#003a5c !important}
  .warn{background:#ffd6d6;border-left:4px solid #c1121f;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c0000 !important}
  .ok  {background:#c3e6cb;border-left:4px solid #28a745;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#0a3d1f !important}
  /* Make all custom div text always visible */
  .note *, .info *, .warn *, .ok * { color: inherit !important; }

  /* ── Main navigation tabs (st.radio) — high visibility ───────────────── */
  /* Each tab rendered as a clear bordered pill so faint tabs are easy to read */
  div[role="radiogroup"] > label{
      background:#1b2436;
      border:1.5px solid #3a4a66;
      border-radius:8px;
      padding:7px 14px !important;
      margin:3px 6px 3px 0 !important;
      transition:all .12s ease;
  }
  div[role="radiogroup"] > label:hover{
      border-color:#5b8def;
      background:#243049;
  }
  /* Tab label text — brighter and bolder than default */
  div[role="radiogroup"] > label p{
      color:#dfe7f5 !important;
      font-weight:700 !important;
      font-size:.92rem !important;
  }
  /* The SELECTED tab — strong amber highlight so the active page is obvious */
  div[role="radiogroup"] > label:has(input:checked){
      background:#f4a000 !important;
      border-color:#ffd166 !important;
      box-shadow:0 0 0 2px rgba(244,160,0,.35);
  }
  div[role="radiogroup"] > label:has(input:checked) p{
      color:#1a1300 !important;
      font-weight:900 !important;
  }

  /* ── Game selector PRIMARY buttons — make selected game very obvious ── */
  /* Secondary (inactive) game buttons — cool dark background, clear text */
  div[data-testid="stHorizontalBlock"] button[kind="secondary"],
  div[data-testid="column"] button[kind="secondary"]{
      background:#1f2d45 !important;
      border:1.5px solid #3a4e6e !important;
      color:#b8cce4 !important;
      font-weight:600 !important;
  }
  /* Primary (active) game button — bright amber, impossible to miss */
  div[data-testid="stHorizontalBlock"] button[kind="primary"],
  div[data-testid="column"] button[kind="primary"]{
      background:linear-gradient(135deg,#f4a000,#ffcc44) !important;
      border:2px solid #ffd166 !important;
      color:#1a1300 !important;
      font-weight:900 !important;
      box-shadow:0 0 10px rgba(244,160,0,.5) !important;
  }

  /* ── Variable Input sub-tabs — make Streamlit tab bar highly visible ── */
  div[data-testid="stTabs"] button[role="tab"]{
      font-weight:700 !important;
      font-size:.88rem !important;
      color:#c5d6f0 !important;
      border-bottom:2px solid transparent !important;
      padding:8px 16px !important;
  }
  div[data-testid="stTabs"] button[role="tab"][aria-selected="true"]{
      color:#ffd166 !important;
      border-bottom:2px solid #f4a000 !important;
      font-weight:900 !important;
  }
  div[data-testid="stTabs"] button[role="tab"]:hover{
      color:#ffffff !important;
      background:#243049 !important;
  }
</style>
""", unsafe_allow_html=True)

_init_state()
S = st.session_state.S

# ═══════════════════════════════════════════════════════════════════════════════
# 11. NAVIGATION
# ═══════════════════════════════════════════════════════════════════════════════
st.title("🎰 Syndicate System")
st.caption(f"Root: `{ROOT}`")

# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL SCRAPER — collapsible, above game selector, standalone
# ═══════════════════════════════════════════════════════════════════════════════
with st.expander("🕷️ Global Scraper — sweep all states, all games", expanded=False):
        st.markdown('<span class="sec-hdr hdr-blue">🕷️ Scraper — Global D Variable Sweep (all games)</span>',
                    unsafe_allow_html=True)
        st.markdown("""
        <div class="info">
        <b>This page is standalone — not game-specific.</b>
        It sweeps thelott.com for ALL syndicate games at once and saves raw state
        files to <code>Global_Scraper/</code>. After sweeping, click
        <b>🔀 Promote All + Split by Game</b> to route each row to its correct
        game folder. The game selector above does not affect this page.
        </div>
        """, unsafe_allow_html=True)
        am_toggle("scraper")

        # ── Connectivity notice (informational only — buttons always active) ───
        online = _is_online()
        if not online:
            st.markdown(
                '<div class="warn">⚠️ <b>Streamlit cannot reach thelott API directly</b> '
                '(Mac SSL restriction in browser process). '
                'Use the <b>terminal commands</b> in Section 7 below to run sweeps — '
                'they work perfectly from terminal. Cached data is shown here.</div>',
                unsafe_allow_html=True)

        # ── Info ───────────────────────────────────────────────────────────────
        st.markdown("""
        <div class="info">
        <b>Confirmed API</b> — two-step: outlets per postcode → syndicates per outlet batch.<br>
        No auth, no cookies, no Playwright required. Saves to <code>Global_Scraper/D_{STATE}.csv</code>.<br>
        Coverage: <b>NSW · VIC · QLD · SA · TAS</b> (ACT inside NSW data · NT=0 syndicates · WA=Lotterywest, in-store only).
        </div>
        """, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — DATA STATUS DASHBOARD
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📊 Data Status")

        status_rows = _data_status()
        col_labels = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
        for label in ["State", "File", "Rows", "Size", "Age", "Last Updated", "Freshness"]:
            col_labels[["State","File","Rows","Size","Age","Last Updated","Freshness"]
                        .index(label)].markdown(f"**{label}**")

        for row in status_rows:
            cols = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
            cols[0].write(row["state"])
            cols[1].write(f"`{row['file']}`")
            cols[2].write(f"{row['rows']:,}" if row["exists"] else "—")
            cols[3].write(f"{row['size_kb']} KB" if row["exists"] else "—")
            cols[4].write(f"{row['age_h']}h" if row["exists"] else "—")
            cols[5].write(row["modified"])
            if not row["exists"]:
                cols[6].markdown('<span style="color:#c1121f">⛔ Missing</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 6:
                cols[6].markdown('<span style="color:#28a745">🟢 Fresh</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 48:
                cols[6].markdown('<span style="color:#e07c00">🟡 Stale</span>',
                                 unsafe_allow_html=True)
            else:
                cols[6].markdown('<span style="color:#c1121f">🔴 Old</span>',
                                 unsafe_allow_html=True)

        total_rows = sum(r["rows"] for r in status_rows if r["exists"])
        freshest   = min((r["age_h"] for r in status_rows if r["exists"]), default=9999)
        oldest     = max((r["age_h"] for r in status_rows if r["exists"]), default=0)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total syndicates cached", f"{total_rows:,}")
        m2.metric("States with data", sum(1 for r in status_rows if r["exists"]))
        m3.metric("Freshest file", f"{freshest}h ago" if freshest < 9999 else "—")
        m4.metric("Oldest file",   f"{oldest}h ago"   if oldest > 0     else "—")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — SWEEP CONTROLS
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 🔄 Refresh Data")

        SWEEP_STATES  = ["NSW", "VIC", "QLD", "SA", "TAS"]
        SWEEP_POSTCODES = {
            "NSW": list(range(2000, 3000)),
            "VIC": list(range(3000, 4000)),
            "QLD": list(range(4000, 5000)),
            "SA":  list(range(5000, 5600)),
            "TAS": list(range(7000, 7800)),
        }

        c1, c2, c3 = st.columns([2, 1, 2])
        with c1:
            max_pc = st.number_input(
                "Max postcodes per state (0 = all)",
                min_value=0, max_value=9999, value=0,
                help="Set to a small number (e.g. 50) for a quick test run. 0 = full sweep."
            )
        with c2:
            st.write("")
            smart_skip = st.checkbox("Smart skip", value=True,
                                     help="Skip states where data is under 6 hours old")
        with c3:
            est_pcs = sum(min(max_pc or len(v), len(v))
                          for v in SWEEP_POSTCODES.values())
            est_min = max(round(est_pcs * 0.8 / 60), 1)
            st.metric("Est. full sweep time", f"~{est_min} min")

        # ── Per-state buttons ─────────────────────────────────────────────────
        st.markdown("**Sweep individual state:**")
        btn_cols = st.columns(len(SWEEP_STATES))
        triggered_state = None
        for i, s in enumerate(SWEEP_STATES):
            with btn_cols[i]:
                row = next(r for r in status_rows if r["state"] == s)
                age_label = f"{row['age_h']}h" if row["exists"] else "no data"
                if st.button(f"{s}\n({age_label})", key=f"sw_{s}",
                             use_container_width=True):
                    triggered_state = s

        # ── Sweep all button ──────────────────────────────────────────────────
        sweep_all = st.button(
            "🚀 SWEEP ALL STATES",
            type="primary", use_container_width=True,
        )

        def _run_sweep(states_to_sweep: list):
            """Launch sweep as a subprocess — opens a visible Terminal window on Mac."""
            import subprocess as _sp, tempfile as _tf, os as _os

            script_dir = Path(__file__).resolve().parent

            # Apply smart_skip before deciding which states to launch
            states_filtered = []
            for _sw_state in states_to_sweep:
                _sw_row = next(r for r in status_rows if r["state"] == _sw_state)
                if smart_skip and _sw_row["exists"] and _sw_row["age_h"] < 6:
                    st.markdown(
                        f'<div class="info">⏭️ <b>{_sw_state}</b> skipped — data is only '
                        f'{_sw_row["age_h"]}h old (smart skip on).</div>',
                        unsafe_allow_html=True)
                    continue
                states_filtered.append(_sw_state)

            if not states_filtered:
                return

            label = ", ".join(states_filtered)
            states_repr = repr(states_filtered)

            # Build a self-deleting temp script so Terminal cleans up after itself
            py_code = "\n".join([
                "import sys, os",
                f"sys.path.insert(0, {repr(str(script_dir))})",
                "from masterapp import sweep_state_picks",
                f"for s in {states_repr}:",
                "    print(f'=== Sweeping {s} ===')",
                "    sweep_state_picks(s)",
                "print('=== All done. ===')",
            ]) + "\n"

            _tmp = _tf.NamedTemporaryFile(
                mode='w', suffix='.py', delete=False, dir='/tmp', prefix='sweep_')
            _tmp.write(py_code)
            _tmp.close()

            # ── Preference 1: visible Terminal window via osascript ───────────
            _apple = (
                f'tell application "Terminal" to do script '
                f'"python3 {_tmp.name}"'
            )
            _launched = False
            try:
                _r = _sp.run(["osascript", "-e", _apple],
                             timeout=5, capture_output=True)
                if _r.returncode == 0:
                    _launched = True
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  [{label}]  launched in Terminal")
                    st.markdown(
                        f'<div class="ok">✅ Sweep launched in Terminal for: <b>{label}</b>. '
                        f'Watch progress there — click <b>🔄 Scan Main_Data folder</b> '
                        f'or switch tabs when done to refresh the status.</div>',
                        unsafe_allow_html=True)
            except Exception:
                pass

            # ── Preference 2: silent background subprocess with spinner ───────
            if not _launched:
                st.markdown(
                    f'<div class="info">⚙️ Terminal unavailable — running sweep for '
                    f'<b>{label}</b> in background…</div>',
                    unsafe_allow_html=True)
                with st.spinner(f"Sweeping {label} — this may take several minutes…"):
                    _proc = _sp.Popen(
                        ["python3", _tmp.name],
                        stdout=_sp.PIPE, stderr=_sp.PIPE)
                    _out, _err = _proc.communicate()
                try:
                    _os.unlink(_tmp.name)
                except Exception:
                    pass
                if _proc.returncode == 0:
                    st.markdown(
                        f'<div class="ok">✅ Sweep complete for: <b>{label}</b>.</div>',
                        unsafe_allow_html=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  [{label}]  background sweep complete")
                else:
                    _err_txt = _err.decode(errors='replace')[:400]
                    st.markdown(
                        f'<div class="warn">⚠️ Sweep for <b>{label}</b> exited with errors: '
                        f'<pre>{_err_txt}</pre></div>',
                        unsafe_allow_html=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  [{label}]  error: {_err_txt[:100]}")
                st.rerun()

        if triggered_state:
            _run_sweep([triggered_state])
        elif sweep_all:
            _run_sweep(SWEEP_STATES)

        # ── Health warning ────────────────────────────────────────────────────
        zero_states = [r["state"] for r in status_rows if r["exists"] and r["rows"] < 10]
        if zero_states:
            st.markdown(
                f'<div class="warn">⚠️ <b>Health check:</b> {", ".join(zero_states)} '
                f'returned very few records. The API may be throttling — '
                f'wait and retry, or check logs below.</div>',
                unsafe_allow_html=True)

        # ── Scrape log ────────────────────────────────────────────────────────
        if S["scrape_log"]:
            with st.expander("🪵 Sweep log"):
                for line in S["scrape_log"][-60:]:
                    st.text(line)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — QUICK API TEST (single postcode)
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("🔍 Quick API test — single postcode"):
            tc1, tc2, tc3 = st.columns([1.5, 1.5, 1])
            with tc1:
                test_state = st.selectbox("State", SWEEP_STATES, key="test_state")
            with tc2:
                defaults = {"NSW": 2000, "VIC": 3000, "QLD": 4000, "SA": 5000, "TAS": 7000}
                test_pc_val = st.number_input("Postcode", value=defaults.get(test_state, 2000),
                                              key="test_pc")
            with tc3:
                st.write(""); st.write("")
                do_test = st.button("▶ Test", key="do_test")

            if do_test:
                with st.spinner(f"Fetching outlets for {test_state} / {test_pc_val}…"):
                    outlets = _fetch_outlets(test_state, int(test_pc_val))
                st.write(f"**Outlets found:** {len(outlets)}  —  IDs: `{', '.join(outlets[:8])}`{'…' if len(outlets)>8 else ''}")
                if outlets:
                    with st.spinner("Fetching syndicates…"):
                        company = _STATE_COMPANY[test_state]
                        syns = _fetch_syndicates_batch(company, outlets[:20])
                    if syns:
                        rows = [_parse_syndicate_row(s, int(test_pc_val), test_state)
                                for s in syns]
                        st.markdown(
                            f'<div class="ok">✅ {len(rows)} syndicates returned.</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(pd.DataFrame(rows), key="quick_api_rows", use_container_width=True)
                    else:
                        st.markdown(
                            '<div class="warn">0 syndicates for these outlets.</div>',
                            unsafe_allow_html=True)
                else:
                    st.markdown(
                        '<div class="warn">No outlets found for this postcode.</div>',
                        unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — PROMOTE TO DIRECT/
        # ══════════════════════════════════════════════════════════════════════


with st.expander("🗂️ Game Breakdown — promote & split by game", expanded=False):
        # SECTION 4 — PROMOTE ALL + SPLIT BY GAME
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### ✅ Promote All → Split by Game")

        # ── Re-split warning if any game folders are empty ────────────────────
        empty_games = []
        for gk in GAME_KEYS:
            gd = game_dirs(gk)["Direct"]
            if not list(gd.glob("D_*.csv")):
                empty_games.append(f"{GAMES_CFG[gk]['emoji']} {GAMES_CFG[gk]['label']}")
        if empty_games:
            st.markdown(
                f'<div class="warn">⚠️ These games have no D files yet: '
                f'<b>{", ".join(empty_games)}</b><br>'
                f'Run <b>🔀 Promote All + Split by Game</b> to populate them. '
                f'Key fix: "Monday &amp; Wednesday Lotto" is now correctly mapped to MWF.</div>',
                unsafe_allow_html=True)

        st.markdown("""
        <div class="info">
        One click: reads all D_*.csv files from <b>Main_Data/</b>, splits every row
        by its <b>Games</b> column, saves game-specific files into each game's
        <code>Games_Breakdown/</code> folder (one per game).<br>
        Multi-game syndicates (e.g. "Oz Lotto | Powerball") → a copy in <b>each</b>
        matching game folder.<br>
        Mapped: TattsLotto/Gold Lotto/X Lotto → SAT &nbsp;·&nbsp;
        Monday &amp; Wednesday Lotto → MWF &nbsp;·&nbsp;
        Super 66/Lucky Lotteries → skipped.
        </div>
        """, unsafe_allow_html=True)

        raw_d_files = sorted(DIRS["Global_Scraper"].glob("D_*.csv"))

        if raw_d_files:
            # Show quick game breakdown from first available file
            with st.expander("📊 Preview game counts before splitting"):
                for src in raw_d_files[:2]:  # show first 2 to avoid slowness
                    try:
                        df_prev = pd.read_csv(src, usecols=["Games"])
                        gc = df_prev["Games"].value_counts().reset_index()
                        gc.columns = ["Games value", "Rows"]
                        st.write(f"**{src.name}:**")
                        show_paginated_df(gc, key=f"game_breakdown_prev_{src.name}", use_container_width=True, height=200)
                    except Exception:
                        pass

            st.write(f"**Files ready to split:** {[f.name for f in raw_d_files]}")

            if st.button("🔀 PROMOTE ALL + SPLIT BY GAME",
                         type="primary", use_container_width=True,
                         key="promote_all_split"):
                all_results = {}
                progress = st.progress(0)
                for fi, src in enumerate(raw_d_files):
                    progress.progress((fi + 1) / len(raw_d_files),
                                      text=f"Splitting {src.name}…")
                    result = split_d_by_game(src, ROOT)
                    all_results[src.name] = result
                progress.empty()

                # ── Combine all states into ONE national file per game ────────
                # Each state's syndicates are distinct; concatenate them so each
                # game's CVI sees the whole country. Writes D_ALL_<game>.csv.
                games_touched = set()
                for res in all_results.values():
                    for gkey, count in res.items():
                        if gkey.startswith("_") or gkey == "error":
                            continue
                        if isinstance(count, int) and count > 0:
                            games_touched.add(gkey)
                combine_rows = []
                for gkey in sorted(games_touched):
                    info = combine_states_for_game(gkey)
                    if info.get("files"):
                        combine_rows.append({
                            "Game":   GAMES_CFG[gkey]["label"],
                            "States combined": ", ".join(info["states"]),
                            "National rows": f"{info['rows']:,}",
                            "File": f"D_ALL_{gkey}.csv",
                        })
                if combine_rows:
                    st.markdown('<div class="ok">✅ National combine complete — '
                                'each game now has a D_ALL_&lt;game&gt;.csv across all '
                                'states (this is the default CVI source).</div>',
                                unsafe_allow_html=True)
                    show_paginated_df(pd.DataFrame(combine_rows), key="combine_rows_tbl", use_container_width=True)

                # Summary table
                st.markdown("**Split results:**")
                summary_rows = []
                all_unknowns = set()
                all_skipped  = set()
                for fname, res in all_results.items():
                    unknowns = res.pop("_unknown_games", [])
                    skipped  = res.pop("_skipped_games", [])
                    all_unknowns.update(unknowns)
                    all_skipped.update(skipped)
                    err_msg = res.pop("error", None)
                    if err_msg:
                        st.error(f"Split failed for {fname}: {err_msg}")
                        continue
                    for gkey, count in res.items():
                        if not isinstance(count, int):
                            continue
                        if count > 0:
                            summary_rows.append({
                                "Source file":   fname,
                                "Game":          GAMES_CFG.get(gkey, {}).get("label", gkey),
                                "Rows written":  f"{count:,}",
                                "Saved to":      f"Games/{gkey.upper()}/Games_Breakdown/",
                            })

                if summary_rows:
                    show_paginated_df(pd.DataFrame(summary_rows), key="split_summary_tbl", use_container_width=True)
                    st.markdown('<div class="ok">✅ Split complete — '
                                'each game folder now has its own D variable files.</div>',
                                unsafe_allow_html=True)
                else:
                    st.warning("No rows were written. Check the Games column values.")

                if all_unknowns:
                    st.markdown(
                        f'<div class="warn">⚠️ Unrecognised game names (skipped): '
                        f'<b>{sorted(all_unknowns)}</b><br>'
                        f'Add these to GAME_NAME_MAP in masterapp.py if needed.</div>',
                        unsafe_allow_html=True)
                if all_skipped:
                    st.markdown(
                        f'<div class="info">ℹ️ Intentionally skipped '
                        f'(supplementary games): {sorted(all_skipped)}</div>',
                        unsafe_allow_html=True)

        else:
            st.info("No D_*.csv files in Main_Data/ yet. Run a sweep first.")

        # ── Promote files to Direct/ ──────────────────────────────────────────
        with st.expander("📤 Promote files to Direct/"):
            promote_candidates = (sorted(DIRS["Global_Scraper"].glob("D_*.csv")) +
                                  sorted(active_game_dirs()["Scraper"].glob("D_*.csv")))
            if promote_candidates:
                import shutil as _shutil_promote

                # ── Button row: ALL states at once OR one at a time ───────────
                st.markdown(
                    '<div class="info">Choose <b>Promote ALL</b> to copy every '
                    'state/territory file in one click, or use the selector below '
                    'to promote a single file.</div>',
                    unsafe_allow_html=True)

                btn_all, btn_single_col = st.columns(2)

                with btn_all:
                    if st.button("📤 Promote ALL States/Territories to Direct/",
                                 use_container_width=True, key="do_promote_all",
                                 type="primary"):
                        _direct = active_game_dirs()["Direct"]
                        _done, _errs = [], []
                        _prog = st.progress(0)
                        for _fi, _fp in enumerate(promote_candidates):
                            try:
                                _dst = _direct / _fp.name
                                _shutil_promote.copy2(_fp, _dst)
                                _done.append(_fp.name)
                            except Exception as _e:
                                _errs.append(f"{_fp.name}: {_e}")
                            _prog.progress((_fi + 1) / len(promote_candidates))
                        _prog.empty()
                        if _done:
                            st.success(
                                f"✅ Promoted {len(_done)} file(s) → "
                                f"{_direct.parent.name}/{_direct.name}/\n\n" +
                                "\n".join(f"• {n}" for n in _done))
                        if _errs:
                            st.error("Errors:\n" + "\n".join(_errs))

                st.markdown("---")

                # ── Single-file promote ───────────────────────────────────────
                st.markdown("**Promote a single file:**")
                chosen = st.selectbox("File to promote:",
                                      [f"{fp.parent.name}/{fp.name}"
                                       for fp in promote_candidates],
                                      key="promote_sel")
                chosen_fp = next(fp for fp in promote_candidates
                                 if f"{fp.parent.name}/{fp.name}" == chosen)
                c1p, c2p = st.columns([2, 1])
                with c1p:
                    df_prev = pd.read_csv(chosen_fp)
                    st.write(f"{len(df_prev):,} rows · {list(df_prev.columns)}")
                with c2p:
                    st.write("")
                    if st.button("📤 Promote to Direct/",
                                 use_container_width=True, key="do_promote"):
                        dst = active_game_dirs()["Direct"] / chosen_fp.name
                        _shutil_promote.copy2(chosen_fp, dst)
                        gs_set("D", _load_file(dst))
                        st.success(f"Promoted → {dst.parent.name}/{chosen_fp.name} "
                                   f"({len(gs('D', pd.DataFrame())):,} rows loaded)")
            else:
                st.info("No D_*.csv files found.")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 5 — DATA BROWSER
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📂 Cached Data Browser")
        for row in status_rows:
            if not row["exists"]:
                continue
            fp = DIRS["Global_Scraper"] / row["file"]
            with st.expander(
                f"📄 {row['file']}  —  {row['rows']:,} rows  ·  {row['size_kb']} KB  "
                f"·  updated {row['modified']}"
            ):
                df_view = pd.read_csv(fp)
                # Show game breakdown
                if "Games" in df_view.columns:
                    st.markdown("**Games breakdown:**")
                    gc = df_view["Games"].value_counts().reset_index()
                    gc.columns = ["Game", "Rows"]
                    show_paginated_df(gc, key=f"cached_browser_gc_{row['state']}", use_container_width=True, height=200)

                # ── Current draw vs future draws ───────────────────────────
                if "Draw_Date" in df_view.columns:
                    try:
                        _dv_dd = pd.to_datetime(df_view["Draw_Date"], errors="coerce")
                        _dv_sorted = sorted(_dv_dd.dropna().unique())
                        if _dv_sorted:
                            _dv_cur  = _dv_sorted[0]
                            _dv_fut  = [d for d in _dv_sorted if d > _dv_cur]
                            _dv_c1, _dv_c2 = st.columns(2)
                            _n_cur_dv = int((_dv_dd == _dv_cur).sum())
                            _n_fut_dv = int((_dv_dd.isin(_dv_fut)).sum())
                            _dv_c1.metric(
                                "📅 Current draw syndicates", f"{_n_cur_dv:,}",
                                help=f"Draw date: {str(_dv_cur.date())}")
                            _dv_c2.metric(
                                "🔮 Future draw syndicates", f"{_n_fut_dv:,}",
                                help=f"{len(_dv_fut)} future date(s)")
                    except Exception:
                        pass

                show_paginated_df(df_view, key=f"cached_browser_view_{row['state']}", use_container_width=True)
                st.download_button(
                    f"⬇ Download {row['file']} — all {len(df_view):,} rows",
                    to_csv_bytes(df_view),
                    row["file"], "text/csv",
                    key=f"dl_view_{row['state']}"
                )

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 6 — UPLOAD PRE-SCRAPED FILE
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("📤 Upload a pre-scraped CSV"):
            up = st.file_uploader("Upload data file (CSV, Excel, HTML, JSON, XML, …)",
                                  type=_UPLOAD_TYPES, key="scraper_upload")
            if up:
                df_up = _read_uploaded(up)
                stem = up.name.rsplit(".", 1)[0]
                sp = active_game_dirs()["Main_Data"] / f"{stem}.csv"
                df_up.to_csv(sp, index=False)
                st.markdown(f'<div class="ok">✅ {sp.name} — {len(df_up):,} rows saved '
                            f'to Main_Data/</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 7 — TERMINAL COMMANDS + SCHEDULER
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("⏰ Run sweeps from terminal + Schedule nightly (Mac)"):
            script_path = Path(__file__).resolve()
            log_path    = ROOT / "logs" / "scraper.log"

            st.markdown(
                '<div class="note">💡 <b>Tip:</b> Use the <b>SWEEP STATES</b> buttons in '
                'Section 2 above for in-app scraping. The scraper is fully inlined in '
                'masterapp.py — <b>no separate thelott_picks_scraper.py needed.</b></div>',
                unsafe_allow_html=True)

            st.markdown("**Copy a command → paste directly into your terminal:**")

            _cmds_state = {
                "NSW": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('NSW')\"",
                "VIC": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('VIC')\"",
                "QLD": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('QLD')\"",
                "SA":  f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('SA')\"",
                "TAS": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('TAS')\"",
            }
            for _state, _cmd in _cmds_state.items():
                st.markdown(f"**Sweep {_state}:**")
                st.code(_cmd, language="bash")

            st.markdown("**Sweep ALL states at once (recommended):**")
            st.code(
                f"cd {script_path.parent} && python3 -c \"\n"
                f"from masterapp import sweep_state_picks\n"
                f"for s in ['NSW','VIC','QLD','SA','TAS']:\n"
                f"    print(f'=== sweeping {{s}} ===')\n"
                f"    sweep_state_picks(s)\n"
                f"\"",
                language="bash")

            st.markdown("**After sweeping — split and combine by game:**")
            st.code(
                f"cd {script_path.parent} && python3 -c \"\n"
                f"import sys; sys.path.insert(0, '.')\n"
                f"from syndicate_core.pipeline import split_d_by_game, combine_states_for_game\n"
                f"from syndicate_core.config import ROOT, DIRS\n"
                f"from pathlib import Path\n"
                f"for f in sorted(DIRS['Global_Scraper'].glob('D_*.csv')):\n"
                f"    r = split_d_by_game(f, ROOT)\n"
                f"    print(f.name, r)\n"
                f"\"",
                language="bash")

            st.markdown("**Schedule nightly at 2 am (Mac crontab):**")
            st.code(
                f"# 1. Open crontab editor:\n"
                f"crontab -e\n\n"
                f"# 2. Add this line, then save:\n"
                f"0 2 * * * cd {script_path.parent} && python3 -c \""
                f"from masterapp import sweep_state_picks; "
                f"[sweep_state_picks(s) for s in ['NSW','VIC','QLD','SA','TAS']]"
                f"\" >> {log_path} 2>&1",
                language="bash")


st.markdown("---")

# ── GAME SELECTOR ─────────────────────────────────────────────────────────────
_game_cols = st.columns(len(GAME_KEYS) + 1)
with _game_cols[0]:
    st.markdown("**Select Game:**")
for _i, _gk in enumerate(GAME_KEYS):
    with _game_cols[_i + 1]:
        _cfg = GAMES_CFG[_gk]
        _active = (active_game() == _gk)
        _btn_type = "primary" if _active else "secondary"
        if st.button(
            f"{_cfg['emoji']} {_cfg['label']}\n{_cfg['draw_day']}",
            key=f"game_btn_{_gk}",
            type=_btn_type,
            use_container_width=True,
        ):
            st.session_state["active_game"] = _gk
            # Auto-reload B for the newly selected game so it's available
            # immediately (before the user clicks the B tab).
            _b_new = _auto_load_b(_gk)
            if not _b_new.empty:
                st.session_state[f"B__{_gk}"] = _b_new
            st.rerun()

_gcfg  = active_game_cfg()
_gdirs = active_game_dirs()
_gkey  = active_game()
st.markdown(
    f'<div class="info">🎮 Active game: <b>{_gcfg["emoji"]} {_gcfg["label"]}</b> '
    f'— Pool: 1–{_gcfg["pool"]} · Pick {_gcfg["pick"]} · Draws: {_gcfg["draw_day"]} '
    f'· Data folder: <code>Games/{active_game().upper()}/</code></div>',
    unsafe_allow_html=True,
)
st.markdown("---")

page = st.radio("Page", [
    "📥 Main Data",
    "🔄 CVI Matrix",
    "🧩 Variable Inputs",
    "📦 Container Formula",
    "🖥️ Container Dashboards",
    "📤 Master Outputs",
    "🗂️ Cluster Manager",
], horizontal=True, label_visibility="collapsed")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MAIN DATA
# ═══════════════════════════════════════════════════════════════════════════════
if page == "📥 Main Data":
    st.markdown('<span class="sec-hdr hdr-teal">📥 Main Data — Manually uploaded each run</span>',
                unsafe_allow_html=True)

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    Main Data for <b>{_gcfg['emoji']} {_gcfg['label']}</b> —
    your historical combination dataset matched against variable inputs.<br>
    Saved to disk immediately — <b>survives browser refresh</b>.<br>
    <b>Large files (63M rows):</b> save CSV to the Main_Data folder directly —
    no upload needed. The app reads from disk.<br>
    <b>accdb files:</b> place them directly in
    <code>Games/{_gkey.upper()}/Main_Data/</code> — Streamlit cannot upload
    Access databases via browser (browser security restriction).
    Install <code>mdb-tools</code> via Homebrew to read them.
    </div>
    """, unsafe_allow_html=True)

    # ── Folder path hint ───────────────────────────────────────────────────
    with st.expander("📁 Drop large files here directly (no upload needed)"):
        st.code(str(_gdirs["Main_Data"]), language=None)
        st.markdown("""
**For 63M row files:**
```bash
# Copy your main data file directly — no size limit
cp /path/to/your/maindata.csv ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/
```
Then click **🔄 Scan folder** below — the app reads it from disk.

**For accdb files:**
```bash
# 1. Install mdb-tools if not done
brew install mdb-tools

# 2. Convert accdb to CSV (no size limit)
mdb-export /path/to/file.accdb TableName > ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/maindata.csv
```
        """)

    # ── Upload (for smaller files under config limit) ──────────────────────
    up = st.file_uploader(
        "Upload Main Data file (CSV, Excel, HTML, JSON, XML, TSV, …)",
        type=_UPLOAD_TYPES,
        key="up_main_data")
    if up:
        path = _gdirs["Main_Data"] / up.name
        path.write_bytes(up.getvalue())
        st.markdown(
            f'<div class="ok">✅ Saved to disk: <code>{path.name}</code> '
            f'({path.stat().st_size / 1024 / 1024:.1f} MB) — '
            f'persists across refresh.</div>',
            unsafe_allow_html=True)

    # ── Scan folder for ALL files (disk-first approach) ────────────────────
    if st.button("🔄 Scan Main_Data folder", key="scan_main_data",
                 use_container_width=True):
        st.rerun()

    existing = sorted(
        list(_gdirs["Main_Data"].glob("*.csv")) +
        list(_gdirs["Main_Data"].glob("*.xlsx")) +
        list(_gdirs["Main_Data"].glob("*.accdb")) +
        list(_gdirs["Main_Data"].glob("*.mdb"))
    )
    # Also check legacy 1n subfolder
    existing += sorted(
        list(_gdirs["Main_Data"].glob("1n")) +
        [_gdirs["Main_Data"] / "1n" / f
         for f in ["Main_Data.xlsx", "Main_Data.csv"]
         if (_gdirs["Main_Data"] / "1n" / f).exists()]
    )
    existing = list(dict.fromkeys(existing))  # deduplicate

    if existing:
        st.markdown(f"**{len(existing)} file(s) found in Main_Data folder:**")
        for fp in existing:
            if not fp.is_file():
                continue
            size_mb = fp.stat().st_size / 1024 / 1024
            # Fast row count for CSV (no load)
            if fp.suffix == ".csv":
                rows = _count_csv_rows(fp)
                row_label = f"{rows:,} rows"
            else:
                row_label = f"{size_mb:.1f} MB"

            c1, c2, c3, c4 = st.columns([3, 1.2, 1, 0.8])
            c1.write(f"`{fp.name}`")
            c2.write(row_label)
            c3.write(f"{size_mb:.1f} MB")
            with c4:
                if st.button("Load", key=f"load_md_{fp.name}"):
                    with st.spinner(f"Loading {fp.name}…"):
                        df_ex = _load_file(fp, numeric=False)
                    if not df_ex.empty:
                        gs_set("main_data",      df_ex)
                        gs_set("main_data_path", str(fp))
                        st.success(f"Loaded {fp.name}: {len(df_ex):,} rows")
                    else:
                        st.error(f"Could not read {fp.name}")

    # ── Auto-load on game switch when exactly one file is present ─────────
    _single_files = [f for f in existing if f.is_file()]
    if S.get("main_data_auto_loaded_game") != _gkey and len(_single_files) == 1:
        _auto_fp = _single_files[0]
        st.markdown(
            f'<div class="info">ℹ️ Auto-loading <b>{_auto_fp.name}</b> for '
            f'{_gcfg["label"]}…</div>', unsafe_allow_html=True)
        with st.spinner(f"Loading {_auto_fp.name}…"):
            df_auto = _load_file(_auto_fp, numeric=False)
        if not df_auto.empty:
            gs_set("main_data",      df_auto)
            gs_set("main_data_path", str(_auto_fp))
            S["main_data_auto_loaded_game"] = _gkey

    # ── Auto-reload from disk if session state lost ────────────────────────
    if gs("main_data", pd.DataFrame()).empty:
        saved_path = gs("main_data_path", "")
        if saved_path and Path(saved_path).exists():
            with st.spinner("Reloading main data from disk…"):
                gs_set("main_data", _load_file(Path(saved_path), numeric=False))
        else:
            # Auto-load most recent file in folder
            auto_files = sorted(
                _gdirs["Main_Data"].glob("*.csv"),
                key=lambda f: f.stat().st_mtime, reverse=True)
            if auto_files:
                newest = auto_files[0]
                st.markdown(
                    f'<div class="info">ℹ️ Auto-loading most recent file: '
                    f'<b>{newest.name}</b></div>', unsafe_allow_html=True)
                with st.spinner(f"Loading {newest.name}…"):
                    gs_set("main_data", _load_file(newest, numeric=False))
                gs_set("main_data_path", str(newest))

    md = gs("main_data", pd.DataFrame())
    if not md.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Rows", f"{len(md):,}")
        c2.metric("Columns", len(md.columns))
        c3.metric("Est. match time (17 formulas)",
                  f"~{max(len(md)*7*17//16//1_000_000, 1)} sec")
        st.markdown("**Preview (first 20 rows):**")
        show_paginated_df(md, key="main_data_preview", use_container_width=True)
        st.download_button("⬇ Export Main Data CSV",
                           to_csv_bytes(md), "main_data.csv", "text/csv",
                           key="dl_main_data")
    else:
        st.info("No Main Data loaded. Upload a file, drop it in the folder, "
                "or click Load on an existing file above.")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CVI MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 CVI Matrix":
    st.markdown('<span class="sec-hdr hdr-teal">🔄 CVI Matrix — Transposition & Variable Slicing</span>',
                unsafe_allow_html=True)
    am_toggle("cvi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b><br>
    Reads D from <code>Games/{_gkey.upper()}/games_breakdown_{_gkey}/</code>
    (run <b>Promote All + Split by Game</b> on the Scraper page first).<br>
    Defaults to <b>D_ALL_{_gkey}.csv</b> — all states combined (national view).
    Each syndicate row → one w-column. Sorted longest→shortest.
    Ep = top 8 w-columns · Sp = top 4 lanes (a,b,c,d) · So = union combis.
    </div>
    """, unsafe_allow_html=True)

    # Game-specific Games_Breakdown/ folder — national D_ALL first (default)
    direct_dir = _gdirs["Games_Breakdown"]
    _all = sorted(direct_dir.glob(f"D_ALL_{_gkey}.csv"))
    _per_state = sorted(p for p in direct_dir.glob("D_*.csv")
                        if not p.name.startswith("D_ALL_"))
    raw_files  = _all + _per_state          # national file listed first

    # Fallback: also show global raw scrapes if game-specific not yet split
    global_files = sorted(DIRS["Global_Scraper"].glob("D_*.csv"))

    if not raw_files and not global_files:
        st.warning("No D files found. Run the Scraper, then Promote All + Split by Game.")
    else:
        if raw_files:
            st.markdown(f'<div class="ok">✅ {len(raw_files)} game-specific D file(s) '
                        f'found in Games/{_gkey.upper()}/Games_Breakdown/</div>',
                        unsafe_allow_html=True)
            all_files = raw_files
            file_labels = [f.name for f in all_files]
        else:
            st.markdown(
                f'<div class="warn">⚠️ No game-specific D files yet for '
                f'{_gcfg["label"]}. Showing global Main_Data files — '
                f'run Promote All + Split by Game first.</div>',
                unsafe_allow_html=True)
            all_files = global_files
            file_labels = [f"Main_Data/{f.name}" for f in all_files]

        chosen = st.selectbox("Source D file:", file_labels, key="cvi_d_sel")
        chosen_fp = all_files[file_labels.index(chosen)]
        df_raw = pd.read_csv(chosen_fp)
        _raw_wcols = [c for c in df_raw.columns if re.match(r'^w\d+$', str(c), re.I)]
        _raw_meta  = [c for c in df_raw.columns if c not in _raw_wcols]
        st.write(f"**{len(df_raw):,} syndicates · {len(_raw_wcols)} number positions**")
        # Restore original source view: all metadata columns, then 'w' label, then numbers.
        # Each ROW = one syndicate. The 'w' column labels it (w1, w2, w3…).
        # The original w1..w27 column headers are renamed to 1, 2, 3… (position numbers).
        _prev_n  = min(50, len(df_raw))
        _raw_disp = df_raw.head(_prev_n).copy()
        # Insert 'w' label column right after the last metadata column
        _insert_at = len(_raw_meta)
        _raw_disp.insert(_insert_at, "w", [f"w{i+1}" for i in range(len(_raw_disp))])
        # Rename w1..w27 → 1, 2, 3…
        _raw_disp = _raw_disp.rename(columns={wc: str(i+1) for i, wc in enumerate(_raw_wcols)})
        st.caption(f"Showing first {_prev_n} of {len(df_raw):,} syndicates — 'w' column labels each row:")
        show_paginated_df(_raw_disp, key="cvi_raw_disp", use_container_width=True)

        if st.button("🔄 BUILD W-MATRIX & SLICE Ep / Sp / So",
                     type="primary", use_container_width=True):
            # Use the already-loaded D so any Active Draw filter is honoured.
            # Fall back to the raw CSV only when nothing has been loaded into
            # session state yet (first visit before the Direct tab is used).
            _d_in_mem = gs("D", pd.DataFrame())
            if not _d_in_mem.empty:
                w_mat = _d_in_mem
                _cur_draw = st.session_state.get(gkey("active_draw"))
                if _cur_draw:
                    st.info(f"🎯 Using Active Draw {_cur_draw} — "
                            f"{len(w_mat):,} rows (filtered). "
                            f"W-Matrix reflects this draw only.")
            else:
                w_mat = df_raw
                gs_set("D", w_mat)   # bootstrap D from raw file (nothing was loaded)
                st.warning("⚠️ No D loaded in session — using raw file directly. "
                           "Active Draw filter not applied. "
                           "Load D via the Direct tab first to use a filtered view.")

            mfname = chosen_fp.name.replace("D_", "CVI_Matrix_")
            cvi_out = _gdirs["CVI"] / mfname
            w_mat.to_csv(cvi_out, index=False)

            _wcols_mat = [c for c in w_mat.columns if re.match(r'^w\d+$', str(c), re.I)]
            st.markdown(
                f'<div class="ok">✅ W-Matrix saved: {len(w_mat):,} syndicates · '
                f'{len(_wcols_mat)} number positions → {cvi_out.name}</div>',
                unsafe_allow_html=True)

            # ── Row-view preview (first 200 rows — no transpose needed) ──────
            st.markdown("---")
            st.markdown("### 📋 W-Matrix — row view (first 200 shown)")
            st.markdown(
                '<div class="info">Each row is one syndicate. '
                'Numbers run across to the right (columns = number positions).</div>',
                unsafe_allow_html=True)
            _wm_preview = w_mat[_wcols_mat].head(200).copy()
            _wm_preview.insert(0, "w", [f"w{i+1}" for i in range(len(_wm_preview))])
            _wm_preview = _wm_preview.rename(
                columns={wc: str(i+1) for i, wc in enumerate(_wcols_mat)})
            show_paginated_df(_wm_preview, key="wmat_rows_view", use_container_width=True)
            st.download_button(
                f"⬇ Download W-Matrix (row view) — {len(w_mat):,} rows",
                to_csv_bytes(w_mat),
                mfname,
                "text/csv",
                key="dl_wmat_rows"
            )

            # ── Sp and So: use only the 4 LONGEST syndicates from D ──────────
            # (not all 355k — just the 4 with the most numbers picked)
            st.markdown("---")
            st.markdown("#### Variable slices (Sp · So)")
            st.markdown(
                '<div class="info">Sp and So use only the <b>4 longest</b> syndicates '
                'from D as input sets — not the full file. Each set is split in half '
                'and set-algebra combinations are computed.</div>',
                unsafe_allow_html=True)
            try:
                _sp_input = prepare_d_input_sets(w_mat, 4)
                if _sp_input.empty:
                    st.warning("Could not extract the 4 longest rows from D. "
                               "Check that the D file has w1…wN columns.")
                else:
                    st.markdown(
                        f'<div class="ok">✅ D top-4 input sets ready: '
                        f'{", ".join(_sp_input.columns.tolist())} — '
                        f'lengths: {", ".join(str(_sp_input[c].notna().sum()) for c in _sp_input.columns)}'
                        f'</div>',
                        unsafe_allow_html=True)

                    # Sp
                    _sp_df = generate_splits(_sp_input)
                    if not _sp_df.empty:
                        gs_set("Sp", _sp_df)
                        _sp_path = _gdirs["Splits"] / f"Sp_{_gkey}.csv"
                        _sets_df_to_rows(_sp_df, set_col="set").to_csv(_sp_path, index=False)
                        st.markdown(f"**Sp — {_sp_df.shape[1]} split-combination sets**")
                        show_paginated_df(
                            _sets_df_to_rows(_sp_df, set_col="set"),
                            key="cvi_slice_sp", use_container_width=True, height=200)

                    # So
                    _so_df = generate_splits_combi(_sp_input)
                    if not _so_df.empty:
                        gs_set("So", _so_df)
                        _so_path = _gdirs["Splits_Combi"] / f"So_{_gkey}.csv"
                        _sets_df_to_rows(_so_df, set_col="set").to_csv(_so_path, index=False)
                        st.markdown(f"**So — {_so_df.shape[1]} union-combination sets**")
                        show_paginated_df(
                            _sets_df_to_rows(_so_df, set_col="set"),
                            key="cvi_slice_so", use_container_width=True, height=200)

                    # Ep — needs R's wt list; skip silently if R not yet loaded
                    _r_wt_btn = gs("_R_wt", pd.DataFrame())
                    if not _r_wt_btn.empty:
                        _ep_objs = prepare_ep_objects(w_mat, mode="pairs")
                        _wt_list_btn = (_r_wt_btn["wt"].dropna().tolist()
                                        if "wt" in _r_wt_btn.columns
                                        else _r_wt_btn.iloc[:, 0].dropna().tolist())
                        if _ep_objs and _wt_list_btn:
                            _ep_df = generate_excelpro(_ep_objs, _wt_list_btn)
                            if not _ep_df.empty:
                                gs_set("Ep", _ep_df)
                                _ep_path = _gdirs["ExcelPro"] / f"Ep_{_gkey}.csv"
                                _sets_df_to_rows(_ep_df, set_col="set").to_csv(
                                    _ep_path, index=False)
                                st.markdown(f"**Ep — {_ep_df.shape[1]} ExcelPro sets**")
                                show_paginated_df(
                                    _sets_df_to_rows(_ep_df, set_col="set"),
                                    key="cvi_slice_ep", use_container_width=True, height=200)
                    else:
                        st.info("ℹ️ Ep skipped — load R (Rainbow) first to supply the wt list, "
                                "then press this button again.")
            except Exception as _slice_ex:
                st.error(f"Slice error: {_slice_ex}")

        # Browse existing CVI matrices for this game
        cvi_files = sorted(_gdirs["CVI"].glob("CVI_Matrix_*.csv"))
        if cvi_files:
            st.markdown("---")
            cm = st.selectbox("Inspect saved matrix:", [f.name for f in cvi_files],
                              key="cvi_inspect_sel")
            df_m = pd.read_csv(_gdirs["CVI"] / cm)
            _wcols_m = [c for c in df_m.columns if re.match(r'^w\d+$', str(c), re.I)]
            _meta_m  = [c for c in df_m.columns if c not in _wcols_m]
            st.write(f"**{len(df_m):,} syndicates · {len(_wcols_m)} number positions**")
            # Each ROW = one syndicate = one w-set.
            # Add 'w' label column, rename w1..wN headers to 1, 2, 3… (position numbers).
            # No transposing needed — just relabel.
            _cap_m = min(50, len(df_m))
            _df_m_disp = df_m[_wcols_m].head(_cap_m).copy()
            _df_m_disp.insert(0, "w", [f"w{i+1}" for i in range(len(_df_m_disp))])
            _df_m_disp = _df_m_disp.rename(columns={wc: str(i+1) for i, wc in enumerate(_wcols_m)})
            st.caption(f"Showing first {_cap_m} of {len(df_m):,} syndicates — each row is one w-set:")
            show_paginated_df(_df_m_disp, key="cvi_inspect_disp", use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: VARIABLE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧩 Variable Inputs":
    st.markdown('<span class="sec-hdr hdr-green">🧩 Variable Inputs — B · R · D · Ep · Sp · So</span>',
                unsafe_allow_html=True)
    am_toggle("vi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b> — Pool 1–{_gcfg['pool']},
    Pick {_gcfg['pick']}, draws {_gcfg['draw_day']}.<br>
    <b>B</b> = pre-loaded w-columns from <b>Base_&lt;game&gt;.xlsx</b> (e.g. Base_sat.xlsx, sheet B_sat;
    auto-loaded from your project folder — rarely changes).<br>
    <b>Ep</b> = ExcelPro output: pairs the 8 longest rows of D (w1–w8) into 4 objects, then
    filters R's wt list through each object's two halves. <b>Auto-runs when D is loaded.</b><br>
    <b>Sp</b> = Splits (task1b): uses the 4 longest rows of ALL D (w1,w2,w3,w4) as input sets —
    splits each in half and computes set-algebra combinations. <b>Auto-runs when D is loaded.</b><br>
    <b>So</b> = SplitsCombi (auto_vba): same 4 longest rows of D → union combis of split halves.
    <b>Auto-runs when D is loaded.</b><br>
    <b>R</b>  = Rainbow (task2): Since Last from lottolyzer → powerset combos.<br>
    <b>D</b>  = Syndicate w-columns (standalone + feeds formula row 11).<br>
    <b>Note:</b> In ALL D, each syndicate IS a row; w1,w2,… label the number positions across.
    The generators peel the N longest rows and re-orient them as columns w1,w2,… (one set per column)
    so the set-algebra code can work on them as independent pools.
    </div>
    """, unsafe_allow_html=True)

    vtabs = st.tabs(["B (Base)", "R (Rainbow)", "D (Direct)",
                     "Ep (ExcelPro)", "Sp (Splits)", "So (SplitsCombi)",
                     "Since Last", "📊 Stats"])

    # ── TAB: B (Base) ──────────────────────────────────────────────────────
    with vtabs[0]:
        st.markdown(f"**B — Base variable for {_gcfg['label']}** "
                    f"(sheet: `{_gcfg['b_sheet']}` in {_gcfg.get('b_file','Base.xlsx')})")

        # Load B: prefer this game's own Base_<game>.xlsx; then a shared Base.xlsx;
        # then the legacy f_rules_Gclaude.xlsx. Non-breaking migration.
        b_file = _gcfg.get("b_file", "Base.xlsx")
        b_rules_candidates = (list(ROOT.rglob(b_file))
                              or list(ROOT.rglob("Base.xlsx"))
                              or list(ROOT.rglob("f_rules_Gclaude.xlsx")))
        if b_rules_candidates:
            b_rules_path = b_rules_candidates[0]
            try:
                xl_b = pd.ExcelFile(b_rules_path, engine="openpyxl")
                # Prefer this game's clean sheet (B_pb/B_sat/…); then the uppercase
                # key (PB/SAT/…); then the legacy cryptic sheet name; finally, if the
                # workbook is a single-game file, just use its only sheet.
                sheet = None
                gk = _gkey.upper()
                for cand in (_gcfg.get("b_sheet"), gk, _gcfg.get("b_sheet_legacy")):
                    if cand and cand in xl_b.sheet_names:
                        sheet = cand
                        break
                if sheet is None and len(xl_b.sheet_names) == 1:
                    sheet = xl_b.sheet_names[0]   # per-game file with one sheet
                if sheet is not None:
                    df_b_raw = xl_b.parse(sheet, header=None)

                    def _nums_from_series(s):
                        """Extract positive integers from a pandas Series."""
                        out = []
                        for v in s.dropna():
                            try:
                                fv = float(v)
                                if fv >= 1:
                                    out.append(int(fv))
                            except (ValueError, TypeError):
                                pass
                        return out

                    df_b = pd.DataFrame()

                    # ── Strategy 0: col 0 has w-prefixed row labels (row-oriented) ──
                    # B_sat layout: col A = "w1"/"w701"…; numbers go right across each row.
                    _col0_ui = [str(df_b_raw.iloc[r, 0]).strip()
                                for r in range(df_b_raw.shape[0])]
                    _w_row_pairs_ui = [(r, _col0_ui[r]) for r in range(len(_col0_ui))
                                       if _col0_ui[r].lower().startswith("w")]
                    if _w_row_pairs_ui:
                        _b_rows = []
                        for _rr, _wlabel in _w_row_pairs_ui:
                            _nums = _nums_from_series(df_b_raw.iloc[_rr, 1:])
                            if _nums:
                                _brow = {"w": _wlabel}
                                _brow.update({f"pos_{i+1}": n
                                              for i, n in enumerate(_nums)})
                                _b_rows.append(_brow)
                        if _b_rows:
                            df_b = pd.DataFrame(_b_rows)

                    # ── Strategy 1: row 0 has w-prefixed column headers ────────────
                    if df_b.empty:
                        b_data = {}
                        w_col_pairs_s1 = [
                            (c, str(df_b_raw.iloc[0, c]).strip())
                            for c in range(df_b_raw.shape[1])
                            if str(df_b_raw.iloc[0, c]).strip().lower().startswith("w")
                        ]
                        if w_col_pairs_s1:
                            for actual_col, wc in w_col_pairs_s1:
                                nums = _nums_from_series(df_b_raw.iloc[1:, actual_col])
                                if nums:
                                    b_data[wc] = pd.Series(nums)
                        if b_data:
                            df_b = pd.DataFrame(b_data)

                    # ── Strategy 2: parsed with header=0, columns named w* ────────
                    if df_b.empty:
                        b_data2 = {}
                        df_b_hdr = xl_b.parse(sheet)
                        w_hdr_cols = [c for c in df_b_hdr.columns
                                      if str(c).strip().lower().startswith("w")]
                        for wc in w_hdr_cols:
                            nums = _nums_from_series(df_b_hdr[wc])
                            if nums:
                                b_data2[str(wc).strip()] = pd.Series(nums)
                        if b_data2:
                            df_b = pd.DataFrame(b_data2)

                    if not df_b.empty:
                        gs_set("B", df_b)
                        n_sets = len(df_b) if "w" in df_b.columns else len(df_b.columns)
                        st.markdown(
                            f'<div class="ok">✅ B loaded: {n_sets} w-sets '
                            f'from sheet <b>{sheet}</b> in {b_rules_path.name}</div>',
                            unsafe_allow_html=True)

                        # B is shown as-is (rows as uploaded — no transpose, no reordering)
                        show_paginated_df(df_b, key="b_cols_view", use_container_width=True)
                        st.download_button("⬇ Download B.csv",
                                           to_csv_bytes(df_b), "B.csv",
                                           "text/csv", key="dl_b_rules")
                    else:
                        st.warning(f"Sheet '{sheet}' found but no numeric w-columns parsed.")
                else:
                    st.warning(f"No B sheet for this game. Looked for "
                               f"'{_gcfg.get('b_sheet')}' (or legacy "
                               f"'{_gcfg.get('b_sheet_legacy')}'). "
                               f"Available: {xl_b.sheet_names}")
            except Exception as ex:
                st.error(f"Error reading {b_rules_path.name}: {ex}")
        else:
            st.markdown(f'<div class="warn">{b_file} not found in project folder '
                        '(shared Base.xlsx or legacy f_rules_Gclaude.xlsx also accepted). '
                        'Upload it below — the app will load it immediately.</div>',
                        unsafe_allow_html=True)
            up_b = st.file_uploader(
                f"Upload {b_file} (or Base.xlsx / f_rules_Gclaude.xlsx)",
                type=_UPLOAD_TYPES, key="up_b_rules")
            if up_b:
                # Save under both the game-specific name AND the legacy name
                dest_game   = ROOT / b_file
                dest_legacy = ROOT / "f_rules_Gclaude.xlsx"
                _raw = up_b.read()
                dest_game.write_bytes(_raw)
                if dest_game != dest_legacy:
                    dest_legacy.write_bytes(_raw)
                st.success(f"Saved to {dest_game.name}. Loading now…")
                # Immediately parse and load B so user doesn't need to refresh
                try:
                    xl_up = pd.ExcelFile(dest_game, engine="openpyxl")
                    _sheet_up = None
                    for _cand in (_gcfg.get("b_sheet"), _gkey.upper(),
                                  _gcfg.get("b_sheet_legacy")):
                        if _cand and _cand in xl_up.sheet_names:
                            _sheet_up = _cand
                            break
                    if _sheet_up is None and len(xl_up.sheet_names) == 1:
                        _sheet_up = xl_up.sheet_names[0]
                    if _sheet_up:
                        _df_b_up = xl_up.parse(_sheet_up, header=None)
                        # FIX: capture the actual column index alongside the header
                        # name so we read data from the CORRECT column even when
                        # non-w columns precede the w-headers (e.g. a row-label col).
                        _w_col_pairs_up = [
                            (c, str(_df_b_up.iloc[0, c]))
                            for c in range(_df_b_up.shape[1])
                            if str(_df_b_up.iloc[0, c]).startswith("w")
                        ]
                        _b_data_up = {}
                        for _actual_col, _wc in _w_col_pairs_up:
                            _col_vals = _df_b_up.iloc[1:, _actual_col].dropna()
                            _nums = [int(float(v)) for v in _col_vals
                                     if str(v).replace(".", "").replace("-", "").isdigit()
                                     and float(v) >= 1]
                            if _nums:
                                _b_data_up[_wc] = pd.Series(_nums)
                        if _b_data_up:
                            gs_set("B", pd.DataFrame(_b_data_up))
                            st.markdown(
                                f'<div class="ok">✅ B loaded immediately: '
                                f'{len(_b_data_up)} w-columns from sheet '
                                f'<b>{_sheet_up}</b>.</div>',
                                unsafe_allow_html=True)
                            show_paginated_df(gs("B", pd.DataFrame()), key="b_uploaded_view", use_container_width=True)
                        else:
                            st.warning("File saved but no w-columns found. "
                                       "Check sheet layout — row 0 must have w1, w2, …")
                    else:
                        st.warning(f"File saved but sheet '{_gcfg.get('b_sheet')}' not found. "
                                   f"Available sheets: {xl_up.sheet_names}")
                except Exception as _ex_up:
                    st.error(f"Saved OK but could not parse: {_ex_up}")

    # ── TAB: R (Rainbow) ───────────────────────────────────────────────────
    with vtabs[1]:
        st.markdown("**R — Rainbow (task2.py): Since Last → powerset combos**")

        # ── load / auto-fetch Since Last (shared by both inner sub-tabs) ──
        sl_file = _gdirs["SinceLast"] / "since_last.json"
        _r_sl_url = (st.session_state.get("sl_url_override")
                     or _gcfg["lottolyzer"])
        if not sl_file.exists():
            with st.spinner(f"Fetching Since Last from lottolyzer for {_gcfg['label']}…"):
                sl_dict_auto = fetch_since_last(_r_sl_url, _gcfg["pool"])
            if sl_dict_auto:
                save_since_last(sl_dict_auto, _gkey, _gcfg["label"],
                                _gcfg["pool"], _r_sl_url, sl_file)
                st.markdown(
                    f'<div class="ok">✅ Auto-fetched Since Last — '
                    f'{len(sl_dict_auto)} numbers from lottolyzer.</div>',
                    unsafe_allow_html=True)

        # ── inner sub-tabs ─────────────────────────────────────────────────
        _r_inner = st.tabs(["▶ Generate R", "🎨 Present Order"])

        # ══ inner sub-tab 0 : Generate R ══════════════════════════════════
        with _r_inner[0]:
            if sl_file.exists():
                try:
                    sl_data = json.loads(sl_file.read_text())
                    since_last_dict = {int(k): int(v)
                                       for k, v in sl_data.get("since_last_dict", {}).items()}
                    all_wt = sl_data.get("all_wt", [])
                    to_keep_list = sl_data.get("to_keep", [])
                    scraped_at = sl_data.get("scraped_at", "unknown")
                    _r_ban_cls, _r_ban_msg = _data_freshness_banner(scraped_at, _gkey)
                    st.markdown(
                        f'<div class="{_r_ban_cls}">📅 {_r_ban_msg} '
                        f'({len(since_last_dict)} numbers)</div>',
                        unsafe_allow_html=True)

                    st.write(f"**Numbers ranked by recency** (most recent → oldest, from Since Last): "
                             f"`{all_wt[:15]}...`")

                    _n_groups = len(set(int(v) + 1 for v in since_last_dict.values()))
                    manual_max = None
                    if st.checkbox("Set max groups manually (else auto safe-max)",
                                   key="r_manual"):
                        manual_max = st.slider("Max Since Last groups to combine:",
                                               1, max(2, _n_groups), min(3, _n_groups),
                                               key="r_max_comb")
                    if st.button("▶ Generate Rainbow (R)", key="gen_R",
                                 type="primary", use_container_width=True):
                        try:
                            # Build sl_df with to_keep set per-number (not by position).
                            # Bug fix: pd.Series(all_wt) assigns by positional index,
                            # corrupting the wt filter when dict order ≠ all_wt order.
                            _all_wt_set = set(all_wt)
                            sl_df = pd.DataFrame({
                                "numbers":    list(since_last_dict.keys()),
                                "Since Last": list(since_last_dict.values()),
                                "to_keep":    [n if n in _all_wt_set else pd.NA
                                               for n in since_last_dict.keys()],
                            })
                            r_df, r_wt, r_info = generate_rainbow(sl_df, max_comb=manual_max)
                            gs_set("R", r_df)
                            gs_set("_R_wt", r_wt)
                            r_path = _gdirs["Rainbow"] / f"R_{_gkey}.csv"
                            r_df.to_csv(r_path, index=False)
                            _cap = " (auto-capped to stay safe)" if r_info["capped"] else ""
                            st.markdown(
                                f'<div class="ok">✅ R generated: {r_info["n_combos"]} combos '
                                f'from {r_info["n_groups"]} groups · max_comb={r_info["max_comb"]}'
                                f'{_cap} → {r_path.name}</div>', unsafe_allow_html=True)
                            show_paginated_df(r_df, key="r_generated_view",
                                             use_container_width=True)
                        except Exception as ex:
                            st.error(f"Rainbow error: {ex}")

                except Exception as ex:
                    st.error(f"Error loading Since Last: {ex}")
            else:
                st.markdown(
                    '<div class="warn">⚠️ Couldn\'t auto-fetch Since Last from lottolyzer '
                    '(page may be unreachable or its layout changed). Use the '
                    '<b>Since Last</b> tab to fetch again or upload it manually.</div>',
                    unsafe_allow_html=True)

            if not gs("R", pd.DataFrame()).empty:
                st.markdown("**Current R in memory:**")
                show_paginated_df(gs("R", pd.DataFrame()), key="r_current_memory",
                                  use_container_width=True, height=200)

                with st.expander("R — transposed preview (as used in collation)", expanded=False):
                    _r_t = _to_w_rows(gs("R", pd.DataFrame()), force_column_oriented=True)
                    if not _r_t.empty:
                        _r_t_val = [c for c in _r_t.columns if c != "Set_Label"]
                        _r_t = _r_t.rename(
                            columns={c: f"w{i+1}" for i, c in enumerate(_r_t_val)})
                        st.caption(
                            f"{len(_r_t)} rows · each row = one combo tuple; "
                            f"Set_Label = original R column name; "
                            f"w1…w{len(_r_t_val)} = the numbers in that combo")
                        show_paginated_df(_r_t, key="r_transposed_preview",
                                          use_container_width=True, height=300)
                    else:
                        st.info("Transposed view is empty — R may not be column-oriented.")

        # ══ inner sub-tab 1 : Present Order ═══════════════════════════════
        with _r_inner[1]:
            st.markdown("**Present Order — all numbers ranked by Since Last (most recent → oldest)**")
            st.markdown(
                '<div class="info">'
                'Numbers are arranged in their <b>present order</b>: '
                'sorted by how many draws ago they last appeared. '
                'Each number is colour-coded by its <b>value range</b>.<br>'
                '<b>Legend:</b> '
                '<span style="background:#FFFF00;color:#000;padding:1px 6px;border-radius:3px">■ 1–9</span> '
                '<span style="background:#00B0F0;color:#000;padding:1px 6px;border-radius:3px">■ 10–19</span> '
                '<span style="background:#A0A0A0;color:#000;padding:1px 6px;border-radius:3px">■ 20–29</span> '
                '<span style="background:#92D050;color:#000;padding:1px 6px;border-radius:3px">■ 30–40</span> '
                '<span style="background:#FF69B4;color:#fff;padding:1px 6px;border-radius:3px">■ 41–49</span>'
                '</div>',
                unsafe_allow_html=True)

            if not sl_file.exists():
                st.markdown(
                    '<div class="warn">⚠️ No Since Last data yet. '
                    'Go to the <b>Since Last</b> tab to fetch or upload it first.</div>',
                    unsafe_allow_html=True)
            else:
                try:
                    _po_data = json.loads(sl_file.read_text())
                    _po_sl   = {int(k): int(v)
                                for k, v in _po_data.get("since_last_dict", {}).items()}
                    _po_wt   = _po_data.get("all_wt", [])
                    _po_at   = _po_data.get("scraped_at", "")[:16]

                    def _po_color(sl_val):
                        if sl_val == 0:   return ("#FFD700", "#000")
                        if sl_val <= 3:   return ("#FFFF55", "#000")
                        if sl_val <= 9:   return ("#92D050", "#000")
                        if sl_val <= 19:  return ("#00B0F0", "#000")
                        if sl_val <= 29:  return ("#CC44FF", "#fff")
                        return ("#FF69B4", "#fff")

                    def _po_label(sl_val):
                        if sl_val == 0:   return "Last Draw"
                        if sl_val <= 3:   return "Very Recent"
                        if sl_val <= 9:   return "Recent"
                        if sl_val <= 19:  return "Moderate"
                        if sl_val <= 29:  return "Old"
                        return "Very Old"

                    # ── View selector ──────────────────────────────────────
                    _po_view = st.radio(
                        "Display style:",
                        ["Colour Grid (grouped by SL value)",
                         "Flat ranked table",
                         "Rainbow combo overlay"],
                        horizontal=True, key="po_view_sel")

                    st.caption(f"Since Last data scraped: {_po_at} · "
                               f"{len(_po_sl)} numbers · game: {_gcfg['label']}")

                    # ── VIEW A : Colour grid ───────────────────────────────
                    if _po_view == "Colour Grid (grouped by SL value)":
                        # Group numbers by their exact SL value, show as rows
                        from collections import defaultdict as _dd
                        _po_groups = _dd(list)
                        for _pn in _po_wt:
                            _psl = _po_sl.get(_pn, 0)
                            _po_groups[_psl].append(_pn)

                        _html_rows = []
                        _html_rows.append(
                            "<table style='border-collapse:collapse;font-size:.85rem;"
                            "width:100%;margin-top:8px'>"
                            "<thead><tr>"
                            "<th style='padding:4px 8px;text-align:center;background:#333;color:#fff;"
                            "width:90px'>Since<br>Last</th>"
                            "<th style='padding:4px 8px;text-align:center;background:#333;color:#fff;"
                            "width:100px'>Group</th>"
                            "<th style='padding:4px 8px;background:#333;color:#fff'>Numbers "
                            "(present order — most recent first)</th>"
                            "</tr></thead><tbody>")

                        for _psl_val in sorted(_po_groups.keys()):
                            _nums_in_grp = _po_groups[_psl_val]
                            _bg, _fg = _po_color(_psl_val)
                            _grp_lbl = _po_label(_psl_val)
                            _cells = []
                            for _pn in _nums_in_grp:
                                _nb, _nf = _num_colour(_pn)
                                _cells.append(
                                    f"<span style='display:inline-block;background:{_nb};"
                                    f"color:{_nf};border-radius:4px;padding:2px 7px;"
                                    f"margin:2px 3px;font-weight:600;min-width:28px;"
                                    f"text-align:center'>{_pn}</span>")
                            _cells = "".join(_cells)
                            _html_rows.append(
                                f"<tr>"
                                f"<td style='padding:4px 8px;text-align:center;"
                                f"background:{_bg};color:{_fg};font-weight:700;"
                                f"border:1px solid #555'>{_psl_val}</td>"
                                f"<td style='padding:4px 8px;text-align:center;"
                                f"background:{_bg};color:{_fg};"
                                f"border:1px solid #555'>{_grp_lbl}</td>"
                                f"<td style='padding:4px 8px;border:1px solid #555'>"
                                f"{_cells}</td>"
                                f"</tr>")
                        _html_rows.append("</tbody></table>")
                        st.markdown("".join(_html_rows), unsafe_allow_html=True)

                    # ── VIEW B : Flat ranked table ─────────────────────────
                    elif _po_view == "Flat ranked table":
                        _po_rows = []
                        for _rank, _pn in enumerate(_po_wt, 1):
                            _psl = _po_sl.get(_pn, 0)
                            _po_rows.append({
                                "Rank":       _rank,
                                "Number":     _pn,
                                "Since Last": _psl,
                                "Group":      _po_label(_psl),
                            })
                        _po_flat = pd.DataFrame(_po_rows)
                        # Colour the "Group" column using background gradient hack
                        # (plain dataframe; actual coloring via HTML caption below)
                        show_paginated_df(_po_flat, key="po_flat_tbl",
                                          use_container_width=True)

                        # Coloured summary strip
                        _strip_parts = []
                        for _rank, _pn in enumerate(_po_wt, 1):
                            _bg, _fg = _num_colour(_pn)
                            _strip_parts.append(
                                f"<span style='display:inline-block;background:{_bg};"
                                f"color:{_fg};border-radius:4px;padding:2px 7px;"
                                f"margin:2px 2px;font-weight:600;min-width:28px;"
                                f"text-align:center'>{_pn}</span>")
                        st.markdown(
                            "<div style='margin-top:10px'><b>Full present order (colour-coded):</b><br>"
                            + "".join(_strip_parts) + "</div>",
                            unsafe_allow_html=True)

                    # ── VIEW C : Rainbow combo overlay ─────────────────────
                    else:
                        _r_mem = gs("R", pd.DataFrame())
                        if _r_mem is None or _r_mem.empty:
                            st.markdown(
                                '<div class="warn">⚠️ R not generated yet. '
                                'Go to <b>▶ Generate R</b> first, then come back here.</div>',
                                unsafe_allow_html=True)
                        else:
                            st.markdown(
                                '<div class="info">Each column below is one Rainbow combination. '
                                'Numbers are shown in <b>present order</b> (most recent → oldest) '
                                'and colour-coded by number value range. '
                                'Red outline = number appears in this combo.</div>',
                                unsafe_allow_html=True)

                            # R is column-oriented: each column is one combo, rows are numbers in it
                            _r_all_combos = list(_r_mem.columns)
                            _max_combos_disp = st.slider(
                                "Number of Rainbow combos to display:",
                                1, min(len(_r_all_combos), 50), min(len(_r_all_combos), 20),
                                key="po_n_combos")
                            _display_combos = _r_all_combos[:_max_combos_disp]

                            # Build HTML table: rows = present-order rank, cols = combo
                            _combo_sets = []
                            for _col in _display_combos:
                                _cset = set()
                                for _v in _r_mem[_col].dropna():
                                    try:
                                        _v = int(_v)
                                        if _v >= 1:
                                            _cset.add(_v)
                                    except Exception:
                                        pass
                                _combo_sets.append(_cset)

                            _hdr = ("<table style='border-collapse:collapse;"
                                    "font-size:.78rem;width:100%'><thead><tr>"
                                    "<th style='background:#333;color:#fff;padding:3px 6px;"
                                    "white-space:nowrap'>Rank</th>"
                                    "<th style='background:#333;color:#fff;padding:3px 6px;"
                                    "white-space:nowrap'>Num / SL</th>")
                            for _ci in range(len(_combo_sets)):
                                _hdr += (f"<th style='background:#444;color:#fff;"
                                         f"padding:3px 4px;text-align:center'>"
                                         f"C{_ci+1}</th>")
                            _hdr += "</tr></thead><tbody>"

                            _body_parts = [_hdr]
                            for _rank, _pn in enumerate(_po_wt, 1):
                                _psl = _po_sl.get(_pn, 0)
                                _bg, _fg = _num_colour(_pn)
                                _row_html = (
                                    f"<tr>"
                                    f"<td style='padding:2px 5px;text-align:center;"
                                    f"color:#aaa'>{_rank}</td>"
                                    f"<td style='padding:2px 5px;background:{_bg};"
                                    f"color:{_fg};font-weight:600;text-align:center;"
                                    f"border-radius:3px'>{_pn}<br>"
                                    f"<span style='font-size:.65rem;font-weight:400'>"
                                    f"SL={_psl}</span></td>")
                                for _cset in _combo_sets:
                                    if _pn in _cset:
                                        _row_html += (
                                            f"<td style='text-align:center;background:{_bg};"
                                            f"color:{_fg};font-weight:700;"
                                            f"border:2px solid #f00;padding:1px 3px'>"
                                            f"{_pn}</td>")
                                    else:
                                        _row_html += (
                                            "<td style='text-align:center;"
                                            "color:#555;padding:1px 3px'>·</td>")
                                _row_html += "</tr>"
                                _body_parts.append(_row_html)
                            _body_parts.append("</tbody></table>")

                            st.markdown("".join(_body_parts), unsafe_allow_html=True)

                except Exception as _po_ex:
                    st.error(f"Present Order error: {_po_ex}")

    # ── TAB: D (Direct) ────────────────────────────────────────────────────
    with vtabs[2]:
        st.markdown(f"**D — Syndicate data for {_gcfg['label']}**")

        # ── Build GAME-SPECIFIC file lists (Q1 fix) ────────────────────────
        # Per-state files for THIS game only  e.g. D_NSW_sat.csv, D_VIC_sat.csv…
        _d_state_files = sorted(
            p for p in _gdirs["Games_Breakdown"].glob("D_*.csv")
            if not p.name.startswith("D_ALL_"))
        # Pre-combined national file for this game (D_ALL_sat.csv)
        _d_all_combined = sorted(_gdirs["Games_Breakdown"].glob(f"D_ALL_{_gkey}.csv"))
        # Additional files already in Direct/
        _d_direct_files = (sorted(_gdirs["Direct"].glob("D*.csv"))
                           + sorted(_gdirs["Direct"].glob("D*.xlsx")))

        # Extract state abbreviations from per-state file names for display
        _state_abbrs = []
        for _sf in _d_state_files:
            _parts = _sf.stem.upper().split("_")   # D_NSW_SAT → ['D','NSW','SAT']
            if len(_parts) >= 2:
                _state_abbrs.append(_parts[1])      # 'NSW', 'VIC' …

        # Selectbox list: pre-combined first (already merged), then per-state, then Direct
        _d_select_files = _d_all_combined + _d_state_files + _d_direct_files
        _seen_fp, _d_select_dedup = set(), []
        for _f in _d_select_files:
            if _f not in _seen_fp:
                _seen_fp.add(_f); _d_select_dedup.append(_f)
        _d_select_files = _d_select_dedup

        # Session state keys for unfiltered D and active draw filter
        _d_full_key = gkey("D_full")
        _d_draw_key = gkey("active_draw")

        st.markdown(
            f'<div class="info">'
            f'<b>D scope — {_gcfg["label"]} only.</b> '
            f'Per-state files: <code>'
            + (", ".join(f"D_{s}_{_gkey}.csv" for s in _state_abbrs) if _state_abbrs
               else f"none found yet in Games_Breakdown/")
            + f'</code><br>'
            f'<b>Load ALL</b> combines only the per-state files for {_gcfg["label"]} '
            f'({len(_d_state_files)} file{"s" if len(_d_state_files)!=1 else ""}). '
            f'Then use <b>Set Active Draw ▼</b> to lock one draw into memory before '
            f'CVI Matrix and all downstream steps.</div>',
            unsafe_allow_html=True)

        if _d_select_files:
            _d_labels = [f.name for f in _d_select_files]
            ch_d = st.selectbox("Select file to load (or use Load ALL below):",
                                _d_labels, key="sel_d_file")

            c_load1, c_loadall = st.columns(2)

            # ── Load single file ────────────────────────────────────────────
            with c_load1:
                if st.button("▶ Load selected file", key="btn_d_load_one",
                             use_container_width=True):
                    _fp = _d_select_files[_d_labels.index(ch_d)]
                    with st.spinner(f"Loading {_fp.name}…"):
                        df_d_loaded = _load_file(_fp)
                        if not df_d_loaded.empty:
                            df_d_loaded = sort_d_longest_first(df_d_loaded)
                    gs_set("D", df_d_loaded)
                    st.session_state[_d_full_key] = df_d_loaded.copy()
                    st.session_state.pop(_d_draw_key, None)  # clear any prior draw filter
                    st.success(f"✅ Loaded {len(df_d_loaded):,} rows from {_fp.name} "
                               f"(sorted longest → shortest)")
                    _auto_wire_generators(_gdirs, _gkey)

            # ── Load ALL per-state files for this game ──────────────────────
            with c_loadall:
                _load_all_lbl = (
                    f"📦 Load ALL {_gcfg['label']} state files "
                    f"({len(_d_state_files)} file{'s' if len(_d_state_files)!=1 else ''}"
                    + (f": {' + '.join(_state_abbrs)}" if _state_abbrs else "")
                    + ")"
                )
                if _d_state_files:
                    if st.button(_load_all_lbl, key="btn_d_load_all",
                                 type="primary", use_container_width=True):
                        _dfs_states = []
                        _prog_d = st.progress(0,
                            text=f"Loading {_gcfg['label']} state files…")
                        for _fi, _fp in enumerate(_d_state_files):
                            _prog_d.progress((_fi + 1) / len(_d_state_files),
                                             text=f"Loading {_fp.name}…")
                            try:
                                _tmp = _load_file(_fp)
                                if not _tmp.empty:
                                    _dfs_states.append(_tmp)
                            except Exception as _ex:
                                st.warning(f"Skipped {_fp.name}: {_ex}")
                        _prog_d.empty()
                        if _dfs_states:
                            _d_combined = pd.concat(_dfs_states, ignore_index=True)
                            _d_combined = sort_d_longest_first(_d_combined)
                            # De-duplicate by Syndicate_ID + Draw_Number
                            _dedup_cols = [c for c in
                                           ["Syndicate_ID", "Draw_Number", "Game"]
                                           if c in _d_combined.columns]
                            if _dedup_cols:
                                _d_combined = _d_combined.drop_duplicates(
                                    subset=_dedup_cols, keep="first")
                            gs_set("D", _d_combined)
                            st.session_state[_d_full_key] = _d_combined.copy()
                            st.session_state.pop(_d_draw_key, None)  # clear prior draw filter
                            st.success(
                                f"✅ Combined {len(_dfs_states)} state file(s) for "
                                f"{_gcfg['label']} "
                                f"({' + '.join(_state_abbrs) if _state_abbrs else ''}) → "
                                f"**{len(_d_combined):,} rows** · "
                                f"{len(_d_combined.columns)} cols "
                                f"(duplicates removed, sorted longest → shortest)")
                            _auto_wire_generators(_gdirs, _gkey)
                        else:
                            st.error("No state D files could be loaded.")
                else:
                    st.warning(
                        f"No per-state D files found for {_gcfg['label']} in "
                        f"Games_Breakdown/. Run Promote All + Split by Game first, "
                        f"then return here.")

        # ── Active Draw Selector (Q2 fix) ───────────────────────────────────
        _df_full = st.session_state.get(_d_full_key, gs("D", pd.DataFrame()))
        if not _df_full.empty:
            st.markdown("---")
            st.markdown("#### 🎯 Active Draw — Lock one draw into memory")
            st.markdown(
                '<div class="info">'
                'Selecting a draw and clicking <b>Set Active Draw</b> filters '
                '<code>S["D"]</code> to <b>only that draw\'s syndicates</b>. '
                'CVI Matrix, Container Formula, and all downstream steps will then '
                'work exclusively on that draw\'s data. '
                'Click <b>Restore all draws</b> to go back to the full dataset.'
                '</div>',
                unsafe_allow_html=True)

            _dn_col = "Draw_Number" if "Draw_Number" in _df_full.columns else None
            _dd_col = "Draw_Date"   if "Draw_Date"   in _df_full.columns else None
            _cur_active_draw = st.session_state.get(_d_draw_key)

            # Build draw options from Draw_Number column
            if _dn_col:
                try:
                    _draw_num_vals = (
                        _df_full[_dn_col]
                        .dropna()
                        .astype(str)
                        .str.extract(r'(\d+)')[0]
                        .dropna()
                        .astype(int)
                        .sort_values()
                        .unique()
                        .tolist()
                    )
                except Exception:
                    _draw_num_vals = []
            else:
                _draw_num_vals = []

            _draw_opts = ["All draws"] + [str(d) for d in _draw_num_vals]
            _def_draw_idx = 0
            if _cur_active_draw and str(_cur_active_draw) in _draw_opts:
                _def_draw_idx = _draw_opts.index(str(_cur_active_draw))

            _adc1, _adc2, _adc3 = st.columns([2, 1, 2])
            with _adc1:
                _draw_sel = st.selectbox(
                    f"Draw # for {_gcfg['label']}:",
                    _draw_opts,
                    index=_def_draw_idx,
                    key=f"d_draw_sel_{_gkey}")
            with _adc2:
                st.write(""); st.write("")
                if st.button("🎯 Set Active Draw",
                             key=f"btn_set_draw_{_gkey}",
                             type="primary",
                             use_container_width=True):
                    if _draw_sel == "All draws":
                        gs_set("D", st.session_state[_d_full_key].copy())
                        st.session_state.pop(_d_draw_key, None)
                        st.success(
                            f"✅ Restored all draws — "
                            f"{len(gs('D', pd.DataFrame())):,} rows in memory.")
                    else:
                        st.session_state[_d_draw_key] = int(_draw_sel)
                        if _dn_col:
                            try:
                                _dn_s = (_df_full[_dn_col]
                                         .astype(str)
                                         .str.extract(r'(\d+)')[0]
                                         .fillna("-1")
                                         .astype(int))
                                _filt = _df_full[_dn_s == int(_draw_sel)].copy()
                            except Exception:
                                _filt = _df_full.copy()
                        elif _dd_col:
                            # fallback: filter by Draw_Date
                            try:
                                _dds = pd.to_datetime(_df_full[_dd_col], errors="coerce")
                                _uniq_dt = sorted(_dds.dropna().unique())
                                _di = min(int(_draw_sel), len(_uniq_dt)) - 1
                                _tgt = _uniq_dt[max(0, _di)]
                                _filt = _df_full[
                                    _dds.dt.normalize() ==
                                    pd.Timestamp(_tgt).normalize()
                                ].copy()
                            except Exception:
                                _filt = _df_full.copy()
                        else:
                            _filt = _df_full.copy()
                        gs_set("D", _filt)
                        st.success(
                            f"✅ Active draw set to **Draw {_draw_sel}** → "
                            f"**{len(_filt):,} rows** now in memory. "
                            f"CVI Matrix and all downstream steps use this draw only.")
                    # Invalidate stale Sp/So/Ep results and persisted split-point
                    # widget keys so they recompute fresh from the new D, not the
                    # previously-loaded full D.
                    for _k in [k for k in st.session_state
                               if k.startswith("sp_split_") or k.startswith("so_split_")]:
                        st.session_state.pop(_k, None)
                    for _var in ("Sp", "So", "Ep"):
                        gs_set(_var, pd.DataFrame())
                    _auto_wire_generators(_gdirs, _gkey)
                    st.rerun()

            # ── Active draw status banner ───────────────────────────────────
            with _adc3:
                if _cur_active_draw:
                    _filt_n = len(gs("D", pd.DataFrame()))
                    st.success(
                        f"🎯 **Active: Draw {_cur_active_draw}** "
                        f"({_filt_n:,} rows in memory)")
                else:
                    _full_n = len(_df_full)
                    st.info(f"📋 No draw filter — all {_full_n:,} rows loaded")

            # ── Per-draw breakdown metrics ──────────────────────────────────
            if _dn_col and _draw_num_vals:
                try:
                    _dn_s2 = (_df_full[_dn_col]
                              .astype(str)
                              .str.extract(r'(\d+)')[0]
                              .dropna()
                              .astype(int))
                    _draw_counts = _dn_s2.value_counts().sort_index()
                    _mc = st.columns(min(len(_draw_counts), 6))
                    for _ci, (_dv, _dc) in enumerate(_draw_counts.items()):
                        _flag = "🎯 " if str(_dv) == str(_cur_active_draw) else ""
                        _mc[_ci % len(_mc)].metric(
                            f"{_flag}Draw {_dv}", f"{_dc:,}",
                            help=f"{_dc:,} syndicates for draw {_dv}")
                except Exception:
                    pass
            elif _dd_col:
                try:
                    _dda = pd.to_datetime(_df_full[_dd_col], errors="coerce")
                    _udates = sorted(_dda.dropna().unique())
                    if _udates:
                        _mc = st.columns(min(len(_udates), 6))
                        for _ci, _ud in enumerate(_udates):
                            _cnt = int((_dda.dt.normalize() ==
                                        pd.Timestamp(_ud).normalize()).sum())
                            _mc[_ci % len(_mc)].metric(
                                str(pd.Timestamp(_ud).date()), f"{_cnt:,}")
                except Exception:
                    pass

        # ── Display current D (filtered or full) ────────────────────────────
        df_d = gs("D", pd.DataFrame())
        if not df_d.empty:
            _d_wcols_tab = [c for c in df_d.columns
                            if re.match(r'^w\d+$', str(c), re.I)]
            _d_col_orient = (len(_d_wcols_tab) > 0 and
                             len(df_d) < len(_d_wcols_tab) and
                             len(_d_wcols_tab) > 20)

            if _d_col_orient:
                _d_norm = df_d[_d_wcols_tab].T.reset_index(drop=True)
            else:
                _d_norm = (df_d[_d_wcols_tab].reset_index(drop=True)
                           if _d_wcols_tab else df_d.reset_index(drop=True))

            _d_norm = _d_norm.apply(pd.to_numeric, errors="coerce")
            _d_sort_key = _d_norm.notna().sum(axis=1)
            _d_norm = _d_norm.loc[
                _d_sort_key.sort_values(ascending=False, kind="stable").index
            ].reset_index(drop=True)
            _d_norm.columns = list(range(1, len(_d_norm.columns) + 1))
            _d_norm.insert(0, "w", [f"w{i+1}" for i in range(len(_d_norm))])

            n_syn = len(_d_norm)
            n_pos = len(_d_norm.columns) - 1
            _draw_lbl = (f" · Draw {st.session_state.get(_d_draw_key)}"
                         if st.session_state.get(_d_draw_key) else " · all draws")
            st.write(f"**{n_syn:,} syndicates · up to {n_pos} number "
                     f"positions{_draw_lbl}**")
            show_paginated_df(_d_norm, key="d_tab_main_view",
                              use_container_width=True)
            st.download_button(
                f"⬇ Download D (current view) as CSV — {n_syn:,} rows",
                to_csv_bytes(_d_norm),
                f"D_{_gkey}_export.csv",
                "text/csv",
                key="dl_d_tab"
            )
        elif _d_select_files:
            st.info("Load a file above to preview syndicates.")
        else:
            st.info(f"No D files in Games/{_gkey.upper()}/. Run the Scraper page "
                    f"→ Promote All + Split by Game.")

    # ── TAB: Ep (ExcelPro) ─────────────────────────────────────────────────
    with vtabs[3]:
        st.markdown("**Ep — ExcelPro** · objects a,b,c,d from D's 8 longest rows "
                    "→ 6 pair blocks filtered through R's wt list")
        st.markdown(
            "Objects: **a** = D rows 1+2 (w1,w2) · **b** = rows 3+4 (w3,w4) · "
            "**c** = rows 5+6 (w5,w6) · **d** = rows 7+8 (w7,w8). "
            "wt list from Since Last (all_wt) when R is loaded; falls back to unique numbers in D's top-8 rows if no Since Last data. "
            "Auto-runs when D loads — use button to re-run manually.")

        d_df = gs("D", pd.DataFrame())
        r_df = gs("R", pd.DataFrame())

        if d_df.empty:
            st.warning("Load D first (Direct tab).")
        else:
            wt_list_ep: list = []
            _wt_source = "from D fallback"
            if not r_df.empty:
                _sl_ep_file = _gdirs["SinceLast"] / "since_last.json"
                if _sl_ep_file.exists():
                    try:
                        _sl_ep_data = json.loads(_sl_ep_file.read_text())
                        wt_list_ep = [int(n) for n in _sl_ep_data.get("all_wt", [])]
                        if wt_list_ep:
                            _wt_source = "from Since Last (all_wt)"
                    except Exception:
                        pass
            if not wt_list_ep:
                _d_top8 = prepare_d_input_sets(d_df, 8)
                wt_list_ep = sorted({int(v) for col in _d_top8.columns
                                     for v in _d_top8[col].dropna()})
                _wt_source = "from D fallback"

            st.write(f"wt list: **{len(wt_list_ep)} numbers** ({_wt_source})")

            if st.button("▶ Run Ep", type="primary",
                         key="run_ep_btn", use_container_width=True):
                try:
                    _ep_objs = prepare_ep_objects(d_df, mode="pairs")
                    _ep_df   = generate_excelpro(_ep_objs, wt_list_ep)
                    gs_set("Ep", _ep_df)
                    _ep_path = _gdirs["ExcelPro"] / f"Ep_{_gkey}.csv"
                    _sets_df_to_rows(_ep_df, set_col="set").to_csv(_ep_path, index=False)
                    st.markdown(
                        f'<div class="ok">✅ Ep: {_ep_df.shape[1]} cols → '
                        f'{_ep_path.name}</div>',
                        unsafe_allow_html=True)
                except Exception as _ep_ex:
                    st.error(f"Ep error: {_ep_ex}")

        ep_df_view = gs("Ep", pd.DataFrame())
        if not ep_df_view.empty:
            st.markdown("**Ep output — row-oriented (one row per set):**")
            show_paginated_df(_sets_df_to_rows(ep_df_view, set_col="set"), key="ep_rows_view", use_container_width=True)

            try:
                import io as _ep_io
                _ep_buf = _ep_io.BytesIO()
                with pd.ExcelWriter(_ep_buf, engine="openpyxl") as _ep_xl:
                    _sets_df_to_rows(ep_df_view, set_col="set").to_excel(
                        _ep_xl, sheet_name="All", index=False)
                    from itertools import combinations as _ep_combos2
                    for _p in ["".join(p) for p in _ep_combos2(["a","b","c","d"], 2)]:
                        _pcols = [c for c in ep_df_view.columns if c.endswith("_" + _p)]
                        if _pcols:
                            _pb = ep_df_view[_pcols].copy()
                            _pb.columns = [c.split("_")[0] for c in _pcols]
                            _sets_df_to_rows(_pb, set_col="obj").to_excel(
                                _ep_xl, sheet_name=_p, index=False)
                _ep_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download Ep_{_gkey}.xlsx  (All + ab ac ad bc bd cd)",
                    data=_ep_buf.getvalue(),
                    file_name=f"Ep_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _ep_dl_ex:
                st.warning(f"Excel export unavailable: {_ep_dl_ex}")

    # ── TAB: Sp (Splits) ───────────────────────────────────────────────────
    with vtabs[4]:
        st.markdown("**Sp — Splits (task1b.py): top 4 w-columns of D + 4 split points**")

        d_df = gs("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)

            top4 = w_cols_d[:4]
            st.write(f"Top 4 w-columns → Sp input: `{top4}`")

            sets_ready = {}
            for col in top4:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready[col] = vals

            if sets_ready:
                st.markdown("**Set split points** (0 = midpoint auto):")
                sp_cols = st.columns(4)
                splitters = []
                for i, (col, vals) in enumerate(sets_ready.items()):
                    with sp_cols[i]:
                        default_split = len(vals) // 2
                        sp = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_split,
                            key=f"sp_split_{col}_{i}")
                        splitters.append(sp if sp > 0 else default_split)

                if st.button("▶ Run Splits (Sp)", type="primary",
                             key="run_sp", use_container_width=True):
                    try:
                        from itertools import combinations as _sp_combos

                        keys = list(sets_ready.keys())
                        split_sets = {}
                        for j, key in enumerate(keys):
                            vals = sets_ready[key]
                            sp_val = splitters[j]
                            split_sets[key + "0"] = vals[:sp_val]
                            split_sets[key + "1"] = vals[sp_val:]

                        universe = set(n for v in sets_ready.values() for n in v)
                        comb3 = [c for r in range(3, 4)
                                 for c in _sp_combos(keys, r)]
                        comb2 = [c for c in _sp_combos(keys, 2)]

                        result_sp = {}
                        # Split sets
                        for i, key in enumerate(sorted(split_sets.keys())):
                            result_sp[key] = split_sets[key]
                            result_sp[f"y{i}"] = list(
                                universe - set(split_sets[key]))

                        # 3-combinations
                        let_3 = ["e", "f", "g", "h"]
                        comb3dict = {}
                        for j, combo in enumerate(comb3):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb3dict[let_3[j % 4] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb3dict.items():
                            result_sp[key] = list(val)

                        # 2-combinations
                        let_2 = ["i", "j", "k", "l", "m", "n"]
                        comb2dict = {}
                        for j, combo in enumerate(comb2):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb2dict[let_2[j % 6] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb2dict.items():
                            result_sp[key] = list(val)

                        sp_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_sp.items()})
                        order_sp = sp_df.isna().sum().sort_values().index
                        sp_df = sp_df[order_sp]
                        gs_set("Sp", sp_df)
                        sp_path = _gdirs["Splits"] / f"Sp_{_gkey}.csv"
                        _sets_df_to_rows(sp_df, set_col="set").to_csv(sp_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ Sp generated: {sp_df.shape[1]} columns '
                            f'→ {sp_path.name}</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(sp_df, key="sp_generated_view", use_container_width=True)
                    except Exception as ex:
                        st.error(f"Splits error: {ex}")
            else:
                st.warning("Could not extract numeric data from D columns.")

        sp_view = gs("Sp", pd.DataFrame())
        if not sp_view.empty:
            st.markdown("**Current Sp in memory (row-oriented — one row per set):**")
            show_paginated_df(_sets_df_to_rows(sp_view, set_col="set"), key="sp_current_memory", use_container_width=True)

            # ── Excel export: row-oriented, one sheet per group ──────────
            st.markdown("#### Download as Excel (one sheet per set group, row-oriented)")
            try:
                import io as _sp_io
                _sp_buf = _sp_io.BytesIO()
                with pd.ExcelWriter(_sp_buf, engine="openpyxl") as _sp_writer:
                    # Sheet "All" — full row-oriented table
                    _sets_df_to_rows(sp_view, set_col="set").to_excel(
                        _sp_writer, sheet_name="All", index=False)
                    # Group columns by prefix, then export each group row-oriented
                    _sp_groups = {
                        "Splits_w":      [c for c in sp_view.columns
                                          if re.match(r'^w\d+[01]$', str(c))],
                        "Complements_y": [c for c in sp_view.columns
                                          if re.match(r'^y\d+$', str(c))],
                        "Combo3_efgh":   [c for c in sp_view.columns
                                          if re.match(r'^[efgh]\d+$', str(c))],
                        "Combo2_ijklmn": [c for c in sp_view.columns
                                          if re.match(r'^[ijklmn]\d+$', str(c))],
                    }
                    for _sheet, _cols in _sp_groups.items():
                        if _cols:
                            _sets_df_to_rows(sp_view[_cols], set_col="set").to_excel(
                                _sp_writer, sheet_name=_sheet, index=False)
                _sp_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download Sp_{_gkey}.xlsx (All + 4 group sheets)",
                    data=_sp_buf.getvalue(),
                    file_name=f"Sp_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _sp_dl_ex:
                st.warning(f"Excel export unavailable: {_sp_dl_ex}")

    # ── TAB: So (SplitsCombi) ──────────────────────────────────────────────
    with vtabs[5]:
        st.markdown("**So — SplitsCombi (automation_vba.py): top 4 w-columns → union combis**")

        d_df = gs("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d_so = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)
            top4_so = w_cols_d_so[:4]
            st.write(f"Top 4 w-columns → So input: `{top4_so}`")

            sets_ready_so = {}
            for col in top4_so:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready_so[col] = vals

            if sets_ready_so and len(sets_ready_so) == 4:
                st.markdown("**Set split points for So** (0 = midpoint auto):")
                so_cols = st.columns(4)
                splitters_so = []
                for i, (col, vals) in enumerate(sets_ready_so.items()):
                    with so_cols[i]:
                        default_s = len(vals) // 2
                        sv = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_s,
                            key=f"so_split_{col}_{i}")
                        splitters_so.append(sv if sv > 0 else default_s)

                if st.button("▶ Run SplitsCombi (So)", type="primary",
                             key="run_so", use_container_width=True):
                    try:
                        from itertools import combinations as _so_combos
                        from collections import OrderedDict

                        keys_so = list(sets_ready_so.keys())
                        split_sets_so = {}
                        for j, key in enumerate(keys_so):
                            vals = sets_ready_so[key]
                            sv = splitters_so[j]
                            split_sets_so[key + "0"] = vals[:sv]
                            split_sets_so[key + "1"] = vals[sv:]

                        universe_so = set(
                            n for v in sets_ready_so.values() for n in v)

                        # Build pair/triple dicts; store actual tuples separately so
                        # membership check can use proper key extraction instead of
                        # fragile string-index tricks (which broke with multi-char
                        # keys like w1,w2,w3,w4).
                        _pair_tuples = list(_so_combos(list(split_sets_so.keys()), 2))
                        _triple_tuples = list(_so_combos(keys_so, 3))

                        comb_pairs = {}
                        _pair_map: dict = {}   # str(tuple) → actual tuple
                        for _pt in _pair_tuples:
                            _k = str(_pt)
                            comb_pairs[_k] = set(split_sets_so[_pt[0]]).union(
                                set(split_sets_so[_pt[1]]))
                            _pair_map[_k] = _pt

                        comb_three = {}
                        _triple_map: dict = {}
                        for _tt in _triple_tuples:
                            _k = str(_tt)
                            comb_three[_k] = set(sets_ready_so[_tt[0]]).union(
                                set(sets_ready_so[_tt[1]]),
                                set(sets_ready_so[_tt[2]]))
                            _triple_map[_k] = _tt

                        result_so: dict = {}
                        result_so["U"] = universe_so

                        for set_i, tset in comb_three.items():
                            triple_base_keys = set(_triple_map[set_i])
                            for set_j, pset in comb_pairs.items():
                                # Extract base key by stripping trailing "0"/"1" suffix.
                                # Works for any key length: 'w10'→'w1', 'a0'→'a'.
                                _pair_t = _pair_map[set_j]
                                base0 = _pair_t[0][:-1]
                                base1 = _pair_t[1][:-1]
                                if (pset.issubset(tset)
                                        and base0 in triple_base_keys
                                        and base1 in triple_base_keys):
                                    result_so[f"U-{set_i}-{set_j}"] = \
                                        universe_so - (tset - pset)
                                    result_so[f"{set_i}-{set_j}"] = tset - pset
                            result_so[str(set_i)] = tset

                        for set_j, pset in comb_pairs.items():
                            result_so[f"U-{set_j}"] = universe_so - pset
                            result_so[str(set_j)] = pset

                        so_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_so.items()})
                        order_so = so_df.isna().sum().sort_values().index
                        so_df = so_df[order_so]
                        gs_set("So", so_df)
                        so_path = _gdirs["Splits_Combi"] / f"So_{_gkey}.csv"
                        _sets_df_to_rows(so_df, set_col="set").to_csv(so_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ So generated: {so_df.shape[1]} columns '
                            f'→ {so_path.name}</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(so_df, key="so_generated_view", use_container_width=True)
                    except Exception as ex:
                        st.error(f"SplitsCombi error: {ex}")
            elif sets_ready_so:
                st.info(f"So engine needs exactly 4 w-columns. Found {len(sets_ready_so)}.")
            else:
                st.warning("Could not extract numeric data from D columns.")

        so_view = gs("So", pd.DataFrame())
        if not so_view.empty:
            st.markdown("**Current So in memory (row-oriented — one row per set):**")
            show_paginated_df(_sets_df_to_rows(so_view, set_col="set"), key="so_current_memory", use_container_width=True)

            # ── Excel export: row-oriented, one sheet per combination type
            st.markdown("#### Download as Excel (one sheet per combination type, row-oriented)")
            try:
                import io as _so_io
                _so_buf = _so_io.BytesIO()

                def _so_group(col: str) -> str:
                    """Classify a So column name into its combination category."""
                    c = str(col)
                    if c == "U":
                        return "Universe"
                    if c.startswith("U-") and c.count("(") >= 2:
                        return "U_minus_triple_pair"
                    if not c.startswith("U-") and c.count("(") >= 2:
                        return "Triple_minus_pair"
                    if c.count("(") == 1 and "," in c and c.count(",") >= 2:
                        return "Triples"
                    if c.startswith("U-"):
                        return "U_minus_pair"
                    return "Pairs"

                from collections import defaultdict as _sodd
                _so_groups: dict = _sodd(list)
                for _c in so_view.columns:
                    _so_groups[_so_group(_c)].append(_c)

                with pd.ExcelWriter(_so_buf, engine="openpyxl") as _so_writer:
                    # Sheet "All" — full row-oriented table
                    _sets_df_to_rows(so_view, set_col="set").to_excel(
                        _so_writer, sheet_name="All", index=False)
                    _sheet_order = ["Universe", "Triples", "Triple_minus_pair",
                                    "U_minus_triple_pair", "Pairs", "U_minus_pair"]
                    for _sh in _sheet_order:
                        _cols = _so_groups.get(_sh, [])
                        if _cols:
                            _sets_df_to_rows(so_view[_cols], set_col="set").to_excel(
                                _so_writer, sheet_name=_sh[:31], index=False)

                _so_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download So_{_gkey}.xlsx (All + 6 type sheets)",
                    data=_so_buf.getvalue(),
                    file_name=f"So_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _so_dl_ex:
                st.warning(f"Excel export unavailable: {_so_dl_ex}")

    # ── TAB: Since Last ────────────────────────────────────────────────────
    with vtabs[6]:
        st.markdown(f"**Since Last — fetch from lottolyzer for {_gcfg['label']}**")

        # ── Draw schedule header ─────────────────────────────────────────────
        _sl_last_draw = _last_draw_date(_gkey)
        if _sl_last_draw:
            from datetime import timedelta as _td
            _sl_next_draw = _sl_last_draw + _td(days=7)
            st.markdown(
                f'<div class="info">📅 <b>Draw schedule</b> — '
                f'last draw: <b>{_sl_last_draw.strftime("%-d %b %Y (%A)")}</b> · '
                f'next draw: <b>{_sl_next_draw.strftime("%-d %b %Y (%A)")}</b> · '
                f'day of week: <b>{_gcfg.get("draw_day","?")}</b></div>',
                unsafe_allow_html=True)

        _default_sl_url = _gcfg["lottolyzer"]
        sl_url = st.text_input(
            "Lottolyzer URL (editable — fix if it points to the wrong game):",
            value=_default_sl_url,
            key="sl_url_override"
        )
        st.markdown(f"→ [Open this link in browser]({sl_url})")

        sl_file = _gdirs["SinceLast"] / "since_last.json"

        _sl_btn_col1, _sl_btn_col2 = st.columns([3, 1])
        with _sl_btn_col1:
            _do_fetch_sl = st.button("⤓ Fetch now from lottolyzer", type="primary",
                                     use_container_width=True, key="fetch_sl_now")
        with _sl_btn_col2:
            _do_clear_sl = st.button("🗑️ Clear cache", use_container_width=True,
                                     key="clear_sl_cache",
                                     help="Delete the cached since_last.json so the next fetch gets fresh data")
        if _do_clear_sl and sl_file.exists():
            sl_file.unlink()
            st.success("Cache cleared. Click 'Fetch now' to get current data from lottolyzer.")
            st.rerun()

        if _do_fetch_sl:
            _stats_freq = gs("Freq")
            _has_freq_cache = (
                _stats_freq is not None
                and not _stats_freq.empty
                and "number" in _stats_freq.columns
                and "since_last" in _stats_freq.columns
            )
            if _has_freq_cache:
                _sl_extract = _stats_freq[["number", "since_last"]].copy()
                _sl_extract["number"]     = pd.to_numeric(_sl_extract["number"],     errors="coerce")
                _sl_extract["since_last"] = pd.to_numeric(_sl_extract["since_last"], errors="coerce")
                _sl_extract = _sl_extract.dropna(subset=["number", "since_last"])
                d = {int(r["number"]): int(r["since_last"]) for _, r in _sl_extract.iterrows()}
            else:
                with st.spinner("Fetching from lottolyzer…"):
                    d = fetch_since_last(sl_url, _gcfg["pool"])
            if d:
                save_since_last(d, _gkey, _gcfg["label"], _gcfg["pool"],
                                sl_url, sl_file)
                st.success(f"Fetched and cached {len(d)} numbers. Refresh to see the table.")
            else:
                st.error("Couldn't parse lottolyzer (unreachable or layout changed). "
                         "Use the manual upload below.")
        if sl_file.exists():
            try:
                sl_data = json.loads(sl_file.read_text())
                scraped_at = sl_data.get("scraped_at", "unknown")
                n_nums = len(sl_data.get("since_last_dict", {}))
                _sl_ban_cls, _sl_ban_msg = _data_freshness_banner(scraped_at, _gkey)
                st.markdown(
                    f'<div class="{_sl_ban_cls}">📅 {_sl_ban_msg} '
                    f'({n_nums} numbers cached)</div>',
                    unsafe_allow_html=True)
                # Show table
                sl_dict = {int(k): int(v)
                           for k, v in sl_data.get("since_last_dict", {}).items()}
                all_wt = sl_data.get("all_wt", [])
                sl_display = pd.DataFrame([
                    {"Rank": i+1, "Number": num,
                     "Since_Last": sl_dict.get(num, "?"),
                     "Group": ("🟡 1–9" if num <= 9
                               else "🔵 10–19" if num <= 19
                               else "⬜ 20–29" if num <= 29
                               else "🟢 30–40" if num <= 40
                               else "🩷 41–49")}
                    for i, num in enumerate(all_wt)
                ])
                show_paginated_df(sl_display, key="sl_display_tbl", use_container_width=True)
            except Exception as ex:
                st.error(f"Error loading Since Last: {ex}")

        st.markdown("---")
        st.markdown("**Manual upload** (paste from lottolyzer CSV export):")
        up_sl = st.file_uploader(
            "Upload Since Last file (columns: Number, Since Last) — CSV, Excel, …",
            type=_UPLOAD_TYPES, key="up_sl")
        if up_sl:
            try:
                df_sl = _read_uploaded(up_sl)
                # Find Since Last column
                sl_col = next((c for c in df_sl.columns
                               if "since" in c.lower()), None)
                num_col = next((c for c in df_sl.columns
                                if "number" in c.lower() or c == "Number"), None)
                if sl_col and num_col:
                    sl_dict_up = {int(row[num_col]): int(row[sl_col])
                                  for _, row in df_sl.iterrows()
                                  if pd.notna(row[sl_col]) and pd.notna(row[num_col])}
                    pool = _gcfg["pool"]
                    all_wt_up = sorted(sl_dict_up.keys(),
                                       key=lambda n: (sl_dict_up[n], n))
                    # Same ordering as all_wt; kept as a separate list so the two
                    # keys can diverge later without surprises.
                    to_keep_up = list(all_wt_up)
                    save_data = {
                        "since_last_dict": {str(k): v
                                             for k, v in sl_dict_up.items()},
                        "all_wt": all_wt_up,
                        "to_keep": to_keep_up,
                        "game": _gkey,
                        "game_name": _gcfg["label"],
                        "pool_size": pool,
                        "scraped_at": datetime.now().isoformat(),
                        "url": sl_url,
                    }
                    _gdirs["SinceLast"].mkdir(parents=True, exist_ok=True)
                    sl_file.write_text(json.dumps(save_data, indent=2))
                    st.success(f"Since Last saved: {len(sl_dict_up)} numbers. "
                               f"Refresh page to see.")
                else:
                    st.error("Could not find 'Number' and 'Since Last' columns.")
            except Exception as ex:
                st.error(f"Upload error: {ex}")

        st.markdown("---")
        st.markdown("""
        **To get Since Last manually from lottolyzer:**
        1. Go to the link above
        2. Export or copy the frequency table
        3. Save as CSV with columns: `Number`, `Since Last`
        4. Upload above
        """)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER FORMULA
# ═══════════════════════════════════════════════════════════════════════════════
    # ── TAB: Stats (Statistics Pipeline) ────────────────────────────────────────
    with vtabs[7]:
        st.markdown(f"**📊 Statistics Pipeline — {_gcfg['label']}**")

        # ── Draw schedule header ─────────────────────────────────────────────
        _stats_last_draw = _last_draw_date(_gkey)
        if _stats_last_draw:
            from datetime import timedelta as _td2
            _stats_next_draw = _stats_last_draw + _td2(days=7)
            st.markdown(
                f'<div class="info">📅 <b>Draw schedule</b> — '
                f'last draw: <b>{_stats_last_draw.strftime("%-d %b %Y (%A)")}</b> · '
                f'next draw: <b>{_stats_next_draw.strftime("%-d %b %Y (%A)")}</b></div>',
                unsafe_allow_html=True)

        st.caption(
            "Fetch number frequency table and draw history from lottolyzer. "
            "Use *Since Last* from here to feed R. Also manually record new draws "
            "to keep B up-to-date.")

        _st_url    = _gcfg.get("lottolyzer","")
        _st_pool   = _gcfg.get("pool", 45)
        _st_freq_path  = _gdirs.get("SinceLast", _gdirs.get("Base", Path("."))) / "number_freq.csv"
        _st_hist_path  = _gdirs.get("SinceLast", _gdirs.get("Base", Path("."))) / "draw_history.csv"

        # ── Clear cached stats files ─────────────────────────────────────────
        if st.button("🗑️ Clear cached frequency & history (force fresh fetch)",
                     key="stats_clear_cache",
                     help="Deletes number_freq.csv and draw_history.csv so the next fetch pulls current data"):
            _cleared = []
            for _p in [_st_freq_path, _st_hist_path]:
                if _p.exists():
                    _p.unlink()
                    _cleared.append(_p.name)
                    S.pop(f"Freq_{_gkey}", None)
                    S.pop(f"DrawHist_{_gkey}", None)
            if _cleared:
                st.success(f"Cleared: {', '.join(_cleared)}. Now click Fetch below to get current {_gcfg['label']} data.")
            else:
                st.info("No cached files to clear.")
            st.rerun()

        # ── Section 1: Number Frequency Table ───────────────────────────────────
        st.markdown("### Number Frequency Table")
        _st_col1, _st_col2 = st.columns([3,1])
        with _st_col1:
            _st_freq_url = st.text_input("Frequency URL", value=_st_url,
                                          key="stats_freq_url",
                                          placeholder="https://en.lottolyzer.com/…")
        with _st_col2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            _do_fetch_freq = st.button("🔄 Fetch Frequencies", key="stats_fetch_freq",
                                        use_container_width=True)

        if _do_fetch_freq and _st_freq_url:
            with st.spinner("Fetching number frequency table…"):
                _freq_df = fetch_number_frequencies(_st_freq_url, _st_pool)
            if _freq_df is not None and not _freq_df.empty:
                # Sanity check: a valid frequency table should have at least half
                # the pool size as rows, and since_last must not be all-None.
                _sl_valid = pd.to_numeric(_freq_df.get("since_last", pd.Series()), errors="coerce").notna().sum()
                _min_rows = max(5, _st_pool // 2)
                if len(_freq_df) < _min_rows or _sl_valid == 0:
                    st.error(
                        f"❌ Fetch returned only {len(_freq_df)} rows with {_sl_valid} valid "
                        f"'since_last' values — this doesn't look like a frequency table. "
                        f"The URL must be in the format: "
                        f"https://en.lottolyzer.com/number-frequencies/australia/[game-slug] "
                        f"(e.g. tattslotto). Check the URL and try again. "
                        f"Data NOT cached."
                    )
                else:
                    _freq_df.to_csv(_st_freq_path, index=False)
                    gs_set("Freq", _freq_df)
                    st.success(f"✅ Fetched {len(_freq_df)} numbers — saved to {_st_freq_path.name}")
            else:
                st.error(
                    "❌ Could not parse frequency table. "
                    "URL must be: https://en.lottolyzer.com/number-frequencies/australia/tattslotto "
                    "(for Saturday Lotto). Check URL and try again."
                )

        # Auto-load cached frequency table
        if gs("Freq") is None or gs("Freq", pd.DataFrame()).empty:
            if _st_freq_path.exists():
                try:
                    gs_set("Freq", pd.read_csv(_st_freq_path))
                except Exception:
                    pass

        _freq_view = gs("Freq", pd.DataFrame())
        if not _freq_view.empty:
            # Show a human-readable "last fetched" and freshness banner
            if _st_freq_path.exists():
                import time as _time
                _mtime = _st_freq_path.stat().st_mtime
                _mtime_str = datetime.fromtimestamp(_mtime).strftime("%-d %b %Y %H:%M")
                _fban_cls, _fban_msg = _data_freshness_banner(
                    datetime.fromtimestamp(_mtime).isoformat(), _gkey)
                st.markdown(
                    f'<div class="{_fban_cls}">📅 {_fban_msg} '
                    f'({len(_freq_view)} numbers)</div>',
                    unsafe_allow_html=True)
            else:
                st.caption(f"{len(_freq_view)} numbers")
            show_paginated_df(_freq_view, key="stats_freq_view", use_container_width=True)
            st.download_button("⬇ Download frequency CSV",
                               _freq_view.to_csv(index=False).encode(),
                               f"freq_{_gkey}.csv", "text/csv", key="dl_stats_freq")
            # Quick Since-Last digest — drop rows where number or since_last is NaN
            if "since_last" in _freq_view.columns and "number" in _freq_view.columns:
                _sl_clean = _freq_view[["number", "since_last"]].copy()
                _sl_clean["number"]     = pd.to_numeric(_sl_clean["number"],     errors="coerce")
                _sl_clean["since_last"] = pd.to_numeric(_sl_clean["since_last"], errors="coerce")
                _sl_clean = _sl_clean.dropna(subset=["number", "since_last"])
                _sl_clean["number"]     = _sl_clean["number"].astype(int)
                _sl_clean["since_last"] = _sl_clean["since_last"].astype(int)
                sl_from_freq = dict(zip(_sl_clean["number"], _sl_clean["since_last"]))
                st.info(f"💡 Since Last extracted for {len(sl_from_freq)} numbers. "
                        f"Use **Save to Since Last** below to make this available to R.")
                if st.button("💾 Save frequency Since Last → R pipeline", key="stats_save_sl"):
                    save_since_last(sl_from_freq, _gkey, _gcfg["label"],
                                    _st_pool, _st_freq_url,
                                    _gdirs["SinceLast"] / "since_last.json")
                    st.success("✅ Since Last saved — reload the R tab to use it.")

        st.markdown("---")

        # ── Section 2: Draw History ──────────────────────────────────────────────
        st.markdown("### Draw History")
        _st_hist_pages = st.slider("Pages to fetch (50 draws/page)", 1, 10, 3,
                                    key="stats_hist_pages")
        _do_fetch_hist = st.button("🔄 Fetch Draw History", key="stats_fetch_hist",
                                    use_container_width=False)

        if _do_fetch_hist:
            with st.spinner(f"Fetching {_st_hist_pages} page(s) of draw history…"):
                _hist_df = fetch_draw_history(_gkey, pages=_st_hist_pages)
            if _hist_df is not None and not _hist_df.empty:
                _hist_df.to_csv(_st_hist_path, index=False)
                gs_set("DrawHist", _hist_df)
                st.success(f"✅ Fetched {len(_hist_df)} draws — saved to {_st_hist_path.name}")
            else:
                st.error("❌ Could not parse draw history. Check game config URL.")

        if gs("DrawHist") is None or gs("DrawHist", pd.DataFrame()).empty:
            if _st_hist_path.exists():
                try:
                    gs_set("DrawHist", pd.read_csv(_st_hist_path))
                except Exception:
                    pass

        _hist_view = gs("DrawHist", pd.DataFrame())
        if not _hist_view.empty:
            # Show latest 10 draws inline
            _disp_hist = _hist_view.head(10).copy()
            st.caption(f"Latest {min(10, len(_hist_view))} of {len(_hist_view)} draws")
            show_paginated_df(_disp_hist, key="stats_hist_view", use_container_width=True)
            st.download_button("⬇ Download draw history CSV",
                               _hist_view.to_csv(index=False).encode(),
                               f"draw_history_{_gkey}.csv", "text/csv", key="dl_stats_hist")

        st.markdown("---")

        # ── Section 3: B Sync / Add Draw ─────────────────────────────────────────
        st.markdown("### B Sync / Add Draw")

        # ── 3a: Sync from draw history ────────────────────────────────────────
        st.markdown("#### 🔄 Sync B to latest")
        _b_hist_start_cfg = _gcfg.get("b_hist_start")
        if _b_hist_start_cfg is None:
            st.info(f"ℹ️ `b_hist_start` not configured for **{_gcfg['label']}** — "
                    f"sync unavailable. Use manual entry below.")
        else:
            st.caption(
                f"Compares B row {_b_hist_start_cfg} (current newest draw) against "
                f"cached draw history to detect missing draws.")

            _hist_for_sync = gs("DrawHist", pd.DataFrame())
            if _hist_for_sync.empty:
                st.warning("⚠️ No draw history in memory — fetch draw history (section 2 above) first.")
            else:
                _do_sync = st.button("🔄 Sync B to latest", key="stats_sync_b")
                if _do_sync:
                    _b_for_sync = gs("B", pd.DataFrame())
                    _sync_res = sync_b_with_latest_draws(_gkey, _b_for_sync, _hist_for_sync)
                    S[f"b_sync_result__{_gkey}"] = _sync_res

                _sync_res = S.get(f"b_sync_result__{_gkey}")
                if _sync_res:
                    _s = _sync_res["status"]
                    if _s == "current":
                        st.success(f"✅ B is current through draw **{_sync_res['draw']}** "
                                   f"— numbers: {_sync_res['numbers']}")
                    elif _s == "behind":
                        st.warning(f"⚠️ B is **{_sync_res['count']} draw(s) behind**.")
                        _missing_df = pd.DataFrame(_sync_res["missing_draws"])
                        st.dataframe(_missing_df, use_container_width=True, hide_index=True)
                        _do_apply = st.button(
                            f"✅ Apply {_sync_res['count']} missing draw(s) to B",
                            key="stats_sync_apply")
                        if _do_apply:
                            try:
                                _b_current = gs("B", pd.DataFrame())
                                _b_updated = append_draws_to_b(
                                    _gkey, _b_current, _sync_res["missing_draws"])
                                gs_set("B", _b_updated)
                                _b_save_path = _gdirs["Base"] / f"B_{_gkey}_updated.csv"
                                _b_updated.to_csv(_b_save_path, index=False)
                                S.pop(f"b_sync_result__{_gkey}", None)
                                st.success(
                                    f"✅ Inserted {_sync_res['count']} draw(s) at row "
                                    f"{_b_hist_start_cfg}. B now has {len(_b_updated)} rows. "
                                    f"Saved to {_b_save_path.name}")
                                st.caption("Re-run generators to reflect the new draw(s).")
                                st.rerun()
                            except Exception as _apply_ex:
                                st.error(f"Error applying draws: {_apply_ex}")
                    elif _s == "gap_too_large":
                        st.error(f"❌ {_sync_res.get('detail', 'Gap too large')}")
                    elif _s == "not_configured":
                        st.info("ℹ️ `b_hist_start` not configured for this game.")

        st.markdown("---")

        # ── 3b: Manual draw entry (fallback) ─────────────────────────────────
        st.markdown("#### ➕ Add draw manually")
        st.caption(
            "Enter the winning numbers for a new draw. They will be added to B "
            "so the pipeline stays current. Numbers are sorted (w1=smallest) before saving.")
        if _b_hist_start_cfg is not None:
            st.caption(f"Inserted at row {_b_hist_start_cfg} (newest-first), "
                       f"matching B's draw-history ordering — same insertion "
                       f"point as the Sync button above.")
        else:
            st.caption("⚠️ `b_hist_start` not configured for this game — the new "
                       "draw will be appended at the bottom of B for now.")

        _draw_label_in = st.text_input("Draw label / number (optional)", value="",
                                        key="stats_draw_label",
                                        placeholder="e.g. 4687  or  2026-06-07")
        _draw_nums_in  = st.text_input("Winning numbers (space or comma separated)",
                                        value="", key="stats_draw_nums",
                                        placeholder="e.g.  4 7 13 21 29 38")

        _do_add_draw = st.button("➕ Add draw to B", key="stats_add_draw")
        if _do_add_draw:
            try:
                _new_nums = [int(x) for x in re.split(r"[,\s]+", _draw_nums_in.strip()) if x]
                if not _new_nums:
                    st.warning("Enter at least one number.")
                else:
                    _b_current = gs("B", pd.DataFrame())
                    _b_updated = append_draw_to_b(_b_current, _new_nums,
                                                   draw_label=_draw_label_in.strip(),
                                                   game=_gkey)
                    gs_set("B", _b_updated)
                    _b_save_path = _gdirs["Base"] / f"B_{_gkey}_updated.csv"
                    _b_updated.to_csv(_b_save_path, index=False)
                    st.success(
                        f"✅ Added draw {_draw_label_in or '(unlabelled)'}: "
                        f"{sorted(_new_nums)} → B now has {len(_b_updated)} rows. "
                        f"Saved to {_b_save_path.name}")
                    st.caption("B in memory has been updated. Re-run generators to reflect the new draw.")
            except Exception as _draw_ex:
                st.error(f"Error adding draw: {_draw_ex}")

        # Show current B tail for confirmation
        _b_for_stats = gs("B", pd.DataFrame())
        if not _b_for_stats.empty:
            if "w" in _b_for_stats.columns:
                _tail_b = _b_for_stats.tail(5)
                st.caption(f"Last 5 w-sets of B ({len(_b_for_stats)} total w-sets):")
            else:
                _wcols_b = [c for c in _b_for_stats.columns
                             if str(c).lower().startswith("w")]
                _tail_b = _b_for_stats[_wcols_b].tail(5)
                st.caption(f"Last 5 rows of B ({len(_b_for_stats)} total rows, "
                           f"{len(_wcols_b)} w-columns):")
            show_paginated_df(_tail_b, key="stats_b_tail", use_container_width=True)
elif page == "📦 Container Formula":
    st.markdown('<span class="sec-hdr hdr-purple">📦 Container Formula — Live Collation</span>',
                unsafe_allow_html=True)
    am_toggle("cf")

    st.markdown("""
    <div class="info">
    Each row = 1 container. 17 rows = 17 containers running in parallel (Python multiprocessing).
    No Docker needed — Python processes handle parallel execution natively.
    Sp = S throughout. Collation: drop col-0 of each component, concatenate, relabel w1…wN.
    </div>
    """, unsafe_allow_html=True)

    # Variable status
    st.markdown("**Variables in memory:**")
    vc = st.columns(6)
    for i, k in enumerate(["B","R","D","Sp","So","Ep"]):
        with vc[i]:
            df = gs(k, pd.DataFrame())
            if isinstance(df, pd.DataFrame) and not df.empty:
                st.markdown(f'<div class="ok">✅ {k} {len(df)}r×{len(df.columns)}c</div>',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="warn">⚠️ {k} empty</div>',
                            unsafe_allow_html=True)

    st.markdown("---")
    cf_df = pd.DataFrame({
        "#": [r[0] for r in CF_ROWS],
        "Name": [r[1] for r in CF_ROWS],
        "Type": [r[2] for r in CF_ROWS],
        "Components": [" + ".join(r[3]) for r in CF_ROWS],
        "Active": [S["cf_active"].get(r[1],True) for r in CF_ROWS],
    })
    edited_cf = st.data_editor(cf_df, key="cf_tbl",
        use_container_width=True, num_rows="fixed", height=530,
        column_config={
            "Active": st.column_config.CheckboxColumn("Active"),
            "Type":   st.column_config.SelectboxColumn("Type",
                          options=["Formed","Ready-made"]),
        })
    for _, row in edited_cf.iterrows():
        S["cf_active"][row["Name"]] = row["Active"]

    st.markdown("---")
    active_names = [r[1] for r in CF_ROWS if S["cf_active"].get(r[1],True)]
    chosen_f = st.selectbox("Collate formula:", active_names)
    comps = COMP_MAP.get(chosen_f, [])
    st.write(f"**Components:** {' + '.join(comps)}")
    # Resolve to base variables (strip trailing digits, de-dup) — must match
    # execute_collation, so D1D2D3 checks for D (not D1/D2/D3), B1B2B3 → B, etc.
    _seen, base_needed = set(), []
    for c in comps:
        b = re.sub(r"\d+$", "", str(c)).strip()
        if b and b not in _seen:
            _seen.add(b); base_needed.append(b)
    missing = [b for b in base_needed if gs(b) is None or
               (isinstance(gs(b), pd.DataFrame) and gs(b).empty)]
    if missing:
        st.markdown(f'<div class="warn">⚠️ Missing/empty: {missing} '
                    f'(load these in Variable Inputs, or Build W-Matrix for D)</div>',
                    unsafe_allow_html=True)

    # ── JOIN DEMO: show how variables stack BEFORE running collation ─────
    with st.expander("🔍 Collation Join Preview — see exactly how variables stack", expanded=False):
        st.markdown(
            '<div class="info"><b>Rule:</b> B rows come first (all of them), then R rows '
            '(if loaded), then D rows underneath. Each block is separated by a divider. '
            'Numbers fill across columns w1, w2, … matching the widest row.</div>',
            unsafe_allow_html=True)

        _demo_pieces = []
        _demo_summary = []
        for _dv in base_needed:
            _df_dv = gs(_dv)
            if _df_dv is None or (isinstance(_df_dv, pd.DataFrame) and _df_dv.empty):
                continue
            _block = _to_w_rows(_df_dv, is_direct=(_dv == "D"),
                                 force_column_oriented=(_dv == "Sp"))
            if _block is None or _block.empty:
                continue
            _block = _block.copy()
            _val_cols = [c for c in _block.columns if c != "Set_Label"]
            _block = _block.rename(
                columns={c: f"w{i+1}" for i, c in enumerate(_val_cols)})
            _block.insert(0, "Source", _dv)
            _demo_pieces.append(_block)
            _demo_summary.append({"Variable": _dv, "Rows": f"{len(_block):,}",
                                   "Widest pick": len(_block.columns) - 2})

        if _demo_summary:
            st.markdown("**Block sizes — in stacking order:**")
            _sum_df = pd.DataFrame(_demo_summary)
            _sum_df["Cumulative rows"] = _sum_df["Rows"].str.replace(",","").astype(int).cumsum().apply(lambda x: f"{x:,}")
            show_paginated_df(_sum_df, key="cf_demo_sum", use_container_width=True, hide_index=True)

            for _pi, _piece in enumerate(_demo_pieces):
                _var_name = _piece["Source"].iloc[0]
                _n_rows = len(_piece)
                _preview_n = min(8, _n_rows)
                st.markdown(
                    f'<div class="ok" style="margin-top:10px">▶ <b>{_var_name}</b> block — '
                    f'{_n_rows:,} row{"s" if _n_rows!=1 else ""} '
                    f'(showing first {_preview_n})</div>',
                    unsafe_allow_html=True)
                show_paginated_df(_piece, key=f"cf_piece_preview_{_pi}", use_container_width=True)
                if _pi < len(_demo_pieces) - 1:
                    _next_var = _demo_pieces[_pi+1]["Source"].iloc[0]
                    st.markdown(
                        f'<div class="note" style="text-align:center;font-size:.85rem">'
                        f'⬇ {_next_var} rows continue below this point</div>',
                        unsafe_allow_html=True)
        else:
            st.info("Load variables in Variable Inputs first to preview the join.")

    st.markdown("---")

    if st.button(f"▶ Collate {chosen_f}", type="primary", use_container_width=True):
        with st.spinner("Collating…"):
            result = execute_collation(comps)
        if result.empty:
            st.error("Result empty — load components in Variable Inputs first.")
        else:
            out = _gdirs["CVI"] / f"CVI_{chosen_f}.csv"
            result.to_csv(out, index=False)
            st.session_state.setdefault(gkey("cvi"), {})[chosen_f] = result
            n_wcols = sum(1 for c in result.columns if str(c).startswith("w"))
            st.markdown(f'<div class="ok">✅ {chosen_f}: {len(result):,} rows '
                        f'(w-sets) × {n_wcols} number columns (w1…w{n_wcols}) '
                        f'→ <code>{out.name}</code></div>',
                        unsafe_allow_html=True)

            # ── Stacked result — show each variable's block clearly ───────
            _res_sources = result["Source"].unique().tolist() if "Source" in result.columns else []
            if _res_sources:
                st.markdown("**Result breakdown by variable block:**")
                _blk_rows = []
                for _src in _res_sources:
                    _blk = result[result["Source"] == _src]
                    _blk_rows.append({"Variable": _src,
                                      "Rows": f"{len(_blk):,}",
                                      "Starts at row": f"{result[result['Source']==_src].index[0]+1:,}"})
                show_paginated_df(pd.DataFrame(_blk_rows), key="cf_blk_rows", use_container_width=True, hide_index=True)

            n_show_c = min(60, result.shape[1])
            st.caption(f"Result — {result.shape[0]:,} rows (full matrix saved to disk):")
            show_paginated_df(result.iloc[:, :n_show_c], key="cf_result_view", use_container_width=True)
            st.download_button(f"⬇ CVI_{chosen_f}.csv", to_csv_bytes(result),
                               f"CVI_{chosen_f}.csv","text/csv")

    if gs("cvi"):
        with st.expander("📋 All collated CVIs in memory"):
            for fname, df in gs("cvi", {}).items():
                st.markdown(f"**{fname}** — {len(df)} rows × {len(df.columns)} cols")
                show_paginated_df(df, key=f"cvi_mem_expander_{fname}", use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER DASHBOARDS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🖥️ Container Dashboards":
    st.markdown('<span class="sec-hdr hdr-red">🖥️ Container Dashboards</span>',
                unsafe_allow_html=True)
    am_toggle("cd")

    # ── PARALLEL RUN — runs all active containers simultaneously ──────────
    st.markdown("### ▶ Run All Active Containers in Parallel")
    st.markdown("""
    <div class="info">
    Fires all CVI files found in <code>Container_Variable_Inputs/</code>
    simultaneously using Python multiprocessing.
    Each container gets its own worker process.
    Carry-forward defaults to all-U for parallel run.
    </div>
    """, unsafe_allow_html=True)

    # Cluster config
    clusters = load_clusters()
    next_id  = next_cluster_id(clusters)
    all_main_files = scan_main_data_files()
    all_cvi_files  = scan_cvi_files()

    _lotto_keys = list(LOTTO_TYPES.keys())
    _default_lotto = _gcfg.get("lotto_type", _lotto_keys[0])
    _default_lotto_idx = _lotto_keys.index(_default_lotto) if _default_lotto in _lotto_keys else 0

    p1,p2,p3,p4 = st.columns(4)
    with p1:
        st.metric("Next Cluster ID", next_id)
    with p2:
        cluster_label = st.text_input("Cluster label:", f"1n",
                                      key="par_cluster_label",
                                      help="e.g. 1n, Draw1567, May28")
    with p3:
        lotto_sel = st.selectbox("Lotto type:",
                                 _lotto_keys,
                                 index=_default_lotto_idx,
                                 format_func=lambda x: f"{x} — {LOTTO_TYPES[x]}",
                                 key="par_lotto")
    with p4:
        draw_date_par = st.text_input("Draw date (YYYY_MM_DD):",
                                      datetime.now().strftime("%Y_%m_%d"),
                                      key="par_date")

    # Auto-scan Main Data
    matching_main = [f for f in all_main_files
                     if f.get("lotto") == lotto_sel or not f.get("lotto")]
    if matching_main:
        main_choice = st.selectbox(
            "Main Data file (auto-scanned from Main_Data/):",
            [f["raw"] for f in matching_main],
            format_func=lambda x: f"{x}  ({next(f['rows'] for f in matching_main if f['raw']==x):,} rows)",
            key="par_main_choice"
        )
        chosen_main_info = next(f for f in matching_main if f["raw"] == main_choice)
        draw_no_par = chosen_main_info.get("draw","D?")
    else:
        st.warning("No Main Data files found in Main_Data/. "
                   f"Expected pattern: {{cluster}}_{lotto_sel}_D{{draw_no}}.csv")
        main_choice  = None
        draw_no_par  = "D?"

    # Show available CVI files
    matching_cvi = [f for f in all_cvi_files
                    if f.get("lotto") == lotto_sel or not f.get("lotto")]
    st.markdown(f"**CVI files found for `{lotto_sel}`: {len(matching_cvi)}**")
    if matching_cvi:
        cvi_preview_df = pd.DataFrame([{
            "Formula": f["formula"], "Date": f["date"], "File": f["raw"]
        } for f in matching_cvi])
        show_paginated_df(cvi_preview_df, key="cd_cvi_preview", use_container_width=True, hide_index=True)
    else:
        st.warning("No CVI files found. Collate formulas in Container Formula first.")

    run_all_btn = st.button(
        f"▶ RUN ALL {len(matching_cvi)} CONTAINERS IN PARALLEL",
        type="primary", use_container_width=True,
        key="run_all_par",
        disabled=(not main_choice or not matching_cvi)
    )

    if run_all_btn and main_choice and matching_cvi:
        main_path  = _gdirs["Main_Data"] / main_choice
        output_dir = _gdirs["Outputs"] / f"Cluster_{next_id}_{cluster_label}"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Build worker args
        worker_args = []
        for cvi_info in matching_cvi:
            formula = cvi_info["formula"]
            sc_path = _gdirs["Selected_Counts"] / f"SC_{formula}.csv"
            worker_args.append((
                formula,
                cvi_info["path"],
                str(main_path),
                str(sc_path) if sc_path.exists() else "",
                next_id,
                cluster_label,
                lotto_sel,
                draw_no_par,
                draw_date_par,
                str(output_dir),
            ))

        st.info(f"Launching {len(worker_args)} workers…")
        prog = st.progress(0, text="Starting workers…")

        # Pool and cpu_count are imported at module top.
        n_workers = min(len(worker_args), cpu_count())
        st.caption(f"Using {n_workers} CPU cores of {cpu_count()} available")

        # On macOS Python 3.8+, the default multiprocessing start method is 'spawn'.
        # Spawn re-imports this module in each worker, which triggers top-level
        # st.set_page_config() and raises StreamlitAPIException — all workers fail
        # silently (result["status"] stays "error").  Use 'fork' explicitly: the
        # worker inherits the already-initialised interpreter state and never
        # re-imports the module.  Forking a multithreaded process (Streamlit has
        # background threads) is generally safe here because workers are CPU-bound
        # and do not touch shared Streamlit state.
        import multiprocessing as _mp_mod
        _pool_ctx = _mp_mod.get_context("fork")
        with _pool_ctx.Pool(processes=n_workers) as pool:
            results = []
            for i, res in enumerate(
                pool.imap_unordered(_parallel_worker, worker_args)
            ):
                results.append(res)
                prog.progress(
                    (i+1)/len(worker_args),
                    text=f"Completed {i+1}/{len(worker_args)}: "
                         f"{res['formula']} — "
                         f"{'✅' if res['status']=='complete' else '❌'} "
                         f"S:{res['selected_n']} U:{res['unselected_n']}"
                )

        prog.empty()

        # Save cluster to registry
        new_cluster = {
            "id":          next_id,
            "label":       cluster_label,
            "lotto":       lotto_sel,
            "draw_no":     draw_no_par,
            "draw_date":   draw_date_par,
            "main_file":   main_choice,
            "output_dir":  str(output_dir),
            "containers":  len(worker_args),
            "status":      "complete",
            "results":     results,
            "timestamp":   datetime.now().isoformat(),
        }
        clusters.append(new_cluster)
        save_clusters(clusters)

        # Show summary
        ok  = [r for r in results if r["status"] == "complete"]
        err = [r for r in results if r["status"] != "complete"]
        st.success(f"✅ Cluster {next_id} '{cluster_label}' complete — "
                   f"{len(ok)} succeeded, {len(err)} errors")
        res_df = pd.DataFrame([{
            "Formula":    r["formula"],
            "Status":     r["status"],
            "Selected":   r["selected_n"],
            "Unselected": r["unselected_n"],
            "Error":      r.get("error",""),
        } for r in results])
        show_paginated_df(res_df, key="cd_parallel_results", use_container_width=True, hide_index=True)
        if err:
            for r in err:
                st.error(f"❌ {r['formula']}: {r['error']}")

    st.markdown("---")

    # ── Dashboard stack ────────────────────────────────────────────────────
    st.markdown("**Individual dashboard (open one at a time):**")
    # Build list from all CVI files + hardcoded defaults
    cvi_found = [parse_cvi_filename(f.name)["formula"]
                 for f in sorted(_gdirs["CVI"].glob("CVI_*.csv"))]
    all_db_names = list(dict.fromkeys(
        [f"1n & {f}" for f in cvi_found] + DASHBOARDS
    ))
    for db_name in all_db_names:
        st.markdown(
            f'<div style="background:#1a1a2e;border:1px solid #444;color:#ccc;'
            f'font-size:.71rem;padding:2px 12px;margin-bottom:2px;'
            f'display:inline-block;min-width:260px;">'
            f'{db_name}…  Container Dashboard</div><br>',
            unsafe_allow_html=True)

    st.markdown("---")
    db = st.selectbox("Open Dashboard:", all_db_names,
                      format_func=lambda x: f"{x}… Container Dashboard")
    formula_name = db.replace("1n & ","")

    st.markdown(f'<div style="font-size:1rem;font-weight:900;border:2px solid #555;'
                f'padding:5px 14px;display:inline-block;margin-bottom:8px;">'
                f'{db}…  Container Dashboard</div>', unsafe_allow_html=True)

    # ── Auto-load CVI for this formula ────────────────────────────────────
    st.session_state.setdefault(gkey("cvi"), {})
    cvi_df = gs("cvi", {}).get(formula_name)
    if cvi_df is None:
        # Try all CVI files for this formula (any lotto type, any date)
        for cvi_fp in sorted(_gdirs["CVI"].glob(f"CVI_*{formula_name}*.csv")):
            cvi_df = _load_file(cvi_fp)
            if not cvi_df.empty:
                st.session_state.setdefault(gkey("cvi"), {})[formula_name] = cvi_df
                break
        if cvi_df is None:
            for search_dir in [_gdirs["CVI"], _gdirs["Rainbow"]]:
                cvi_path = search_dir / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_df = _load_file(cvi_path)
                    break
    if cvi_df is None:
        cvi_df = pd.DataFrame()

    # ── Auto-load SC ───────────────────────────────────────────────────────
    sc_auto = {}
    sc_folder = _gdirs["Selected_Counts"]
    for sc_file in (sorted(sc_folder.glob(f"SC_{formula_name}*.csv")) +
                    sorted(sc_folder.glob(f"SC_{formula_name}*.xlsx"))):
        try:
            df_sc = (pd.read_excel(sc_file, engine="openpyxl")
                     if sc_file.suffix.lower() == ".xlsx"
                     else pd.read_csv(sc_file))
            if "w" in df_sc.columns and "Selected Count" in df_sc.columns:
                for _, row in df_sc.iterrows():
                    sc_auto[str(row["w"])] = str(row["Selected Count"])
        except Exception as _e:
            logging.warning("SC auto-load: could not read %s: %s", sc_file, _e)

    # ── Auto-scan and load Main Data ──────────────────────────────────────
    main_df = gs("main_data", pd.DataFrame())
    avail_main = scan_main_data_files()

    # ── Preview panel ──────────────────────────────────────────────────────
    st.markdown("#### Preview — Loaded Data")
    pv1, pv2, pv3 = st.columns(3)
    with pv1:
        if not cvi_df.empty:
            w_cols_prev = [c for c in cvi_df.columns
                           if re.match(r'^w\d+$', c)]
            st.success(f"✅ **CVI** — {len(w_cols_prev)} w-columns × "
                       f"{len(cvi_df)} rows")
            with st.expander("Preview CVI"):
                show_paginated_df(cvi_df, key="cd_cvi_preview_auto", use_container_width=True,
                             hide_index=True, height=220)
        else:
            st.warning("⚠️ CVI not loaded")

    with pv2:
        if sc_auto:
            st.success(f"✅ **SC** — {len(sc_auto)} w-columns")
            with st.expander("Preview SC"):
                show_paginated_df(pd.DataFrame([
                    {"w": k, "SC": v} for k,v in sc_auto.items()
                ]), key="cd_sc_preview", use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ SC not loaded")
            up_sc = st.file_uploader(
                "Upload SC (cols: w, Selected Count) — CSV, Excel, …",
                type=_UPLOAD_TYPES,
                key="dash_sc_upload", label_visibility="collapsed")
            if up_sc is not None:
                try:
                    df_up = _read_uploaded(up_sc)
                    if "w" in df_up.columns and "Selected Count" in df_up.columns:
                        sc_folder.mkdir(parents=True, exist_ok=True)
                        dest = sc_folder / f"SC_{formula_name}.csv"
                        df_up.to_csv(dest, index=False)
                        st.success(f"Saved → {dest.name}. Refresh to load.")
                    else:
                        st.error("Need columns: 'w' and 'Selected Count'.")
                except Exception as ex:
                    st.error(f"Upload error: {ex}")

    with pv3:
        if not main_df.empty:
            # Validate it's actually main data (has n-columns, not w-columns)
            n_check = [c for c in main_df.columns if re.match(r'^n\d+$', c, re.I)]
            w_check = [c for c in main_df.columns if re.match(r'^w\d+$', c)]
            if w_check and not n_check:
                st.error("⚠️ Loaded file looks like a CVI file (w-columns). "
                         "Please load correct Main Data below.")
                gs_set("main_data", pd.DataFrame())
                main_df = pd.DataFrame()
            else:
                st.success(f"✅ **Main Data** — {len(main_df):,} rows × "
                           f"{len(main_df.columns)} cols")
            with st.expander("Preview Main Data"):
                if avail_main:
                    sel_main = st.selectbox(
                        "Switch Main Data file:",
                        [f["raw"] for f in avail_main],
                        key=f"switch_main_{db}"
                    )
                    if st.button("Load selected", key=f"load_main_{db}"):
                        chosen = next(
                            f for f in avail_main if f["raw"] == sel_main)
                        main_df = _load_file(Path(chosen["path"]))
                        gs_set("main_data",      main_df)
                        gs_set("main_data_path", chosen["path"])
                        st.success(f"Loaded {sel_main}")
                        st.rerun()
                show_paginated_df(main_df, key="cd_main_df_preview", use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ Main Data not loaded")
            if avail_main:
                sel_main2 = st.selectbox(
                    "Available Main Data files:",
                    [f["raw"] for f in avail_main],
                    key=f"sel_main2_{db}"
                )
                if st.button("Load", key=f"load_main2_{db}", type="primary"):
                    chosen2 = next(
                        f for f in avail_main if f["raw"] == sel_main2)
                    main_df = _load_file(Path(chosen2["path"]))
                    gs_set("main_data",      main_df)
                    gs_set("main_data_path", chosen2["path"])
                    st.rerun()

    st.markdown("---")

    # ── Data Status + Re-upload (always visible, always replaceable) ──────
    st.markdown("#### Data Status")
    col_s1, col_s2 = st.columns(2)

    # ── CVI ───────────────────────────────────────────────────────────────
    with col_s1:
        if not cvi_df.empty:
            st.success(f"✅ CVI loaded: {len(cvi_df.columns)} columns, "
                       f"{len(cvi_df)} rows")
            # Always show replace button even when loaded
            if st.button("🔄 Replace CVI", key=f"clr_cvi_{db}",
                         use_container_width=True):
                gs("cvi", {}).pop(formula_name, None)
                # Remove from disk so it won't auto-reload
                cvi_path = _gdirs["CVI"] / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_path.unlink()
                cvi_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning(f"⚠️ No CVI loaded for {formula_name}")

        # Upload always visible — key uses a counter so it resets after clear
        upload_key_cvi = f"up_cvi_{db}_{S.get('cvi_upload_v', {}).get(db, 0)}"
        up_cvi = st.file_uploader(
            "Upload CVI file (CSV, Excel, HTML, JSON, …):",
            type=_UPLOAD_TYPES, key=upload_key_cvi
        )
        if up_cvi:
            df_cvi_up = _read_uploaded(up_cvi)
            cvi_save = _gdirs["CVI"] / f"CVI_{formula_name}.csv"
            df_cvi_up.to_csv(cvi_save, index=False)
            st.session_state.setdefault(gkey("cvi"), {})[formula_name] = df_cvi_up
            # Bump version so key resets next time Replace is clicked
            if "cvi_upload_v" not in S: S["cvi_upload_v"] = {}
            S["cvi_upload_v"][db] = S["cvi_upload_v"].get(db, 0) + 1
            cvi_df = df_cvi_up
            st.success(f"✅ CVI saved → Container_Variable_Inputs/CVI_{formula_name}.csv")
            st.rerun()

    # ── Main Data ─────────────────────────────────────────────────────────
    with col_s2:
        if not main_df.empty:
            st.success(f"✅ Main Data: {len(main_df):,} rows × "
                       f"{len(main_df.columns)} cols")
            if st.button("🔄 Replace Main Data", key=f"clr_md_{db}",
                         use_container_width=True):
                gs_set("main_data",      pd.DataFrame())
                gs_set("main_data_path", "")   # drop stale path so DuckDB can't misfire
                main_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning("⚠️ No Main Data loaded")

        upload_key_md = f"up_md_{db}_{S.get('md_upload_v', {}).get(db, 0)}"
        up_md = st.file_uploader(
            "Upload Main Data file (CSV, Excel, HTML, JSON, …):",
            type=_UPLOAD_TYPES, key=upload_key_md
        )
        if up_md:
            df_md_up = _read_uploaded(up_md)
            for col in df_md_up.columns:
                df_md_up[col] = pd.to_numeric(df_md_up[col], errors="coerce")
            gs_set("main_data",      df_md_up)
            # Uploaded in-memory data has no stable on-disk source; clear the
            # path so the DuckDB pre-pass guard won't match it to another file.
            gs_set("main_data_path", "")
            if "md_upload_v" not in S: S["md_upload_v"] = {}
            S["md_upload_v"][db] = S["md_upload_v"].get(db, 0) + 1
            main_df = df_md_up
            st.success(f"✅ Main Data: {len(df_md_up):,} rows × "
                       f"{len(df_md_up.columns)} cols")
            st.rerun()

    st.markdown("---")
    # Selected Count
    st.markdown("**Selected Count input**")

    # Show if auto-loaded from file
    if sc_auto:
        st.success(f"✅ Selected Counts loaded from SC_{formula_name}*.csv")
        st.caption("  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))

    # SC upload — always replaceable
    with st.expander("📂 Upload / Replace Selected Count file"):
        st.caption("File must have columns: w, Selected Count")
        upload_key_sc = f"up_sc_{db}_{S.get('sc_upload_v', {}).get(db, 0)}"
        up_sc = st.file_uploader("Upload SC file (CSV, Excel, HTML, JSON, …)",
                                  type=_UPLOAD_TYPES, key=upload_key_sc)
        if up_sc:
            df_sc_up = _read_uploaded(up_sc)
            sp_sc = _gdirs["Selected_Counts"] / f"SC_{formula_name}.csv"
            df_sc_up.to_csv(sp_sc, index=False)
            if "sc_upload_v" not in S: S["sc_upload_v"] = {}
            S["sc_upload_v"][db] = S["sc_upload_v"].get(db, 0) + 1
            st.success(f"✅ Saved → SC_{formula_name}.csv")
            st.rerun()

    c1,c2 = st.columns(2)
    with c1:
        sc_method = st.radio("Method:", ["Custom","Same for all","Range"],
                             horizontal=True, key=f"scm_{db}")
    with c2:
        # Show per-column summary — do NOT merge into one string
        if sc_auto:
            st.caption("Per-column SC (from file): " +
                       "  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))
        fallback_sc = st.text_input(
            "Fallback Count (used if no SC file loaded):", "7",
            key=f"scr_{db}")

    # Build sc_dict: per w-column, specific threshold
    # sc_auto keys look like "w1","w2",… or "1","2",… normalise to "w1","w2"
    sc_dict = {}
    for k, v in sc_auto.items():
        key = k if k.startswith("w") else f"w{k}"
        sc_dict[key] = [int(x.strip()) for x in str(v).split(",")
                        if x.strip().lstrip('-').isdigit()]

    # If no SC file loaded, use fallback for all w-columns
    if not sc_dict and fallback_sc:
        fallback_list = [int(x.strip()) for x in fallback_sc.split(",")
                         if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = fallback_list

    if sc_method == "Same for all" and fallback_sc:
        same_list = [int(x.strip()) for x in fallback_sc.split(",")
                     if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = same_list

    # ── Carry-forward direction toggles ───────────────────────────────────
    st.markdown("---")
    st.markdown("**Carry-forward direction — which pool feeds the next row**")
    st.caption("U = Unselected carries forward (default)  |  "
               "S = Selected carries forward  |  "
               "Toggle at Row N controls what Row N receives from Row N-1.")

    # Initialise carry_fwd in session state per dashboard
    cf_key = f"carry_fwd_{db}"
    if cf_key not in S:
        S[cf_key] = {}

    # Get w-columns from loaded CVI or fallback
    w_keys_cf = sorted(
        [c for c in (cvi_df.columns if not cvi_df.empty else [])
         if re.match(r'^w\d+$', c)],
        key=lambda x: int(x[1:])
    )

    # ── Compact carry-forward control ─────────────────────────────────────
    # Global default + per-row overrides via a single editable table.
    # No per-column button grid — works cleanly with any number of w-columns.
    gd1, gd2, gd3 = st.columns([1.5, 1.5, 5])
    with gd1:
        global_dir = st.selectbox(
            "Default direction:",
            ["U — Unselected", "S — Selected"],
            key=f"gdir_{db}",
            help="Sets carry-forward direction for ALL rows at once."
        )
        _gd = "U" if global_dir.startswith("U") else "S"
    with gd2:
        if st.button("Apply to all", key=f"gdir_apply_{db}",
                     use_container_width=True):
            for wk in w_keys_cf:
                S[cf_key][wk] = _gd
            st.rerun()
    with gd3:
        st.caption(
            f"{len(w_keys_cf)} w-columns loaded. "
            "Edit the Direction column in the table below to override individual rows. "
            "U = Unselected carries forward · S = Selected carries forward."
        )

    if w_keys_cf:
        # Build override table: one row per w-column
        _cf_rows = [{"Row": f"Row {int(wk[1:])}  ·  {wk}",
                     "Direction": S[cf_key].get(wk, "U")}
                    for wk in w_keys_cf]
        _cf_edit_df = pd.DataFrame(_cf_rows)
        _cf_edited = st.data_editor(
            _cf_edit_df,
            key=f"cf_tbl_{db}",
            use_container_width=True,
            num_rows="fixed",
            height=min(40 * len(w_keys_cf) + 50, 340),
            column_config={
                "Row":       st.column_config.TextColumn("Row", disabled=True,
                                                          width="medium"),
                "Direction": st.column_config.SelectboxColumn(
                                 "Direction", options=["U", "S"], width="small"),
            },
            hide_index=True,
        )
        # Persist edits back into session state
        for _idx, _row in _cf_edited.iterrows():
            _wk = w_keys_cf[_idx]
            S[cf_key][_wk] = _row["Direction"]

    # Build final carry_fwd dict (defaults to U)
    carry_fwd = {wk: S[cf_key].get(wk, "U") for wk in w_keys_cf}

    # ── Is SC Available toggle ────────────────────────────────────────────
    sc_avail_key = f"sc_avail_{db}"
    if sc_avail_key not in S:
        S[sc_avail_key] = "YES"

    # ── Individual run cluster label ───────────────────────────────────────
    ind_cl1, ind_cl2, ind_cl3 = st.columns([2, 2, 4])
    with ind_cl1:
        ind_cluster = st.text_input(
            "Cluster label for this run:",
            value="1n",
            key=f"ind_cluster_{db}",
            help="Used in output filenames: e.g. 1n, Draw1567, Test01"
        )
    with ind_cl2:
        ind_lotto = st.selectbox(
            "Lotto type:",
            _lotto_keys,
            index=_default_lotto_idx,
            format_func=lambda x: f"{x}",
            key=f"ind_lotto_{db}"
        )
    with ind_cl3:
        ind_draw = st.text_input(
            "Draw number:",
            value="D????",
            key=f"ind_draw_{db}",
            help="e.g. D1567"
        )

    # File prefix for individual run outputs
    ind_prefix = f"{ind_cluster}_{ind_lotto}_{ind_draw}_{formula_name}"
    st.markdown("---")
    sa1, sa2, _ = st.columns([1,1,6])
    with sa1:
        if st.button("🟢 SC: YES (Auto)", key=f"sca_y_{db}",
                     type="primary" if S[sc_avail_key]=="YES" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "YES"
            S.pop(f"step_state_{db}", None)
            S.pop(f"step_pending_{db}", None)
    with sa2:
        if st.button("🔴 SC: NO (Manual)", key=f"sca_n_{db}",
                     type="primary" if S[sc_avail_key]=="NO" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "NO"
            S.pop(f"step_state_{db}", None)
            S.pop(f"step_pending_{db}", None)
    if S[sc_avail_key] == "NO":
        st.caption("Manual mode — engine pauses after each stage; "
                   "pick SC values, then Continue.")
        st.markdown("---")
        can_run = not main_df.empty and not cvi_df.empty
        if not can_run:
            st.info("📤 Upload CVI and Main Data above to enable matching.")

        if S.get(f"step_state_{db}") is None:
            # ── No step-run active: show START button ─────────────────────
            if st.button(f"▶ START MATCHING (Step Mode) — {db}",
                         type="primary", use_container_width=True,
                         key=f"start_step_{db}", disabled=not can_run):
                with st.spinner("Running to first SC decision…"):
                    _sr = run_matching_step(
                        resume_state=None,
                        main_df=main_df, cvi_df=cvi_df,
                        carry_fwd=carry_fwd,
                        main_path=gs("main_data_path"))
                if _sr["paused"]:
                    S[f"step_state_{db}"]   = _sr["resume_state"]
                    S[f"step_pending_{db}"] = {
                        "w":                     _sr["w"],
                        "count_dist":            _sr["count_dist"],
                        "awaiting_sc_for_stage": _sr["awaiting_sc_for_stage"],
                    }
                else:
                    st.session_state.setdefault(gkey("results"), {})[db] = _sr
                st.rerun()

        else:
            # ── Paused mid-run: show distribution + SC picker + Continue ──
            _pending = S[f"step_pending_{db}"]
            _stage_w = _pending["w"]
            _stage_i = _pending["awaiting_sc_for_stage"]
            _cd      = _pending["count_dist"]

            # ── All-stages preview table ────────────────────────────────
            _rs        = S[f"step_state_{db}"]
            _AWAIT     = "— (awaiting SC)"
            _COL_ORDER = [
                "Row", "Main\nData", "CVI", "Dir",
                "Main\nCount", "Main\nBreakdown",
                "Present\nData", "Present\nCount",
                "SC", "Selected", "Sel\nBreakdown",
                "Unsel\nCount", "Unselected", "Unsel\nBreakdown",
            ]
            _preview_rows = []
            for _pi, _pw in enumerate(_rs["w_cols"]):
                _pw_num = int(_pw[1:])
                if _pi < _rs["stage_idx"]:
                    _preview_rows.append(_rs["fig9_rows"][_pi])
                elif _pi == _rs["stage_idx"]:
                    _preview_rows.append({
                        "Row":             f"Row {_pw_num}",
                        "Main\nData":      f"M:{_rs['M']}",
                        "CVI":             _pw,
                        "Dir":             _rs["carry_fwd"].get(_pw, "U"),
                        "Main\nCount":     _rs["main_count_map"].get(_pw, "—"),
                        "Main\nBreakdown": _rs["main_bd_map"].get(_pw, "—"),
                        "Present\nData":   _rs["cached_stage_info"]["present_label"],
                        "Present\nCount":  ",".join(
                            str(int(k[1:])) for k in sorted(_cd)
                        ),
                        "SC":              _AWAIT,
                        "Selected":        _AWAIT,
                        "Sel\nBreakdown":  _AWAIT,
                        "Unsel\nCount":    _AWAIT,
                        "Unselected":      _AWAIT,
                        "Unsel\nBreakdown":_AWAIT,
                    })
                else:
                    _preview_rows.append({
                        "Row":             f"Row {_pw_num}",
                        "Main\nData":      f"M:{_rs['M']}",
                        "CVI":             _pw,
                        "Dir":             _rs["carry_fwd"].get(_pw, "U"),
                        "Main\nCount":     _rs["main_count_map"].get(_pw, "—"),
                        "Main\nBreakdown": _rs["main_bd_map"].get(_pw, "—"),
                        "Present\nData":   "—",
                        "Present\nCount":  "—",
                        "SC":              "—",
                        "Selected":        "—",
                        "Sel\nBreakdown":  "—",
                        "Unsel\nCount":    "—",
                        "Unselected":      "—",
                        "Unsel\nBreakdown":"—",
                    })
            show_paginated_df(
                pd.DataFrame(_preview_rows)[_COL_ORDER],
                key=f"step_preview_{db}_{_stage_i}",
                use_container_width=True, hide_index=True,
            )
            st.markdown("---")

            st.write(f"**Paused at {_stage_w}** — choose SC for this stage")

            if _cd:
                st.dataframe(
                    pd.DataFrame([
                        {"Count k": int(k[1:]), "Rows": v}
                        for k, v in sorted(_cd.items())
                    ]),
                    use_container_width=False, hide_index=True,
                )
            else:
                st.info("No rows enter this stage (pool exhausted — "
                        "any SC choice will complete the run).")

            _available_ks = sorted(int(k[1:]) for k in _cd)
            _sc_chosen = st.multiselect(
                f"Select count values for SC (stage {_stage_i}, {_stage_w}):",
                _available_ks,
                key=f"step_sc_choice_{db}_{_stage_i}",
            )

            _cc1, _cc2 = st.columns([2, 1])
            with _cc1:
                if st.button(f"▶ Continue — {db}", type="primary",
                             use_container_width=True,
                             key=f"continue_step_{db}"):
                    with st.spinner(f"Applying SC for {_stage_w}…"):
                        _sr = run_matching_step(
                            S[f"step_state_{db}"],
                            sc_for_stage=_sc_chosen)
                    if _sr["paused"]:
                        S[f"step_state_{db}"]   = _sr["resume_state"]
                        S[f"step_pending_{db}"] = {
                            "w":                     _sr["w"],
                            "count_dist":            _sr["count_dist"],
                            "awaiting_sc_for_stage": _sr["awaiting_sc_for_stage"],
                        }
                    else:
                        st.session_state.setdefault(
                            gkey("results"), {})[db] = _sr
                        S.pop(f"step_state_{db}", None)
                        S.pop(f"step_pending_{db}", None)
                    st.rerun()
            with _cc2:
                if st.button("↺ Cancel step run", use_container_width=True,
                             key=f"cancel_step_{db}"):
                    S.pop(f"step_state_{db}", None)
                    S.pop(f"step_pending_{db}", None)
                    st.rerun()

    else:
        st.caption("Automated mode: all stages run sequentially without stopping.")

        # RUN MATCHING
        st.markdown("---")
        can_run = not main_df.empty and not cvi_df.empty
        if not can_run:
            st.info("📤 Upload CVI and Main Data above to enable matching.")

        if st.button(f"▶ RUN MATCHING — {db}", type="primary",
                     use_container_width=True, key=f"run_{db}",
                     disabled=not can_run):
            n = len(main_df)
            st.info(f"Matching M={n:,} rows × {len(w_keys_cf)} w-columns…")
            with st.spinner("Matching engine running…"):
                # Pass the source path (if known). run_matching only uses it when
                # it is safe — large CSV, DuckDB present, row-count == M, contiguous
                # n-columns — and falls back to pandas otherwise.
                res = run_matching(main_df, cvi_df, sc_dict, carry_fwd,
                                   main_path=gs("main_data_path"))
            st.session_state.setdefault(gkey("results"), {})[db] = res

    # ── Display results (persists after run) ──────────────────────────────
    if db in gs("results", {}):
        res   = gs("results", {})[db]
        fig9  = res["fig9_table"]
        sel   = res["selected"]
        unsel = res["unselected"]
        bd    = res["breakdown"]
        dbg   = res.get("debug_rows", [])
        small = res.get("small_enough", True)
        n_cols_res = res.get("n_cols", [])

        st.success(f"✅ Final stage — Selected: {len(sel):,}  ·  "
                   f"Unselected: {len(unsel):,}")

        # ── Helper: show df with count filter + highlights inline ──────
        def show_filtered_highlighted(df_in, cvi_set_in, n_cols_in,
                                      section_key, default_counts=None):
            if df_in is None or df_in.empty:
                st.info("No data.")
                return
            count_col = "Count" if "Count" in df_in.columns else None
            avail_counts = []
            if count_col:
                avail_counts = sorted(df_in[count_col].dropna()
                                      .astype(int).unique().tolist())
            if avail_counts:
                filt = st.multiselect(
                    "Filter by count:", avail_counts,
                    default=default_counts or avail_counts,
                    key=f"flt_{db}_{section_key}"
                )
                df_show = df_in[df_in[count_col].isin(filt)] if filt else df_in
            else:
                df_show = df_in
            if df_show.empty:
                st.info("No rows match the selected count filter.")
                return
            # Highlight matching cells
            nc = [c for c in n_cols_in if c in df_show.columns]
            if nc and cvi_set_in:
                def _style(val):
                    try:
                        if int(round(float(val))) in cvi_set_in:
                            return ("background-color:#d4a0a0;"
                                    "color:#000;font-weight:bold")
                    except (ValueError, TypeError):
                        pass
                    return ""
                def _cnt_style(val):
                    return "background-color:#ffffaa;color:#000;font-weight:bold"
                try:
                    styled = df_show.style.applymap(_style, subset=nc)
                    if count_col:
                        styled = styled.applymap(_cnt_style,
                                                 subset=[count_col])
                    show_paginated_df(styled, key="cd_formula_result_styled", use_container_width=True, hide_index=True)
                except Exception:
                    show_paginated_df(df_show, key="cd_formula_result_plain", use_container_width=True, hide_index=True)
            else:
                show_paginated_df(df_show, key="cd_formula_result_unstyled", use_container_width=True, hide_index=True)

        # ── Interactive Matching Table — per-cell popover detail ──────
        st.markdown("#### Matching Table")
        if not fig9.empty:
            _dbg_by_w = {d["w"]: d for d in dbg}

            # Column config: (header label, relative width)
            _MT = [
                ("Row",        1.5), ("M",          1.2), ("CVI",       1.0),
                ("Dir",        0.6), ("Main Ct",    2.5), ("Main Bkdn", 3.5),
                ("Present",    1.2), ("Pres Ct",    2.5), ("SC",        0.8),
                ("Selected",   1.2), ("Sel Bkdn",   3.5),
                ("Unsel Ct",   1.2), ("Unselected", 1.2), ("Unsel Bkdn",3.5),
            ]
            _widths = [c[1] for c in _MT]

            # Header row
            _hdrs = st.columns(_widths)
            for _hc, (_hn, _) in zip(_hdrs, _MT):
                _hc.markdown(f"<small><b>{_hn}</b></small>",
                             unsafe_allow_html=True)
            st.divider()

            def _trunc(s, n=16):
                s = str(s)
                return (s[:n] + "…") if len(s) > n else s

            for _ri, _fr in fig9.reset_index(drop=True).iterrows():
                _w       = _fr["CVI"]
                _d       = _dbg_by_w.get(_w, {})
                _cvi_set = _d.get("cvi_set", set())
                _sc_set  = set(_d.get("sc", []))
                _rc      = st.columns(_widths)

                # 0 · Row — plain
                _rc[0].write(_fr["Row"])

                # 1 · M:29 — popover → main data table
                _fr_main_data = _fr["Main\nData"]
                with _rc[1].popover(_fr_main_data,
                                    use_container_width=True):
                    st.markdown(f"**Main Data — {_fr_main_data}**")
                    _md_wc = _d.get("main_df_wc")
                    if _md_wc is not None and not _md_wc.empty:
                        show_paginated_df(_md_wc, key=f"pop_md_wc_{_w}_{_ri}", use_container_width=True)
                    else:
                        st.info("Full data not stored "
                                "(M > threshold or no CVI).")

                # 2 · CVI (w1…) — popover → CVI numbers list
                with _rc[2].popover(_fr["CVI"],
                                    use_container_width=True):
                    st.markdown(f"**CVI numbers for {_w}** "
                                f"({len(_cvi_set)} values):")
                    st.write(sorted(_cvi_set) if _cvi_set else "— None —")

                # 3 · Dir — plain
                _rc[3].write(_fr["Dir"])

                # 4 · Main Count — popover → full count string
                with _rc[4].popover(_trunc(_fr["Main\nCount"]),
                                    use_container_width=True):
                    st.markdown("**Main Count distribution:**")
                    st.write(str(_fr["Main\nCount"]))

                # 5 · Main Breakdown — popover → breakdown + dist table
                with _rc[5].popover(_trunc(_fr["Main\nBreakdown"]),
                                    use_container_width=True):
                    st.markdown("**Main Breakdown:**")
                    st.write(str(_fr["Main\nBreakdown"]))
                    _cd = _d.get("count_dist", {})
                    if _cd:
                        show_paginated_df(pd.DataFrame([
                            {"Count": int(k[1:]), "Rows": v,
                             "Status": "✅ Sel" if int(k[1:]) in _sc_set
                                       else "— Unsel"}
                            for k, v in sorted(_cd.items())
                        ]), key=f"pop_cd_{_w}_{_ri}", hide_index=True, use_container_width=True)

                # 6 · Present Data — popover → row count entering stage
                _fr_present_data = _fr["Present\nData"]
                with _rc[6].popover(_fr_present_data,
                                    use_container_width=True):
                    st.markdown(f"**Present at this stage:** "
                                f"{_fr_present_data}")
                    st.write(f"Rows entering: "
                             f"{_d.get('present_in', '—')}")

                # 7 · Present Count — popover → count breakdown string
                with _rc[7].popover(_trunc(_fr["Present\nCount"]),
                                    use_container_width=True):
                    st.markdown("**Present Count distribution:**")
                    st.write(_d.get("pres_count_str",
                                    str(_fr["Present\nCount"])))

                # 8 · SC — plain
                _rc[8].write(_fr["SC"])

                # 9 · Selected — popover → sel_df table + CSV download
                with _rc[9].popover(_fr["Selected"],
                                    use_container_width=True):
                    st.markdown(f"**Selected rows — {_fr['Selected']}**")
                    _sel_df_p = _d.get("sel_df")
                    if _sel_df_p is not None and not _sel_df_p.empty:
                        show_paginated_df(_sel_df_p, key=f"pop_sel_{db}_{_ri}_df", use_container_width=True)
                        st.download_button(
                            f"⬇ {_w} Selected CSV",
                            to_csv_bytes(_sel_df_p),
                            f"{ind_prefix}_{_w}_sel.csv",
                            "text/csv",
                            key=f"pop_sel_{db}_{_ri}",
                        )
                    elif _d.get("selected_n", 0) == 0:
                        st.info("No rows selected at this stage.")
                    else:
                        st.info(f"S:{_d['selected_n']} — not stored "
                                f"(M > {DISPLAY_THRESHOLD:,})")

                # 10 · Sel Breakdown — popover → breakdown text
                with _rc[10].popover(_trunc(_fr["Sel\nBreakdown"]),
                                     use_container_width=True):
                    st.markdown("**Selected Breakdown:**")
                    st.write(str(_fr["Sel\nBreakdown"]))

                # 11 · Unsel Count — popover → unsel count string
                with _rc[11].popover(_trunc(_fr["Unsel\nCount"]),
                                     use_container_width=True):
                    st.markdown("**Unselected Count distribution:**")
                    st.write(str(_fr["Unsel\nCount"]))

                # 12 · Unselected — popover → unsel_df table + CSV download
                with _rc[12].popover(_fr["Unselected"],
                                     use_container_width=True):
                    st.markdown(f"**Unselected rows — {_fr['Unselected']}**")
                    _unsel_df_p = _d.get("unsel_df")
                    if _unsel_df_p is not None and not _unsel_df_p.empty:
                        show_paginated_df(_unsel_df_p, key=f"pop_unsel_{db}_{_ri}_df", use_container_width=True)
                        st.download_button(
                            f"⬇ {_w} Unselected CSV",
                            to_csv_bytes(_unsel_df_p),
                            f"{ind_prefix}_{_w}_unsel.csv",
                            "text/csv",
                            key=f"pop_unsel_{db}_{_ri}",
                        )
                    elif _d.get("unselected_n", 0) == 0:
                        st.info("No unselected rows remaining.")
                    else:
                        st.info(f"U:{_d['unselected_n']} — not stored "
                                f"(M > {DISPLAY_THRESHOLD:,})")

                # 13 · Unsel Breakdown — popover → breakdown text
                with _rc[13].popover(_trunc(_fr["Unsel\nBreakdown"]),
                                     use_container_width=True):
                    st.markdown("**Unselected Breakdown:**")
                    st.write(str(_fr["Unsel\nBreakdown"]))

            # Download the full table as CSV
            st.download_button("⬇ Matching Table CSV",
                               to_csv_bytes(fig9),
                               f"{ind_prefix}_matching_table.csv",
                               "text/csv", key=f"dl_fig9_{db}")

        # ── Per-row summary table + single-row inspector ──────────────
        # Show all rows as a compact summary table first, then let the
        # user pick ONE row to drill into — avoids rendering hundreds of
        # expanders at once.
        st.markdown("#### Row-by-Row Summary")
        _summary_rows = []
        for _d in dbg:
            _icon = "✅" if _d["selected_n"] > 0 else (
                    "⚪" if _d["present_in"] == 0 else "🔵")
            _note = f"  ⚠️{_d.get('note','')}" if _d.get("note") else ""
            _summary_rows.append({
                "":        _icon,
                "Row":     int(_d["w"][1:]),
                "w":       _d["w"],
                "Dir":     _d.get("direction", "U"),
                "Present": _d["present_in"],
                "SC":      str(list(_d["sc"])),
                "S":       _d["selected_n"],
                "U":       _d["unselected_n"],
                "Note":    _note.strip(),
            })
        if _summary_rows:
            _sum_df = pd.DataFrame(_summary_rows)
            show_paginated_df(_sum_df, key=f"cd_row_summary_{db}", use_container_width=True, hide_index=True)

            # ── One-click export of the full summary ───────────────────
            _bdl1, _bdl2 = st.columns(2)
            with _bdl1:
                st.download_button(
                    "⬇ Download Row Summary CSV",
                    to_csv_bytes(_sum_df.drop(columns=[""])),
                    f"{ind_prefix}_row_summary.csv",
                    "text/csv",
                    key=f"dl_sum_csv_{db}",
                    use_container_width=True,
                )
            with _bdl2:
                try:
                    import io as _sum_io
                    _sum_buf = _sum_io.BytesIO()
                    with pd.ExcelWriter(_sum_buf, engine="openpyxl") as _sum_xl:
                        _export_df = _sum_df.drop(columns=[""]).copy()
                        _export_df.to_excel(_sum_xl, sheet_name="Row_Summary",
                                            index=False)
                        # Conditional formatting: green S>0, grey S=0
                        from openpyxl.styles import PatternFill as _PF, Font as _Fnt
                        _ws = _sum_xl.sheets["Row_Summary"]
                        _green = _PF("solid", fgColor="C6EFCE")
                        _red   = _PF("solid", fgColor="FFC7CE")
                        _s_col = list(_export_df.columns).index("S") + 1
                        _u_col = list(_export_df.columns).index("U") + 1
                        for _r in range(2, len(_export_df) + 2):
                            _sc = _ws.cell(row=_r, column=_s_col)
                            _uc = _ws.cell(row=_r, column=_u_col)
                            try:
                                if int(_sc.value or 0) > 0:
                                    _sc.fill = _green
                                else:
                                    _sc.fill = _red
                            except (ValueError, TypeError):
                                pass
                    _sum_buf.seek(0)
                    st.download_button(
                        "⬇ Download Row Summary Excel (S highlighted)",
                        _sum_buf.getvalue(),
                        f"{ind_prefix}_row_summary.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_sum_xl_{db}",
                        use_container_width=True,
                    )
                except Exception as _sum_xl_ex:
                    st.warning(f"Excel export unavailable: {_sum_xl_ex}")

        # ── Single-row inspector ───────────────────────────────────────
        st.markdown("#### Inspect a Row")
        _row_labels = [
            f"Row {int(d['w'][1:])}  ·  {d['w']}  ·  "
            f"Dir:{d.get('direction','U')}  ·  Present:{d['present_in']}  ·  "
            f"S:{d['selected_n']}  ·  U:{d['unselected_n']}"
            for d in dbg
        ]
        if _row_labels:
            _sel_row_label = st.selectbox(
                "Select row to inspect:",
                _row_labels,
                key=f"row_inspect_{db}"
            )
            _sel_i = _row_labels.index(_sel_row_label)
            d = dbg[_sel_i]
            i = _sel_i
            w_lbl     = d["w"]
            cvi_set_d = d.get("cvi_set", set(d.get("cvi_numbers", [])))
            n_cols_d  = d.get("n_cols", [])

            if not cvi_set_d:
                st.warning(f"{w_lbl}: No CVI numbers — column is empty. "
                           "All present rows carry forward.")
            else:
                # CVI numbers + distribution
                cca, ccb = st.columns(2)
                with cca:
                    st.markdown(f"**{w_lbl} numbers ({len(cvi_set_d)}):**")
                    st.write(sorted(cvi_set_d))
                with ccb:
                    st.markdown("**Count distribution (present rows):**")
                    if d["count_dist"]:
                        dist_df = pd.DataFrame([
                            {"Count": int(k[1:]),
                             "Rows": v,
                             "": "✅ Sel" if int(k[1:]) in d["sc"]
                                 else "— Unsel"}
                            for k, v in sorted(d["count_dist"].items())
                        ])
                        show_paginated_df(dist_df, key=f"cd_dist_df_{_w}_{_ri}", hide_index=True, use_container_width=True)
                    else:
                        st.info("—")

                st.markdown("---")

                # Main Data breakdown (collapsible)
                with st.expander(
                    f"📊 Main Data Breakdown  ·  M:{len(main_df)}"
                    f"  ·  {d.get('main_bd_str','—')}"
                ):
                    md_wc = d.get("main_df_wc")
                    if md_wc is not None and not md_wc.empty:
                        show_filtered_highlighted(
                            md_wc, cvi_set_d, n_cols_d, f"main_{i}")
                        c1e, c2e = st.columns(2)
                        with c1e:
                            st.download_button(
                                "⬇ Main Breakdown CSV",
                                to_csv_bytes(md_wc),
                                f"{ind_prefix}_{w_lbl}_main.csv",
                                "text/csv", key=f"dl_main_{db}_{i}")
                        with c2e:
                            st.download_button(
                                "⬇ Main Breakdown Excel (highlighted)",
                                to_styled_excel(md_wc, cvi_set_d, n_cols_d,
                                                f"{w_lbl}_Main"),
                                f"{ind_prefix}_{w_lbl}_main.xlsx",
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                                key=f"dl_main_xl_{db}_{i}")
                    else:
                        st.info("Main data breakdown not available "
                                f"(M>{DISPLAY_THRESHOLD:,} or no CVI).")

                st.markdown("---")

                # Selected + Unselected side by side (collapsible)
                ts, tu = st.columns(2)
                with ts:
                    d_sel = d.get("sel_df")
                    with st.expander(
                        f"✅ Selected  S:{d['selected_n']}", expanded=True
                    ):
                        if d_sel is not None and not d_sel.empty:
                            show_filtered_highlighted(
                                d_sel, cvi_set_d, n_cols_d,
                                f"sel_{i}",
                                default_counts=list(d["sc"]))
                            c1s, c2s = st.columns(2)
                            with c1s:
                                st.download_button(
                                    f"⬇ {w_lbl} Selected CSV",
                                    to_csv_bytes(d_sel),
                                    f"{ind_prefix}_{w_lbl}_sel.csv",
                                    "text/csv", key=f"dl_sel_{db}_{i}")
                            with c2s:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(d_sel, cvi_set_d,
                                                    n_cols_d, f"{w_lbl}_Sel"),
                                    f"{ind_prefix}_{w_lbl}_sel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_sel_xl_{db}_{i}")
                        elif d["selected_n"] == 0:
                            st.info("No rows selected at this stage.")
                        else:
                            st.info(f"S:{d['selected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

                with tu:
                    d_unsel = d.get("unsel_df")
                    with st.expander(
                        f"🔵 Unselected  U:{d['unselected_n']} → fwd",
                        expanded=True
                    ):
                        if d_unsel is not None and not d_unsel.empty:
                            show_filtered_highlighted(
                                d_unsel, cvi_set_d, n_cols_d, f"unsel_{i}")
                            c1u, c2u = st.columns(2)
                            with c1u:
                                st.download_button(
                                    f"⬇ {w_lbl} Unselected CSV",
                                    to_csv_bytes(d_unsel),
                                    f"{ind_prefix}_{w_lbl}_unsel.csv",
                                    "text/csv", key=f"dl_unsel_{db}_{i}")
                            with c2u:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(d_unsel, cvi_set_d,
                                                    n_cols_d, f"{w_lbl}_Unsel"),
                                    f"{ind_prefix}_{w_lbl}_unsel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_unsel_xl_{db}_{i}")
                        elif d["unselected_n"] == 0:
                            st.info("No rows remaining.")
                        else:
                            st.info(f"U:{d['unselected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

        # ── Final stage permanent tabs ─────────────────────────────────
        st.markdown("---")
        st.markdown("#### Final Stage Output")

        # Get final w CVI set
        final_cvi_set = set()
        if dbg:
            last_d = next((d for d in reversed(dbg)
                           if d.get("cvi_set")), None)
            if last_d:
                final_cvi_set = last_d.get("cvi_set", set())

        ft1, ft2, ft3 = st.tabs([
            f"Selected (final)  S:{len(sel)}",
            f"Unselected (final)  U:{len(unsel)}",
            "Breakdown S0/S1/S2… (all stages)",
        ])

        with ft1:
            if sel.empty:
                st.info("No rows selected at final stage.")
            else:
                n_f = [c for c in n_cols_res if c in sel.columns]
                show_filtered_highlighted(
                    sel, final_cvi_set, n_f, "final_sel",
                    default_counts=None)
                fpath = _gdirs["Outputs"] / f"{ind_prefix}_selected.csv"
                sel.to_csv(fpath, index=False)
                c1f, c2f = st.columns(2)
                with c1f:
                    st.download_button(
                        "⬇ Selected CSV", to_csv_bytes(sel),
                        f"{ind_prefix}_selected.csv","text/csv",
                        key=f"dl_s_{db}")
                with c2f:
                    st.download_button(
                        "⬇ Selected Excel (highlighted)",
                        to_styled_excel(sel, final_cvi_set, n_f, "Selected"),
                        f"{ind_prefix}_selected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_s_xl_{db}")

        with ft2:
            if unsel.empty:
                st.info("No rows remaining at final stage.")
            else:
                n_u = [c for c in n_cols_res if c in unsel.columns]
                show_filtered_highlighted(
                    unsel, final_cvi_set, n_u, "final_unsel")
                c1u2, c2u2 = st.columns(2)
                with c1u2:
                    st.download_button(
                        "⬇ Unselected CSV", to_csv_bytes(unsel),
                        f"{ind_prefix}_unselected.csv","text/csv",
                        key=f"dl_u_{db}")
                with c2u2:
                    st.download_button(
                        "⬇ Unselected Excel (highlighted)",
                        to_styled_excel(unsel, final_cvi_set, n_u,
                                        "Unselected"),
                        f"{ind_prefix}_unselected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_u_xl_{db}")

        with ft3:
            st.caption("S0=0 matches · S1=1 match … per w-column")
            if bd.empty:
                st.info("No breakdown data.")
            else:
                show_paginated_df(bd, key=f"cd_breakdown_{db}", use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇ Breakdown CSV", to_csv_bytes(bd),
                    f"{ind_prefix}_breakdown.csv","text/csv",
                    key=f"dl_bd_{db}")

    # Controls
    st.markdown("---")
    if db not in S["container_status"]:
        S["container_status"][db] = pd.DataFrame({
            "Name":[db],"Status":["Stopped"],"Memory Usage":[""],
            "CPU Usage":[""],"Percent progress %":[0],
            "Time Guestimate Min":[""],"Time Guestimate Max":[""],
        })

    b1,b2,b3,b4,b5,b6,b7 = st.columns(7)
    with b1:
        if st.button("▶ Start", type="primary", use_container_width=True, key=f"s1_{db}"):
            set_container_status(db,"Running"); st.toast("Started!")
    with b2:
        if st.button("■ Stop", use_container_width=True, key=f"s2_{db}"):
            set_container_status(db,"Stopped")
    with b3:
        if st.button("⚡ Kill", use_container_width=True, key=f"s3_{db}"):
            set_container_status(db,"Killed")
    with b4:
        if st.button("↺ Restart", use_container_width=True, key=f"s4_{db}"):
            st.toast("Restarting…")
    with b5:
        if st.button("⏸ Pause", use_container_width=True, key=f"s5_{db}"):
            set_container_status(db,"Paused")
    with b6:
        if st.button("▶ Resume", use_container_width=True, key=f"s6_{db}"):
            set_container_status(db,"Running")
    with b7:
        if st.button("🗑 Remove", use_container_width=True, key=f"s7_{db}"):
            st.toast("Removed.")

    S["container_status"][db] = st.data_editor(
        S["container_status"][db], key=f"stbl_{db}",
        use_container_width=True, num_rows="dynamic",
        column_config={
            "Status": st.column_config.SelectboxColumn(
                "Status", options=["Stopped","Running","Paused","Killed","Error"]),
            "Percent progress %": st.column_config.ProgressColumn(
                "Percent progress %", min_value=0, max_value=100),
        })

    c1,c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {db} Status CSV",
            to_csv_bytes(S["container_status"][db]),
            f"{db.replace(' ','_')}_status.csv","text/csv",key=f"dls_{db}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MASTER OUTPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Master Outputs":
    st.markdown('<span class="sec-hdr hdr-brown">📤 Master Outputs — Aggregated Results</span>',
                unsafe_allow_html=True)
    am_toggle("out")

    # Collect from all dashboards
    all_sel, all_unsel, all_bd = [], [], []
    for db in DASHBOARDS:
        res = gs("results", {}).get(db)
        if res:
            if not res["selected"].empty:
                s = res["selected"].copy(); s.insert(0,"Dashboard",db)
                all_sel.append(s)
            if not res["unselected"].empty:
                u = res["unselected"].copy(); u.insert(0,"Dashboard",db)
                all_unsel.append(u)
            if not res["breakdown"].empty:
                b = res["breakdown"].copy(); b.insert(0,"Dashboard",db)
                all_bd.append(b)

    n_done = len(all_sel)
    if n_done:
        st.markdown(f'<div class="ok">✅ Results from {n_done} dashboard(s).</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="warn">⚠️ No results yet. Run matching in Container Dashboards.</div>',
                    unsafe_allow_html=True)

    def cc(dfs):
        if not dfs: return pd.DataFrame()
        return pd.concat(dfs, ignore_index=True).drop_duplicates()

    def show_output_tab(df: pd.DataFrame, label: str, fname: str):
        if df.empty:
            st.info("No data yet.")
            return
        st.write(f"**{len(df):,} unique rows**")
        show_paginated_df(df, key=f"master_out_{fname}", use_container_width=True)
        df.to_csv(_gdirs["Outputs"]/fname, index=False)
        st.download_button(f"⬇ Export", to_csv_bytes(df), fname,"text/csv",
                           key=f"dl_out_{fname}")

    t1,t2,t3,t4,t5 = st.tabs([
        "Cluster List Combined Selected",
        "Cluster Combined Selected",
        "Cluster List Combined Unselected",
        "Cluster Combined Unselected",
        "Selected Breakdown S0/S1/S2…",
    ])

    with t1:
        show_output_tab(cc(all_sel), "Cluster List Combined Selected",
                        "cluster_list_combined_selected.csv")
    with t2:
        df_t2 = cc(all_sel)
        nc = [c for c in df_t2.columns if re.match(r'^n\d+$',c)] if not df_t2.empty else []
        show_output_tab(df_t2[nc].drop_duplicates() if nc else pd.DataFrame(),
                        "Cluster Combined Selected", "cluster_combined_selected.csv")
    with t3:
        show_output_tab(cc(all_unsel), "Cluster List Combined Unselected",
                        "cluster_list_combined_unselected.csv")
    with t4:
        df_t4 = cc(all_unsel)
        nc4 = [c for c in df_t4.columns if re.match(r'^n\d+$',c)] if not df_t4.empty else []
        show_output_tab(df_t4[nc4].drop_duplicates() if nc4 else pd.DataFrame(),
                        "Cluster Combined Unselected", "cluster_combined_unselected.csv")
    with t5:
        show_output_tab(cc(all_bd), "Selected Breakdown",
                        "selected_breakdown.csv")

# ════════════════════════════════════════════════════════════════════════════
# PAGE: CLUSTER MANAGER
# ════════════════════════════════════════════════════════════════════════════
elif page == "🗂️ Cluster Manager":
    st.markdown('<span class="sec-hdr hdr-purple">🗂️ Cluster Manager</span>',
                unsafe_allow_html=True)

    st.markdown("""
    <div class="info">
    Each cluster = one complete run of all active containers against one Main Data file.
    <b>1n</b> = current run · <b>1n+1</b> = next run · <b>1n-1</b> = previous run.
    Labels are user-assigned. System auto-assigns numeric ID (001, 002…).
    </div>
    """, unsafe_allow_html=True)

    clusters = load_clusters()

    if not clusters:
        st.info("No clusters run yet. Use the parallel run button on Container Dashboards.")
    else:
        # ── Cluster registry table ─────────────────────────────────────
        st.markdown("### All Clusters")
        reg_df = pd.DataFrame([{
            "ID":         c["id"],
            "Label":      c["label"],
            "Lotto":      c.get("lotto",""),
            "Draw":       c.get("draw_no",""),
            "Date":       c.get("draw_date",""),
            "Main Data":  c.get("main_file",""),
            "Containers": c.get("containers",0),
            "Status":     c.get("status",""),
            "Run at":     c.get("timestamp","")[:16].replace("T"," "),
        } for c in clusters])
        reg_df.index = range(1, len(reg_df)+1)
        show_paginated_df(reg_df, key="cm_reg_df", use_container_width=True)

        # ── Select cluster to inspect ──────────────────────────────────
        st.markdown("---")
        st.markdown("### Inspect a Cluster")
        cluster_ids = [f"{c['id']} — {c['label']}" for c in clusters]
        chosen_c = st.selectbox("Select cluster:", cluster_ids,
                                key="cm_chosen")
        chosen_idx = cluster_ids.index(chosen_c)
        c_data = clusters[chosen_idx]

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Cluster ID",   c_data["id"])
        c2.metric("Label",        c_data["label"])
        c3.metric("Lotto",        c_data.get("lotto",""))
        c4.metric("Containers",   c_data.get("containers",0))

        # ── Per-container results for this cluster ─────────────────────
        results = c_data.get("results", [])
        if results:
            st.markdown("**Container results:**")
            res_df = pd.DataFrame([{
                "Formula":    r["formula"],
                "Status":     "✅" if r["status"]=="complete" else "❌",
                "Selected":   r["selected_n"],
                "Unselected": r["unselected_n"],
                "Error":      r.get("error",""),
            } for r in results])
            res_df.index = range(1, len(res_df)+1)
            show_paginated_df(res_df, key=f"cm_res_df_{chosen_c}", use_container_width=True, hide_index=False)

        # ── Load and browse cluster output files ───────────────────────
        out_dir = Path(c_data.get("output_dir", ""))
        if out_dir.exists():
            st.markdown("---")
            st.markdown("### Browse Cluster Output Files")
            out_files = sorted(out_dir.glob("*.csv"))
            if out_files:
                chosen_file = st.selectbox(
                    "File:", [f.name for f in out_files],
                    key="cm_file"
                )
                fp = out_dir / chosen_file
                df_view = pd.read_csv(fp)
                st.write(f"**{len(df_view):,} rows · {len(df_view.columns)} cols**")
                show_paginated_df(df_view, key=f"cm_file_view_{chosen_file}", use_container_width=True, hide_index=True)
                st.download_button(f"⬇ {chosen_file}",
                                   df_view.to_csv(index=False).encode(),
                                   chosen_file,"text/csv",
                                   key=f"dl_cm_{chosen_file}")
            else:
                st.info("No output files found in this cluster's output folder.")

        # ── Cross-cluster comparison ───────────────────────────────────
        if len(clusters) >= 2:
            st.markdown("---")
            st.markdown("### Cross-cluster Comparison")
            st.caption("Compare selected/unselected counts across clusters "
                       "for the same formula.")

            all_formulas = sorted(set(
                r["formula"]
                for c in clusters for r in c.get("results",[])
            ))
            if all_formulas:
                comp_formula = st.selectbox("Formula to compare:",
                                            all_formulas, key="cm_comp")
                comp_rows = []
                for c in clusters:
                    res = next((r for r in c.get("results",[])
                                if r["formula"] == comp_formula), None)
                    if res:
                        comp_rows.append({
                            "Cluster ID":  c["id"],
                            "Label":       c["label"],
                            "Lotto":       c.get("lotto",""),
                            "Draw":        c.get("draw_no",""),
                            "Date":        c.get("draw_date",""),
                            "Selected":    res["selected_n"],
                            "Unselected":  res["unselected_n"],
                            "Status":      res["status"],
                        })
                if comp_rows:
                    comp_df = pd.DataFrame(comp_rows)
                    comp_df.index = range(1, len(comp_df)+1)
                    show_paginated_df(comp_df, key="cm_comp_df", use_container_width=True)
                    st.download_button(
                        f"⬇ Cross-cluster comparison CSV",
                        comp_df.to_csv(index=False).encode(),
                        f"cross_cluster_{comp_formula}.csv","text/csv",
                        key="dl_cm_comp"
                    )

        # ── Rename / delete cluster ────────────────────────────────────
        st.markdown("---")
        st.markdown("### Manage This Cluster")
        new_label = st.text_input("Rename cluster label:",
                                  value=c_data["label"],
                                  key="cm_rename")
        rc1,rc2,_ = st.columns([1,1,4])
        with rc1:
            if st.button("💾 Save label", key="cm_save_label"):
                clusters[chosen_idx]["label"] = new_label
                save_clusters(clusters)
                st.success("Label updated.")
                st.rerun()
        with rc2:
            if st.button("🗑 Delete cluster", key="cm_delete"):
                clusters.pop(chosen_idx)
                save_clusters(clusters)
                st.success("Cluster deleted from registry.")
                st.rerun()

