"""
masterapp.py — Syndicate System, Single File
All logic, scraping, matching, collation, and UI in one place.
Run: streamlit run masterapp.py
Requires: pip install streamlit pandas numpy openpyxl playwright requests beautifulsoup4
          playwright install chromium
"""

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════
import streamlit as st
import pandas as pd
import numpy as np
import re, json, time, random, asyncio, os
from pathlib import Path
from datetime import datetime
from multiprocessing import Pool, cpu_count

try:
    from playwright.async_api import async_playwright
    PW_OK = True
except ImportError:
    PW_OK = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQ_OK = True
except ImportError:
    REQ_OK = False

# ═══════════════════════════════════════════════════════════════════════════════
# 1. PATH AUTO-DETECTION
# ═══════════════════════════════════════════════════════════════════════════════
def find_root() -> Path:
    for base in [Path.home()/"Desktop", Path.home()/"Documents",
                 Path.home(), Path.cwd()]:
        p = base / "o_Automation_Suite"
        if p.is_dir():
            return p
        if base.is_dir():
            for sub in base.iterdir():
                try:
                    q = sub / "o_Automation_Suite"
                    if q.is_dir():
                        return q
                except Exception:
                    pass
    fb = Path(__file__).parent / "o_Automation_Suite"
    fb.mkdir(parents=True, exist_ok=True)
    return fb

ROOT = find_root()

# ── Auto-create .streamlit/config.toml to raise upload limit ──────────────
_streamlit_dir = Path(__file__).parent / ".streamlit"
_streamlit_dir.mkdir(exist_ok=True)
_config_path = _streamlit_dir / "config.toml"
if not _config_path.exists():
    _config_path.write_text(
        "[server]\n"
        "maxUploadSize = 10000\n"
        "maxMessageSize = 10000\n\n"
        "[browser]\n"
        "gatherUsageStats = false\n\n"
        "[runner]\n"
        "fastReruns = true\n"
    )

DIRS = {
    # ── Top-level ──────────────────────────────────────────────────────────
    "Main_Data":    ROOT / "Main_Data",          # user's manually-loaded main data
    "Outputs":      ROOT / "Outputs",            # final matched outputs
    "Formulas":     ROOT / "Formulas",           # Container Formula.xlsx lives here
    "Containers":   ROOT / "Containers",
    # ── Scraper pipeline ───────────────────────────────────────────────────
    # Step 1: raw scraper output lands here first
    "Scraper":      ROOT / "Variables" / "Scraper",
    # Step 2: after processing, moved here (the D variable)
    "Direct":       ROOT / "Variables" / "Variable_Elements" / "Direct",
    # ── Variable Elements ──────────────────────────────────────────────────
    "Base":         ROOT / "Variables" / "Variable_Elements" / "Base",
    "Splits":       ROOT / "Variables" / "Variable_Elements" / "Splits",
    "Splits_Combi": ROOT / "Variables" / "Variable_Elements" / "Splits_Combi",
    "Rainbow":      ROOT / "Variables" / "Variable_Elements" / "Rainbow",
    "ExcelPro":     ROOT / "Variables" / "Variable_Elements" / "ExcelPro",
    # ── CVI: collated outputs land here, dashboards load from here ─────────
    "CVI":          ROOT / "Variables" / "Container_Variable_Inputs",
    # ── Selected Counts: one file per formula row ──────────────────────────
    "Selected_Counts": ROOT / "Formulas" / "Selected_Counts",
    # ── Debug ─────────────────────────────────────────────────────────────
    "debug_html":   ROOT / "debug_html",
}
for d in DIRS.values():
    d.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
# 2. CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
TARGET_STATES = ["NSW","VIC","QLD","WA","SA","TAS","ACT","NT"]

STATE_POSTCODES = {
    "NSW": list(range(2000,3000)),
    "VIC": list(range(3000,4000)),
    "QLD": list(range(4000,5000)),
    "WA":  list(range(6000,7000)),
    "SA":  list(range(5000,5600)),
    "TAS": list(range(7000,7800)),
    "ACT": list(range(2600,2619)),
    "NT":  list(range(800, 1000)),
}

SCRAPE_URL = "https://www.thelott.com/syndicates?postcode={pc}"
SCRAPE_URL_WA = "https://www.lotterywest.wa.gov.au/play-online/syndicate-games?postcode={pc}"

SCRAPE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-AU,en;q=0.9",
    "Referer": "https://www.thelott.com/syndicates",
}

CF_ROWS = [
    (1,  "BRD",       "Formed",    ["B","R","D"]),
    (2,  "BSD",       "Formed",    ["B","Sp","D"]),
    (3,  "BSoD",      "Formed",    ["B","So","D"]),
    (4,  "SD",        "Formed",    ["Sp","D"]),
    (5,  "SoD",       "Formed",    ["So","D"]),
    (6,  "BD",        "Formed",    ["B","D"]),
    (7,  "BSSoD",     "Formed",    ["B","Sp","So","D"]),
    (8,  "BRDSSo",    "Formed",    ["B","R","D","Sp","So"]),
    (9,  "B1B2B3",    "Formed",    ["B1","B2","B3"]),
    (10, "R1R2R3",    "Formed",    ["R1","R2","R3"]),
    (11, "D1D2D3",    "Formed",    ["D1","D2","D3"]),
    (12, "S1S2S3",    "Formed",    ["Sp1","Sp2","Sp3"]),
    (13, "So1So2So3", "Formed",    ["So1","So2","So3"]),
    (14, "Xn",        "Formed",    ["Xn"]),
    (15, "RVI1",      "Ready-made",["RVI1"]),
    (16, "RVI2",      "Ready-made",["RVI2"]),
    (17, "Xnn",       "Ready-made",["Xnn"]),
]
DASHBOARDS = [f"1n & {r[1]}" for r in CF_ROWS]
COMP_MAP   = {r[1]: r[3] for r in CF_ROWS}

CHUNK_SIZE = 500_000  # rows per processing chunk

# ── Lotto type codes (worldwide — extendable) ─────────────────────────────
LOTTO_TYPES = {
    "oz":  "Oz Lotto (AUS 7/47)",
    "pb":  "Powerball (AUS 7+1/35+20)",
    "sfl": "Set for Life (AUS)",
    "Mon": "Monday Lotto (AUS 6/45)",
    "Wed": "Wednesday Lotto (AUS 6/45)",
    "Fri": "Friday Lotto (AUS 6/45)",
    "Sat": "Saturday Lotto (AUS 6/45)",
    # Add more as needed — no cap
}

# ═══════════════════════════════════════════════════════════════════════════════
# GAME CONFIGURATION — one entry per game
# ═══════════════════════════════════════════════════════════════════════════════
GAMES_CFG = {
    "pb": {
        "label":       "Powerball",
        "emoji":       "🔵",
        "pool":        35,
        "pick":        7,
        "draw_day":    "Thursday",
        "lottolyzer":  "https://en.lottolyzer.com/number-frequencies/australia/powerball",
        "b_sheet":     "w values Pb A (2)",
        "thelott_key": "pb",
    },
    "oz": {
        "label":       "Oz Lotto",
        "emoji":       "🟠",
        "pool":        47,
        "pick":        7,
        "draw_day":    "Tuesday",
        "lottolyzer":  "https://en.lottolyzer.com/number-frequencies/australia/oz-lotto",
        "b_sheet":     "oz (2)",
        "thelott_key": "oz",
    },
    "sat": {
        "label":       "Saturday Lotto",
        "emoji":       "🟡",
        "pool":        45,
        "pick":        6,
        "draw_day":    "Saturday",
        "lottolyzer":  "https://en.lottolyzer.com/number-frequencies/australia/saturday-lotto",
        "b_sheet":     "Ta (2)",
        "thelott_key": "sat",
    },
    "sfl": {
        "label":       "Set for Life",
        "emoji":       "🟢",
        "pool":        44,
        "pick":        7,
        "draw_day":    "Daily",
        "lottolyzer":  "https://en.lottolyzer.com/number-frequencies/australia/set-for-life",
        "b_sheet":     "sfl",
        "thelott_key": "sfl",
    },
    "mwf": {
        "label":       "Mon/Wed/Fri",
        "emoji":       "🟣",
        "pool":        45,
        "pick":        6,
        "draw_day":    "Mon, Wed, Fri",
        "lottolyzer":  "https://en.lottolyzer.com/number-frequencies/australia/monday-lotto",
        "b_sheet":     "Ta (2)",
        "thelott_key": "mwf",
    },
}

GAME_KEYS   = list(GAMES_CFG.keys())
GAME_LABELS = {k: f"{v['emoji']} {v['label']}" for k, v in GAMES_CFG.items()}

# ── Game name → folder key mapping (handles brand name variations per state) ──
GAME_NAME_MAP = {
    # Saturday Lotto — multiple brand names across states
    "TattsLotto":               "sat",   # NSW, VIC, QLD brand
    "Saturday Lotto":           "sat",
    "Gold Lotto":               "sat",   # QLD brand
    "X Lotto":                  "sat",   # SA brand
    "Lotto":                    "sat",   # generic fallback

    # Powerball
    "Powerball":                "pb",

    # Oz Lotto
    "Oz Lotto":                 "oz",

    # Mon/Wed/Fri — THE KEY FIX: actual name in data is "Monday & Wednesday Lotto"
    "Monday & Wednesday Lotto": "mwf",   # ← actual name found in scraped data
    "Monday Lotto":             "mwf",   # alternate
    "Wednesday Lotto":          "mwf",   # alternate
    "Friday Lotto":             "mwf",   # alternate
    "Mon & Wed Lotto":          "mwf",   # safety variant

    # Set for Life — likely no syndicates sold, but map in case
    "Set for Life":             "sfl",

    # Skip these — supplementary games, not main pipelines
    "Super 66":                 None,
    "Lucky Lotteries":          None,
    "Lucky Lotteries Mega Jackpot": None,
    "Lucky Lotteries Super Jackpot": None,
}


def split_d_by_game(src_csv: Path, root: Path) -> dict:
    """
    Split a raw D_*.csv file by the Games column into per-game Direct/ folders.

    Handles:
    - Single game rows       → copied to that game's Direct/ folder
    - Pipe-separated rows    → a copy goes to EACH matching game folder
    - None-mapped games      → skipped intentionally (Super 66, Lucky Lotteries)
    - Unknown game names     → logged in _unknown_games for inspection

    Returns dict: {game_key: row_count_written, '_unknown_games': [...]}
    """
    try:
        df = pd.read_csv(src_csv)
    except Exception as ex:
        return {"error": str(ex)}

    if "Games" not in df.columns:
        return {"error": "No 'Games' column found in file"}

    # Accumulate rows per game key
    game_rows: dict[str, list] = {k: [] for k in GAME_KEYS}
    unknown_games: set = set()
    skipped_games: set = set()

    for _, row in df.iterrows():
        raw_games = str(row.get("Games", "")).strip()
        if not raw_games or raw_games == "nan":
            continue

        # Split pipe-separated multi-game entries
        parts = [g.strip() for g in raw_games.split("|")]

        matched_keys = set()
        for part in parts:
            if part not in GAME_NAME_MAP:
                unknown_games.add(part)
            else:
                gkey = GAME_NAME_MAP[part]
                if gkey is None:
                    skipped_games.add(part)  # intentionally skipped
                else:
                    matched_keys.add(gkey)

        for gkey in matched_keys:
            game_rows[gkey].append(row)

    # Save each game's rows to its Direct/ folder
    state_tag = src_csv.stem   # e.g. "D_NSW_NSW"
    results = {}

    # Columns the downstream pipeline is allowed to see. The metadata columns
    # (Syndicate_ID kept for ref, but Draw_Number, Outlet_ID, Postcode, Share_Cost,
    # System_Number, CompanyId, Product, Total_Shares, Available_Shares, …) were
    # leaking into CVI/BRD numeric output. Keep ONLY: picks (w1…wN), powerball (PB),
    # and a few harmless identity/label columns for display. This fixes the leak at
    # the SOURCE so EVERY reader (Variable Inputs, CVI Matrix, collation, dashboards)
    # gets clean data — not just the build-matrix path.
    def _clean_for_pipeline(gdf: pd.DataFrame) -> pd.DataFrame:
        w_cols = sorted([c for c in gdf.columns if re.match(r'^w\d+$', str(c), re.I)],
                        key=lambda x: int(str(x)[1:]))
        keep = [c for c in ("Syndicate_ID", "Syndicate_Name", "Game", "Games")
                if c in gdf.columns]
        if "PB" in gdf.columns:
            keep.append("PB")
        keep += w_cols
        return gdf[keep]

    for gkey, rows in game_rows.items():
        if not rows:
            results[gkey] = 0
            continue
        gdf = pd.DataFrame(rows).reset_index(drop=True)
        gdf = _clean_for_pipeline(gdf)
        dest_dir = game_dirs(gkey)["Direct"]
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_file = dest_dir / f"{state_tag}_{gkey}.csv"
        gdf.to_csv(dest_file, index=False)
        results[gkey] = len(gdf)

    if unknown_games:
        results["_unknown_games"] = sorted(unknown_games)
    if skipped_games:
        results["_skipped_games"] = sorted(skipped_games)

    return results


def game_dirs(game_key: str) -> dict:
    """Return DIRS-like dict scoped to a specific game folder."""
    g = ROOT / "Games" / game_key.upper()
    d = {
        "Game":            g,
        "Main_Data":       g / "Main_Data",
        "Outputs":         g / "Outputs",
        "Formulas":        g / "Formulas",
        "Containers":      g / "Containers",
        "Scraper":         g / "Variables" / "Scraper",
        "Direct":          g / "Variables" / "Variable_Elements" / "Direct",
        "Base":            g / "Variables" / "Variable_Elements" / "Base",
        "Splits":          g / "Variables" / "Variable_Elements" / "Splits",
        "Splits_Combi":    g / "Variables" / "Variable_Elements" / "Splits_Combi",
        "Rainbow":         g / "Variables" / "Variable_Elements" / "Rainbow",
        "ExcelPro":        g / "Variables" / "Variable_Elements" / "ExcelPro",
        "CVI":             g / "Variables" / "Container_Variable_Inputs",
        "Selected_Counts": g / "Formulas" / "Selected_Counts",
        "SinceLast":       g / "SinceLast",
    }
    for p in d.values():
        p.mkdir(parents=True, exist_ok=True)
    return d


def active_game() -> str:
    return st.session_state.get("active_game", "sat")


def active_game_dirs() -> dict:
    return game_dirs(active_game())


def active_game_cfg() -> dict:
    return GAMES_CFG[active_game()]

# ── Naming convention helpers ──────────────────────────────────────────────
def parse_main_filename(fname: str) -> dict:
    """
    Parse `{cluster_label}_{lotto_type}_D{draw_no}.csv`
    e.g. `1n_oz_D1567.csv` → {cluster:'1n', lotto:'oz', draw:'D1567'}
    """
    stem = Path(fname).stem           # strip .csv
    parts = stem.split("_")
    result = {"cluster": "", "lotto": "", "draw": "", "raw": fname}
    if len(parts) >= 3:
        result["cluster"] = parts[0]
        result["lotto"]   = parts[1]
        result["draw"]    = "_".join(parts[2:])
    elif len(parts) == 2:
        result["lotto"]  = parts[0]
        result["draw"]   = parts[1]
    return result


def parse_cvi_filename(fname: str) -> dict:
    """
    Parse `CVI_{lotto_type}_{formula}_{date}.csv`
    e.g. `CVI_oz_BRD_2026_05_28.csv`
    """
    stem  = Path(fname).stem
    parts = stem.split("_")
    result = {"lotto": "", "formula": "", "date": "", "raw": fname}
    if len(parts) >= 3 and parts[0] == "CVI":
        result["lotto"]   = parts[1]
        result["formula"] = parts[2]
        result["date"]    = "_".join(parts[3:])
    return result


def scan_main_data_files() -> list[dict]:
    """
    Scan Main_Data/ and return only files matching the naming convention:
    {cluster}_{lotto_type}_D{draw_no}.csv
    e.g. 1n_oz_D1567.csv
    Rejects CVI files, SC files, or any file without a draw number (_D prefix).
    """
    lotto_codes = set(LOTTO_TYPES.keys())
    files = []
    for fp in sorted(DIRS["Main_Data"].glob("*.csv")):
        name = fp.stem   # without .csv
        parts = name.split("_")

        # Must have at least 3 parts: cluster, lotto, draw
        if len(parts) < 3:
            continue

        # Reject files that start with known non-main-data prefixes
        if parts[0] in ("CVI", "SC", "CF", "Container", "D", "MAIN"):
            # Allow MAIN_ prefix (old style) but check further
            if parts[0] == "MAIN":
                pass
            else:
                continue

        # Must contain a lotto type code
        lotto_found = next((p for p in parts if p in lotto_codes), None)
        if not lotto_found:
            continue

        # Must contain a draw part starting with D followed by digits
        draw_found = next((p for p in parts
                          if p.startswith("D") and p[1:].isdigit()), None)
        if not draw_found:
            continue

        # Quick column check: reject if columns look like w1,w2,w3 (CVI file)
        try:
            sample = pd.read_csv(fp, nrows=2)
            w_cols = [c for c in sample.columns if re.match(r'^w\d+$', c)]
            n_cols_found = [c for c in sample.columns if re.match(r'^n\d+$', c, re.I)]
            if w_cols and not n_cols_found:
                continue   # looks like a CVI file, not main data
        except Exception:
            continue

        info = parse_main_filename(fp.name)
        info["path"] = str(fp)
        info["rows"] = 0
        try:
            info["rows"] = sum(1 for _ in open(fp)) - 1
        except Exception:
            pass
        files.append(info)
    return files


def scan_cvi_files(lotto_type: str = "") -> list[dict]:
    """Scan Container_Variable_Inputs/ for CVI files."""
    files = []
    for fp in sorted(DIRS["CVI"].glob("CVI_*.csv")):
        info = parse_cvi_filename(fp.name)
        if not lotto_type or info["lotto"] == lotto_type:
            info["path"] = str(fp)
            files.append(info)
    return files


# ── Cluster registry ──────────────────────────────────────────────────────
CLUSTER_REGISTRY = ROOT / "clusters.json"

def load_clusters() -> list[dict]:
    if CLUSTER_REGISTRY.exists():
        try:
            return json.loads(CLUSTER_REGISTRY.read_text())
        except Exception:
            pass
    return []

def save_clusters(clusters: list[dict]):
    CLUSTER_REGISTRY.write_text(json.dumps(clusters, indent=2))

def next_cluster_id(clusters: list[dict]) -> str:
    if not clusters:
        return "001"
    last = max(int(c["id"]) for c in clusters if str(c.get("id","0")).isdigit())
    return f"{last+1:03d}"


# ── Parallel matching worker (must be top-level for multiprocessing) ───────
def _parallel_worker(args: tuple) -> dict:
    """
    Standalone worker — no Streamlit state.
    Loads CVI, SC, CF from disk; runs matching; saves results.
    Returns summary dict.
    """
    import pandas as pd, numpy as np, re, json
    from pathlib import Path

    (formula_name, cvi_path, main_data_path,
     sc_path, cluster_id, cluster_label,
     lotto_type, draw_no, draw_date,
     output_dir) = args

    result = {
        "formula":    formula_name,
        "status":     "error",
        "selected_n": 0,
        "unselected_n": 0,
        "error":      "",
    }

    try:
        cvi_df  = pd.read_csv(cvi_path)
        main_df = pd.read_csv(main_data_path)

        # Coerce numeric
        for col in main_df.columns:
            main_df[col] = pd.to_numeric(main_df[col], errors="coerce")

        # Load SC
        sc_dict = {}
        if sc_path and Path(sc_path).exists():
            sc_df = pd.read_csv(sc_path)
            if "w" in sc_df.columns and "Selected Count" in sc_df.columns:
                for _, row in sc_df.iterrows():
                    k = str(row["w"]).strip()
                    key = k if k.startswith("w") else f"w{k}"
                    sc_dict[key] = [int(x.strip()) for x in
                                    str(row["Selected Count"]).split(",")
                                    if x.strip().isdigit()]

        # CF: always fresh all-U for parallel run
        w_cols = sorted([c for c in cvi_df.columns
                         if re.match(r'^w\d+$', c)], key=lambda x: int(x[1:]))
        carry_fwd = {w: "U" for w in w_cols}

        # Re-import run_matching locally (worker process)
        import sys
        sys.path.insert(0, str(Path(__file__).parent))
        # Run matching inline to avoid circular imports
        # (simplified version for worker — same logic as main engine)

        # Detect n_cols — prefer explicit n1…nN; only use the ≥90%-numeric
        # heuristic when there are none (mirrors run_matching, so the parallel
        # path and the sequential engine agree and neither counts metadata
        # columns like Postcode/Length/Draw Number as lottery numbers).
        _explicit_n = [c for c in main_df.columns
                       if re.match(r'^n\d+$', c, re.I)]
        if _explicit_n:
            n_cols = sorted(
                _explicit_n,
                key=lambda x: int(re.sub(r'\D', '', x) or 0)
            )[:20]
        else:
            n_cols = sorted(
                [c for c in main_df.columns
                 if pd.to_numeric(main_df[c], errors="coerce").notna().mean() > 0.9],
                key=lambda x: int(re.sub(r'\D', '', x) or 0)
            )[:20]

        M = len(main_df)
        prev_sel = pd.DataFrame()
        prev_unsel = main_df.copy().reset_index(drop=True)
        final_sel = pd.DataFrame()
        final_unsel = prev_unsel.copy()

        CSIZE = 500_000

        def parse_col(series):
            r = []
            for v in series.dropna():
                try:
                    n = float(str(v).strip())
                    if not pd.isna(n) and n >= 1:
                        r.append(int(round(n)))
                except Exception: pass
            return r

        def match_chunk(arr, cvi_arr):
            counts = np.zeros(arr.shape[0], dtype=np.int32)
            for j in range(arr.shape[1]):
                counts += np.isin(arr[:,j], cvi_arr).astype(np.int32)
            return counts

        for idx, w in enumerate(w_cols):
            raw_cvi = parse_col(cvi_df[w])
            if not raw_cvi:
                continue
            cvi_arr = np.array(raw_cvi, dtype=np.float32)
            sc_this = sc_dict.get(w, [])
            sc_set  = set(sc_this)

            direction = carry_fwd.get(w, "U").upper()
            if idx == 0:
                present = main_df.copy().reset_index(drop=True)
            elif direction == "S":
                present = prev_sel.reset_index(drop=True)
            else:
                present = prev_unsel.reset_index(drop=True)

            if present.empty:
                break

            all_counts = np.empty(0, dtype=np.int32)
            for start in range(0, len(present), CSIZE):
                chunk = present.iloc[start:start+CSIZE]
                arr   = chunk[n_cols].to_numpy(dtype=np.float32)
                arr   = np.nan_to_num(arr, nan=-1.0)
                all_counts = np.concatenate(
                    [all_counts, match_chunk(arr, cvi_arr)])

            sel_mask = np.isin(all_counts, list(sc_set))
            sel_df   = present[sel_mask].reset_index(drop=True)
            unsel_df = present[~sel_mask].reset_index(drop=True)

            prev_sel   = sel_df
            prev_unsel = unsel_df
            final_sel   = sel_df
            final_unsel = unsel_df

        # Save results — naming: {cluster_label}_{lotto}_{draw}_{date}_{formula}
        out    = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        prefix = f"{cluster_label}_{lotto_type}_{draw_no}_{draw_date}"

        sel_path   = out / f"{prefix}_{formula_name}_selected.csv"
        unsel_path = out / f"{prefix}_{formula_name}_unselected.csv"
        final_sel.to_csv(sel_path,   index=False)
        final_unsel.to_csv(unsel_path, index=False)

        result["status"]      = "complete"
        result["selected_n"]  = len(final_sel)
        result["unselected_n"]= len(final_unsel)
        result["sel_path"]    = str(sel_path)
        result["unsel_path"]  = str(unsel_path)

    except Exception as ex:
        result["error"] = str(ex)

    return result

# ═══════════════════════════════════════════════════════════════════════════════
# 3. SESSION STATE — single shared dict S
# ═══════════════════════════════════════════════════════════════════════════════
def _init_state():
    if "S" in st.session_state:
        return
    st.session_state.S = {
        # Variable DataFrames
        "B":   _load_file(DIRS["Base"]         / "B.xlsx"),
        "R":   pd.DataFrame({"w": range(1,9), "w1":[49,22,2,13,35,27,6,12]}),
        "D":   _load_file(DIRS["Direct"]       / "D.xlsx"),
        "Sp":  _load_file(DIRS["Splits"]       / "data_1b.xlsx"),
        "So":  _load_file(DIRS["Splits_Combi"] / "data.xlsx"),
        # Main Data (user uploads manually each run)
        "main_data": pd.DataFrame(),
        # Collated CVI results per formula
        "cvi": {},
        # Matching results per dashboard
        "results": {},
        # Container formula active flags
        "cf_active": {r[1]: True for r in CF_ROWS},
        # Auto mode per section
        "auto": {},
        # Scraper state
        "confirmed_api_url": "",
        "cookie_str": "",
        "scrape_log": [],
        # Container status per dashboard
        "container_status": {},
    }
    # Init container status
    for db in DASHBOARDS:
        st.session_state.S["container_status"][db] = pd.DataFrame({
            "Name": [db], "Status": ["Stopped"],
            "Memory Usage": [""], "CPU Usage": [""],
            "Percent progress %": [0],
            "Time Guestimate Min": [""], "Time Guestimate Max": [""],
        })

S = None  # set after _init_state() call in main

# ═══════════════════════════════════════════════════════════════════════════════
# 4. FILE I/O — large-file safe, all formats
# ═══════════════════════════════════════════════════════════════════════════════
LARGE_FILE_THRESHOLD = 50 * 1024 * 1024   # 50 MB

def _count_csv_rows(path: Path) -> int:
    """Fast row count without loading into memory."""
    try:
        with open(path, "rb") as f:
            return sum(1 for _ in f) - 1
    except Exception:
        return 0

def _load_file(path: Path, max_rows: int = 0,
               numeric: bool = True) -> pd.DataFrame:
    """
    Load any supported file format.
    Large CSVs (>50 MB) are read in 500K-row chunks to avoid freezing.
    accdb/mdb files must be placed directly in the project folder —
    Streamlit cannot upload them via the browser (browser restriction, not ours).
    Use mdb-tools on Mac: brew install mdb-tools
    """
    if not path or not path.exists():
        return pd.DataFrame()
    try:
        ext  = path.suffix.lower()
        size = path.stat().st_size

        if ext == ".csv":
            if max_rows > 0:
                df = pd.read_csv(path, nrows=max_rows)
            elif size > LARGE_FILE_THRESHOLD:
                chunks = pd.read_csv(path, chunksize=500_000)
                df = pd.concat(chunks, ignore_index=True)
            else:
                df = pd.read_csv(path)
        elif ext in [".xlsx", ".xlsm"]:
            df = pd.read_excel(path, engine="openpyxl",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext == ".xls":
            df = pd.read_excel(path, engine="xlrd",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext in [".accdb", ".mdb"]:
            try:
                import pyodbc
                conn = pyodbc.connect(
                    f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={path};")
                cursor = conn.cursor()
                tables = [t.table_name for t in cursor.tables(tableType="TABLE")]
                q = (f"SELECT TOP {max_rows} * FROM [{tables[0]}]"
                     if max_rows > 0 else f"SELECT * FROM [{tables[0]}]")
                df = pd.read_sql(q, conn)
                conn.close()
            except Exception:
                import subprocess, io
                try:
                    tables_raw = subprocess.check_output(
                        ["mdb-tables", "-1", str(path)], timeout=30).decode()
                    tables = [t for t in tables_raw.strip().split("\n") if t]
                    csv_data = subprocess.check_output(
                        ["mdb-export", str(path), tables[0]],
                        timeout=120).decode()
                    df = pd.read_csv(io.StringIO(csv_data))
                    if max_rows > 0:
                        df = df.head(max_rows)
                except Exception:
                    return pd.DataFrame()
        elif ext == ".txt":
            for sep in ["\t", ",", " ", "|"]:
                try:
                    df = pd.read_csv(path, sep=sep,
                                     nrows=max_rows if max_rows > 0 else None)
                    if len(df.columns) >= 3:
                        break
                except Exception:
                    pass
            else:
                return pd.DataFrame()
        elif ext == ".xml":
            df = pd.read_xml(path)
        else:
            return pd.DataFrame()

        if numeric:
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()

def _load_file_preview(path: Path, n: int = 20) -> pd.DataFrame:
    """Load first N rows only — never freezes."""
    return _load_file(path, max_rows=n, numeric=False)

def _save_df(df: pd.DataFrame, path: Path):
    """Save DataFrame to CSV, padding n-columns to consistent width.

    Operates on a copy so the caller's DataFrame is never mutated. (Previously
    this added missing n-columns in place, silently altering the object the
    caller still held.)
    """
    if df.empty:
        return
    out = df.copy()
    nc = sorted([c for c in out.columns if re.match(r'^n\d+$', c)],
                key=lambda x: int(x[1:]))
    if nc:
        max_n = int(nc[-1][1:])
        for i in range(1, max_n+1):
            if f"n{i}" not in out.columns:
                out[f"n{i}"] = None
    out.to_csv(path, index=False)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_styled_excel(df: pd.DataFrame, cvi_set: set,
                    n_cols: list, sheet_name: str = "Data") -> bytes:
    """
    Export DataFrame to Excel with highlighted cells.
    Cells whose value is in cvi_set get a salmon/pink background.
    Count column gets a yellow background.
    Returns bytes suitable for st.download_button.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    PINK   = PatternFill("solid", fgColor="FFB3B3")
    YELLOW = PatternFill("solid", fgColor="FFFF99")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    cols = list(df.columns)
    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = GREY
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = 8

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci,
                           value=None if pd.isna(val) else val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            if col == "Count":
                cell.fill = YELLOW
            elif col in n_cols:
                try:
                    if int(round(float(val))) in cvi_set:
                        cell.fill = PINK
                except (ValueError, TypeError):
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# 5. HIGH-PERFORMANCE MATCHING ENGINE (numpy vectorised, chunked, parallel)
# ═══════════════════════════════════════════════════════════════════════════════

# ── DuckDB engine — for files >50MB / 63M+ rows ───────────────────────────────
DUCKDB_THRESHOLD = 50 * 1024 * 1024   # 50 MB CSV → use DuckDB

def _duckdb_available() -> bool:
    try:
        import duckdb
        return True
    except ImportError:
        return False


def _match_duckdb(csv_path: Path, cvi_nums: list[int],
                  n_col_count: int) -> np.ndarray:
    """
    DuckDB-based matching for very large CSV files.
    Reads CSV directly from disk — never loads full file into RAM.
    Works on 63M rows in ~30 seconds vs pandas which would freeze.

    csv_path:     path to the main data CSV file
    cvi_nums:     list of CVI numbers to match against
    n_col_count:  number of n-columns (n1..nN)

    Returns np.ndarray of match counts per row (length = row count in file).
    """
    import duckdb

    n_cols_sql = " + ".join(
        f"(CASE WHEN n{i} IN ({','.join(str(x) for x in cvi_nums)}) "
        f"THEN 1 ELSE 0 END)"
        for i in range(1, n_col_count + 1)
    )

    query = f"""
        SELECT {n_cols_sql} AS match_count
        FROM read_csv_auto('{csv_path}',
            ignore_errors=true,
            sample_size=10000)
        ORDER BY rowid
    """

    try:
        con = duckdb.connect(database=":memory:")
        result = con.execute(query).fetchnumpy()
        con.close()
        counts = result["match_count"].astype(np.int32)
        return counts
    except Exception:
        return np.array([], dtype=np.int32)


def _smart_match(main_df_or_path, n_cols: list,
                 cvi_nums: np.ndarray) -> np.ndarray:
    """
    Route to DuckDB (large files) or numpy (in-memory).
    main_df_or_path: pd.DataFrame already loaded, OR Path to CSV file.
    """
    if isinstance(main_df_or_path, Path):
        size = main_df_or_path.stat().st_size
        if size > DUCKDB_THRESHOLD and _duckdb_available():
            return _match_duckdb(
                main_df_or_path, cvi_nums.tolist(), len(n_cols))
        else:
            # Load in chunks if large
            chunks = pd.read_csv(main_df_or_path, chunksize=500_000)
            df = pd.concat(chunks, ignore_index=True)
            return _count_matches(df, n_cols, cvi_nums)
    else:
        return _count_matches(main_df_or_path, n_cols, cvi_nums)


def _match_chunk(chunk_arr: np.ndarray, cvi_arr: np.ndarray) -> np.ndarray:
    """
    Vectorised intersection count.
    chunk_arr: (N, K) float32 — main data chunk
    cvi_arr:   (M,)   float32 — CVI numbers
    returns:   (N,)   int32   — match count per row
    Handles 500K rows in ~0.05s on M3 Pro Max.

    NOTE (restored): the `def` line for this function had been lost in a
    previous edit. That left this body as unreachable code *inside*
    `_smart_match`, so `_match_chunk` was never actually defined — and every
    call to `_count_matches` / `run_matching` raised `NameError`. It is now a
    proper module-level function again. Logic is unchanged.
    """
    counts = np.zeros(chunk_arr.shape[0], dtype=np.int32)
    for j in range(chunk_arr.shape[1]):
        counts += np.isin(chunk_arr[:, j], cvi_arr).astype(np.int32)
    return counts


DISPLAY_THRESHOLD = 500_000   # rows — above this, don't store per-row DataFrames


def _parse_cvi_col(series: pd.Series) -> list[int]:
    """
    Extract valid integers from a CVI column.
    Handles floats (1.0→1), strings ("1"), NaN safely.
    No upper range cap — CVI values are not raw lottery balls
    and can legitimately exceed 45 depending on the collation formula.
    """
    result = []
    for v in series.dropna():
        try:
            n = float(str(v).strip())
            if not pd.isna(n) and n >= 1:          # only require positive integer
                result.append(int(round(n)))
        except (ValueError, TypeError):
            pass
    return result


def _count_matches(df: pd.DataFrame, n_cols: list,
                   cvi_nums: np.ndarray) -> np.ndarray:
    """Vectorised match count — chunks handled internally.

    Accumulates each chunk's counts in a list and concatenates once at the end.
    (Previously this re-allocated and copied a growing array on every chunk —
    O(n²) total copying for many chunks. Output is identical.)
    """
    parts = []
    for start in range(0, len(df), CHUNK_SIZE):
        chunk  = df.iloc[start:start+CHUNK_SIZE]
        arr    = chunk[n_cols].to_numpy(dtype=np.float32)
        arr    = np.nan_to_num(arr, nan=-1.0)
        parts.append(_match_chunk(arr, cvi_nums))
    return np.concatenate(parts) if parts else np.empty(0, dtype=np.int32)


def _count_str(counts: np.ndarray) -> str:
    return ",".join(str(int(k)) for k in sorted(np.unique(counts)))


def _breakdown_str(counts: np.ndarray, sc_set: set,
                   prefix_sel="S", prefix_unsel="U") -> tuple[str, str, str, str]:
    """
    Returns (sel_bd_str, unsel_bd_str, unsel_count_str, count_dist_dict_str).
    """
    unique = sorted(int(k) for k in np.unique(counts))
    sel_parts, unsel_parts = [], []
    for k in unique:
        n_k = int((counts == k).sum())
        if k in sc_set:
            sel_parts.append(f"{prefix_sel}{k}:{n_k}")
        else:
            unsel_parts.append(f"{prefix_unsel}{k}:{n_k}")
    unsel_counts = [k for k in unique if k not in sc_set]
    return (
        "  ".join(sel_parts)   or "No breakdown",
        "  ".join(unsel_parts) or "No breakdown",
        ",".join(str(k) for k in unsel_counts) or "No count",
        {f"S{k}": int((counts==k).sum()) for k in unique},
    )


def run_matching(main_df: pd.DataFrame,
                 cvi_df:  pd.DataFrame,
                 sc_dict: dict,
                 carry_fwd: dict,
                 main_path=None) -> dict:
    """
    Sequential matching engine — corrected carry-forward logic.

    carry_fwd[w] = "U" or "S"  (default "U" for any missing key)
    Meaning of toggle at Row N:
        "U" → Row N's present = Row N-1's UNSELECTED
        "S" → Row N's present = Row N-1's SELECTED

    Main Count: pre-computed once against ORIGINAL main_df for every w-column.
                Never recalculated regardless of carry-forward state.

    Memory: if M > DISPLAY_THRESHOLD, per-row DataFrames are NOT stored —
            only count distributions are kept. Final stage data always stored.

    main_path (optional): path to the CSV that `main_df` was loaded from. When
        supplied AND it is a large CSV AND DuckDB is installed AND it is safe
        (file row-count == M and the n-columns are a contiguous n1…nN block),
        the heavy *main-count pre-pass* (one full-M scan per w-column) is run in
        DuckDB instead of pandas. If DuckDB returns an unexpected length for any
        column, that column silently falls back to the pandas path, so results
        are always identical to the pure-pandas engine. Passing None (the
        default) keeps the original all-pandas behavior.

        NOTE on scope: only the pre-pass uses DuckDB. The staged carry-forward
        matching still runs in memory on `main_df`, because each stage filters a
        shrinking pool of rows (selected vs unselected) — that is row-level and
        stateful, not a single SQL aggregate. So this lowers the cost of the
        most expensive *uniform* passes but does NOT remove the need to hold
        `main_df` in RAM. Fully streaming the staged engine is a larger,
        separate change.
    """
    if main_df.empty or cvi_df.empty:
        return {"selected": pd.DataFrame(), "unselected": pd.DataFrame(),
                "fig9_table": pd.DataFrame(), "breakdown": pd.DataFrame(),
                "debug_rows": []}

    # ── Detect number columns ──────────────────────────────────────────────
    # Prefer explicit n1…nN columns when present. Only fall back to the
    # "≥90% numeric" heuristic when there are NO explicit n-columns — otherwise
    # numeric metadata columns (Postcode, Length, Draw Number, …) get treated
    # as lottery numbers and corrupt every match count.
    _explicit_n = [c for c in main_df.columns if re.match(r'^n\d+$', c, re.I)]
    if _explicit_n:
        n_cols = sorted(
            _explicit_n,
            key=lambda x: int(re.sub(r'\D', '', x) or 0)
        )[:20]
    else:
        n_cols = sorted(
            [c for c in main_df.columns
             if pd.to_numeric(main_df[c], errors="coerce").notna().mean() > 0.9],
            key=lambda x: int(re.sub(r'\D', '', x) or 0)
        )[:20]

    w_cols = sorted(
        [c for c in cvi_df.columns if re.match(r'^w\d+$', c)],
        key=lambda x: int(x[1:])
    )

    M = len(main_df)
    small_enough = (M <= DISPLAY_THRESHOLD)

    # ── PRE-COMPUTE Main Count AND Main Breakdown for every w-column ──────
    # Computed once against original M rows. Never recalculated.
    main_count_map: dict[str, str] = {}
    main_bd_map:    dict[str, str] = {}   # full distribution e.g. "S0:1 S1:3 S2:9..."
    main_counts_map: dict[str, np.ndarray] = {}  # raw counts for filter UI

    # ── Decide whether the pre-pass may use DuckDB (safe, opt-in) ──────────
    # Only when: a path was given, it's a large CSV, DuckDB is installed, the
    # file's row count matches M (so DuckDB row order lines up with main_df),
    # and the n-columns form a contiguous n1…nN block (so _match_duckdb's
    # generated SQL references columns that actually exist).
    _duck_path = None
    try:
        if main_path is not None:
            _p = Path(main_path)
            _contiguous_n = (
                bool(n_cols)
                and all(c.lower() == f"n{i}" for i, c in enumerate(n_cols, 1))
            )
            if (_p.exists() and _p.suffix.lower() == ".csv"
                    and _duckdb_available()
                    and _p.stat().st_size > DUCKDB_THRESHOLD
                    and _contiguous_n
                    and _count_csv_rows(_p) == M):
                _duck_path = _p
    except Exception:
        _duck_path = None

    for w in w_cols:
        raw = _parse_cvi_col(cvi_df[w])
        if raw:
            mc = None
            if _duck_path is not None:
                # DuckDB streams the file from disk for this one w-column.
                duck_mc = _match_duckdb(_duck_path, raw, len(n_cols))
                # Trust it only if it returned exactly M counts; otherwise the
                # pandas path below guarantees a correct result.
                if duck_mc.shape[0] == M:
                    mc = duck_mc
            if mc is None:
                mc = _count_matches(
                    main_df, n_cols, np.array(raw, dtype=np.float32))
            main_counts_map[w] = mc
            main_count_map[w]  = _count_str(mc)
            # Full breakdown (all M rows, no SC filter — just distribution)
            unique_k = sorted(int(k) for k in np.unique(mc))
            main_bd_map[w] = "  ".join(
                f"S{k}:{int((mc==k).sum())}" for k in unique_k
            )
        else:
            main_count_map[w]   = "—"
            main_bd_map[w]      = "—"
            main_counts_map[w]  = np.array([], dtype=np.int32)

    # ── Stage state ────────────────────────────────────────────────────────
    # prev_sel / prev_unsel are the two pools produced by the PREVIOUS stage.
    # The toggle at the CURRENT stage picks which one becomes "present".
    prev_sel   = pd.DataFrame()        # Row 0 has no previous selected
    prev_unsel = main_df.copy().reset_index(drop=True)  # Row 1 default = all M

    fig9_rows      = []
    breakdown_rows = []
    debug_rows     = []

    final_sel   = pd.DataFrame()
    final_unsel = prev_unsel.copy()

    for idx, w in enumerate(w_cols):
        w_num     = int(w[1:])
        direction = carry_fwd.get(w, carry_fwd.get(f"w{w_num}", "U")).upper()

        # ── Determine present_df for this stage ───────────────────────
        if idx == 0:
            # First row always receives full original main data
            present_df    = main_df.copy().reset_index(drop=True)
            present_label = f"M:{M}"
        elif direction == "S":
            present_df    = prev_sel.reset_index(drop=True)
            present_label = f"S:{len(prev_sel)}"
        else:  # "U"
            present_df    = prev_unsel.reset_index(drop=True)
            present_label = f"U:{len(prev_unsel)}"

        present_count = len(present_df)
        # Defensive: sync label with actual count (catches any reference drift)
        if present_label.startswith("U:"):
            present_label = f"U:{present_count}"
        elif present_label.startswith("S:") and not present_label.startswith("S:0"):
            present_label = f"S:{present_count}"
        raw_cvi       = _parse_cvi_col(cvi_df[w])
        cvi_nums      = np.array(raw_cvi, dtype=np.float32)
        sc_this       = sc_dict.get(w, sc_dict.get(str(w_num), []))
        if isinstance(sc_this, (int, float)):
            sc_this = [int(sc_this)]
        elif isinstance(sc_this, str):
            sc_this = [int(x.strip()) for x in sc_this.split(",")
                       if x.strip().lstrip('-').isdigit()]
        sc_set  = set(sc_this)
        sc_str  = ",".join(str(s) for s in sorted(sc_set)) if sc_set else "—"

        # ── Empty CVI column ──────────────────────────────────────────
        if not len(cvi_nums):
            fig9_rows.append({
                "Row":            f"Row {w_num}",
                "Main\nData":     f"M:{M}",
                "CVI":            w,
                "Dir":            direction,
                "Main\nCount":    main_count_map[w],
                "Main\nBreakdown":main_bd_map[w],
                "Present\nData":  present_label,
                "Present\nCount": "— No CVI",
                "SC":             sc_str,
                "Selected":       "S:0",
                "Sel\nBreakdown": "—",
                "Unsel\nCount":   "—",
                "Unselected":     f"U:{present_count} (fwd)",
                "Unsel\nBreakdown":"—",
            })
            debug_rows.append({
                "w": w, "direction": direction,
                "cvi_numbers": [], "cvi_set": set(),
                "present_in": present_count, "present_label": present_label,
                "sc": list(sc_set), "selected_n": 0,
                "unselected_n": present_count,
                "main_count_str": main_count_map[w],
                "main_bd_str":    main_bd_map[w],
                "main_counts":    main_counts_map[w],
                "pres_count_str": "—", "count_dist": {},
                "note": "No CVI data — all carry forward",
                "sel_df": None,
                "unsel_df": present_df.copy() if small_enough else None,
                "n_cols": n_cols,
            })
            prev_sel   = pd.DataFrame()
            prev_unsel = present_df.copy()
            final_sel   = pd.DataFrame()
            final_unsel = present_df.copy()
            continue

        # ── Match present_df against this w-column ────────────────────
        # Via the _smart_match dispatcher: with an in-memory DataFrame this is
        # exactly _count_matches(...). present_df is a shrinking pool (it cannot
        # be re-read from disk), so this stage stays in RAM by design.
        pres_counts = _smart_match(present_df, n_cols, cvi_nums)
        sel_bd_str, unsel_bd_str, unsel_count_str, count_dist = \
            _breakdown_str(pres_counts, sc_set)
        pres_count_str = _count_str(pres_counts)

        sel_mask   = np.isin(pres_counts, list(sc_set))
        sel_df     = present_df[sel_mask].reset_index(drop=True)
        unsel_df   = present_df[~sel_mask].reset_index(drop=True)

        # Add count column to sel/unsel for display
        if small_enough and not sel_df.empty:
            sel_counts_arr = pres_counts[sel_mask]
            sel_df = sel_df.copy()
            sel_df["Count"] = sel_counts_arr
        if small_enough and not unsel_df.empty:
            unsel_df = unsel_df.copy()
            unsel_df["Count"] = pres_counts[~sel_mask]

        fig9_rows.append({
            "Row":            f"Row {w_num}",
            "Main\nData":     f"M:{M}",
            "CVI":            w,
            "Dir":            direction,
            "Main\nCount":    main_count_map[w],
            "Main\nBreakdown":main_bd_map[w],
            "Present\nData":  present_label,
            "Present\nCount": pres_count_str,
            "SC":             sc_str,
            "Selected":       f"S:{len(sel_df)}" if not sel_df.empty else "S:0",
            "Sel\nBreakdown": sel_bd_str,
            "Unsel\nCount":   unsel_count_str,
            "Unselected":     f"U:{len(unsel_df)}" if not unsel_df.empty else "U:0",
            "Unsel\nBreakdown":unsel_bd_str,
        })

        bd = {**count_dist, "w_col": w}
        breakdown_rows.append(bd)

        # Also add count to main_df for display purposes
        main_df_with_count = None
        if small_enough:
            mc_arr = main_counts_map.get(w, np.array([], dtype=np.int32))
            if len(mc_arr) == M:
                main_df_with_count = main_df.copy()
                main_df_with_count["Count"] = mc_arr

        debug_rows.append({
            "w":              w,
            "direction":      direction,
            "cvi_numbers":    raw_cvi,
            "cvi_set":        set(raw_cvi),
            "present_in":     present_count,
            "present_label":  present_label,
            "sc":             sorted(sc_set),
            "selected_n":     len(sel_df),
            "unselected_n":   len(unsel_df),
            "main_count_str": main_count_map[w],
            "main_bd_str":    main_bd_map[w],
            "main_counts":    main_counts_map[w],
            "pres_count_str": pres_count_str,
            "count_dist":     count_dist,
            "note":           "",
            "sel_df":         sel_df   if small_enough else None,
            "unsel_df":       unsel_df if small_enough else None,
            "main_df_wc":     main_df_with_count,
            "n_cols":         n_cols,
        })

        # ── Carry forward: store both pools for next stage ────────────
        prev_sel   = sel_df.drop(columns=["Count"], errors="ignore")
        prev_unsel = unsel_df.drop(columns=["Count"], errors="ignore")
        final_sel   = sel_df
        final_unsel = unsel_df

        # ── If present exhausted, fill remaining rows ─────────────────
        if present_df.empty or (sel_df.empty and unsel_df.empty):
            for rw in w_cols[idx+1:]:
                rw_num = int(rw[1:])
                fig9_rows.append({
                    "Row": f"Row {rw_num}", "Main\nData": f"M:{M}",
                    "CVI": rw, "Dir": carry_fwd.get(rw,"U"),
                    "Main\nCount": main_count_map.get(rw,"—"),
                    "Main\nBreakdown": main_bd_map.get(rw,"—"),
                    "Present\nData": "U:0", "Present\nCount": "—",
                    "SC": sc_dict.get(rw,"—"), "Selected": "S:0",
                    "Sel\nBreakdown": "—", "Unsel\nCount": "—",
                    "Unselected": "U:0", "Unsel\nBreakdown": "—",
                })
                debug_rows.append({
                    "w": rw, "direction": carry_fwd.get(rw,"U"),
                    "cvi_numbers": [], "cvi_set": set(),
                    "present_in": 0, "present_label": "U:0",
                    "sc": [], "selected_n": 0, "unselected_n": 0,
                    "main_count_str": main_count_map.get(rw,"—"),
                    "main_bd_str": main_bd_map.get(rw,"—"),
                    "main_counts": np.array([], dtype=np.int32),
                    "pres_count_str": "—", "count_dist": {},
                    "note": "No present data — exhausted",
                    "sel_df": None, "unsel_df": None,
                    "main_df_wc": None, "n_cols": n_cols,
                })
            break

    return {
        "selected":    final_sel,
        "unselected":  final_unsel,
        "fig9_table":  pd.DataFrame(fig9_rows),
        "breakdown":   pd.DataFrame(breakdown_rows).fillna(0),
        "debug_rows":  debug_rows,
        "n_cols":      n_cols,
        "small_enough": small_enough,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# 6. COLLATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
def execute_collation(components: list[str]) -> pd.DataFrame:
    """
    For each component: take ONLY its number columns (w1…wN, plus PB if present),
    concatenate, relabel w1, w2, w3… Typed ints throughout.

    NOTE: previously this did `df.columns[1:]` ("drop col-0, keep the rest"), which
    is correct for clean B/Ep/Sp/So matrices but WRONG for D — a D file carries
    metadata (Syndicate_Name, Draw_Number, Outlet_ID, Postcode, Share_Cost, …)
    after col-0, and those values were being coerced to numbers and concatenated
    into the CVI/BRD (the 1400151 / 2000 / 1571 leak). Selecting w-columns
    explicitly fixes the leak for D and hardens every component.
    """
    pieces = []
    for comp in components:
        df = S.get(comp) if S else None
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        # Prefer explicit number columns: w1…wN (the picks) + optional PB.
        w_cols = sorted([c for c in df.columns if re.match(r'^w\d+$', str(c), re.I)],
                        key=lambda x: int(str(x)[1:]))
        if w_cols:
            data_cols = w_cols + (["PB"] if "PB" in df.columns else [])
        else:
            # No w-columns (e.g. a clean matrix with a leading index col only):
            # fall back to the original "everything after col-0" behaviour.
            data_cols = list(df.columns[1:])
        if not data_cols:
            continue
        piece = df[data_cols].copy().reset_index(drop=True)
        for c in piece.columns:
            piece[c] = pd.to_numeric(piece[c], errors="coerce")
        pieces.append(piece)

    if not pieces:
        return pd.DataFrame()

    max_rows = max(len(p) for p in pieces)
    padded = []
    for p in pieces:
        if len(p) < max_rows:
            pad = pd.DataFrame({c: [None]*(max_rows-len(p)) for c in p.columns})
            p = pd.concat([p, pad], ignore_index=True)
        padded.append(p)

    combined = pd.concat(padded, axis=1, ignore_index=True)
    combined.columns = [f"w{i+1}" for i in range(len(combined.columns))]
    combined.insert(0, "Row", range(1, len(combined)+1))
    return combined

# ═══════════════════════════════════════════════════════════════════════════════
# 7. SCRAPER ENGINE (Playwright + requests)
# ═══════════════════════════════════════════════════════════════════════════════
def _extract_nums(text) -> list[int]:
    """Extract positive integers from any text. No upper cap — supports any jurisdiction."""
    return sorted(set(int(n) for n in re.findall(r'\b(\d{1,3})\b', str(text))
                  if int(n) >= 1))

def _valid_combo(nums: list) -> bool:
    if not (6 <= len(nums) <= 25): return False
    if max(nums) - min(nums) < 10: return False
    return True

def _parse_json(body: str, pc: int, state: str) -> list[dict]:
    """Walk any JSON shape to extract syndicate number combinations."""
    try:
        data = json.loads(body)
    except Exception:
        return []
    records, seen = [], set()

    def walk(obj, depth=0):
        if depth > 8: return
        if isinstance(obj, list):
            for item in obj: walk(item, depth+1)
        elif isinstance(obj, dict):
            nums = []
            for f in ["numbers","combination","selections","picks","balls",
                      "gameNumbers","game_numbers","selectedNumbers","ticketNumbers"]:
                if f in obj and isinstance(obj[f], list):
                    nums = sorted(set(int(x) for x in obj[f]
                                  if str(x).isdigit() and int(x) >= 1))
                    if nums: break
            if _valid_combo(nums):
                key = tuple(nums)
                if key not in seen:
                    seen.add(key)
                    def g(*keys):
                        for k in keys:
                            v = obj.get(k)
                            if v is not None and str(v).strip(): return str(v).strip()
                        return ""
                    row = {
                        "Postcode":        g("postcode","post_code") or str(pc),
                        "State":           g("state","jurisdiction") or state,
                        "Syndicate Number": g("id","syndicateId","syndicateNumber"),
                        "Agent":           g("agentName","retailer","outlet"),
                        "Title":           g("title","name","description"),
                        "Price per Share": g("price","sharePrice"),
                        "Shares":          g("shares","sharesAvailable"),
                        "Draw Number":     g("drawNumber","draw_number","drawId","drawNo"),
                        "Draw Date":       g("drawDate","draw_date","scheduledDate"),
                        "Game Type":       g("gameType","type","product","gameName") or
                                          ("Normal" if len(nums)<=7 else f"System {len(nums)}"),
                        "Length":          len(nums),
                    }
                    for j,n in enumerate(nums,1): row[f"n{j}"] = n
                    records.append(row)
            for v in obj.values(): walk(v, depth+1)

    walk(data)
    return records


def _parse_html(html: str, pc: int, state: str) -> list[dict]:
    """HTML fallback — extracts number clusters from rendered page."""
    if not REQ_OK or not html: return []
    soup = BeautifulSoup(html, "html.parser")
    for t in soup(["script","style","noscript","head"]): t.decompose()
    records, seen = [], set()
    for el in soup.find_all(True):
        if len(list(el.children)) > 15: continue
        nums = _extract_nums(el.get_text(" ",strip=True))
        if _valid_combo(nums):
            key = tuple(nums)
            if key in seen: continue
            seen.add(key)
            row = {"Postcode":pc,"State":state,
                   "Syndicate Number":"","Agent":"","Title":"",
                   "Price per Share":"","Shares":"",
                   "Draw Number":"","Draw Date":"",
                   "Game Type":"Normal" if len(nums)<=7 else f"System {len(nums)}",
                   "Length":len(nums)}
            for j,n in enumerate(nums,1): row[f"n{j}"] = n
            records.append(row)
    return records


async def _pw_scrape_state(state: str, postcodes: list,
                            api_url: str, cookie_str: str,
                            pb, stx) -> list[dict]:
    """Playwright scrape — intercepts JSON API responses per postcode."""
    all_recs = []
    total = len(postcodes)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            user_agent=SCRAPE_HEADERS["User-Agent"],
            locale="en-AU", timezone_id="Australia/Sydney",
            ignore_https_errors=True,
        )
        if cookie_str:
            cookies = [{"name":k.strip(),"value":v.strip(),
                        "domain":".thelott.com","path":"/"}
                       for kv in cookie_str.split(";")
                       for k,_,v in [kv.partition("=")]]
            await ctx.add_cookies(cookies)
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        page = await ctx.new_page()

        for i, pc in enumerate(postcodes):
            pb.progress((i+1)/total,
                text=f"[{state}] {pc} ({i+1}/{total}) — {len(all_recs)} records")
            stx.caption(f"Scraping {state} · postcode {pc}")

            captured = []
            async def on_resp(resp):
                ct = resp.headers.get("content-type","")
                if resp.status == 200 and \
                   ("json" in ct or any(k in resp.url.lower()
                    for k in ["syndic","api","play","search"])):
                    try:
                        body = (await resp.body()).decode("utf-8","ignore")
                        if body.strip()[:1] in ["{","["]:
                            captured.append({"url":resp.url,"body":body})
                    except Exception: pass

            page.on("response", on_resp)
            try:
                url = api_url.format(pc=pc) if "{pc}" in api_url \
                      else f"{api_url}?postcode={pc}"
                await page.goto(url, timeout=25000, wait_until="load")
                await asyncio.sleep(3)
                for sel in ["input[placeholder*='postcode' i]",
                            "input[type='search']"]:
                    try:
                        el = page.locator(sel).first
                        if await el.is_visible(timeout=1000):
                            await el.fill(str(pc))
                            await el.press("Enter")
                            await asyncio.sleep(2)
                            break
                    except Exception: pass
                await asyncio.sleep(2)

                recs = []
                for cap in captured:
                    recs.extend(_parse_json(cap["body"], pc, state))
                if not recs:
                    html = await page.content()
                    recs = _parse_html(html, pc, state)
                all_recs.extend(recs)

            except Exception as ex:
                st.session_state.S["scrape_log"].append(f"{state} {pc}: {ex}")
            finally:
                page.remove_listener("response", on_resp)

            await asyncio.sleep(random.uniform(1.0, 2.5))

        await browser.close()

    return all_recs


def _run_sync(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try: return loop.run_until_complete(coro)
    finally: loop.close()

# ═══════════════════════════════════════════════════════════════════════════════
# 7b. THELOTT CONFIRMED API — two-step outlet → syndicate scraper
# ═══════════════════════════════════════════════════════════════════════════════
import ssl as _ssl
import urllib.request as _urlreq

# Company IDs confirmed by testing (NT=no syndicates, WA=Lotterywest no API)
_STATE_COMPANY = {"NSW": 3, "ACT": 3, "VIC": 1, "TAS": 1, "QLD": 2, "SA": 6}

_TLOTT_HEADERS = {
    "Accept": "application/json",
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Origin":  "https://www.thelott.com",
    "Referer": "https://www.thelott.com/play/syndicates/syndicate-shares",
}

def _ssl_ctx():
    ctx = _ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl.CERT_NONE
    return ctx

def _api_get(url: str):
    """Raw GET with SSL bypass (required on Mac). Returns parsed JSON or None."""
    req = _urlreq.Request(url, headers=_TLOTT_HEADERS)
    try:
        with _urlreq.urlopen(req, timeout=15, context=_ssl_ctx()) as r:
            return json.loads(r.read().decode("utf-8"))
    except Exception:
        return None

def _is_online() -> bool:
    """Quick connectivity check to thelott API."""
    try:
        req = _urlreq.Request(
            "https://api.thelott.com/outlet/outlets?state=NSW&postcode_or_locality=2000",
            headers=_TLOTT_HEADERS)
        with _urlreq.urlopen(req, timeout=8, context=_ssl_ctx()):
            return True
    except Exception:
        return False

def _fetch_outlets(state: str, postcode: int) -> list:
    """Step 1: Get all outlet IDs for a postcode."""
    url = (f"https://api.thelott.com/outlet/outlets"
           f"?state={state}&postcode_or_locality={postcode}")
    data = _api_get(url)
    if not data:
        return []
    outlets = data if isinstance(data, list) else data.get("outlets", [])
    ids = []
    for o in (outlets if isinstance(outlets, list) else []):
        oid = str(o.get("id") or o.get("outletId") or "").strip()
        if oid:
            ids.append(oid)
    return ids

def _fetch_syndicates_batch(company: int, outlet_ids: list) -> list:
    """Step 2: Get syndicates for a comma-separated batch of outlet IDs."""
    if not outlet_ids:
        return []
    url = (f"https://api.thelott.com/syndicates/api/search"
           f"?company={company}&outlets={','.join(outlet_ids)}&limit=100")
    data = _api_get(url)
    if not data:
        return []
    items = (data if isinstance(data, list)
             else data.get("syndicates", data.get("items", data.get("data", []))))
    return items if isinstance(items, list) else []

def _parse_syndicate_row(syn: dict, postcode: int, state: str) -> dict:
    """Map a thelott syndicate API object to the standard CSV schema."""
    def g(*keys):
        for k in keys:
            v = syn.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""
    outlet = syn.get("outlet") or syn.get("store") or {}
    if not isinstance(outlet, dict):
        outlet = {}
    # Collect draw number strings from games/combinations arrays
    draw_numbers = ""
    for gkey in ("games", "combinations", "entries"):
        games = syn.get(gkey, [])
        if isinstance(games, list) and games:
            parts = []
            for gm in games:
                nums = (gm.get("numbers") or gm.get("selections") or []
                        if isinstance(gm, dict) else gm if isinstance(gm, list) else [])
                if nums:
                    parts.append(" ".join(str(n) for n in sorted(nums)))
            if parts:
                draw_numbers = " | ".join(parts)
            break
    return {
        "Postcode":         postcode,
        "State":            state,
        "Syndicate_ID":     g("syndicateId", "id", "syndicateNumber"),
        "Syndicate_Name":   g("syndicateName", "title", "name", "description"),
        "Draw_Date":        g("drawDate", "draw_date", "scheduledDate", "closeDate"),
        "Share_Cost":       g("sharePrice", "price", "costPerShare", "shareCost"),
        "Available_Shares": g("availableShares", "sharesAvailable", "sharesRemaining"),
        "Total_Shares":     g("totalShares", "shareCount", "shares"),
        "Games":            g("gameType", "gameName", "product", "type"),
        "Draw_Numbers":     draw_numbers,
        "Outlet_ID":        g("outletId") or str(outlet.get("id", "")),
        "Outlet_Name":      g("agentName", "storeName") or str(outlet.get("name", "")),
        "Address":          str(outlet.get("address", g("address"))),
        "Suburb":           str(outlet.get("suburb", g("suburb"))),
    }

def _sweep_state_thelott(state: str, postcodes: list,
                          pb, stx) -> pd.DataFrame:
    """
    Full sweep for one state using confirmed two-step thelott API.
    Deduplicates by Syndicate_ID. Returns DataFrame.
    """
    company = _STATE_COMPANY.get(state)
    if company is None:
        return pd.DataFrame()

    seen_ids = {}   # syndicate_id → row dict
    total = len(postcodes)

    for i, pc in enumerate(postcodes):
        pb.progress(
            (i + 1) / total,
            text=f"[{state}] {pc}  ({i+1}/{total}) — {len(seen_ids)} syndicates so far"
        )
        stx.caption(f"🔍 {state} · postcode {pc}")

        outlet_ids = _fetch_outlets(state, pc)
        if not outlet_ids:
            time.sleep(random.uniform(0.1, 0.25))
            continue

        # Batch outlets in groups of 20 (comma-separated — critical format)
        for b in range(0, len(outlet_ids), 20):
            batch = outlet_ids[b:b + 20]
            syns  = _fetch_syndicates_batch(company, batch)
            for syn in syns:
                row = _parse_syndicate_row(syn, pc, state)
                sid = row["Syndicate_ID"] or f"{pc}_{b}_{syns.index(syn)}"
                if sid not in seen_ids:
                    seen_ids[sid] = row
            time.sleep(random.uniform(0.08, 0.20))

        time.sleep(random.uniform(0.20, 0.55))

    stx.empty()
    return pd.DataFrame(list(seen_ids.values())) if seen_ids else pd.DataFrame()

def _data_status() -> list:
    """Return status info for each D_*.csv file in Main_Data/."""
    rows = []
    for state in ["NSW", "VIC", "QLD", "SA", "TAS"]:
        fp = DIRS["Main_Data"] / f"D_{state}_{state}.csv"
        if fp.exists():
            stat = fp.stat()
            age_h = (datetime.now().timestamp() - stat.st_mtime) / 3600
            try:
                n = sum(1 for _ in open(fp)) - 1
            except Exception:
                n = 0
            rows.append({
                "state": state, "file": fp.name, "rows": n,
                "size_kb": round(stat.st_size / 1024, 1),
                "age_h": round(age_h, 1),
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%d %b %H:%M"),
                "exists": True,
            })
        else:
            rows.append({"state": state, "file": f"D_{state}_{state}.csv",
                         "rows": 0, "size_kb": 0, "age_h": 9999,
                         "modified": "—", "exists": False})
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# 8. CVI MATRIX — transposition + variable slicing
# ═══════════════════════════════════════════════════════════════════════════════
def build_w_matrix(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Sort rows longest→shortest, transpose each into a w-column, pad with None.

    The source 'number' columns of a D file are the syndicate picks, labelled
    **w1…wN** (project convention: syndicate = w, main data = n). Older/main-data
    files may instead use n1…nN. Accept EITHER prefix as the input columns so D
    (w-columns) loads correctly. (Previously this only looked for n1…nN, so D —
    which is w-columns — produced an empty matrix and the "No n-columns found"
    error.) The OUTPUT is always w-columns regardless of input naming.
    """
    src_cols = sorted(
        [c for c in df_raw.columns if re.match(r'^[wn]\d+$', c, re.I)],
        key=lambda x: int(x[1:]))
    if not src_cols:
        return pd.DataFrame()
    rows_as_lists = []
    for _, row in df_raw.iterrows():
        nums = [int(row[c]) for c in src_cols
                if pd.notna(row[c]) and str(row[c]).strip().lstrip('-').isdigit()]
        if nums: rows_as_lists.append(sorted(nums))
    if not rows_as_lists: return pd.DataFrame()
    rows_as_lists.sort(key=len, reverse=True)
    max_len = max(len(r) for r in rows_as_lists)
    w_dict = {f"w{i+1}": (r + [None]*(max_len-len(r)))
              for i, r in enumerate(rows_as_lists)}
    return pd.DataFrame(w_dict)


def d_to_w_only(df: pd.DataFrame) -> pd.DataFrame:
    """Return the D variable as CLEAN w-columns only.

    A freshly-loaded D_*.csv carries metadata columns (Syndicate_ID, Draw_Number,
    Outlet_ID, Postcode, Share_Cost, …) alongside the per-pick w-columns. If those
    metadata columns reach the collation, their values get treated as numbers and
    contaminate the output (the 4687 / 1400286 / 2000 leak). This reduces any D
    frame to just its number content:
      • if it already looks like a built matrix (only w-columns) → return as-is;
      • otherwise rebuild it via build_w_matrix (one w-column per syndicate line).
    """
    if df is None or df.empty:
        return df
    wcols = [c for c in df.columns if re.match(r'^w\d+$', str(c), re.I)]
    meta = [c for c in df.columns if not re.match(r'^[wn]\d+$', str(c), re.I)]
    # Already clean (a built W-matrix: only w-columns, no metadata)
    if wcols and not meta:
        return df
    # Raw D file (w-columns + metadata, or n-columns): rebuild to w-only matrix
    return build_w_matrix(df)


def slice_variables(w_mat: pd.DataFrame) -> dict:
    """Auto-slice Ep (top 8), Sp (top 4 lanes a-d), So (union of Sp lanes)."""
    w_keys = sorted([c for c in w_mat.columns if re.match(r'^w\d+$',c)],
                    key=lambda x: int(x[1:]))
    def col_ints(k):
        result = []
        for v in w_mat[k].dropna():
            try:
                n = float(str(v))
                if not pd.isna(n) and n >= 1:
                    result.append(int(round(n)))
            except (ValueError, TypeError):
                pass
        return result

    ep_keys = w_keys[:8]
    sp_keys = w_keys[:4]
    ep_raw = {k: col_ints(k) for k in ep_keys}
    sp_raw = {"lane_a": col_ints(sp_keys[0]) if len(sp_keys)>0 else [],
              "lane_b": col_ints(sp_keys[1]) if len(sp_keys)>1 else [],
              "lane_c": col_ints(sp_keys[2]) if len(sp_keys)>2 else [],
              "lane_d": col_ints(sp_keys[3]) if len(sp_keys)>3 else []}

    def to_df(d):
        ml = max((len(v) for v in d.values()), default=0)
        return pd.DataFrame({k: v+[None]*(ml-len(v)) for k,v in d.items()})

    so_union = sorted(set(sum(sp_raw.values(),[])))
    return {"Ep": to_df(ep_raw), "Sp": to_df(sp_raw),
            "So": pd.DataFrame({"So_union": so_union})}

# ═══════════════════════════════════════════════════════════════════════════════
# 9. UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def am_toggle(section: str) -> str:
    """Auto/Manual toggle — every page."""
    if section not in S["auto"]:
        S["auto"][section] = "Manual"
    c1, c2, _ = st.columns([1,1,8])
    with c1:
        if st.button("▶ AUTO", key=f"_a_{section}",
                     type="primary" if S["auto"][section]=="Auto" else "secondary",
                     use_container_width=True):
            S["auto"][section] = "Auto"
    with c2:
        if st.button("✋ MANUAL", key=f"_m_{section}",
                     type="primary" if S["auto"][section]=="Manual" else "secondary",
                     use_container_width=True):
            S["auto"][section] = "Manual"
    return S["auto"][section]


def edit_var(var_key: str, folder: Path, filename: str, label: str, note=""):
    """Editable variable input table — reads from and saves to real folder."""
    st.markdown(f"**{label}**")
    if note: st.caption(note)
    p = folder / filename
    if var_key not in S or (isinstance(S.get(var_key), pd.DataFrame) and S[var_key].empty and p.exists()):
        S[var_key] = _load_file(p)
    df = S.get(var_key, pd.DataFrame())
    edited = st.data_editor(df, key=f"ed_{var_key}",
                            use_container_width=True, num_rows="dynamic", height=280)
    for col in edited.columns:
        edited[col] = pd.to_numeric(edited[col], errors="coerce")
    S[var_key] = edited
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {label} CSV", to_csv_bytes(edited),
                           f"{label}.csv","text/csv",key=f"dl_{var_key}")
    with c2:
        if st.button(f"💾 Save to {filename}", key=f"sv_{var_key}"):
            edited.to_csv(p, index=False); st.success(f"Saved to {p}")


def set_container_status(db: str, val: str):
    for i in range(len(S["container_status"][db])):
        S["container_status"][db].at[i,"Status"] = val

# ═══════════════════════════════════════════════════════════════════════════════
# 10. PAGE CONFIG & CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Syndicate System", layout="wide",
                   initial_sidebar_state="collapsed")
st.markdown("""
<style>
  .block-container{padding-top:.4rem}
  .sec-hdr{font-size:.9rem;font-weight:900;color:#fff;padding:5px 14px;
            border-radius:5px;margin:4px 0 6px 0;display:inline-block}
  .hdr-blue{background:#1d3557} .hdr-green{background:#2d6a4f}
  .hdr-orange{background:#e07c00} .hdr-purple{background:#6d3b8e}
  .hdr-red{background:#c1121f} .hdr-teal{background:#0077b6}
  .hdr-brown{background:#6b4226}
  /* Status boxes — fixed foreground colour so they show on dark AND light themes */
  .note{background:#fff3cd;border-left:4px solid #e07c00;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c3d00 !important}
  .info{background:#cce5ff;border-left:4px solid #0077b6;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#003a5c !important}
  .warn{background:#ffd6d6;border-left:4px solid #c1121f;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c0000 !important}
  .ok  {background:#c3e6cb;border-left:4px solid #28a745;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#0a3d1f !important}
  /* Make all custom div text always visible */
  .note *, .info *, .warn *, .ok * { color: inherit !important; }

  /* ── Main navigation tabs (st.radio) — high visibility ───────────────── */
  /* Each tab rendered as a clear bordered pill so faint tabs are easy to read */
  div[role="radiogroup"] > label{
      background:#1b2436;
      border:1.5px solid #3a4a66;
      border-radius:8px;
      padding:7px 14px !important;
      margin:3px 6px 3px 0 !important;
      transition:all .12s ease;
  }
  div[role="radiogroup"] > label:hover{
      border-color:#5b8def;
      background:#243049;
  }
  /* Tab label text — brighter and bolder than default */
  div[role="radiogroup"] > label p{
      color:#dfe7f5 !important;
      font-weight:700 !important;
      font-size:.92rem !important;
  }
  /* The SELECTED tab — strong amber highlight so the active page is obvious */
  div[role="radiogroup"] > label:has(input:checked){
      background:#f4a000 !important;
      border-color:#ffd166 !important;
      box-shadow:0 0 0 2px rgba(244,160,0,.35);
  }
  div[role="radiogroup"] > label:has(input:checked) p{
      color:#1a1300 !important;
      font-weight:900 !important;
  }
</style>
""", unsafe_allow_html=True)

_init_state()
S = st.session_state.S

# ═══════════════════════════════════════════════════════════════════════════════
# 11. NAVIGATION
# ═══════════════════════════════════════════════════════════════════════════════
st.title("🎰 Syndicate System")
st.caption(f"Root: `{ROOT}`")

# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL SCRAPER — collapsible, above game selector, standalone
# ═══════════════════════════════════════════════════════════════════════════════
with st.expander("🕷️ Global Scraper — sweep all states, all games", expanded=False):
        st.markdown('<span class="sec-hdr hdr-blue">🕷️ Scraper — Global D Variable Sweep (all games)</span>',
                    unsafe_allow_html=True)
        st.markdown("""
        <div class="info">
        <b>This page is standalone — not game-specific.</b>
        It sweeps thelott.com for ALL syndicate games at once and saves raw state
        files to <code>Main_Data/</code>. After sweeping, click
        <b>🔀 Promote All + Split by Game</b> to route each row to its correct
        game folder. The game selector above does not affect this page.
        </div>
        """, unsafe_allow_html=True)
        am_toggle("scraper")

        # ── Connectivity notice (informational only — buttons always active) ───
        online = _is_online()
        if not online:
            st.markdown(
                '<div class="warn">⚠️ <b>Streamlit cannot reach thelott API directly</b> '
                '(Mac SSL restriction in browser process). '
                'Use the <b>terminal commands</b> in Section 7 below to run sweeps — '
                'they work perfectly from terminal. Cached data is shown here.</div>',
                unsafe_allow_html=True)

        # ── Info ───────────────────────────────────────────────────────────────
        st.markdown("""
        <div class="info">
        <b>Confirmed API</b> — two-step: outlets per postcode → syndicates per outlet batch.<br>
        No auth, no cookies, no Playwright required. Saves directly to <code>Main_Data/D_{STATE}_{STATE}.csv</code>.<br>
        Coverage: <b>NSW · VIC · QLD · SA · TAS</b> (ACT inside NSW data · NT=0 syndicates · WA=Lotterywest, in-store only).
        </div>
        """, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — DATA STATUS DASHBOARD
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📊 Data Status")

        status_rows = _data_status()
        col_labels = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
        for label in ["State", "File", "Rows", "Size", "Age", "Last Updated", "Freshness"]:
            col_labels[["State","File","Rows","Size","Age","Last Updated","Freshness"]
                        .index(label)].markdown(f"**{label}**")

        for row in status_rows:
            cols = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
            cols[0].write(row["state"])
            cols[1].write(f"`{row['file']}`")
            cols[2].write(f"{row['rows']:,}" if row["exists"] else "—")
            cols[3].write(f"{row['size_kb']} KB" if row["exists"] else "—")
            cols[4].write(f"{row['age_h']}h" if row["exists"] else "—")
            cols[5].write(row["modified"])
            if not row["exists"]:
                cols[6].markdown('<span style="color:#c1121f">⛔ Missing</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 6:
                cols[6].markdown('<span style="color:#28a745">🟢 Fresh</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 48:
                cols[6].markdown('<span style="color:#e07c00">🟡 Stale</span>',
                                 unsafe_allow_html=True)
            else:
                cols[6].markdown('<span style="color:#c1121f">🔴 Old</span>',
                                 unsafe_allow_html=True)

        total_rows = sum(r["rows"] for r in status_rows if r["exists"])
        freshest   = min((r["age_h"] for r in status_rows if r["exists"]), default=9999)
        oldest     = max((r["age_h"] for r in status_rows if r["exists"]), default=0)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total syndicates cached", f"{total_rows:,}")
        m2.metric("States with data", sum(1 for r in status_rows if r["exists"]))
        m3.metric("Freshest file", f"{freshest}h ago" if freshest < 9999 else "—")
        m4.metric("Oldest file",   f"{oldest}h ago"   if oldest > 0     else "—")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — SWEEP CONTROLS
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 🔄 Refresh Data")

        SWEEP_STATES  = ["NSW", "VIC", "QLD", "SA", "TAS"]
        SWEEP_POSTCODES = {
            "NSW": list(range(2000, 3000)),
            "VIC": list(range(3000, 4000)),
            "QLD": list(range(4000, 5000)),
            "SA":  list(range(5000, 5600)),
            "TAS": list(range(7000, 7800)),
        }

        c1, c2, c3 = st.columns([2, 1, 2])
        with c1:
            max_pc = st.number_input(
                "Max postcodes per state (0 = all)",
                min_value=0, max_value=9999, value=0,
                help="Set to a small number (e.g. 50) for a quick test run. 0 = full sweep."
            )
        with c2:
            st.write("")
            smart_skip = st.checkbox("Smart skip", value=True,
                                     help="Skip states where data is under 6 hours old")
        with c3:
            est_pcs = sum(min(max_pc or len(v), len(v))
                          for v in SWEEP_POSTCODES.values())
            est_min = max(round(est_pcs * 0.8 / 60), 1)
            st.metric("Est. full sweep time", f"~{est_min} min")

        # ── Per-state buttons ─────────────────────────────────────────────────
        st.markdown("**Sweep individual state:**")
        btn_cols = st.columns(len(SWEEP_STATES))
        triggered_state = None
        for i, s in enumerate(SWEEP_STATES):
            with btn_cols[i]:
                row = next(r for r in status_rows if r["state"] == s)
                age_label = f"{row['age_h']}h" if row["exists"] else "no data"
                if st.button(f"{s}\n({age_label})", key=f"sw_{s}",
                             use_container_width=True):
                    triggered_state = s

        # ── Sweep all button ──────────────────────────────────────────────────
        sweep_all = st.button(
            "🚀 SWEEP ALL STATES",
            type="primary", use_container_width=True,
        )

        def _run_sweep(states_to_sweep: list):
            """Execute sweep for given list of states, saving each immediately."""
            for state in states_to_sweep:
                row = next(r for r in status_rows if r["state"] == state)
                if smart_skip and row["exists"] and row["age_h"] < 6:
                    st.markdown(
                        f'<div class="info">⏭️ <b>{state}</b> skipped — data is only '
                        f'{row["age_h"]}h old (smart skip on).</div>',
                        unsafe_allow_html=True)
                    continue

                pcs = list(SWEEP_POSTCODES[state])
                if max_pc:
                    pcs = pcs[:max_pc]

                out_path = DIRS["Main_Data"] / f"D_{state}_{state}.csv"
                st.markdown(f"**▶ Sweeping {state} — {len(pcs):,} postcodes…**")
                pb  = st.progress(0)
                stx = st.empty()

                df_result = _sweep_state_thelott(state, pcs, pb, stx)
                pb.empty()

                if not df_result.empty:
                    df_result.to_csv(out_path, index=False)
                    st.markdown(
                        f'<div class="ok">✅ <b>{state}</b>: {len(df_result):,} syndicates '
                        f'→ <code>{out_path.name}</code></div>',
                        unsafe_allow_html=True)
                    st.dataframe(df_result.head(5), use_container_width=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  {state}  {len(df_result):,} rows")
                else:
                    st.markdown(
                        f'<div class="warn">⚠️ <b>{state}</b>: 0 syndicates found. '
                        f'Check connectivity or try again later.</div>',
                        unsafe_allow_html=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  {state}  0 rows — possible block")

        if triggered_state:
            _run_sweep([triggered_state])
        elif sweep_all:
            _run_sweep(SWEEP_STATES)

        # ── Health warning ────────────────────────────────────────────────────
        zero_states = [r["state"] for r in status_rows if r["exists"] and r["rows"] < 10]
        if zero_states:
            st.markdown(
                f'<div class="warn">⚠️ <b>Health check:</b> {", ".join(zero_states)} '
                f'returned very few records. The API may be throttling — '
                f'wait and retry, or check logs below.</div>',
                unsafe_allow_html=True)

        # ── Scrape log ────────────────────────────────────────────────────────
        if S["scrape_log"]:
            with st.expander("🪵 Sweep log"):
                for line in S["scrape_log"][-60:]:
                    st.text(line)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — QUICK API TEST (single postcode)
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("🔍 Quick API test — single postcode"):
            tc1, tc2, tc3 = st.columns([1.5, 1.5, 1])
            with tc1:
                test_state = st.selectbox("State", SWEEP_STATES, key="test_state")
            with tc2:
                defaults = {"NSW": 2000, "VIC": 3000, "QLD": 4000, "SA": 5000, "TAS": 7000}
                test_pc_val = st.number_input("Postcode", value=defaults.get(test_state, 2000),
                                              key="test_pc")
            with tc3:
                st.write(""); st.write("")
                do_test = st.button("▶ Test", key="do_test")

            if do_test:
                with st.spinner(f"Fetching outlets for {test_state} / {test_pc_val}…"):
                    outlets = _fetch_outlets(test_state, int(test_pc_val))
                st.write(f"**Outlets found:** {len(outlets)}  —  IDs: `{', '.join(outlets[:8])}`{'…' if len(outlets)>8 else ''}")
                if outlets:
                    with st.spinner("Fetching syndicates…"):
                        company = _STATE_COMPANY[test_state]
                        syns = _fetch_syndicates_batch(company, outlets[:20])
                    if syns:
                        rows = [_parse_syndicate_row(s, int(test_pc_val), test_state)
                                for s in syns]
                        st.markdown(
                            f'<div class="ok">✅ {len(rows)} syndicates returned.</div>',
                            unsafe_allow_html=True)
                        st.dataframe(pd.DataFrame(rows), use_container_width=True)
                    else:
                        st.markdown(
                            '<div class="warn">0 syndicates for these outlets.</div>',
                            unsafe_allow_html=True)
                else:
                    st.markdown(
                        '<div class="warn">No outlets found for this postcode.</div>',
                        unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — PROMOTE TO DIRECT/
        # ══════════════════════════════════════════════════════════════════════


with st.expander("🗂️ Game Breakdown — promote & split by game", expanded=False):
        # SECTION 4 — PROMOTE ALL + SPLIT BY GAME
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### ✅ Promote All → Split by Game")

        # ── Re-split warning if any game folders are empty ────────────────────
        empty_games = []
        for gk in GAME_KEYS:
            gd = game_dirs(gk)["Direct"]
            if not list(gd.glob("D_*.csv")):
                empty_games.append(f"{GAMES_CFG[gk]['emoji']} {GAMES_CFG[gk]['label']}")
        if empty_games:
            st.markdown(
                f'<div class="warn">⚠️ These games have no D files yet: '
                f'<b>{", ".join(empty_games)}</b><br>'
                f'Run <b>🔀 Promote All + Split by Game</b> to populate them. '
                f'Key fix: "Monday &amp; Wednesday Lotto" is now correctly mapped to MWF.</div>',
                unsafe_allow_html=True)

        st.markdown("""
        <div class="info">
        One click: reads all D_*.csv files from <b>Main_Data/</b>, splits every row
        by its <b>Games</b> column, saves game-specific files into each game's
        <code>Variables/Variable_Elements/Direct/</code> folder.<br>
        Multi-game syndicates (e.g. "Oz Lotto | Powerball") → a copy in <b>each</b>
        matching game folder.<br>
        Mapped: TattsLotto/Gold Lotto/X Lotto → SAT &nbsp;·&nbsp;
        Monday &amp; Wednesday Lotto → MWF &nbsp;·&nbsp;
        Super 66/Lucky Lotteries → skipped.
        </div>
        """, unsafe_allow_html=True)

        raw_d_files = sorted(DIRS["Main_Data"].glob("D_*.csv"))

        if raw_d_files:
            # Show quick game breakdown from first available file
            with st.expander("📊 Preview game counts before splitting"):
                for src in raw_d_files[:2]:  # show first 2 to avoid slowness
                    try:
                        df_prev = pd.read_csv(src, usecols=["Games"])
                        gc = df_prev["Games"].value_counts().reset_index()
                        gc.columns = ["Games value", "Rows"]
                        st.write(f"**{src.name}:**")
                        st.dataframe(gc, use_container_width=True, height=200)
                    except Exception:
                        pass

            st.write(f"**Files ready to split:** {[f.name for f in raw_d_files]}")

            if st.button("🔀 PROMOTE ALL + SPLIT BY GAME",
                         type="primary", use_container_width=True,
                         key="promote_all_split"):
                all_results = {}
                progress = st.progress(0)
                for fi, src in enumerate(raw_d_files):
                    progress.progress((fi + 1) / len(raw_d_files),
                                      text=f"Splitting {src.name}…")
                    result = split_d_by_game(src, ROOT)
                    all_results[src.name] = result
                progress.empty()

                # Summary table
                st.markdown("**Split results:**")
                summary_rows = []
                all_unknowns = set()
                all_skipped  = set()
                for fname, res in all_results.items():
                    unknowns = res.pop("_unknown_games", [])
                    skipped  = res.pop("_skipped_games", [])
                    all_unknowns.update(unknowns)
                    all_skipped.update(skipped)
                    for gkey, count in res.items():
                        if count > 0:
                            summary_rows.append({
                                "Source file":   fname,
                                "Game":          GAMES_CFG[gkey]["label"],
                                "Rows written":  f"{count:,}",
                                "Saved to":      f"Games/{gkey.upper()}/…/Direct/",
                            })

                if summary_rows:
                    st.dataframe(pd.DataFrame(summary_rows),
                                 use_container_width=True, height=300)
                    st.markdown('<div class="ok">✅ Split complete — '
                                'each game folder now has its own D variable files.</div>',
                                unsafe_allow_html=True)
                else:
                    st.warning("No rows were written. Check the Games column values.")

                if all_unknowns:
                    st.markdown(
                        f'<div class="warn">⚠️ Unrecognised game names (skipped): '
                        f'<b>{sorted(all_unknowns)}</b><br>'
                        f'Add these to GAME_NAME_MAP in masterapp.py if needed.</div>',
                        unsafe_allow_html=True)
                if all_skipped:
                    st.markdown(
                        f'<div class="info">ℹ️ Intentionally skipped '
                        f'(supplementary games): {sorted(all_skipped)}</div>',
                        unsafe_allow_html=True)

        else:
            st.info("No D_*.csv files in Main_Data/ yet. Run a sweep first.")

        # ── Single file promote (legacy) ──────────────────────────────────────
        with st.expander("📤 Promote single file to Direct/ (legacy)"):
            promote_candidates = (sorted(DIRS["Main_Data"].glob("D_*.csv")) +
                                  sorted(DIRS["Scraper"].glob("D_*.csv")))
            if promote_candidates:
                chosen = st.selectbox("File to promote:",
                                      [f"{fp.parent.name}/{fp.name}"
                                       for fp in promote_candidates],
                                      key="promote_sel")
                chosen_fp = next(fp for fp in promote_candidates
                                 if f"{fp.parent.name}/{fp.name}" == chosen)
                c1p, c2p = st.columns([2, 1])
                with c1p:
                    df_prev = pd.read_csv(chosen_fp)
                    st.write(f"{len(df_prev):,} rows · {list(df_prev.columns)}")
                with c2p:
                    st.write("")
                    if st.button("📤 Promote to Direct/",
                                 use_container_width=True, key="do_promote"):
                        import shutil
                        dst = DIRS["Direct"] / chosen_fp.name
                        shutil.copy2(chosen_fp, dst)
                        S["D"] = d_to_w_only(_load_file(dst))
                        st.success(f"Promoted → Direct/{chosen_fp.name} "
                                   f"({len(S['D']):,} rows loaded)")
            else:
                st.info("No D_*.csv files found.")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 5 — DATA BROWSER
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📂 Cached Data Browser")
        for row in status_rows:
            if not row["exists"]:
                continue
            fp = DIRS["Main_Data"] / row["file"]
            with st.expander(
                f"📄 {row['file']}  —  {row['rows']:,} rows  ·  {row['size_kb']} KB  "
                f"·  updated {row['modified']}"
            ):
                df_view = pd.read_csv(fp)
                # Show game breakdown
                if "Games" in df_view.columns:
                    st.markdown("**Games breakdown:**")
                    gc = df_view["Games"].value_counts().reset_index()
                    gc.columns = ["Game", "Rows"]
                    st.dataframe(gc, use_container_width=True, height=200)
                st.dataframe(df_view.head(10), use_container_width=True)
                st.download_button(f"⬇ Download {row['file']}", to_csv_bytes(df_view),
                                   row["file"], "text/csv",
                                   key=f"dl_view_{row['state']}")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 6 — UPLOAD PRE-SCRAPED FILE
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("📤 Upload a pre-scraped CSV"):
            up = st.file_uploader("CSV or Excel", type=["csv", "xlsx"],
                                  key="scraper_upload")
            if up:
                df_up = (pd.read_excel(up) if up.name.endswith(".xlsx")
                         else pd.read_csv(up))
                sp = DIRS["Main_Data"] / (up.name if up.name.endswith(".csv")
                                          else up.name.replace(".xlsx", ".csv"))
                df_up.to_csv(sp, index=False)
                st.markdown(f'<div class="ok">✅ {sp.name} — {len(df_up):,} rows saved '
                            f'to Main_Data/</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 7 — TERMINAL COMMANDS + SCHEDULER
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("⏰ Run sweeps from terminal + Schedule nightly (Mac)"):
            script_path = Path(__file__).resolve()
            scraper_py  = script_path.parent / "thelott_syndicate_scraper.py"
            log_path    = ROOT / "logs" / "scraper.log"
            st.markdown(f"""
        **Why use terminal?** Streamlit on Mac has SSL restrictions that can block
        outbound API calls. Running from terminal bypasses this completely.

        **Sweep individual states:**
        ```bash
        cd {script_path.parent}
        python3 thelott_syndicate_scraper.py sweep NSW
        python3 thelott_syndicate_scraper.py sweep VIC
        python3 thelott_syndicate_scraper.py sweep QLD
        python3 thelott_syndicate_scraper.py sweep SA
        python3 thelott_syndicate_scraper.py sweep TAS
        ```

        **Sweep all states at once:**
        ```bash
        python3 thelott_syndicate_scraper.py sweep ALL
        ```

        **After sweeping — split by game (run in terminal):**
        ```bash
        python3 -c "
        import sys; sys.path.insert(0, '{script_path.parent}')
        from masterapp import split_d_by_game, ROOT, DIRS
        from pathlib import Path
        for f in sorted(DIRS['Main_Data'].glob('D_*.csv')):
        r = split_d_by_game(f, ROOT)
        print(f.name, r)
        "
        ```

        **Schedule nightly at 2am (crontab):**
        ```bash
        crontab -e
        # Add this line:
        0 2 * * * cd {script_path.parent} && python3 thelott_syndicate_scraper.py sweep ALL >> {log_path} 2>&1
        ```
            """)


st.markdown("---")

# ── GAME SELECTOR ─────────────────────────────────────────────────────────────
_game_cols = st.columns(len(GAME_KEYS) + 1)
with _game_cols[0]:
    st.markdown("**Select Game:**")
for _i, _gk in enumerate(GAME_KEYS):
    with _game_cols[_i + 1]:
        _cfg = GAMES_CFG[_gk]
        _active = (active_game() == _gk)
        _btn_type = "primary" if _active else "secondary"
        if st.button(
            f"{_cfg['emoji']} {_cfg['label']}\n{_cfg['draw_day']}",
            key=f"game_btn_{_gk}",
            type=_btn_type,
            use_container_width=True,
        ):
            st.session_state["active_game"] = _gk
            st.rerun()

_gcfg  = active_game_cfg()
_gdirs = active_game_dirs()
st.markdown(
    f'<div class="info">🎮 Active game: <b>{_gcfg["emoji"]} {_gcfg["label"]}</b> '
    f'— Pool: 1–{_gcfg["pool"]} · Pick {_gcfg["pick"]} · Draws: {_gcfg["draw_day"]} '
    f'· Data folder: <code>Games/{active_game().upper()}/</code></div>',
    unsafe_allow_html=True,
)
st.markdown("---")

page = st.radio("", [
    "📥 Main Data",
    "🔄 CVI Matrix",
    "🧩 Variable Inputs",
    "📦 Container Formula",
    "🖥️ Container Dashboards",
    "📤 Master Outputs",
    "🗂️ Cluster Manager",
], horizontal=True, label_visibility="collapsed")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MAIN DATA
# ═══════════════════════════════════════════════════════════════════════════════
if page == "📥 Main Data":
    st.markdown('<span class="sec-hdr hdr-teal">📥 Main Data — Manually uploaded each run</span>',
                unsafe_allow_html=True)

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    Main Data for <b>{_gcfg['emoji']} {_gcfg['label']}</b> —
    your historical combination dataset matched against variable inputs.<br>
    Saved to disk immediately — <b>survives browser refresh</b>.<br>
    <b>Large files (63M rows):</b> save CSV to the Main_Data folder directly —
    no upload needed. The app reads from disk.<br>
    <b>accdb files:</b> place them directly in
    <code>Games/{_gkey.upper()}/Main_Data/</code> — Streamlit cannot upload
    Access databases via browser (browser security restriction).
    Install <code>mdb-tools</code> via Homebrew to read them.
    </div>
    """, unsafe_allow_html=True)

    # ── Folder path hint ───────────────────────────────────────────────────
    with st.expander("📁 Drop large files here directly (no upload needed)"):
        st.code(str(_gdirs["Main_Data"]), language=None)
        st.markdown("""
**For 63M row files:**
```bash
# Copy your main data file directly — no size limit
cp /path/to/your/maindata.csv ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/
```
Then click **🔄 Scan folder** below — the app reads it from disk.

**For accdb files:**
```bash
# 1. Install mdb-tools if not done
brew install mdb-tools

# 2. Convert accdb to CSV (no size limit)
mdb-export /path/to/file.accdb TableName > ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/maindata.csv
```
        """)

    # ── Upload (for smaller files under config limit) ──────────────────────
    up = st.file_uploader(
        "Upload Main Data file (CSV/Excel — for files under 10GB after config update)",
        type=["csv", "xlsx", "xls", "txt", "xml"],
        key="up_main_data")
    if up:
        path = _gdirs["Main_Data"] / f"MAIN_{up.name}"
        path.write_bytes(up.getvalue())
        st.markdown(
            f'<div class="ok">✅ Saved to disk: <code>{path.name}</code> '
            f'({path.stat().st_size / 1024 / 1024:.1f} MB) — '
            f'persists across refresh.</div>',
            unsafe_allow_html=True)

    # ── Scan folder for ALL files (disk-first approach) ────────────────────
    if st.button("🔄 Scan Main_Data folder", key="scan_main_data",
                 use_container_width=True):
        st.rerun()

    existing = sorted(
        list(_gdirs["Main_Data"].glob("MAIN_*")) +
        list(_gdirs["Main_Data"].glob("*.csv")) +
        list(_gdirs["Main_Data"].glob("*.xlsx")) +
        list(_gdirs["Main_Data"].glob("*.accdb")) +
        list(_gdirs["Main_Data"].glob("*.mdb"))
    )
    # Also check legacy root Main_Data
    existing += sorted(
        list(DIRS["Main_Data"].glob("1n")) +
        [DIRS["Main_Data"] / "1n" / f
         for f in ["Main_Data.xlsx", "Main_Data.csv"]
         if (DIRS["Main_Data"] / "1n" / f).exists()]
    )
    existing = list(dict.fromkeys(existing))  # deduplicate

    if existing:
        st.markdown(f"**{len(existing)} file(s) found in Main_Data folder:**")
        for fp in existing:
            if not fp.is_file():
                continue
            size_mb = fp.stat().st_size / 1024 / 1024
            # Fast row count for CSV (no load)
            if fp.suffix == ".csv":
                rows = _count_csv_rows(fp)
                row_label = f"{rows:,} rows"
            else:
                row_label = f"{size_mb:.1f} MB"

            c1, c2, c3, c4 = st.columns([3, 1.2, 1, 0.8])
            c1.write(f"`{fp.name}`")
            c2.write(row_label)
            c3.write(f"{size_mb:.1f} MB")
            with c4:
                if st.button("Load", key=f"load_md_{fp.name}"):
                    with st.spinner(f"Loading {fp.name}…"):
                        df_ex = _load_file(fp, numeric=False)
                    if not df_ex.empty:
                        S["main_data"]      = df_ex
                        S["main_data_path"] = str(fp)
                        st.success(f"Loaded {fp.name}: {len(df_ex):,} rows")
                    else:
                        st.error(f"Could not read {fp.name}")

    # ── Auto-reload from disk if session state lost ────────────────────────
    if S.get("main_data", pd.DataFrame()).empty:
        saved_path = S.get("main_data_path", "")
        if saved_path and Path(saved_path).exists():
            with st.spinner("Reloading main data from disk…"):
                S["main_data"] = _load_file(Path(saved_path), numeric=False)
        else:
            # Auto-load most recent file in folder
            auto_files = sorted(
                _gdirs["Main_Data"].glob("*.csv"),
                key=lambda f: f.stat().st_mtime, reverse=True)
            if auto_files:
                newest = auto_files[0]
                st.markdown(
                    f'<div class="info">ℹ️ Auto-loading most recent file: '
                    f'<b>{newest.name}</b></div>', unsafe_allow_html=True)
                with st.spinner(f"Loading {newest.name}…"):
                    S["main_data"] = _load_file(newest, numeric=False)
                S["main_data_path"] = str(newest)

    md = S.get("main_data", pd.DataFrame())
    if not md.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Rows", f"{len(md):,}")
        c2.metric("Columns", len(md.columns))
        c3.metric("Est. match time (17 formulas)",
                  f"~{max(len(md)*7*17//16//1_000_000, 1)} sec")
        st.markdown("**Preview (first 20 rows):**")
        st.dataframe(md.head(20), use_container_width=True)
        st.download_button("⬇ Export Main Data CSV",
                           to_csv_bytes(md), "main_data.csv", "text/csv",
                           key="dl_main_data")
    else:
        st.info("No Main Data loaded. Upload a file, drop it in the folder, "
                "or click Load on an existing file above.")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CVI MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 CVI Matrix":
    st.markdown('<span class="sec-hdr hdr-teal">🔄 CVI Matrix — Transposition & Variable Slicing</span>',
                unsafe_allow_html=True)
    am_toggle("cvi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b><br>
    Reads D from <code>Games/{_gkey.upper()}/Variables/Variable_Elements/Direct/</code>
    (game-specific — run <b>Promote All + Split by Game</b> on the Scraper page first).<br>
    Each syndicate row → one w-column. Sorted longest→shortest.
    Ep = top 8 w-columns · Sp = top 4 lanes (a,b,c,d) · So = union combis.
    </div>
    """, unsafe_allow_html=True)

    # Game-specific Direct/ folder
    direct_dir = _gdirs["Direct"]
    raw_files  = sorted(direct_dir.glob("D_*.csv"))

    # Fallback: also show global Main_Data files if game-specific not yet split
    global_files = sorted(DIRS["Main_Data"].glob("D_*.csv"))

    if not raw_files and not global_files:
        st.warning("No D files found. Run the Scraper, then Promote All + Split by Game.")
    else:
        if raw_files:
            st.markdown(f'<div class="ok">✅ {len(raw_files)} game-specific D file(s) '
                        f'found in Games/{_gkey.upper()}/Direct/</div>',
                        unsafe_allow_html=True)
            all_files = raw_files
            file_labels = [f.name for f in all_files]
        else:
            st.markdown(
                f'<div class="warn">⚠️ No game-specific D files yet for '
                f'{_gcfg["label"]}. Showing global Main_Data files — '
                f'run Promote All + Split by Game first.</div>',
                unsafe_allow_html=True)
            all_files = global_files
            file_labels = [f"Main_Data/{f.name}" for f in all_files]

        chosen = st.selectbox("Source D file:", file_labels, key="cvi_d_sel")
        chosen_fp = all_files[file_labels.index(chosen)]
        df_raw = pd.read_csv(chosen_fp)
        st.write(f"**{len(df_raw):,} rows · {len(df_raw.columns)} columns**")
        st.dataframe(df_raw.head(5), use_container_width=True)

        if st.button("🔄 BUILD W-MATRIX & SLICE Ep / Sp / So",
                     type="primary", use_container_width=True):
            with st.spinner("Transposing…"):
                w_mat = build_w_matrix(df_raw)
            if w_mat.empty:
                st.error("No number columns found. A D (syndicate) file needs "
                         "w1…wN columns (main-data n1…nN is also accepted).")
            else:
                mfname = chosen_fp.name.replace("D_", "CVI_Matrix_")
                # Save CVI matrix to game-specific CVI folder
                cvi_out = _gdirs["CVI"] / mfname
                w_mat.to_csv(cvi_out, index=False)
                st.markdown(f'<div class="ok">✅ W-Matrix: {len(w_mat.columns)} w-columns × '
                            f'{len(w_mat)} rows → saved to Games/{_gkey.upper()}/</div>',
                            unsafe_allow_html=True)

                slices = slice_variables(w_mat)
                # The D variable used by the formula must be the CLEAN w-only
                # matrix — NOT df_raw, which still carries metadata columns
                # (Syndicate_ID, Draw_Number, Outlet_ID, Postcode, Share_Cost…).
                # Without this, collation pulled those metadata values in as if
                # they were numbers (the 4687/1400286/2000 leak seen in BRD).
                S["D"] = w_mat
                # Save Ep, Sp, So to game-specific folders
                stem = chosen_fp.stem.split("_", 1)[1] if "_" in chosen_fp.stem else chosen_fp.stem
                slices["Ep"].to_csv(_gdirs["ExcelPro"] / f"Ep_{stem}.csv", index=False)
                slices["Sp"].to_csv(_gdirs["Splits"]   / f"Sp_{stem}.csv", index=False)
                slices["So"].to_csv(_gdirs["Splits_Combi"] / f"So_{stem}.csv", index=False)

                S["Ep"] = slices["Ep"]
                S["Sp"] = slices["Sp"]
                S["So"] = slices["So"]

                st.markdown(
                    f'<div class="ok">✅ Ep, Sp, So saved to '
                    f'Games/{_gkey.upper()}/Variables/Variable_Elements/</div>',
                    unsafe_allow_html=True)

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown("**Ep (top 8)**")
                    st.dataframe(slices["Ep"], height=200, use_container_width=True)
                with c2:
                    st.markdown("**Sp lanes a–d** *(Sp=S renamed)*")
                    st.dataframe(slices["Sp"], height=200, use_container_width=True)
                with c3:
                    st.markdown("**So (union)**")
                    st.dataframe(slices["So"], height=200, use_container_width=True)

        # Browse existing CVI matrices for this game
        cvi_files = sorted(_gdirs["CVI"].glob("CVI_Matrix_*.csv"))
        if cvi_files:
            st.markdown("---")
            cm = st.selectbox("Inspect matrix:", [f.name for f in cvi_files],
                              key="cvi_inspect_sel")
            df_m = pd.read_csv(_gdirs["CVI"] / cm)
            st.write(f"{len(df_m.columns)} w-columns · depth: {len(df_m)} rows")
            st.dataframe(df_m, use_container_width=True, height=300)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: VARIABLE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧩 Variable Inputs":
    st.markdown('<span class="sec-hdr hdr-green">🧩 Variable Inputs — B · R · D · Ep · Sp · So</span>',
                unsafe_allow_html=True)
    am_toggle("vi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b> — Pool 1–{_gcfg['pool']},
    Pick {_gcfg['pick']}, draws {_gcfg['draw_day']}.<br>
    <b>B</b> = pre-loaded w-columns from f_rules_Gclaude.xlsx (rarely changes).<br>
    <b>Ep</b> = ExcelPro: top 8 w-columns of D → new w-sets.<br>
    <b>Sp</b> = Splits (task1b): top 4 w-columns of D + 4 split points → split sets.<br>
    <b>So</b> = SplitsCombi (auto_vba): same top 4 w-columns → union combinations.<br>
    <b>R</b>  = Rainbow (task2): Since Last from lottolyzer → powerset combos.<br>
    <b>D</b>  = Syndicate w-columns (standalone + feeds formula row 11).
    </div>
    """, unsafe_allow_html=True)

    vtabs = st.tabs(["B (Base)", "R (Rainbow)", "D (Direct)",
                     "Ep (ExcelPro)", "Sp (Splits)", "So (SplitsCombi)",
                     "Since Last"])

    # ── TAB: B (Base) ──────────────────────────────────────────────────────
    with vtabs[0]:
        st.markdown(f"**B — Base variable for {_gcfg['label']}** "
                    f"(sheet: `{_gcfg['b_sheet']}` in f_rules_Gclaude.xlsx)")

        # Load B from uploaded rules file
        b_rules_candidates = list(ROOT.rglob("f_rules_Gclaude.xlsx"))
        if b_rules_candidates:
            b_rules_path = b_rules_candidates[0]
            try:
                xl_b = pd.ExcelFile(b_rules_path, engine="openpyxl")
                sheet = _gcfg["b_sheet"]
                if sheet in xl_b.sheet_names:
                    df_b_raw = xl_b.parse(sheet, header=None)
                    # Row 0 = w-column headers, rows 1+ = data
                    w_cols_b = [str(df_b_raw.iloc[0, c])
                                for c in range(df_b_raw.shape[1])
                                if str(df_b_raw.iloc[0, c]).startswith("w")]
                    b_data = {}
                    for i, wc in enumerate(w_cols_b):
                        col_vals = df_b_raw.iloc[1:, i].dropna()
                        nums = [int(float(v)) for v in col_vals
                                if str(v).replace(".", "").replace("-", "").isdigit()
                                and float(v) >= 1]
                        if nums:
                            b_data[wc] = pd.Series(nums)
                    if b_data:
                        df_b = pd.DataFrame(b_data)
                        S["B"] = df_b
                        st.markdown(
                            f'<div class="ok">✅ B loaded: {len(b_data)} w-columns '
                            f'from sheet <b>{sheet}</b></div>',
                            unsafe_allow_html=True)
                        # Show w-column lengths
                        len_info = {wc: len(b_data[wc].dropna()) for wc in b_data}
                        st.write("Column lengths:", len_info)
                        st.dataframe(df_b, use_container_width=True, height=280)
                        st.download_button("⬇ Download B.csv",
                                           to_csv_bytes(df_b), "B.csv",
                                           "text/csv", key="dl_b_rules")
                    else:
                        st.warning(f"Sheet '{sheet}' found but no numeric w-columns parsed.")
                else:
                    st.warning(f"Sheet '{sheet}' not found. Available: {xl_b.sheet_names}")
            except Exception as ex:
                st.error(f"Error reading f_rules_Gclaude.xlsx: {ex}")
        else:
            st.markdown('<div class="warn">f_rules_Gclaude.xlsx not found in project folder. '
                        'Upload it below.</div>', unsafe_allow_html=True)
            up_b = st.file_uploader("Upload f_rules_Gclaude.xlsx",
                                    type=["xlsx"], key="up_b_rules")
            if up_b:
                dest = ROOT / "f_rules_Gclaude.xlsx"
                dest.write_bytes(up_b.read())
                st.success(f"Saved to {dest}. Refresh the page.")

    # ── TAB: R (Rainbow) ───────────────────────────────────────────────────
    with vtabs[1]:
        st.markdown("**R — Rainbow (task2.py): Since Last → powerset combos**")

        sl_file = _gdirs["SinceLast"] / "since_last.json"
        if sl_file.exists():
            try:
                sl_data = json.loads(sl_file.read_text())
                since_last_dict = {int(k): int(v)
                                   for k, v in sl_data.get("since_last_dict", {}).items()}
                all_wt = sl_data.get("all_wt", [])
                to_keep_list = sl_data.get("to_keep", [])
                scraped_at = sl_data.get("scraped_at", "unknown")

                st.markdown(
                    f'<div class="ok">✅ Since Last loaded — {len(since_last_dict)} numbers '
                    f'| scraped: {scraped_at[:16]}</div>',
                    unsafe_allow_html=True)

                st.write(f"**all_wt** (first 15 / most recent → oldest): "
                         f"`{all_wt[:15]}...`")

                max_comb = st.slider("Max Since Last groups to combine:", 1, 8, 3,
                                     key="r_max_comb")

                if st.button("▶ Generate Rainbow (R)", key="gen_R",
                             type="primary", use_container_width=True):
                    from itertools import chain, combinations as _combos

                    def _powerset(iterable):
                        s = list(iterable)
                        return chain.from_iterable(
                            _combos(s, r) for r in range(1, len(s)+1))

                    # Group by Since Last (+1 as in task2.py)
                    grouped: dict[int, list] = {}
                    for num, sl in since_last_dict.items():
                        grouped.setdefault(int(sl) + 1, []).append(num)

                    sl_keys = sorted(grouped.keys())
                    combos = [c for c in _powerset(sl_keys)
                              if 1 <= len(c) <= max_comb]
                    to_keep_set = set(all_wt)
                    result = {}
                    for combo in combos:
                        ref = []
                        for sl_val in combo:
                            ref += grouped[sl_val]
                        keep_ = [n for n in ref if n in to_keep_set]
                        result[str(combo)] = keep_

                    r_df = pd.DataFrame(
                        {k: pd.Series(v) for k, v in result.items()})
                    order = r_df.isna().sum().sort_values().index
                    r_df = r_df[order]
                    S["R"] = r_df

                    r_path = _gdirs["Rainbow"] / f"R_{_gkey}.csv"
                    r_df.to_csv(r_path, index=False)
                    st.markdown(
                        f'<div class="ok">✅ R generated: {r_df.shape[1]} columns '
                        f'→ saved to {r_path.name}</div>',
                        unsafe_allow_html=True)
                    st.dataframe(r_df.head(20), use_container_width=True, height=280)

            except Exception as ex:
                st.error(f"Error loading Since Last: {ex}")
        else:
            st.markdown(
                '<div class="warn">⚠️ Since Last data not found for this game. '
                'Go to the <b>Since Last</b> tab to fetch it from lottolyzer.</div>',
                unsafe_allow_html=True)

        if not S.get("R", pd.DataFrame()).empty:
            st.markdown("**Current R in memory:**")
            st.dataframe(S["R"].head(10), use_container_width=True, height=200)

    # ── TAB: D (Direct) ────────────────────────────────────────────────────
    with vtabs[2]:
        st.markdown(f"**D — Syndicate data for {_gcfg['label']}**")
        d_files = (sorted(_gdirs["Main_Data"].glob("D_*.csv")) +
                   sorted(_gdirs["Direct"].glob("D*.csv")) +
                   sorted(_gdirs["Direct"].glob("D*.xlsx")))
        if d_files:
            ch_d = st.selectbox("File:", [f.name for f in d_files], key="sel_d_file")
            df_d = _load_file(d_files[[f.name for f in d_files].index(ch_d)])
            S["D"] = df_d
            st.write(f"{len(df_d):,} rows · {len(df_d.columns)} cols")
            st.dataframe(df_d.head(40), use_container_width=True, height=300)
        else:
            st.info(f"No D files in Games/{_gkey.upper()}/. Run the Scraper and split by game.")

    # ── TAB: Ep (ExcelPro) ─────────────────────────────────────────────────
    with vtabs[3]:
        st.markdown("**Ep — ExcelPro: top 8 w-columns of D → new w-sets**")

        # Load objects from B (the w-columns)
        b_df = S.get("B", pd.DataFrame())
        d_df = S.get("D", pd.DataFrame())

        if b_df.empty:
            st.warning("Load B variable first (Base tab).")
        elif d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            # Get all w-columns from D, sorted longest → shortest
            w_cols_d = [c for c in d_df.columns
                        if re.match(r'^w\d+$', c) or
                        re.match(r'^n\d+$', c)]
            if not w_cols_d:
                # D from scraper — try to extract number columns
                n_cols_d = sorted(
                    [c for c in d_df.columns if re.match(r'^n\d+$', c, re.I)],
                    key=lambda x: int(re.sub(r'\D', '', x) or 0))
                if n_cols_d:
                    w_cols_d = n_cols_d

            st.write(f"D has {len(w_cols_d)} w/n columns. "
                     f"Top 8 will feed ExcelPro.")

            if len(w_cols_d) >= 1:
                # Sort by column length descending
                col_lens = {c: d_df[c].dropna().shape[0] for c in w_cols_d}
                sorted_cols = sorted(col_lens, key=col_lens.get, reverse=True)
                top8 = sorted_cols[:8]
                st.write(f"Top 8 (longest→shortest): `{top8}`")

                if st.button("▶ Run ExcelPro (Ep)", type="primary",
                             key="run_ep", use_container_width=True):
                    try:
                        # Build objects from B w-columns (4 pairs)
                        b_w_cols = [c for c in b_df.columns
                                    if re.match(r'^w\d+$', c)]
                        # Pair them: (w1,w2)=a, (w3,w4)=b, (w5,w6)=c, (w7,w8)=d
                        objects_dict = {}
                        pair_labels = ["a", "b", "c", "d"]
                        for pi, label in enumerate(pair_labels):
                            i1, i2 = pi * 2, pi * 2 + 1
                            if i1 < len(b_w_cols):
                                arr1 = b_df[b_w_cols[i1]].dropna().astype(int).tolist()
                                arr2 = (b_df[b_w_cols[i2]].dropna().astype(int).tolist()
                                        if i2 < len(b_w_cols) else [])
                                objects_dict[label] = {
                                    "title": label,
                                    "sub1": f"{label}1",
                                    "sub2": f"{label}2",
                                    "arr1": arr1,
                                    "arr2": arr2,
                                }

                        # Build all_wt from top8 D columns
                        d_nums = []
                        seen_ep = set()
                        for col in top8:
                            for v in d_df[col].dropna():
                                try:
                                    n = int(float(v))
                                    if n >= 1 and n not in seen_ep:
                                        d_nums.append(n)
                                        seen_ep.add(n)
                                except Exception: pass

                        if objects_dict and d_nums:
                            from itertools import combinations as _ep_combos

                            # Build all_wt from D numbers
                            all_wt_ep = d_nums

                            # Build occurrence matrix
                            result_cols = {}
                            for label, obj in objects_dict.items():
                                set1 = set(obj["arr1"])
                                set2 = set(obj["arr2"])
                                lane_a = [n for n in all_wt_ep if n in set1]
                                lane_b = [n for n in all_wt_ep if n in set2]
                                result_cols[f"wt_{label}_1"] = pd.Series(lane_a)
                                result_cols[f"wt_{label}_2"] = pd.Series(lane_b)

                            ep_df = pd.DataFrame(result_cols)
                            S["Ep"] = ep_df
                            ep_path = _gdirs["ExcelPro"] / f"Ep_{_gkey}.csv"
                            ep_df.to_csv(ep_path, index=False)
                            st.markdown(
                                f'<div class="ok">✅ Ep generated: {ep_df.shape[1]} columns '
                                f'→ {ep_path.name}</div>',
                                unsafe_allow_html=True)
                            st.dataframe(ep_df.head(20),
                                         use_container_width=True, height=280)
                    except Exception as ex:
                        st.error(f"ExcelPro error: {ex}")
            else:
                st.warning("D variable has no numeric columns to process.")

        if not S.get("Ep", pd.DataFrame()).empty:
            st.markdown("**Current Ep in memory:**")
            st.dataframe(S["Ep"].head(10), use_container_width=True, height=180)

    # ── TAB: Sp (Splits) ───────────────────────────────────────────────────
    with vtabs[4]:
        st.markdown("**Sp — Splits (task1b.py): top 4 w-columns of D + 4 split points**")

        d_df = S.get("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)

            top4 = w_cols_d[:4]
            st.write(f"Top 4 w-columns → Sp input: `{top4}`")

            sets_ready = {}
            for col in top4:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready[col] = vals

            if sets_ready:
                st.markdown("**Set split points** (0 = midpoint auto):")
                sp_cols = st.columns(4)
                splitters = []
                for i, (col, vals) in enumerate(sets_ready.items()):
                    with sp_cols[i]:
                        default_split = len(vals) // 2
                        sp = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_split,
                            key=f"sp_split_{col}_{i}")
                        splitters.append(sp if sp > 0 else default_split)

                if st.button("▶ Run Splits (Sp)", type="primary",
                             key="run_sp", use_container_width=True):
                    try:
                        from itertools import combinations as _sp_combos

                        keys = list(sets_ready.keys())
                        split_sets = {}
                        for j, key in enumerate(keys):
                            vals = sets_ready[key]
                            sp_val = splitters[j]
                            split_sets[key + "0"] = vals[:sp_val]
                            split_sets[key + "1"] = vals[sp_val:]

                        universe = set(n for v in sets_ready.values() for n in v)
                        comb3 = [c for r in range(3, 4)
                                 for c in _sp_combos(keys, r)]
                        comb2 = [c for c in _sp_combos(keys, 2)]

                        result_sp = {}
                        # Split sets
                        for i, key in enumerate(sorted(split_sets.keys())):
                            result_sp[key] = split_sets[key]
                            result_sp[f"y{i}"] = list(
                                universe - set(split_sets[key]))

                        # 3-combinations
                        let_3 = ["e", "f", "g", "h"]
                        comb3dict = {}
                        for j, combo in enumerate(comb3):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb3dict[let_3[j % 4] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb3dict.items():
                            result_sp[key] = list(val)

                        # 2-combinations
                        let_2 = ["i", "j", "k", "l", "m", "n"]
                        comb2dict = {}
                        for j, combo in enumerate(comb2):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb2dict[let_2[j % 6] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb2dict.items():
                            result_sp[key] = list(val)

                        sp_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_sp.items()})
                        order_sp = sp_df.isna().sum().sort_values().index
                        sp_df = sp_df[order_sp]
                        S["Sp"] = sp_df
                        sp_path = _gdirs["Splits"] / f"Sp_{_gkey}.csv"
                        sp_df.to_csv(sp_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ Sp generated: {sp_df.shape[1]} columns '
                            f'→ {sp_path.name}</div>',
                            unsafe_allow_html=True)
                        st.dataframe(sp_df.head(20),
                                     use_container_width=True, height=280)
                    except Exception as ex:
                        st.error(f"Splits error: {ex}")
            else:
                st.warning("Could not extract numeric data from D columns.")

        if not S.get("Sp", pd.DataFrame()).empty:
            st.markdown("**Current Sp in memory:**")
            st.dataframe(S["Sp"].head(10), use_container_width=True, height=160)

    # ── TAB: So (SplitsCombi) ──────────────────────────────────────────────
    with vtabs[5]:
        st.markdown("**So — SplitsCombi (automation_vba.py): top 4 w-columns → union combis**")

        d_df = S.get("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d_so = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)
            top4_so = w_cols_d_so[:4]
            st.write(f"Top 4 w-columns → So input: `{top4_so}`")

            sets_ready_so = {}
            for col in top4_so:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready_so[col] = vals

            if sets_ready_so and len(sets_ready_so) == 4:
                st.markdown("**Set split points for So** (0 = midpoint auto):")
                so_cols = st.columns(4)
                splitters_so = []
                for i, (col, vals) in enumerate(sets_ready_so.items()):
                    with so_cols[i]:
                        default_s = len(vals) // 2
                        sv = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_s,
                            key=f"so_split_{col}_{i}")
                        splitters_so.append(sv if sv > 0 else default_s)

                if st.button("▶ Run SplitsCombi (So)", type="primary",
                             key="run_so", use_container_width=True):
                    try:
                        from itertools import combinations as _so_combos
                        from collections import OrderedDict

                        keys_so = list(sets_ready_so.keys())
                        split_sets_so = {}
                        for j, key in enumerate(keys_so):
                            vals = sets_ready_so[key]
                            sv = splitters_so[j]
                            split_sets_so[key + "0"] = vals[:sv]
                            split_sets_so[key + "1"] = vals[sv:]

                        universe_so = set(
                            n for v in sets_ready_so.values() for n in v)

                        comb_pairs = {
                            str(elem): set(split_sets_so[elem[0]]).union(
                                set(split_sets_so[elem[1]]))
                            for elem in _so_combos(split_sets_so, 2)
                        }
                        comb_three = {
                            str(elem): set(sets_ready_so[elem[0]]).union(
                                set(sets_ready_so[elem[1]]),
                                set(sets_ready_so[elem[2]]))
                            for elem in _so_combos(keys_so, 3)
                        }

                        result_so: dict = {}
                        result_so["U"] = universe_so

                        for set_i in comb_three:
                            for set_j in comb_pairs:
                                if (comb_pairs[set_j].issubset(comb_three[set_i])
                                        and set_j[2] in set_i
                                        and set_j[8] in set_i):
                                    result_so[f"U-{set_i}-{set_j}"] = \
                                        universe_so - (comb_three[set_i] -
                                                       comb_pairs[set_j])
                                    result_so[f"{set_i}-{set_j}"] = \
                                        comb_three[set_i] - comb_pairs[set_j]
                            result_so[str(set_i)] = comb_three[set_i]

                        for set_j in comb_pairs:
                            result_so[f"U-{set_j}"] = universe_so - comb_pairs[set_j]
                            result_so[str(set_j)] = comb_pairs[set_j]

                        so_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_so.items()})
                        order_so = so_df.isna().sum().sort_values().index
                        so_df = so_df[order_so]
                        S["So"] = so_df
                        so_path = _gdirs["Splits_Combi"] / f"So_{_gkey}.csv"
                        so_df.to_csv(so_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ So generated: {so_df.shape[1]} columns '
                            f'→ {so_path.name}</div>',
                            unsafe_allow_html=True)
                        st.dataframe(so_df.head(20),
                                     use_container_width=True, height=280)
                    except Exception as ex:
                        st.error(f"SplitsCombi error: {ex}")
            elif sets_ready_so:
                st.info(f"So engine needs exactly 4 w-columns. Found {len(sets_ready_so)}.")
            else:
                st.warning("Could not extract numeric data from D columns.")

        if not S.get("So", pd.DataFrame()).empty:
            st.markdown("**Current So in memory:**")
            st.dataframe(S["So"].head(10), use_container_width=True, height=160)

    # ── TAB: Since Last ────────────────────────────────────────────────────
    with vtabs[6]:
        st.markdown(f"**Since Last — fetch from lottolyzer for {_gcfg['label']}**")
        sl_url = _gcfg["lottolyzer"]
        st.markdown(f"Source: [{sl_url}]({sl_url})")

        sl_file = _gdirs["SinceLast"] / "since_last.json"
        if sl_file.exists():
            try:
                sl_data = json.loads(sl_file.read_text())
                scraped_at = sl_data.get("scraped_at", "unknown")[:16]
                n_nums = len(sl_data.get("since_last_dict", {}))
                st.markdown(
                    f'<div class="ok">✅ Since Last cached — {n_nums} numbers '
                    f'| scraped: {scraped_at}</div>',
                    unsafe_allow_html=True)
                # Show table
                sl_dict = {int(k): int(v)
                           for k, v in sl_data.get("since_last_dict", {}).items()}
                all_wt = sl_data.get("all_wt", [])
                sl_display = pd.DataFrame([
                    {"Rank": i+1, "Number": num,
                     "Since_Last": sl_dict.get(num, "?"),
                     "Group": ("🔴 Last Draw" if sl_dict.get(num,99)==0
                               else "🟠 Very Recent" if sl_dict.get(num,99)<=3
                               else "🟡 Recent" if sl_dict.get(num,99)<=9
                               else "🟢 Moderate" if sl_dict.get(num,99)<=19
                               else "🔵 Old" if sl_dict.get(num,99)<=29
                               else "⚫ Very Old")}
                    for i, num in enumerate(all_wt)
                ])
                st.dataframe(sl_display, use_container_width=True, height=400)
            except Exception as ex:
                st.error(f"Error loading Since Last: {ex}")

        st.markdown("---")
        st.markdown("**Manual upload** (paste from lottolyzer CSV export):")
        up_sl = st.file_uploader(
            "Upload Since Last CSV (columns: Number, Since Last)",
            type=["csv", "xlsx"], key="up_sl")
        if up_sl:
            try:
                df_sl = (pd.read_excel(up_sl, engine="openpyxl")
                         if up_sl.name.endswith(".xlsx")
                         else pd.read_csv(up_sl))
                # Find Since Last column
                sl_col = next((c for c in df_sl.columns
                               if "since" in c.lower()), None)
                num_col = next((c for c in df_sl.columns
                                if "number" in c.lower() or c == "Number"), None)
                if sl_col and num_col:
                    sl_dict_up = {int(row[num_col]): int(row[sl_col])
                                  for _, row in df_sl.iterrows()
                                  if pd.notna(row[sl_col]) and pd.notna(row[num_col])}
                    pool = _gcfg["pool"]
                    all_wt_up = sorted(sl_dict_up.keys(),
                                       key=lambda n: (sl_dict_up[n], n))
                    # Same ordering as all_wt; kept as a separate list so the two
                    # keys can diverge later without surprises.
                    to_keep_up = list(all_wt_up)
                    save_data = {
                        "since_last_dict": {str(k): v
                                             for k, v in sl_dict_up.items()},
                        "all_wt": all_wt_up,
                        "to_keep": to_keep_up,
                        "game": _gkey,
                        "game_name": _gcfg["label"],
                        "pool_size": pool,
                        "scraped_at": datetime.now().isoformat(),
                        "url": sl_url,
                    }
                    _gdirs["SinceLast"].mkdir(parents=True, exist_ok=True)
                    sl_file.write_text(json.dumps(save_data, indent=2))
                    st.success(f"Since Last saved: {len(sl_dict_up)} numbers. "
                               f"Refresh page to see.")
                else:
                    st.error("Could not find 'Number' and 'Since Last' columns.")
            except Exception as ex:
                st.error(f"Upload error: {ex}")

        st.markdown("---")
        st.markdown("""
        **To get Since Last manually from lottolyzer:**
        1. Go to the link above
        2. Export or copy the frequency table
        3. Save as CSV with columns: `Number`, `Since Last`
        4. Upload above
        """)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER FORMULA
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📦 Container Formula":
    st.markdown('<span class="sec-hdr hdr-purple">📦 Container Formula — Live Collation</span>',
                unsafe_allow_html=True)
    am_toggle("cf")

    st.markdown("""
    <div class="info">
    Each row = 1 container. 17 rows = 17 containers running in parallel (Python multiprocessing).
    No Docker needed — Python processes handle parallel execution natively.
    Sp = S throughout. Collation: drop col-0 of each component, concatenate, relabel w1…wN.
    </div>
    """, unsafe_allow_html=True)

    # Variable status
    st.markdown("**Variables in memory:**")
    vc = st.columns(6)
    for i, k in enumerate(["B","R","D","Sp","So","Ep"]):
        with vc[i]:
            df = S.get(k, pd.DataFrame())
            if isinstance(df, pd.DataFrame) and not df.empty:
                st.markdown(f'<div class="ok">✅ {k} {len(df)}r×{len(df.columns)}c</div>',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="warn">⚠️ {k} empty</div>',
                            unsafe_allow_html=True)

    st.markdown("---")
    cf_df = pd.DataFrame({
        "#": [r[0] for r in CF_ROWS],
        "Name": [r[1] for r in CF_ROWS],
        "Type": [r[2] for r in CF_ROWS],
        "Components": [" + ".join(r[3]) for r in CF_ROWS],
        "Active": [S["cf_active"].get(r[1],True) for r in CF_ROWS],
    })
    edited_cf = st.data_editor(cf_df, key="cf_tbl",
        use_container_width=True, num_rows="fixed", height=530,
        column_config={
            "Active": st.column_config.CheckboxColumn("Active"),
            "Type":   st.column_config.SelectboxColumn("Type",
                          options=["Formed","Ready-made"]),
        })
    for _, row in edited_cf.iterrows():
        S["cf_active"][row["Name"]] = row["Active"]

    st.markdown("---")
    active_names = [r[1] for r in CF_ROWS if S["cf_active"].get(r[1],True)]
    chosen_f = st.selectbox("Collate formula:", active_names)
    comps = COMP_MAP.get(chosen_f, [])
    st.write(f"**Components:** {' + '.join(comps)}")
    missing = [c for c in comps if c not in S or
               (isinstance(S.get(c), pd.DataFrame) and S[c].empty)]
    if missing:
        st.markdown(f'<div class="warn">⚠️ Missing/empty: {missing}</div>',
                    unsafe_allow_html=True)

    if st.button(f"▶ Collate {chosen_f}", type="primary", use_container_width=True):
        with st.spinner("Collating…"):
            result = execute_collation(comps)
        if result.empty:
            st.error("Result empty — load components in Variable Inputs first.")
        else:
            out = DIRS["CVI"] / f"CVI_{chosen_f}.csv"
            result.to_csv(out, index=False)
            S["cvi"][chosen_f] = result
            st.markdown(f'<div class="ok">✅ {chosen_f}: {len(result)} rows × '
                        f'{len(result.columns)} cols → <code>{out.name}</code></div>',
                        unsafe_allow_html=True)
            st.dataframe(result, use_container_width=True, height=260)
            st.download_button(f"⬇ CVI_{chosen_f}.csv", to_csv_bytes(result),
                               f"CVI_{chosen_f}.csv","text/csv")

    if S["cvi"]:
        with st.expander("📋 All collated CVIs in memory"):
            for fname, df in S["cvi"].items():
                st.markdown(f"**{fname}** — {len(df)} rows × {len(df.columns)} cols")
                st.dataframe(df.head(5), use_container_width=True, height=120)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER DASHBOARDS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🖥️ Container Dashboards":
    st.markdown('<span class="sec-hdr hdr-red">🖥️ Container Dashboards</span>',
                unsafe_allow_html=True)
    am_toggle("cd")

    # ── PARALLEL RUN — runs all active containers simultaneously ──────────
    st.markdown("### ▶ Run All Active Containers in Parallel")
    st.markdown("""
    <div class="info">
    Fires all CVI files found in <code>Container_Variable_Inputs/</code>
    simultaneously using Python multiprocessing.
    Each container gets its own worker process.
    Carry-forward defaults to all-U for parallel run.
    </div>
    """, unsafe_allow_html=True)

    # Cluster config
    clusters = load_clusters()
    next_id  = next_cluster_id(clusters)
    all_main_files = scan_main_data_files()
    all_cvi_files  = scan_cvi_files()

    p1,p2,p3,p4 = st.columns(4)
    with p1:
        st.metric("Next Cluster ID", next_id)
    with p2:
        cluster_label = st.text_input("Cluster label:", f"1n",
                                      key="par_cluster_label",
                                      help="e.g. 1n, Draw1567, May28")
    with p3:
        lotto_sel = st.selectbox("Lotto type:",
                                 list(LOTTO_TYPES.keys()),
                                 format_func=lambda x: f"{x} — {LOTTO_TYPES[x]}",
                                 key="par_lotto")
    with p4:
        draw_date_par = st.text_input("Draw date (YYYY_MM_DD):",
                                      datetime.now().strftime("%Y_%m_%d"),
                                      key="par_date")

    # Auto-scan Main Data
    matching_main = [f for f in all_main_files
                     if f.get("lotto") == lotto_sel or not f.get("lotto")]
    if matching_main:
        main_choice = st.selectbox(
            "Main Data file (auto-scanned from Main_Data/):",
            [f["raw"] for f in matching_main],
            format_func=lambda x: f"{x}  ({next(f['rows'] for f in matching_main if f['raw']==x):,} rows)",
            key="par_main_choice"
        )
        chosen_main_info = next(f for f in matching_main if f["raw"] == main_choice)
        draw_no_par = chosen_main_info.get("draw","D?")
    else:
        st.warning("No Main Data files found in Main_Data/. "
                   f"Expected pattern: {{cluster}}_{lotto_sel}_D{{draw_no}}.csv")
        main_choice  = None
        draw_no_par  = "D?"

    # Show available CVI files
    matching_cvi = [f for f in all_cvi_files
                    if f.get("lotto") == lotto_sel or not f.get("lotto")]
    st.markdown(f"**CVI files found for `{lotto_sel}`: {len(matching_cvi)}**")
    if matching_cvi:
        cvi_preview_df = pd.DataFrame([{
            "Formula": f["formula"], "Date": f["date"], "File": f["raw"]
        } for f in matching_cvi])
        st.dataframe(cvi_preview_df, use_container_width=True,
                     hide_index=True, height=min(40*len(cvi_preview_df)+50,220))
    else:
        st.warning("No CVI files found. Collate formulas in Container Formula first.")

    run_all_btn = st.button(
        f"▶ RUN ALL {len(matching_cvi)} CONTAINERS IN PARALLEL",
        type="primary", use_container_width=True,
        key="run_all_par",
        disabled=(not main_choice or not matching_cvi)
    )

    if run_all_btn and main_choice and matching_cvi:
        main_path  = DIRS["Main_Data"] / main_choice
        output_dir = DIRS["Outputs"] / f"Cluster_{next_id}_{cluster_label}"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Build worker args
        worker_args = []
        for cvi_info in matching_cvi:
            formula = cvi_info["formula"]
            sc_path = DIRS["Selected_Counts"] / f"SC_{formula}.csv"
            worker_args.append((
                formula,
                cvi_info["path"],
                str(main_path),
                str(sc_path) if sc_path.exists() else "",
                next_id,
                cluster_label,
                lotto_sel,
                draw_no_par,
                draw_date_par,
                str(output_dir),
            ))

        st.info(f"Launching {len(worker_args)} workers…")
        prog = st.progress(0, text="Starting workers…")

        # Pool and cpu_count are imported at module top.
        n_workers = min(len(worker_args), cpu_count())
        st.caption(f"Using {n_workers} CPU cores of {cpu_count()} available")

        with Pool(processes=n_workers) as pool:
            results = []
            for i, res in enumerate(
                pool.imap_unordered(_parallel_worker, worker_args)
            ):
                results.append(res)
                prog.progress(
                    (i+1)/len(worker_args),
                    text=f"Completed {i+1}/{len(worker_args)}: "
                         f"{res['formula']} — "
                         f"{'✅' if res['status']=='complete' else '❌'} "
                         f"S:{res['selected_n']} U:{res['unselected_n']}"
                )

        prog.empty()

        # Save cluster to registry
        new_cluster = {
            "id":          next_id,
            "label":       cluster_label,
            "lotto":       lotto_sel,
            "draw_no":     draw_no_par,
            "draw_date":   draw_date_par,
            "main_file":   main_choice,
            "output_dir":  str(output_dir),
            "containers":  len(worker_args),
            "status":      "complete",
            "results":     results,
            "timestamp":   datetime.now().isoformat(),
        }
        clusters.append(new_cluster)
        save_clusters(clusters)

        # Show summary
        ok  = [r for r in results if r["status"] == "complete"]
        err = [r for r in results if r["status"] != "complete"]
        st.success(f"✅ Cluster {next_id} '{cluster_label}' complete — "
                   f"{len(ok)} succeeded, {len(err)} errors")
        res_df = pd.DataFrame([{
            "Formula":    r["formula"],
            "Status":     r["status"],
            "Selected":   r["selected_n"],
            "Unselected": r["unselected_n"],
            "Error":      r.get("error",""),
        } for r in results])
        st.dataframe(res_df, use_container_width=True,
                     hide_index=True, height=min(40*len(res_df)+50,320))
        if err:
            for r in err:
                st.error(f"❌ {r['formula']}: {r['error']}")

    st.markdown("---")

    # ── Dashboard stack ────────────────────────────────────────────────────
    st.markdown("**Individual dashboard (open one at a time):**")
    # Build list from all CVI files + hardcoded defaults
    cvi_found = [parse_cvi_filename(f.name)["formula"]
                 for f in sorted(DIRS["CVI"].glob("CVI_*.csv"))]
    all_db_names = list(dict.fromkeys(
        [f"1n & {f}" for f in cvi_found] + DASHBOARDS
    ))
    for db_name in all_db_names:
        st.markdown(
            f'<div style="background:#1a1a2e;border:1px solid #444;color:#ccc;'
            f'font-size:.71rem;padding:2px 12px;margin-bottom:2px;'
            f'display:inline-block;min-width:260px;">'
            f'{db_name}…  Container Dashboard</div><br>',
            unsafe_allow_html=True)

    st.markdown("---")
    db = st.selectbox("Open Dashboard:", all_db_names,
                      format_func=lambda x: f"{x}… Container Dashboard")
    formula_name = db.replace("1n & ","")

    st.markdown(f'<div style="font-size:1rem;font-weight:900;border:2px solid #555;'
                f'padding:5px 14px;display:inline-block;margin-bottom:8px;">'
                f'{db}…  Container Dashboard</div>', unsafe_allow_html=True)

    # ── Auto-load CVI for this formula ────────────────────────────────────
    cvi_df = S["cvi"].get(formula_name)
    if cvi_df is None:
        # Try all CVI files for this formula (any lotto type, any date)
        for cvi_fp in sorted(DIRS["CVI"].glob(f"CVI_*{formula_name}*.csv")):
            cvi_df = _load_file(cvi_fp)
            if not cvi_df.empty:
                S["cvi"][formula_name] = cvi_df
                break
        if cvi_df is None:
            for search_dir in [DIRS["CVI"], DIRS["Rainbow"]]:
                cvi_path = search_dir / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_df = _load_file(cvi_path)
                    break
    if cvi_df is None:
        cvi_df = pd.DataFrame()

    # ── Auto-load SC ───────────────────────────────────────────────────────
    sc_auto = {}
    sc_folder = DIRS["Selected_Counts"]
    for sc_file in sc_folder.glob(f"SC_{formula_name}*.csv"):
        try:
            df_sc = pd.read_csv(sc_file)
            if "w" in df_sc.columns and "Selected Count" in df_sc.columns:
                for _, row in df_sc.iterrows():
                    sc_auto[str(row["w"])] = str(row["Selected Count"])
        except Exception: pass

    # ── Auto-scan and load Main Data ──────────────────────────────────────
    main_df = S.get("main_data", pd.DataFrame())
    avail_main = scan_main_data_files()

    # ── Preview panel ──────────────────────────────────────────────────────
    st.markdown("#### Preview — Loaded Data")
    pv1, pv2, pv3 = st.columns(3)
    with pv1:
        if not cvi_df.empty:
            w_cols_prev = [c for c in cvi_df.columns
                           if re.match(r'^w\d+$', c)]
            st.success(f"✅ **CVI** — {len(w_cols_prev)} w-columns × "
                       f"{len(cvi_df)} rows")
            with st.expander("Preview CVI"):
                st.dataframe(cvi_df.head(10), use_container_width=True,
                             hide_index=True, height=220)
        else:
            st.warning("⚠️ CVI not loaded")

    with pv2:
        if sc_auto:
            st.success(f"✅ **SC** — {len(sc_auto)} w-columns")
            with st.expander("Preview SC"):
                st.dataframe(pd.DataFrame([
                    {"w": k, "SC": v} for k,v in sc_auto.items()
                ]), use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ SC not loaded")

    with pv3:
        if not main_df.empty:
            # Validate it's actually main data (has n-columns, not w-columns)
            n_check = [c for c in main_df.columns if re.match(r'^n\d+$', c, re.I)]
            w_check = [c for c in main_df.columns if re.match(r'^w\d+$', c)]
            if w_check and not n_check:
                st.error("⚠️ Loaded file looks like a CVI file (w-columns). "
                         "Please load correct Main Data below.")
                S["main_data"] = pd.DataFrame()
                main_df = pd.DataFrame()
            else:
                st.success(f"✅ **Main Data** — {len(main_df):,} rows × "
                           f"{len(main_df.columns)} cols")
            with st.expander("Preview Main Data"):
                if avail_main:
                    sel_main = st.selectbox(
                        "Switch Main Data file:",
                        [f["raw"] for f in avail_main],
                        key=f"switch_main_{db}"
                    )
                    if st.button("Load selected", key=f"load_main_{db}"):
                        chosen = next(
                            f for f in avail_main if f["raw"] == sel_main)
                        main_df = _load_file(Path(chosen["path"]))
                        S["main_data"] = main_df
                        S["main_data_path"] = chosen["path"]
                        st.success(f"Loaded {sel_main}")
                        st.rerun()
                st.dataframe(main_df.head(10), use_container_width=True,
                             hide_index=True, height=200)
        else:
            st.warning("⚠️ Main Data not loaded")
            if avail_main:
                sel_main2 = st.selectbox(
                    "Available Main Data files:",
                    [f["raw"] for f in avail_main],
                    key=f"sel_main2_{db}"
                )
                if st.button("Load", key=f"load_main2_{db}", type="primary"):
                    chosen2 = next(
                        f for f in avail_main if f["raw"] == sel_main2)
                    main_df = _load_file(Path(chosen2["path"]))
                    S["main_data"] = main_df
                    S["main_data_path"] = chosen2["path"]
                    st.rerun()

    st.markdown("---")

    # ── Data Status + Re-upload (always visible, always replaceable) ──────
    st.markdown("#### Data Status")
    col_s1, col_s2 = st.columns(2)

    # ── CVI ───────────────────────────────────────────────────────────────
    with col_s1:
        if not cvi_df.empty:
            st.success(f"✅ CVI loaded: {len(cvi_df.columns)} columns, "
                       f"{len(cvi_df)} rows")
            # Always show replace button even when loaded
            if st.button("🔄 Replace CVI", key=f"clr_cvi_{db}",
                         use_container_width=True):
                S["cvi"].pop(formula_name, None)
                # Remove from disk so it won't auto-reload
                cvi_path = DIRS["CVI"] / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_path.unlink()
                cvi_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning(f"⚠️ No CVI loaded for {formula_name}")

        # Upload always visible — key uses a counter so it resets after clear
        upload_key_cvi = f"up_cvi_{db}_{S.get('cvi_upload_v', {}).get(db, 0)}"
        up_cvi = st.file_uploader(
            "Upload CVI file (csv/xlsx):",
            type=["csv","xlsx"], key=upload_key_cvi
        )
        if up_cvi:
            df_cvi_up = pd.read_csv(up_cvi) if up_cvi.name.endswith(".csv") \
                        else pd.read_excel(up_cvi)
            cvi_save = DIRS["CVI"] / f"CVI_{formula_name}.csv"
            df_cvi_up.to_csv(cvi_save, index=False)
            S["cvi"][formula_name] = df_cvi_up
            # Bump version so key resets next time Replace is clicked
            if "cvi_upload_v" not in S: S["cvi_upload_v"] = {}
            S["cvi_upload_v"][db] = S["cvi_upload_v"].get(db, 0) + 1
            cvi_df = df_cvi_up
            st.success(f"✅ CVI saved → Container_Variable_Inputs/CVI_{formula_name}.csv")
            st.rerun()

    # ── Main Data ─────────────────────────────────────────────────────────
    with col_s2:
        if not main_df.empty:
            st.success(f"✅ Main Data: {len(main_df):,} rows × "
                       f"{len(main_df.columns)} cols")
            if st.button("🔄 Replace Main Data", key=f"clr_md_{db}",
                         use_container_width=True):
                S["main_data"] = pd.DataFrame()
                S["main_data_path"] = ""   # drop stale path so DuckDB can't misfire
                main_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning("⚠️ No Main Data loaded")

        upload_key_md = f"up_md_{db}_{S.get('md_upload_v', {}).get(db, 0)}"
        up_md = st.file_uploader(
            "Upload Main Data file (csv/xlsx):",
            type=["csv","xlsx"], key=upload_key_md
        )
        if up_md:
            df_md_up = pd.read_csv(up_md) if up_md.name.endswith(".csv") \
                       else pd.read_excel(up_md)
            for col in df_md_up.columns:
                df_md_up[col] = pd.to_numeric(df_md_up[col], errors="coerce")
            S["main_data"] = df_md_up
            # Uploaded in-memory data has no stable on-disk source; clear the
            # path so the DuckDB pre-pass guard won't match it to another file.
            S["main_data_path"] = ""
            if "md_upload_v" not in S: S["md_upload_v"] = {}
            S["md_upload_v"][db] = S["md_upload_v"].get(db, 0) + 1
            main_df = df_md_up
            st.success(f"✅ Main Data: {len(df_md_up):,} rows × "
                       f"{len(df_md_up.columns)} cols")
            st.rerun()

    st.markdown("---")
    # Selected Count
    st.markdown("**Selected Count input**")

    # Show if auto-loaded from file
    if sc_auto:
        st.success(f"✅ Selected Counts loaded from SC_{formula_name}*.csv")
        st.caption("  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))

    # SC upload — always replaceable
    with st.expander("📂 Upload / Replace Selected Count file"):
        st.caption("File must have columns: w, Selected Count")
        upload_key_sc = f"up_sc_{db}_{S.get('sc_upload_v', {}).get(db, 0)}"
        up_sc = st.file_uploader("Upload SC file", type=["csv","xlsx"],
                                  key=upload_key_sc)
        if up_sc:
            df_sc_up = pd.read_csv(up_sc) if up_sc.name.endswith(".csv") \
                       else pd.read_excel(up_sc)
            sp_sc = DIRS["Selected_Counts"] / f"SC_{formula_name}.csv"
            df_sc_up.to_csv(sp_sc, index=False)
            if "sc_upload_v" not in S: S["sc_upload_v"] = {}
            S["sc_upload_v"][db] = S["sc_upload_v"].get(db, 0) + 1
            st.success(f"✅ Saved → SC_{formula_name}.csv")
            st.rerun()

    c1,c2 = st.columns(2)
    with c1:
        sc_method = st.radio("Method:", ["Custom","Same for all","Range"],
                             horizontal=True, key=f"scm_{db}")
    with c2:
        # Show per-column summary — do NOT merge into one string
        if sc_auto:
            st.caption("Per-column SC (from file): " +
                       "  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))
        fallback_sc = st.text_input(
            "Fallback Count (used if no SC file loaded):", "7",
            key=f"scr_{db}")

    # Build sc_dict: per w-column, specific threshold
    # sc_auto keys look like "w1","w2",… or "1","2",… normalise to "w1","w2"
    sc_dict = {}
    for k, v in sc_auto.items():
        key = k if k.startswith("w") else f"w{k}"
        sc_dict[key] = [int(x.strip()) for x in str(v).split(",")
                        if x.strip().lstrip('-').isdigit()]

    # If no SC file loaded, use fallback for all w-columns
    if not sc_dict and fallback_sc:
        fallback_list = [int(x.strip()) for x in fallback_sc.split(",")
                         if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = fallback_list

    if sc_method == "Same for all" and fallback_sc:
        same_list = [int(x.strip()) for x in fallback_sc.split(",")
                     if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = same_list

    # ── Carry-forward direction toggles ───────────────────────────────────
    st.markdown("---")
    st.markdown("**Carry-forward direction — which pool feeds the next row**")
    st.caption("U = Unselected carries forward (default)  |  "
               "S = Selected carries forward  |  "
               "Toggle at Row N controls what Row N receives from Row N-1.")

    # Initialise carry_fwd in session state per dashboard
    cf_key = f"carry_fwd_{db}"
    if cf_key not in S:
        S[cf_key] = {}

    # Get w-columns from loaded CVI or fallback
    w_keys_cf = sorted(
        [c for c in (cvi_df.columns if not cvi_df.empty else [])
         if re.match(r'^w\d+$', c)],
        key=lambda x: int(x[1:])
    )

    # Global toggle
    g1, g2, _ = st.columns([1, 1, 6])
    with g1:
        if st.button("🌐 All → U", key=f"gall_u_{db}", use_container_width=True):
            for wk in w_keys_cf:
                S[cf_key][wk] = "U"
    with g2:
        if st.button("🌐 All → S", key=f"gall_s_{db}", use_container_width=True):
            for wk in w_keys_cf:
                S[cf_key][wk] = "S"

    # Per-row toggles
    if w_keys_cf:
        row_cols = st.columns(len(w_keys_cf))
        for i, wk in enumerate(w_keys_cf):
            cur = S[cf_key].get(wk, "U")
            with row_cols[i]:
                st.caption(wk)
                if st.button(
                    f"{'🔵 U' if cur=='U' else '🟢 S'}",
                    key=f"cf_{db}_{wk}",
                    use_container_width=True,
                    help=f"{wk}: currently carrying {'Unselected' if cur=='U' else 'Selected'} forward"
                ):
                    S[cf_key][wk] = "S" if cur == "U" else "U"
                    st.rerun()

    # Build final carry_fwd dict (defaults to U)
    carry_fwd = {wk: S[cf_key].get(wk, "U") for wk in w_keys_cf}

    # ── Is SC Available toggle ────────────────────────────────────────────
    sc_avail_key = f"sc_avail_{db}"
    if sc_avail_key not in S:
        S[sc_avail_key] = "YES"

    # ── Individual run cluster label ───────────────────────────────────────
    ind_cl1, ind_cl2, ind_cl3 = st.columns([2, 2, 4])
    with ind_cl1:
        ind_cluster = st.text_input(
            "Cluster label for this run:",
            value="1n",
            key=f"ind_cluster_{db}",
            help="Used in output filenames: e.g. 1n, Draw1567, Test01"
        )
    with ind_cl2:
        ind_lotto = st.selectbox(
            "Lotto type:",
            list(LOTTO_TYPES.keys()),
            format_func=lambda x: f"{x}",
            key=f"ind_lotto_{db}"
        )
    with ind_cl3:
        ind_draw = st.text_input(
            "Draw number:",
            value="D????",
            key=f"ind_draw_{db}",
            help="e.g. D1567"
        )

    # File prefix for individual run outputs
    ind_prefix = f"{ind_cluster}_{ind_lotto}_{ind_draw}_{formula_name}"
    st.markdown("---")
    sa1, sa2, _ = st.columns([1,1,6])
    with sa1:
        if st.button("🟢 SC: YES (Auto)", key=f"sca_y_{db}",
                     type="primary" if S[sc_avail_key]=="YES" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "YES"
    with sa2:
        if st.button("🔴 SC: NO (Manual)", key=f"sca_n_{db}",
                     type="primary" if S[sc_avail_key]=="NO" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "NO"
    if S[sc_avail_key] == "NO":
        st.info("Manual mode: SC not provided upfront. "
                "Process pauses after each stage for researcher to provide SC.")
    else:
        st.caption("Automated mode: all stages run sequentially without stopping.")

    # RUN MATCHING
    st.markdown("---")
    can_run = not main_df.empty and not cvi_df.empty
    if not can_run:
        st.info("📤 Upload CVI and Main Data above to enable matching.")

    if st.button(f"▶ RUN MATCHING — {db}", type="primary",
                 use_container_width=True, key=f"run_{db}",
                 disabled=not can_run):
        n = len(main_df)
        st.info(f"Matching M={n:,} rows × {len(w_keys_cf)} w-columns…")
        with st.spinner("Matching engine running…"):
            # Pass the source path (if known). run_matching only uses it when
            # it is safe — large CSV, DuckDB present, row-count == M, contiguous
            # n-columns — and falls back to pandas otherwise.
            res = run_matching(main_df, cvi_df, sc_dict, carry_fwd,
                               main_path=S.get("main_data_path"))
        S["results"][db] = res

    # ── Display results (persists after run) ──────────────────────────────
    if db in S["results"]:
        res   = S["results"][db]
        fig9  = res["fig9_table"]
        sel   = res["selected"]
        unsel = res["unselected"]
        bd    = res["breakdown"]
        dbg   = res.get("debug_rows", [])
        small = res.get("small_enough", True)
        n_cols_res = res.get("n_cols", [])

        st.success(f"✅ Final stage — Selected: {len(sel):,}  ·  "
                   f"Unselected: {len(unsel):,}")

        # ── Helper: show df with count filter + highlights inline ──────
        def show_filtered_highlighted(df_in, cvi_set_in, n_cols_in,
                                      section_key, default_counts=None):
            if df_in is None or df_in.empty:
                st.info("No data.")
                return
            count_col = "Count" if "Count" in df_in.columns else None
            avail_counts = []
            if count_col:
                avail_counts = sorted(df_in[count_col].dropna()
                                      .astype(int).unique().tolist())
            if avail_counts:
                filt = st.multiselect(
                    "Filter by count:", avail_counts,
                    default=default_counts or avail_counts,
                    key=f"flt_{db}_{section_key}"
                )
                df_show = df_in[df_in[count_col].isin(filt)] if filt else df_in
            else:
                df_show = df_in
            if df_show.empty:
                st.info("No rows match the selected count filter.")
                return
            # Highlight matching cells
            nc = [c for c in n_cols_in if c in df_show.columns]
            if nc and cvi_set_in:
                def _style(val):
                    try:
                        if int(round(float(val))) in cvi_set_in:
                            return ("background-color:#d4a0a0;"
                                    "color:#000;font-weight:bold")
                    except (ValueError, TypeError):
                        pass
                    return ""
                def _cnt_style(val):
                    return "background-color:#ffffaa;color:#000;font-weight:bold"
                try:
                    styled = df_show.style.applymap(_style, subset=nc)
                    if count_col:
                        styled = styled.applymap(_cnt_style,
                                                 subset=[count_col])
                    st.dataframe(styled, use_container_width=True,
                                 height=min(40*len(df_show)+50, 340),
                                 hide_index=True)
                except Exception:
                    st.dataframe(df_show, use_container_width=True,
                                 height=min(40*len(df_show)+50, 340),
                                 hide_index=True)
            else:
                st.dataframe(df_show, use_container_width=True,
                             height=min(40*len(df_show)+50, 340),
                             hide_index=True)

        # ── Compact summary table ──────────────────────────────────────
        st.markdown("#### Matching Table")
        if not fig9.empty:
            # Rename index to start from 1
            fig9_disp = fig9.copy()
            fig9_disp.index = range(1, len(fig9_disp)+1)
            st.dataframe(fig9_disp, use_container_width=True, height=310)
            c_dl1, c_dl2 = st.columns(2)
            with c_dl1:
                st.download_button("⬇ Matching Table CSV",
                                   to_csv_bytes(fig9),
                                   f"{ind_prefix}_matching_table.csv",
                                   "text/csv", key=f"dl_fig9_{db}")

        # ── Inline expanders — one per w-column ───────────────────────
        for i, d in enumerate(dbg):
            w_lbl    = d["w"]
            cvi_set_d = d.get("cvi_set", set(d.get("cvi_numbers", [])))
            n_cols_d  = d.get("n_cols", [])
            icon      = "✅" if d["selected_n"] > 0 else (
                        "⚪" if d["present_in"] == 0 else "🔵")
            carry_d   = d.get("direction","U")
            note_d    = d.get("note","")

            exp_header = (
                f"{icon} Row {int(w_lbl[1:])}  ·  {w_lbl}  ·  "
                f"Dir:{carry_d}  ·  Present:{d['present_in']}  ·  "
                f"SC:{d['sc']}  ·  "
                f"S:{d['selected_n']}  ·  U:{d['unselected_n']}"
                + (f"  ·  ⚠️{note_d}" if note_d else "")
            )

            with st.expander(exp_header):
                if not cvi_set_d:
                    st.warning("No CVI numbers — column is empty. "
                               "All present rows carry forward.")
                    continue

                # CVI numbers + distribution side by side
                cca, ccb = st.columns(2)
                with cca:
                    st.markdown(f"**{w_lbl} numbers ({len(cvi_set_d)}):**")
                    st.write(sorted(cvi_set_d))

                with ccb:
                    st.markdown("**Count distribution (present rows):**")
                    if d["count_dist"]:
                        dist_df = pd.DataFrame([
                            {"Count": int(k[1:]),
                             "Rows": v,
                             "": "✅ Sel" if int(k[1:]) in d["sc"]
                                 else "— Unsel"}
                            for k,v in sorted(d["count_dist"].items())
                        ])
                        st.dataframe(dist_df, hide_index=True,
                                     use_container_width=True, height=200)
                    else:
                        st.info("—")

                st.markdown("---")

                # ── Main Data breakdown ────────────────────────────────
                with st.expander(
                    f"📊 Main Data Breakdown  ·  M:{len(main_df)}"
                    f"  ·  {d.get('main_bd_str','—')}"
                ):
                    md_wc = d.get("main_df_wc")
                    if md_wc is not None and not md_wc.empty:
                        show_filtered_highlighted(
                            md_wc, cvi_set_d, n_cols_d,
                            f"main_{i}")
                        c1e,c2e = st.columns(2)
                        with c1e:
                            st.download_button(
                                "⬇ Main Breakdown CSV",
                                to_csv_bytes(md_wc),
                                f"{ind_prefix}_{w_lbl}_main.csv",
                                "text/csv", key=f"dl_main_{db}_{i}"
                            )
                        with c2e:
                            st.download_button(
                                "⬇ Main Breakdown Excel (highlighted)",
                                to_styled_excel(
                                    md_wc, cvi_set_d, n_cols_d,
                                    f"{w_lbl}_Main"
                                ),
                                f"{ind_prefix}_{w_lbl}_main.xlsx",
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                                key=f"dl_main_xl_{db}_{i}"
                            )
                    else:
                        st.info("Main data breakdown not available "
                                f"(M>{DISPLAY_THRESHOLD:,} or no CVI).")

                st.markdown("---")

                # ── Selected + Unselected side by side ─────────────────
                ts, tu = st.columns(2)

                with ts:
                    d_sel = d.get("sel_df")
                    with st.expander(
                        f"✅ Selected  S:{d['selected_n']}  ·  "
                        f"{d.get('count_dist',{}) and d.get('sel_bd_str', '') or '—'}"
                        if False else
                        f"✅ Selected  S:{d['selected_n']}", expanded=True
                    ):
                        if d_sel is not None and not d_sel.empty:
                            show_filtered_highlighted(
                                d_sel, cvi_set_d, n_cols_d,
                                f"sel_{i}",
                                default_counts=list(d["sc"]))
                            c1s, c2s = st.columns(2)
                            with c1s:
                                st.download_button(
                                    f"⬇ {w_lbl} Selected CSV",
                                    to_csv_bytes(d_sel),
                                    f"{ind_prefix}_{w_lbl}_sel.csv",
                                    "text/csv",
                                    key=f"dl_sel_{db}_{i}"
                                )
                            with c2s:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(
                                        d_sel, cvi_set_d, n_cols_d,
                                        f"{w_lbl}_Sel"
                                    ),
                                    f"{ind_prefix}_{w_lbl}_sel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_sel_xl_{db}_{i}"
                                )
                        elif d["selected_n"] == 0:
                            st.info("No rows selected at this stage.")
                        else:
                            st.info(f"S:{d['selected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

                with tu:
                    d_unsel = d.get("unsel_df")
                    with st.expander(
                        f"🔵 Unselected  U:{d['unselected_n']} → fwd",
                        expanded=True
                    ):
                        if d_unsel is not None and not d_unsel.empty:
                            show_filtered_highlighted(
                                d_unsel, cvi_set_d, n_cols_d,
                                f"unsel_{i}")
                            c1u, c2u = st.columns(2)
                            with c1u:
                                st.download_button(
                                    f"⬇ {w_lbl} Unselected CSV",
                                    to_csv_bytes(d_unsel),
                                    f"{ind_prefix}_{w_lbl}_unsel.csv",
                                    "text/csv",
                                    key=f"dl_unsel_{db}_{i}"
                                )
                            with c2u:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(
                                        d_unsel, cvi_set_d, n_cols_d,
                                        f"{w_lbl}_Unsel"
                                    ),
                                    f"{ind_prefix}_{w_lbl}_unsel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_unsel_xl_{db}_{i}"
                                )
                        elif d["unselected_n"] == 0:
                            st.info("No rows remaining.")
                        else:
                            st.info(f"U:{d['unselected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

        # ── Final stage permanent tabs ─────────────────────────────────
        st.markdown("---")
        st.markdown("#### Final Stage Output")

        # Get final w CVI set
        final_cvi_set = set()
        if dbg:
            last_d = next((d for d in reversed(dbg)
                           if d.get("cvi_set")), None)
            if last_d:
                final_cvi_set = last_d.get("cvi_set", set())

        ft1, ft2, ft3 = st.tabs([
            f"Selected (final)  S:{len(sel)}",
            f"Unselected (final)  U:{len(unsel)}",
            "Breakdown S0/S1/S2… (all stages)",
        ])

        with ft1:
            if sel.empty:
                st.info("No rows selected at final stage.")
            else:
                n_f = [c for c in n_cols_res if c in sel.columns]
                show_filtered_highlighted(
                    sel, final_cvi_set, n_f, "final_sel",
                    default_counts=None)
                fpath = DIRS["Outputs"] / f"{ind_prefix}_selected.csv"
                sel.to_csv(fpath, index=False)
                c1f, c2f = st.columns(2)
                with c1f:
                    st.download_button(
                        "⬇ Selected CSV", to_csv_bytes(sel),
                        f"{ind_prefix}_selected.csv","text/csv",
                        key=f"dl_s_{db}")
                with c2f:
                    st.download_button(
                        "⬇ Selected Excel (highlighted)",
                        to_styled_excel(sel, final_cvi_set, n_f, "Selected"),
                        f"{ind_prefix}_selected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_s_xl_{db}")

        with ft2:
            if unsel.empty:
                st.info("No rows remaining at final stage.")
            else:
                n_u = [c for c in n_cols_res if c in unsel.columns]
                show_filtered_highlighted(
                    unsel, final_cvi_set, n_u, "final_unsel")
                c1u2, c2u2 = st.columns(2)
                with c1u2:
                    st.download_button(
                        "⬇ Unselected CSV", to_csv_bytes(unsel),
                        f"{ind_prefix}_unselected.csv","text/csv",
                        key=f"dl_u_{db}")
                with c2u2:
                    st.download_button(
                        "⬇ Unselected Excel (highlighted)",
                        to_styled_excel(unsel, final_cvi_set, n_u,
                                        "Unselected"),
                        f"{ind_prefix}_unselected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_u_xl_{db}")

        with ft3:
            st.caption("S0=0 matches · S1=1 match … per w-column")
            if bd.empty:
                st.info("No breakdown data.")
            else:
                st.dataframe(bd, use_container_width=True,
                             height=280, hide_index=True)
                st.download_button(
                    "⬇ Breakdown CSV", to_csv_bytes(bd),
                    f"{ind_prefix}_breakdown.csv","text/csv",
                    key=f"dl_bd_{db}")

    # Controls
    st.markdown("---")
    if db not in S["container_status"]:
        S["container_status"][db] = pd.DataFrame({
            "Name":[db],"Status":["Stopped"],"Memory Usage":[""],
            "CPU Usage":[""],"Percent progress %":[0],
            "Time Guestimate Min":[""],"Time Guestimate Max":[""],
        })

    b1,b2,b3,b4,b5,b6,b7 = st.columns(7)
    with b1:
        if st.button("▶ Start", type="primary", use_container_width=True, key=f"s1_{db}"):
            set_container_status(db,"Running"); st.toast("Started!")
    with b2:
        if st.button("■ Stop", use_container_width=True, key=f"s2_{db}"):
            set_container_status(db,"Stopped")
    with b3:
        if st.button("⚡ Kill", use_container_width=True, key=f"s3_{db}"):
            set_container_status(db,"Killed")
    with b4:
        if st.button("↺ Restart", use_container_width=True, key=f"s4_{db}"):
            st.toast("Restarting…")
    with b5:
        if st.button("⏸ Pause", use_container_width=True, key=f"s5_{db}"):
            set_container_status(db,"Paused")
    with b6:
        if st.button("▶ Resume", use_container_width=True, key=f"s6_{db}"):
            set_container_status(db,"Running")
    with b7:
        if st.button("🗑 Remove", use_container_width=True, key=f"s7_{db}"):
            st.toast("Removed.")

    S["container_status"][db] = st.data_editor(
        S["container_status"][db], key=f"stbl_{db}",
        use_container_width=True, num_rows="dynamic",
        column_config={
            "Status": st.column_config.SelectboxColumn(
                "Status", options=["Stopped","Running","Paused","Killed","Error"]),
            "Percent progress %": st.column_config.ProgressColumn(
                "Percent progress %", min_value=0, max_value=100),
        })

    c1,c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {db} Status CSV",
            to_csv_bytes(S["container_status"][db]),
            f"{db.replace(' ','_')}_status.csv","text/csv",key=f"dls_{db}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MASTER OUTPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Master Outputs":
    st.markdown('<span class="sec-hdr hdr-brown">📤 Master Outputs — Aggregated Results</span>',
                unsafe_allow_html=True)
    am_toggle("out")

    # Collect from all dashboards
    all_sel, all_unsel, all_bd = [], [], []
    for db in DASHBOARDS:
        res = S["results"].get(db)
        if res:
            if not res["selected"].empty:
                s = res["selected"].copy(); s.insert(0,"Dashboard",db)
                all_sel.append(s)
            if not res["unselected"].empty:
                u = res["unselected"].copy(); u.insert(0,"Dashboard",db)
                all_unsel.append(u)
            if not res["breakdown"].empty:
                b = res["breakdown"].copy(); b.insert(0,"Dashboard",db)
                all_bd.append(b)

    n_done = len(all_sel)
    if n_done:
        st.markdown(f'<div class="ok">✅ Results from {n_done} dashboard(s).</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="warn">⚠️ No results yet. Run matching in Container Dashboards.</div>',
                    unsafe_allow_html=True)

    def cc(dfs):
        if not dfs: return pd.DataFrame()
        return pd.concat(dfs, ignore_index=True).drop_duplicates()

    def show_output_tab(df: pd.DataFrame, label: str, fname: str):
        if df.empty:
            st.info("No data yet.")
            return
        st.write(f"**{len(df):,} unique rows**")
        st.dataframe(df, use_container_width=True, height=380)
        df.to_csv(DIRS["Outputs"]/fname, index=False)
        st.download_button(f"⬇ Export", to_csv_bytes(df), fname,"text/csv",
                           key=f"dl_out_{fname}")

    t1,t2,t3,t4,t5 = st.tabs([
        "Cluster List Combined Selected",
        "Cluster Combined Selected",
        "Cluster List Combined Unselected",
        "Cluster Combined Unselected",
        "Selected Breakdown S0/S1/S2…",
    ])

    with t1:
        show_output_tab(cc(all_sel), "Cluster List Combined Selected",
                        "cluster_list_combined_selected.csv")
    with t2:
        df_t2 = cc(all_sel)
        nc = [c for c in df_t2.columns if re.match(r'^n\d+$',c)] if not df_t2.empty else []
        show_output_tab(df_t2[nc].drop_duplicates() if nc else pd.DataFrame(),
                        "Cluster Combined Selected", "cluster_combined_selected.csv")
    with t3:
        show_output_tab(cc(all_unsel), "Cluster List Combined Unselected",
                        "cluster_list_combined_unselected.csv")
    with t4:
        df_t4 = cc(all_unsel)
        nc4 = [c for c in df_t4.columns if re.match(r'^n\d+$',c)] if not df_t4.empty else []
        show_output_tab(df_t4[nc4].drop_duplicates() if nc4 else pd.DataFrame(),
                        "Cluster Combined Unselected", "cluster_combined_unselected.csv")
    with t5:
        show_output_tab(cc(all_bd), "Selected Breakdown",
                        "selected_breakdown.csv")

# ════════════════════════════════════════════════════════════════════════════
# PAGE: CLUSTER MANAGER
# ════════════════════════════════════════════════════════════════════════════
elif page == "🗂️ Cluster Manager":
    st.markdown('<span class="sec-hdr hdr-purple">🗂️ Cluster Manager</span>',
                unsafe_allow_html=True)

    st.markdown("""
    <div class="info">
    Each cluster = one complete run of all active containers against one Main Data file.
    <b>1n</b> = current run · <b>1n+1</b> = next run · <b>1n-1</b> = previous run.
    Labels are user-assigned. System auto-assigns numeric ID (001, 002…).
    </div>
    """, unsafe_allow_html=True)

    clusters = load_clusters()

    if not clusters:
        st.info("No clusters run yet. Use the parallel run button on Container Dashboards.")
    else:
        # ── Cluster registry table ─────────────────────────────────────
        st.markdown("### All Clusters")
        reg_df = pd.DataFrame([{
            "ID":         c["id"],
            "Label":      c["label"],
            "Lotto":      c.get("lotto",""),
            "Draw":       c.get("draw_no",""),
            "Date":       c.get("draw_date",""),
            "Main Data":  c.get("main_file",""),
            "Containers": c.get("containers",0),
            "Status":     c.get("status",""),
            "Run at":     c.get("timestamp","")[:16].replace("T"," "),
        } for c in clusters])
        reg_df.index = range(1, len(reg_df)+1)
        st.dataframe(reg_df, use_container_width=True, height=280)

        # ── Select cluster to inspect ──────────────────────────────────
        st.markdown("---")
        st.markdown("### Inspect a Cluster")
        cluster_ids = [f"{c['id']} — {c['label']}" for c in clusters]
        chosen_c = st.selectbox("Select cluster:", cluster_ids,
                                key="cm_chosen")
        chosen_idx = cluster_ids.index(chosen_c)
        c_data = clusters[chosen_idx]

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Cluster ID",   c_data["id"])
        c2.metric("Label",        c_data["label"])
        c3.metric("Lotto",        c_data.get("lotto",""))
        c4.metric("Containers",   c_data.get("containers",0))

        # ── Per-container results for this cluster ─────────────────────
        results = c_data.get("results", [])
        if results:
            st.markdown("**Container results:**")
            res_df = pd.DataFrame([{
                "Formula":    r["formula"],
                "Status":     "✅" if r["status"]=="complete" else "❌",
                "Selected":   r["selected_n"],
                "Unselected": r["unselected_n"],
                "Error":      r.get("error",""),
            } for r in results])
            res_df.index = range(1, len(res_df)+1)
            st.dataframe(res_df, use_container_width=True,
                         hide_index=False, height=320)

        # ── Load and browse cluster output files ───────────────────────
        out_dir = Path(c_data.get("output_dir", ""))
        if out_dir.exists():
            st.markdown("---")
            st.markdown("### Browse Cluster Output Files")
            out_files = sorted(out_dir.glob("*.csv"))
            if out_files:
                chosen_file = st.selectbox(
                    "File:", [f.name for f in out_files],
                    key="cm_file"
                )
                fp = out_dir / chosen_file
                df_view = pd.read_csv(fp)
                st.write(f"**{len(df_view):,} rows · {len(df_view.columns)} cols**")
                st.dataframe(df_view.head(30), use_container_width=True,
                             hide_index=True, height=300)
                st.download_button(f"⬇ {chosen_file}",
                                   df_view.to_csv(index=False).encode(),
                                   chosen_file,"text/csv",
                                   key=f"dl_cm_{chosen_file}")
            else:
                st.info("No output files found in this cluster's output folder.")

        # ── Cross-cluster comparison ───────────────────────────────────
        if len(clusters) >= 2:
            st.markdown("---")
            st.markdown("### Cross-cluster Comparison")
            st.caption("Compare selected/unselected counts across clusters "
                       "for the same formula.")

            all_formulas = sorted(set(
                r["formula"]
                for c in clusters for r in c.get("results",[])
            ))
            if all_formulas:
                comp_formula = st.selectbox("Formula to compare:",
                                            all_formulas, key="cm_comp")
                comp_rows = []
                for c in clusters:
                    res = next((r for r in c.get("results",[])
                                if r["formula"] == comp_formula), None)
                    if res:
                        comp_rows.append({
                            "Cluster ID":  c["id"],
                            "Label":       c["label"],
                            "Lotto":       c.get("lotto",""),
                            "Draw":        c.get("draw_no",""),
                            "Date":        c.get("draw_date",""),
                            "Selected":    res["selected_n"],
                            "Unselected":  res["unselected_n"],
                            "Status":      res["status"],
                        })
                if comp_rows:
                    comp_df = pd.DataFrame(comp_rows)
                    comp_df.index = range(1, len(comp_df)+1)
                    st.dataframe(comp_df, use_container_width=True,
                                 height=min(40*len(comp_df)+50, 300))
                    st.download_button(
                        f"⬇ Cross-cluster comparison CSV",
                        comp_df.to_csv(index=False).encode(),
                        f"cross_cluster_{comp_formula}.csv","text/csv",
                        key="dl_cm_comp"
                    )

        # ── Rename / delete cluster ────────────────────────────────────
        st.markdown("---")
        st.markdown("### Manage This Cluster")
        new_label = st.text_input("Rename cluster label:",
                                  value=c_data["label"],
                                  key="cm_rename")
        rc1,rc2,_ = st.columns([1,1,4])
        with rc1:
            if st.button("💾 Save label", key="cm_save_label"):
                clusters[chosen_idx]["label"] = new_label
                save_clusters(clusters)
                st.success("Label updated.")
                st.rerun()
        with rc2:
            if st.button("🗑 Delete cluster", key="cm_delete"):
                clusters.pop(chosen_idx)
                save_clusters(clusters)
                st.success("Cluster deleted from registry.")
                st.rerun()