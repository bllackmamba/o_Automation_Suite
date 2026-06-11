"""
masterapp.py — Syndicate System, Single File
All logic, scraping, matching, collation, and UI in one place.
Run: streamlit run masterapp.py
Requires: pip install streamlit pandas numpy openpyxl playwright requests beautifulsoup4
          playwright install chromium
"""

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════
import streamlit as st
import pandas as pd
import numpy as np
import re, json, time, random, asyncio, os
from pathlib import Path
from datetime import datetime
from multiprocessing import Pool, cpu_count

# ── 1-based row numbering for ALL tables ─────────────────────────────────────
# Users count rows from 1, not 0.
# Streamlit's st.dataframe does NOT reliably render a custom pandas .index —
# it shows its own internal 1-to-N counter for the visible slice regardless of
# what the pandas index is set to.  The portable fix is to insert an explicit
# "#" column with the correct row numbers and hide the pandas index.
# Fully guarded — never breaks rendering.
_orig_st_dataframe = st.dataframe
def _dataframe_1based(data=None, *args, **kwargs):
    try:
        if isinstance(data, pd.DataFrame) and not kwargs.get("hide_index", False):
            data = data.copy()
            _col = "#"
            if _col in data.columns:
                _col = "_row_"
            data.insert(0, _col, range(1, len(data) + 1))
            kwargs["hide_index"] = True
    except Exception:
        pass
    return _orig_st_dataframe(data, *args, **kwargs)
st.dataframe = _dataframe_1based

# ── Paginated dataframe helper ────────────────────────────────────────────────
# Navigation bar above AND below every table:
#   [⏮ First] [◀ Back]  Rows X–Y of N | Page P/T  [goto#]  [rows/pg▼]  [▶ Next] [⏭ Last]
# key must be unique per call site so each table has independent page state.
_PAGE_SIZE         = 50
_PAGE_SIZE_OPTIONS = [25, 50, 100, 200]

def show_paginated_df(df, key, use_container_width=True, height=None, hide_index=False, **kwargs):
    try:
        if not isinstance(df, pd.DataFrame):
            return _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)

        n_rows = len(df)
        if n_rows == 0:
            return _orig_st_dataframe(df, use_container_width=use_container_width,
                                      hide_index=hide_index, **kwargs)

        sk       = f"_pg_{key}"
        sk_ps    = f"_ps_{key}"
        goto_key = f"_goto_{key}"
        ps_key   = f"_pssel_{key}"

        # ── initialise session state ──────────────────────────────────────
        if sk    not in st.session_state: st.session_state[sk]    = 0
        if sk_ps not in st.session_state: st.session_state[sk_ps] = _PAGE_SIZE

        page_size = int(st.session_state[sk_ps])
        n_pages   = max(1, (n_rows + page_size - 1) // page_size)
        cur       = max(0, min(int(st.session_state[sk]), n_pages - 1))

        start_r = cur * page_size + 1
        end_r   = min((cur + 1) * page_size, n_rows)

        # ── Sync goto widget to current page whenever buttons change the page.
        # This must happen before rendering the number_input so Streamlit
        # uses the updated value (not the stale cached widget value).
        st.session_state[goto_key] = cur + 1

        # on_change callbacks — called by Streamlit before the next rerun
        def _on_goto():
            val = st.session_state.get(goto_key, 1)
            st.session_state[sk] = max(0, min(int(val) - 1, n_pages - 1))

        def _on_ps():
            new_ps = st.session_state.get(ps_key, _PAGE_SIZE)
            st.session_state[sk_ps] = int(new_ps)
            st.session_state[sk]    = 0

        # ── layout: table on the left, vertical nav panel on the right ───
        col_table, col_nav = st.columns([6, 1])

        with col_table:
            page_df = df.iloc[cur * page_size : (cur + 1) * page_size].copy()
            page_df = page_df.reset_index(drop=True)
            if not hide_index:
                # Insert explicit global row-number column so the correct page-aware
                # numbers are always shown.  Custom pandas .index is not reliably
                # rendered by st.dataframe across Streamlit versions — an explicit
                # column is the only portable solution.
                _row_col = "#"
                if _row_col in page_df.columns:
                    _row_col = "_row_"
                page_df.insert(0, _row_col,
                               range(cur * page_size + 1,
                                     cur * page_size + len(page_df) + 1))
            disp_kwargs = dict(use_container_width=True,
                               hide_index=True, **kwargs)
            if height is not None:
                disp_kwargs["height"] = height
            _orig_st_dataframe(page_df, **disp_kwargs)

        with col_nav:
            # ── page info ─────────────────────────────────────────────
            st.markdown(
                f"<div style='font-size:.75rem;color:#aaa;text-align:center;"
                f"padding:4px 0 6px 0;line-height:1.5'>"
                f"<b>{start_r:,}</b>–<b>{end_r:,}</b><br>"
                f"of <b>{n_rows:,}</b><br>"
                f"pg <b>{cur+1}</b>/<b>{n_pages}</b>"
                f"</div>",
                unsafe_allow_html=True)

            # ── First ─────────────────────────────────────────────────
            # Use text labels instead of double-chevron emojis (⏫/⏬) because
            # those glyphs do not render on all OS / Streamlit versions and
            # the button shows blank or the help-text string instead.
            if st.button("« First", key=f"{sk}_first",
                         use_container_width=True, disabled=(cur == 0),
                         help="Jump to first page"):
                st.session_state[sk] = 0
                st.rerun()

            # ── Back ──────────────────────────────────────────────────
            if st.button("▲ Prev", key=f"{sk}_back",
                         use_container_width=True, disabled=(cur == 0),
                         help="Previous page"):
                st.session_state[sk] = cur - 1
                st.rerun()

            # ── Go to page (on_change handles navigation; no inline check) ──
            st.number_input(
                "pg", min_value=1, max_value=n_pages, step=1,
                key=goto_key,
                on_change=_on_goto,
                label_visibility="collapsed",
                help=f"Go to page (1–{n_pages})")

            # ── Next ──────────────────────────────────────────────────
            if st.button("Next ▼", key=f"{sk}_next",
                         use_container_width=True, disabled=(cur >= n_pages - 1),
                         help="Next page"):
                st.session_state[sk] = cur + 1
                st.rerun()

            # ── Last ──────────────────────────────────────────────────
            if st.button("Last »", key=f"{sk}_last",
                         use_container_width=True, disabled=(cur >= n_pages - 1),
                         help="Jump to last page"):
                st.session_state[sk] = n_pages - 1
                st.rerun()

            # ── Rows per page ─────────────────────────────────────────
            st.markdown(
                "<div style='font-size:.7rem;color:#888;text-align:center;"
                "padding:6px 0 2px 0'>rows/pg</div>",
                unsafe_allow_html=True)
            ps_idx = _PAGE_SIZE_OPTIONS.index(page_size) \
                     if page_size in _PAGE_SIZE_OPTIONS else 1
            st.selectbox(
                "rows", _PAGE_SIZE_OPTIONS,
                index=ps_idx,
                key=ps_key,
                on_change=_on_ps,
                label_visibility="collapsed",
                help="Rows per page")

    except Exception:
        try:
            _fb = df.copy() if isinstance(df, pd.DataFrame) else df
            if isinstance(_fb, pd.DataFrame) and not kwargs.get("hide_index", False):
                _fc = "#"
                if _fc in _fb.columns:
                    _fc = "_row_"
                _fb.insert(0, _fc, range(1, len(_fb) + 1))
                _orig_st_dataframe(_fb, use_container_width=use_container_width,
                                   hide_index=True, **kwargs)
            else:
                _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)
        except Exception:
            _orig_st_dataframe(df, use_container_width=use_container_width, **kwargs)

try:
    from playwright.async_api import async_playwright
    PW_OK = True
except ImportError:
    PW_OK = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQ_OK = True
except ImportError:
    REQ_OK = False

# ═══════════════════════════════════════════════════════════════════════════════
# 1. PATH AUTO-DETECTION
# ═══════════════════════════════════════════════════════════════════════════════
def find_root() -> Path:
    for base in [Path.home()/"Desktop", Path.home()/"Documents",
                 Path.home(), Path.cwd()]:
        p = base / "o_Automation_Suite"
        if p.is_dir():
            return p
        if base.is_dir():
            for sub in base.iterdir():
                try:
                    q = sub / "o_Automation_Suite"
                    if q.is_dir():
                        return q
                except Exception:
                    pass
    fb = Path(__file__).parent / "o_Automation_Suite"
    fb.mkdir(parents=True, exist_ok=True)
    return fb

ROOT = find_root()

# ── Auto-create .streamlit/config.toml to raise upload limit ──────────────
_streamlit_dir = Path(__file__).parent / ".streamlit"
_streamlit_dir.mkdir(exist_ok=True)
_config_path = _streamlit_dir / "config.toml"
if not _config_path.exists():
    _config_path.write_text(
        "[server]\n"
        "maxUploadSize = 10000\n"
        "maxMessageSize = 10000\n\n"
        "[browser]\n"
        "gatherUsageStats = false\n\n"
        "[runner]\n"
        "fastReruns = true\n"
    )

# ── Global-only DIRS (shared across all games) ───────────────────────────────
# Game-specific paths (Main_Data, Formulas, Containers, Outputs, Variables, CVI,
# Selected_Counts, Base, Splits, etc.) are accessed via game_dirs() / active_game_dirs()
# so they are always scoped to the correct game and suffixed with _{game}.
# Only truly shared paths live here.
DIRS = {
    "Global_Scraper": ROOT / "Global_Scraper",  # raw scrapes: D_<STATE>.csv (all games)
    "debug_html":     ROOT / "debug_html",       # debug output
}
for d in DIRS.values():
    d.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STARTUP MIGRATION — runs once per session to enforce the folder convention.
#
# Problem: old code created root-level folders (Containers, Formulas, Main_Data,
# Variables, Outputs) and game subfolders without the _{game} suffix. Every time
# the app started those folders were re-created by the old DIRS mkdir loop.
#
# Fix embedded here so no separate script is needed:
#   1. Remove stale root-level duplicate folders (only if empty — no data lost).
#   2. Rename game subfolders to the _{game} suffix convention.
#   3. Flatten old Variables/ tree → variable_inputs_{game}/ and
#      container_variable_inputs_{game}/.
# ═══════════════════════════════════════════════════════════════════════════════
import shutil as _shutil

_ROOT_LEVEL_STALE = [
    "Containers", "Formulas", "Main_Data", "Variables", "Outputs",
]

_SUBFOLDER_CANON = {
    "containers":                "containers",
    "container":                 "containers",
    "formulas":                  "formulas",
    "formula":                   "formulas",
    "main_data":                 "main_data",
    "maindata":                  "main_data",
    "outputs":                   "outputs",
    "output":                    "outputs",
    "sincelast":                 "sincelast",
    "since_last":                "sincelast",
    "games_breakdown":           "games_breakdown",
    "gamesbreakdown":            "games_breakdown",
    "cvi_matrix":                "container_variable_inputs",
    "cvimatrix":                 "container_variable_inputs",
    "container_variable_inputs": "container_variable_inputs",
    "containervariableinputs":   "container_variable_inputs",
    "variable_inputs":           "variable_inputs",
    "variableinputs":            "variable_inputs",
}


def _norm(name: str) -> str:
    return name.lower().replace(" ", "_").replace("-", "_")


def _startup_migrate(root: Path):
    """Idempotent startup migration — safe to run every launch."""

    # ── Step 1: remove stale empty root-level folders ──────────────────────
    for name in _ROOT_LEVEL_STALE:
        p = root / name
        if p.is_dir():
            contents = [x for x in p.iterdir() if not x.name.startswith(".")]
            if not contents:
                try:
                    _shutil.rmtree(str(p))
                except Exception:
                    pass
            # Non-empty root-level folders are handled in steps 4 & 5 below

    games_dir = root / "Games"
    if not games_dir.is_dir():
        return

    for game in ["SAT", "OZ", "PB", "MWF", "SFL"]:
        gp = games_dir / game
        if not gp.is_dir():
            continue
        gk = game.lower()
        suffix = f"_{gk}"

        # ── Step 2: rename top-level game subfolders ────────────────────────
        for entry in sorted(gp.iterdir()):
            if not entry.is_dir():
                continue
            key = _norm(entry.name)
            if key == "variables":
                continue  # handled in step 3
            # strip existing suffix before lookup
            if key.endswith(suffix):
                key = key[: -len(suffix)]
            canonical = _SUBFOLDER_CANON.get(key)
            if canonical is None:
                continue
            desired = gp / f"{canonical}{suffix}"
            if entry != desired and not desired.exists():
                try:
                    entry.rename(desired)
                except Exception:
                    pass

        # ── Step 3: flatten old Variables/ tree ────────────────────────────
        old_vars = gp / "Variables"
        if not old_vars.is_dir():
            continue

        var_inputs_dst = gp / f"variable_inputs_{gk}"
        cvi_dst        = gp / f"container_variable_inputs_{gk}"

        # Variables/Scraper → variable_inputs_{gk}/Scraper
        src = old_vars / "Scraper"
        dst = var_inputs_dst / "Scraper"
        if src.is_dir() and not dst.exists():
            var_inputs_dst.mkdir(parents=True, exist_ok=True)
            try:
                src.rename(dst)
            except Exception:
                pass

        # Variables/Variable_Elements/* → variable_inputs_{gk}/ (flatten)
        var_elem = old_vars / "Variable_Elements"
        if var_elem.is_dir():
            var_inputs_dst.mkdir(parents=True, exist_ok=True)
            for item in sorted(var_elem.iterdir()):
                dst_item = var_inputs_dst / item.name
                if not dst_item.exists():
                    try:
                        item.rename(dst_item)
                    except Exception:
                        pass
            # Remove shell if now empty
            remaining = [x for x in var_elem.iterdir()
                         if not x.name.startswith(".")]
            if not remaining:
                try:
                    var_elem.rmdir()
                except Exception:
                    pass

        # Variables/Container_Variable_Inputs → container_variable_inputs_{gk}
        src_cvi = old_vars / "Container_Variable_Inputs"
        if src_cvi.is_dir() and not cvi_dst.exists():
            try:
                src_cvi.rename(cvi_dst)
            except Exception:
                pass

        # Remove empty Variables/ shell
        remaining = [x for x in old_vars.iterdir()
                     if not x.name.startswith(".")]
        if not remaining:
            try:
                old_vars.rmdir()
            except Exception:
                pass

    # ── Step 4: flatten root-level Variables/ → Games/SAT/variable_inputs_sat/
    #    (old global Variables/ before game-specific structure was introduced)
    root_vars = root / "Variables"
    if root_vars.is_dir():
        sat_path = root / "Games" / "SAT"
        if sat_path.is_dir():
            vi_dst  = sat_path / "variable_inputs_sat"
            cvi_dst2 = sat_path / "container_variable_inputs_sat"

            src_sc = root_vars / "Scraper"
            dst_sc = vi_dst / "Scraper"
            if src_sc.is_dir() and not dst_sc.exists():
                vi_dst.mkdir(parents=True, exist_ok=True)
                try:
                    src_sc.rename(dst_sc)
                except Exception:
                    pass

            ve_dir = root_vars / "Variable_Elements"
            if ve_dir.is_dir():
                vi_dst.mkdir(parents=True, exist_ok=True)
                for item in sorted(ve_dir.iterdir()):
                    if item.name.startswith("."):
                        continue
                    dst_item = vi_dst / item.name
                    if not dst_item.exists():
                        try:
                            item.rename(dst_item)
                        except Exception:
                            pass
                rem = [x for x in ve_dir.iterdir() if not x.name.startswith(".")]
                if not rem:
                    try:
                        ve_dir.rmdir()
                    except Exception:
                        pass

            src_cvi2 = root_vars / "Container_Variable_Inputs"
            if src_cvi2.is_dir() and not cvi_dst2.exists():
                try:
                    src_cvi2.rename(cvi_dst2)
                except Exception:
                    pass

            rem_rv = [x for x in root_vars.iterdir() if not x.name.startswith(".")]
            if not rem_rv:
                try:
                    root_vars.rmdir()
                except Exception:
                    pass

    # ── Step 5: move root-level Formulas/ → Games/SAT/formulas_sat/
    #    (old global Formulas/ before game-specific structure)
    root_formulas = root / "Formulas"
    if root_formulas.is_dir():
        sat_path = root / "Games" / "SAT"
        if sat_path.is_dir():
            sat_formulas_dst = sat_path / "formulas_sat"
            sat_formulas_dst.mkdir(parents=True, exist_ok=True)
            for item in sorted(root_formulas.iterdir()):
                if item.name.startswith("."):
                    continue
                dst_item = sat_formulas_dst / item.name
                if not dst_item.exists():
                    try:
                        item.rename(dst_item)
                    except Exception:
                        pass
            rem_rf = [x for x in root_formulas.iterdir() if not x.name.startswith(".")]
            if not rem_rf:
                try:
                    root_formulas.rmdir()
                except Exception:
                    pass


_startup_migrate(ROOT)

# ═══════════════════════════════════════════════════════════════════════════════
# 2. CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
TARGET_STATES = ["NSW","VIC","QLD","WA","SA","TAS","ACT","NT"]

STATE_POSTCODES = {
    "NSW": list(range(2000,3000)),
    "VIC": list(range(3000,4000)),
    "QLD": list(range(4000,5000)),
    "WA":  list(range(6000,7000)),
    "SA":  list(range(5000,5600)),
    "TAS": list(range(7000,7800)),
    "ACT": list(range(2600,2619)),
    "NT":  list(range(800, 1000)),
}

SCRAPE_URL = "https://www.thelott.com/syndicates?postcode={pc}"
SCRAPE_URL_WA = "https://www.lotterywest.wa.gov.au/play-online/syndicate-games?postcode={pc}"

SCRAPE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-AU,en;q=0.9",
    "Referer": "https://www.thelott.com/syndicates",
}

CF_ROWS = [
    (1,  "BRD",       "Formed",    ["B","R","D"]),
    (2,  "BSD",       "Formed",    ["B","Sp","D"]),
    (3,  "BSoD",      "Formed",    ["B","So","D"]),
    (4,  "SD",        "Formed",    ["Sp","D"]),
    (5,  "SoD",       "Formed",    ["So","D"]),
    (6,  "BD",        "Formed",    ["B","D"]),
    (7,  "BSSoD",     "Formed",    ["B","Sp","So","D"]),
    (8,  "BRDSSo",    "Formed",    ["B","R","D","Sp","So"]),
    (9,  "B1B2B3",    "Formed",    ["B1","B2","B3"]),
    (10, "R1R2R3",    "Formed",    ["R1","R2","R3"]),
    (11, "D1D2D3",    "Formed",    ["D1","D2","D3"]),
    (12, "S1S2S3",    "Formed",    ["Sp1","Sp2","Sp3"]),
    (13, "So1So2So3", "Formed",    ["So1","So2","So3"]),
    (14, "Xn",        "Formed",    ["Xn"]),
    (15, "RVI1",      "Ready-made",["RVI1"]),
    (16, "RVI2",      "Ready-made",["RVI2"]),
    (17, "Xnn",       "Ready-made",["Xnn"]),
]
DASHBOARDS = [f"1n & {r[1]}" for r in CF_ROWS]
COMP_MAP   = {r[1]: r[3] for r in CF_ROWS}

CHUNK_SIZE = 500_000  # rows per processing chunk

# ── Lotto type codes (worldwide — extendable) ─────────────────────────────
LOTTO_TYPES = {
    "oz":  "Oz Lotto (AUS 7/47)",
    "pb":  "Powerball (AUS 7+1/35+20)",
    "sfl": "Set for Life (AUS)",
    "Mon": "Monday Lotto (AUS 6/45)",
    "Wed": "Wednesday Lotto (AUS 6/45)",
    "Fri": "Friday Lotto (AUS 6/45)",
    "Sat": "Saturday Lotto (AUS 6/45)",
    # Add more as needed — no cap
}

# ═══════════════════════════════════════════════════════════════════════════════
# GAME CONFIGURATION — one entry per game
# ═══════════════════════════════════════════════════════════════════════════════
GAMES_CFG = {
    "pb": {
        "label": "Powerball", "emoji": "🔵", "pool": 35, "pick": 7,
        "draw_day": "Thursday",
        "lottolyzer": "https://en.lottolyzer.com/number-frequencies/australia/powerball",
        "b_file": "Base_pb.xlsx", "b_sheet": "B_pb",
        "b_sheet_legacy": "w values Pb A (2)", "thelott_key": "pb",
    },
    "oz": {
        "label": "Oz Lotto", "emoji": "🟠", "pool": 47, "pick": 7,
        "draw_day": "Tuesday",
        "lottolyzer": "https://en.lottolyzer.com/number-frequencies/australia/oz-lotto",
        "b_file": "Base_oz.xlsx", "b_sheet": "B_oz",
        "b_sheet_legacy": "oz (2)", "thelott_key": "oz",
    },
    "sat": {
        "label": "Saturday Lotto", "emoji": "🟡", "pool": 45, "pick": 6,
        "draw_day": "Saturday",
        # NOTE: If this URL opens Set for Life instead of Saturday Lotto, override
        # it in the Since Last tab (Variable Inputs → Since Last → URL field).
        # Known alternative: https://en.lottolyzer.com/number-frequencies/australia/tatts-lotto
        "lottolyzer": "https://en.lottolyzer.com/number-frequencies/australia/saturday-lotto",
        "b_file": "Base_sat.xlsx", "b_sheet": "B_sat",
        "b_sheet_legacy": "Ta (2)", "thelott_key": "sat",
    },
    "sfl": {
        "label": "Set for Life", "emoji": "🟢", "pool": 44, "pick": 7,
        "draw_day": "Daily",
        "lottolyzer": "https://en.lottolyzer.com/number-frequencies/australia/set-for-life",
        "b_file": "Base_sfl.xlsx", "b_sheet": "B_sfl",
        "b_sheet_legacy": "sfl", "thelott_key": "sfl",
    },
    "mwf": {
        "label": "Mon/Wed/Fri", "emoji": "🟣", "pool": 45, "pick": 6,
        "draw_day": "Mon, Wed, Fri",
        "lottolyzer": "https://en.lottolyzer.com/number-frequencies/australia/monday-lotto",
        "b_file": "Base_mwf.xlsx", "b_sheet": "B_mwf",      # own file+sheet; was colliding via "Ta (2)"
        "b_sheet_legacy": "Ta (2)", "thelott_key": "mwf",
    },
}

GAME_KEYS   = list(GAMES_CFG.keys())
GAME_LABELS = {k: f"{v['emoji']} {v['label']}" for k, v in GAMES_CFG.items()}

# ── Game name → folder key mapping (handles brand name variations per state) ──
GAME_NAME_MAP = {
    # Saturday Lotto — multiple brand names across states
    "TattsLotto":               "sat",   # NSW, VIC, QLD brand
    "Saturday Lotto":           "sat",
    "Gold Lotto":               "sat",   # QLD brand
    "X Lotto":                  "sat",   # SA brand
    "Lotto":                    "sat",   # generic fallback

    # Powerball
    "Powerball":                "pb",

    # Oz Lotto
    "Oz Lotto":                 "oz",

    # Mon/Wed/Fri — THE KEY FIX: actual name in data is "Monday & Wednesday Lotto"
    "Monday & Wednesday Lotto": "mwf",   # ← actual name found in scraped data
    "Monday Lotto":             "mwf",   # alternate
    "Wednesday Lotto":          "mwf",   # alternate
    "Friday Lotto":             "mwf",   # alternate
    "Mon & Wed Lotto":          "mwf",   # safety variant

    # Set for Life — likely no syndicates sold, but map in case
    "Set for Life":             "sfl",

    # Skip these — supplementary games, not main pipelines
    "Super 66":                 None,
    "Lucky Lotteries":          None,
    "Lucky Lotteries Mega Jackpot": None,
    "Lucky Lotteries Super Jackpot": None,
}


def split_d_by_game(src_csv: Path, root: Path) -> dict:
    """
    Split a raw D_*.csv file by the Games column into per-game Direct/ folders.

    Handles:
    - Single game rows       → copied to that game's Direct/ folder
    - Pipe-separated rows    → a copy goes to EACH matching game folder
    - None-mapped games      → skipped intentionally (Super 66, Lucky Lotteries)
    - Unknown game names     → logged in _unknown_games for inspection

    Returns dict: {game_key: row_count_written, '_unknown_games': [...]}
    """
    try:
        df = pd.read_csv(src_csv)
    except Exception as ex:
        return {"error": str(ex)}

    if "Games" not in df.columns:
        return {"error": "No 'Games' column found in file"}

    # Accumulate rows per game key
    game_rows: dict[str, list] = {k: [] for k in GAME_KEYS}
    unknown_games: set = set()
    skipped_games: set = set()

    for _, row in df.iterrows():
        raw_games = str(row.get("Games", "")).strip()
        if not raw_games or raw_games == "nan":
            continue

        # Split pipe-separated multi-game entries
        parts = [g.strip() for g in raw_games.split("|")]

        matched_keys = set()
        for part in parts:
            if part not in GAME_NAME_MAP:
                unknown_games.add(part)
            else:
                gkey = GAME_NAME_MAP[part]
                if gkey is None:
                    skipped_games.add(part)  # intentionally skipped
                else:
                    matched_keys.add(gkey)

        for gkey in matched_keys:
            game_rows[gkey].append(row)

    # Save each game's rows to its Direct/ folder
    state_tag = src_csv.stem   # e.g. "D_NSW_NSW"
    results = {}

    # Columns the downstream pipeline keeps. The row-based collation reads ONLY
    # w-columns for D (via _to_w_rows), so identity/label columns can no longer leak
    # into numeric output — which means we can safely retain Draw_Number (+ Draw_Date)
    # here. The draw number is essential for per-draw coverage analysis.
    def _clean_for_pipeline(gdf: pd.DataFrame) -> pd.DataFrame:
        w_cols = sorted([c for c in gdf.columns if re.match(r'^w\d+$', str(c), re.I)],
                        key=lambda x: int(str(x)[1:]))
        keep = [c for c in ("Syndicate_ID", "Syndicate_Name", "Game", "Games",
                            "Draw_Number", "Draw_Date")
                if c in gdf.columns]
        if "PB" in gdf.columns:
            keep.append("PB")
        keep += w_cols
        return gdf[keep]

    for gkey, rows in game_rows.items():
        if not rows:
            results[gkey] = 0
            continue
        gdf = pd.DataFrame(rows).reset_index(drop=True)
        gdf = _clean_for_pipeline(gdf)
        dest_dir = game_dirs(gkey)["Direct"]
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_file = dest_dir / f"{state_tag}_{gkey}.csv"
        gdf.to_csv(dest_file, index=False)
        results[gkey] = len(gdf)

    if unknown_games:
        results["_unknown_games"] = sorted(unknown_games)
    if skipped_games:
        results["_skipped_games"] = sorted(skipped_games)

    return results


def combine_states_for_game(game_key: str) -> dict:
    """Merge every per-state split file for a game into ONE national file.

    Reads all D_<STATE>_<game>.csv in the game's Games_Breakdown folder (e.g.
    D_NSW_pb.csv, D_VIC_pb.csv, …), concatenates them — each state's syndicates
    are distinct, so we keep them all (a "national view") — drops only exact
    duplicate rows (guards against an accidental re-run), and writes
    D_ALL_<game>.csv. That combined file becomes the default CVI source.
    Returns {"states": [...], "files": n, "rows": total}.
    """
    gb = game_dirs(game_key)["Games_Breakdown"]
    parts, states = [], []
    for fp in sorted(gb.glob(f"D_*_{game_key}.csv")):
        if fp.name.startswith("D_ALL_"):
            continue
        try:
            df = pd.read_csv(fp)
        except Exception:
            continue
        if df.empty:
            continue
        # state tag sits between "D_" and f"_{game_key}.csv"
        tag = fp.stem[2:]
        if tag.endswith(f"_{game_key}"):
            tag = tag[: -(len(game_key) + 1)]
        states.append(tag)
        parts.append(df)
    if not parts:
        return {"states": [], "files": 0, "rows": 0}
    combined = pd.concat(parts, ignore_index=True)
    before = len(combined)
    combined = combined.drop_duplicates().reset_index(drop=True)
    out = gb / f"D_ALL_{game_key}.csv"
    combined.to_csv(out, index=False)
    return {"states": states, "files": len(parts),
            "rows": len(combined), "dropped_dups": before - len(combined),
            "path": str(out)}


def game_dirs(game_key: str) -> dict:
    """Return DIRS-like dict scoped to a specific game folder.

    All top-level subfolders are suffixed with _{game_key} so each game's
    folders are visually distinct and never confused with another game's data.
    e.g. Games/SAT/main_data_sat/, Games/OZ/formulas_oz/, etc.
    """
    g  = ROOT / "Games" / game_key.upper()
    gk = game_key.lower()

    # Named intermediate paths used by multiple sub-entries
    var_inputs = g / f"variable_inputs_{gk}"   # was Variables/Variable_Elements
    gb         = g / f"games_breakdown_{gk}"   # per-game split D lives here
    formulas   = g / f"formulas_{gk}"

    d = {
        "Game":            g,
        "Main_Data":       g / f"main_data_{gk}",
        "Outputs":         g / f"outputs_{gk}",
        "Formulas":        formulas,
        "Containers":      g / f"containers_{gk}",
        "Scraper":         var_inputs / "Scraper",
        "Games_Breakdown": gb,
        "Direct":          gb,      # alias (legacy key) → same Games_Breakdown folder
        "Base":            var_inputs / "Base",
        "Splits":          var_inputs / "Splits",
        "Splits_Combi":    var_inputs / "Splits_Combi",
        "Rainbow":         var_inputs / "Rainbow",
        "ExcelPro":        var_inputs / "ExcelPro",
        "CVI":             g / f"container_variable_inputs_{gk}",
        "Selected_Counts": formulas / "Selected_Counts",
        "SinceLast":       g / f"sincelast_{gk}",
    }
    for p in d.values():
        p.mkdir(parents=True, exist_ok=True)
    return d


def active_game() -> str:
    return st.session_state.get("active_game", "sat")


def active_game_dirs() -> dict:
    return game_dirs(active_game())


def active_game_cfg() -> dict:
    return GAMES_CFG[active_game()]

# ── Naming convention helpers ──────────────────────────────────────────────
def parse_main_filename(fname: str) -> dict:
    """
    Parse `{cluster_label}_{lotto_type}_D{draw_no}.csv`
    e.g. `1n_oz_D1567.csv` → {cluster:'1n', lotto:'oz', draw:'D1567'}
    """
    stem = Path(fname).stem           # strip .csv
    parts = stem.split("_")
    result = {"cluster": "", "lotto": "", "draw": "", "raw": fname}
    if len(parts) >= 3:
        result["cluster"] = parts[0]
        result["lotto"]   = parts[1]
        result["draw"]    = "_".join(parts[2:])
    elif len(parts) == 2:
        result["lotto"]  = parts[0]
        result["draw"]   = parts[1]
    return result


def parse_cvi_filename(fname: str) -> dict:
    """
    Parse `CVI_{lotto_type}_{formula}_{date}.csv`
    e.g. `CVI_oz_BRD_2026_05_28.csv`
    """
    stem  = Path(fname).stem
    parts = stem.split("_")
    result = {"lotto": "", "formula": "", "date": "", "raw": fname}
    if len(parts) >= 3 and parts[0] == "CVI":
        result["lotto"]   = parts[1]
        result["formula"] = parts[2]
        result["date"]    = "_".join(parts[3:])
    return result


def scan_main_data_files() -> list[dict]:
    """
    Scan the active game's main_data_{game}/ folder and return only files
    matching the naming convention: {cluster}_{lotto_type}_D{draw_no}.csv
    e.g. 1n_oz_D1567.csv
    Rejects CVI files, SC files, or any file without a draw number (_D prefix).
    """
    lotto_codes = set(LOTTO_TYPES.keys())
    files = []
    for fp in sorted(active_game_dirs()["Main_Data"].glob("*.csv")):
        name = fp.stem   # without .csv
        parts = name.split("_")

        # Must have at least 3 parts: cluster, lotto, draw
        if len(parts) < 3:
            continue

        # Reject files that start with known non-main-data prefixes
        if parts[0] in ("CVI", "SC", "CF", "Container", "D", "MAIN"):
            continue

        # Must contain a lotto type code
        lotto_found = next((p for p in parts if p in lotto_codes), None)
        if not lotto_found:
            continue

        # Must contain a draw part starting with D followed by digits
        draw_found = next((p for p in parts
                          if p.startswith("D") and p[1:].isdigit()), None)
        if not draw_found:
            continue

        # Quick column check: reject if columns look like w1,w2,w3 (CVI file)
        try:
            sample = pd.read_csv(fp, nrows=2)
            w_cols = [c for c in sample.columns if re.match(r'^w\d+$', c)]
            n_cols_found = [c for c in sample.columns if re.match(r'^n\d+$', c, re.I)]
            if w_cols and not n_cols_found:
                continue   # looks like a CVI file, not main data
        except Exception:
            continue

        info = parse_main_filename(fp.name)
        info["path"] = str(fp)
        info["rows"] = 0
        try:
            info["rows"] = sum(1 for _ in open(fp)) - 1
        except Exception:
            pass
        files.append(info)
    return files


def scan_cvi_files(lotto_type: str = "") -> list[dict]:
    """Scan the active game's container_variable_inputs_{game}/ for CVI files."""
    files = []
    for fp in sorted(active_game_dirs()["CVI"].glob("CVI_*.csv")):
        info = parse_cvi_filename(fp.name)
        if not lotto_type or info["lotto"] == lotto_type:
            info["path"] = str(fp)
            files.append(info)
    return files


# ── Cluster registry ──────────────────────────────────────────────────────
CLUSTER_REGISTRY = ROOT / "clusters.json"

def load_clusters() -> list[dict]:
    if CLUSTER_REGISTRY.exists():
        try:
            return json.loads(CLUSTER_REGISTRY.read_text())
        except Exception:
            pass
    return []

def save_clusters(clusters: list[dict]):
    CLUSTER_REGISTRY.write_text(json.dumps(clusters, indent=2))

def next_cluster_id(clusters: list[dict]) -> str:
    if not clusters:
        return "001"
    last = max(int(c["id"]) for c in clusters if str(c.get("id","0")).isdigit())
    return f"{last+1:03d}"


# ── Parallel matching worker (must be top-level for multiprocessing) ───────
def _parallel_worker(args: tuple) -> dict:
    """
    Standalone worker — no Streamlit state.
    Loads CVI, SC, CF from disk; runs matching; saves results.
    Returns summary dict.
    """
    import pandas as pd, numpy as np, re, json
    from pathlib import Path

    (formula_name, cvi_path, main_data_path,
     sc_path, cluster_id, cluster_label,
     lotto_type, draw_no, draw_date,
     output_dir) = args

    result = {
        "formula":    formula_name,
        "status":     "error",
        "selected_n": 0,
        "unselected_n": 0,
        "error":      "",
    }

    try:
        cvi_df  = pd.read_csv(cvi_path)
        main_df = pd.read_csv(main_data_path)

        # Coerce numeric
        for col in main_df.columns:
            main_df[col] = pd.to_numeric(main_df[col], errors="coerce")

        # Load SC
        sc_dict = {}
        if sc_path and Path(sc_path).exists():
            sc_df = pd.read_csv(sc_path)
            if "w" in sc_df.columns and "Selected Count" in sc_df.columns:
                for _, row in sc_df.iterrows():
                    k = str(row["w"]).strip()
                    key = k if k.startswith("w") else f"w{k}"
                    sc_dict[key] = [int(x.strip()) for x in
                                    str(row["Selected Count"]).split(",")
                                    if x.strip().isdigit()]

        # CF: always fresh all-U for parallel run
        w_cols = sorted([c for c in cvi_df.columns
                         if re.match(r'^w\d+$', c)], key=lambda x: int(x[1:]))
        carry_fwd = {w: "U" for w in w_cols}

        # Re-import run_matching locally (worker process)
        import sys
        sys.path.insert(0, str(Path(__file__).parent))
        # Run matching inline to avoid circular imports
        # (simplified version for worker — same logic as main engine)

        # Detect n_cols — prefer explicit n1…nN; only use the ≥90%-numeric
        # heuristic when there are none (mirrors run_matching, so the parallel
        # path and the sequential engine agree and neither counts metadata
        # columns like Postcode/Length/Draw Number as lottery numbers).
        _explicit_n = [c for c in main_df.columns
                       if re.match(r'^n\d+$', c, re.I)]
        if _explicit_n:
            n_cols = sorted(
                _explicit_n,
                key=lambda x: int(re.sub(r'\D', '', x) or 0)
            )[:20]
        else:
            n_cols = sorted(
                [c for c in main_df.columns
                 if pd.to_numeric(main_df[c], errors="coerce").notna().mean() > 0.9],
                key=lambda x: int(re.sub(r'\D', '', x) or 0)
            )[:20]

        M = len(main_df)
        prev_sel = pd.DataFrame()
        prev_unsel = main_df.copy().reset_index(drop=True)
        final_sel = pd.DataFrame()
        final_unsel = prev_unsel.copy()

        CSIZE = 500_000

        def parse_col(series):
            r = []
            for v in series.dropna():
                try:
                    n = float(str(v).strip())
                    if not pd.isna(n) and n >= 1:
                        r.append(int(round(n)))
                except Exception: pass
            return r

        def match_chunk(arr, cvi_arr):
            counts = np.zeros(arr.shape[0], dtype=np.int32)
            for j in range(arr.shape[1]):
                counts += np.isin(arr[:,j], cvi_arr).astype(np.int32)
            return counts

        for idx, w in enumerate(w_cols):
            raw_cvi = parse_col(cvi_df[w])
            if not raw_cvi:
                continue
            cvi_arr = np.array(raw_cvi, dtype=np.float32)
            sc_this = sc_dict.get(w, [])
            sc_set  = set(sc_this)

            direction = carry_fwd.get(w, "U").upper()
            if idx == 0:
                present = main_df.copy().reset_index(drop=True)
            elif direction == "S":
                present = prev_sel.reset_index(drop=True)
            else:
                present = prev_unsel.reset_index(drop=True)

            if present.empty:
                break

            all_counts = np.empty(0, dtype=np.int32)
            for start in range(0, len(present), CSIZE):
                chunk = present.iloc[start:start+CSIZE]
                arr   = chunk[n_cols].to_numpy(dtype=np.float32)
                arr   = np.nan_to_num(arr, nan=-1.0)
                all_counts = np.concatenate(
                    [all_counts, match_chunk(arr, cvi_arr)])

            sel_mask = np.isin(all_counts, list(sc_set))
            sel_df   = present[sel_mask].reset_index(drop=True)
            unsel_df = present[~sel_mask].reset_index(drop=True)

            prev_sel   = sel_df
            prev_unsel = unsel_df
            final_sel   = sel_df
            final_unsel = unsel_df

        # Save results — naming: {cluster_label}_{lotto}_{draw}_{date}_{formula}
        out    = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        prefix = f"{cluster_label}_{lotto_type}_{draw_no}_{draw_date}"

        sel_path   = out / f"{prefix}_{formula_name}_selected.csv"
        unsel_path = out / f"{prefix}_{formula_name}_unselected.csv"
        final_sel.to_csv(sel_path,   index=False)
        final_unsel.to_csv(unsel_path, index=False)

        result["status"]      = "complete"
        result["selected_n"]  = len(final_sel)
        result["unselected_n"]= len(final_unsel)
        result["sel_path"]    = str(sel_path)
        result["unsel_path"]  = str(unsel_path)

    except Exception as ex:
        result["error"] = str(ex)

    return result

# ═══════════════════════════════════════════════════════════════════════════════
# 3. SESSION STATE — single shared dict S
# ═══════════════════════════════════════════════════════════════════════════════
def _auto_load_b(game_key: str = "sat") -> pd.DataFrame:
    """Try to auto-load the Base file for game_key from the project folder tree.

    Search order:
      1. Game-specific file under ROOT/Games/<GAME>/…/Base_<game>.xlsx
      2. Anywhere under ROOT matching Base_<game>.xlsx
      3. Shared Base.xlsx anywhere under ROOT
      4. Legacy f_rules_Gclaude.xlsx anywhere under ROOT
    Returns a DataFrame of w-columns (column-oriented), or empty DataFrame.
    """
    gcfg = GAMES_CFG.get(game_key, {})
    b_file = gcfg.get("b_file", f"Base_{game_key}.xlsx")
    b_sheet = gcfg.get("b_sheet", f"B_{game_key}")
    b_sheet_legacy = gcfg.get("b_sheet_legacy", "")

    candidates = (list(ROOT.rglob(b_file))
                  or list(ROOT.rglob("Base.xlsx"))
                  or list(ROOT.rglob("f_rules_Gclaude.xlsx")))
    if not candidates:
        return pd.DataFrame()

    for path in candidates:
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
            sheet = None
            for cand in (b_sheet, game_key.upper(), b_sheet_legacy):
                if cand and cand in xl.sheet_names:
                    sheet = cand
                    break
            if sheet is None and len(xl.sheet_names) == 1:
                sheet = xl.sheet_names[0]
            if sheet is None:
                continue
            raw = xl.parse(sheet, header=None)
            w_cols = [str(raw.iloc[0, c]) for c in range(raw.shape[1])
                      if str(raw.iloc[0, c]).startswith("w")]
            b_data = {}
            for i, wc in enumerate(w_cols):
                col_vals = raw.iloc[1:, i].dropna()
                nums = [int(float(v)) for v in col_vals
                        if str(v).replace(".", "").replace("-", "").isdigit()
                        and float(v) >= 1]
                if nums:
                    b_data[wc] = pd.Series(nums)
            if b_data:
                return pd.DataFrame(b_data)
        except Exception:
            continue
    return pd.DataFrame()


def _init_state():
    if "S" in st.session_state:
        return
    # Auto-load B for the default game (sat) at startup.
    # The active game can change via the game selector; B will reload in the B tab.
    _b_init = _auto_load_b("sat")
    _sat_dirs = game_dirs("sat")
    if _b_init.empty:
        _b_init = _load_file(_sat_dirs["Base"] / "B.xlsx")   # legacy fallback

    st.session_state.S = {
        # Variable DataFrames
        "B":   _b_init,
        "R":   pd.DataFrame(),
        "D":   _load_file(_sat_dirs["Direct"]       / "D.xlsx"),
        "Sp":  _load_file(_sat_dirs["Splits"]       / "data_1b.xlsx"),
        "So":  _load_file(_sat_dirs["Splits_Combi"] / "data.xlsx"),
        # Main Data (user uploads manually each run)
        "main_data": pd.DataFrame(),
        # Collated CVI results per formula
        "cvi": {},
        # Matching results per dashboard
        "results": {},
        # Container formula active flags
        "cf_active": {r[1]: True for r in CF_ROWS},
        # Auto mode per section
        "auto": {},
        # Scraper state
        "confirmed_api_url": "",
        "cookie_str": "",
        "scrape_log": [],
        # Container status per dashboard
        "container_status": {},
    }
    # Init container status
    for db in DASHBOARDS:
        st.session_state.S["container_status"][db] = pd.DataFrame({
            "Name": [db], "Status": ["Stopped"],
            "Memory Usage": [""], "CPU Usage": [""],
            "Percent progress %": [0],
            "Time Guestimate Min": [""], "Time Guestimate Max": [""],
        })

S = None  # set after _init_state() call in main

# ═══════════════════════════════════════════════════════════════════════════════
# 4. FILE I/O — large-file safe, all formats
# ═══════════════════════════════════════════════════════════════════════════════
LARGE_FILE_THRESHOLD = 50 * 1024 * 1024   # 50 MB

def _count_csv_rows(path: Path) -> int:
    """Fast row count without loading into memory."""
    try:
        with open(path, "rb") as f:
            return sum(1 for _ in f) - 1
    except Exception:
        return 0

def _load_file(path: Path, max_rows: int = 0,
               numeric: bool = True) -> pd.DataFrame:
    """
    Load any supported file format.
    Large CSVs (>50 MB) are read in 500K-row chunks to avoid freezing.
    accdb/mdb files must be placed directly in the project folder —
    Streamlit cannot upload them via the browser (browser restriction, not ours).
    Use mdb-tools on Mac: brew install mdb-tools
    """
    if not path or not path.exists():
        return pd.DataFrame()
    try:
        ext  = path.suffix.lower()
        size = path.stat().st_size

        if ext == ".csv":
            if max_rows > 0:
                df = pd.read_csv(path, nrows=max_rows)
            elif size > LARGE_FILE_THRESHOLD:
                chunks = pd.read_csv(path, chunksize=500_000)
                df = pd.concat(chunks, ignore_index=True)
            else:
                df = pd.read_csv(path)
        elif ext in [".xlsx", ".xlsm"]:
            df = pd.read_excel(path, engine="openpyxl",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext == ".xls":
            df = pd.read_excel(path, engine="xlrd",
                               nrows=max_rows if max_rows > 0 else None)
        elif ext in [".accdb", ".mdb"]:
            try:
                import pyodbc
                conn = pyodbc.connect(
                    f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={path};")
                cursor = conn.cursor()
                tables = [t.table_name for t in cursor.tables(tableType="TABLE")]
                q = (f"SELECT TOP {max_rows} * FROM [{tables[0]}]"
                     if max_rows > 0 else f"SELECT * FROM [{tables[0]}]")
                df = pd.read_sql(q, conn)
                conn.close()
            except Exception:
                import subprocess, io
                try:
                    tables_raw = subprocess.check_output(
                        ["mdb-tables", "-1", str(path)], timeout=30).decode()
                    tables = [t for t in tables_raw.strip().split("\n") if t]
                    csv_data = subprocess.check_output(
                        ["mdb-export", str(path), tables[0]],
                        timeout=120).decode()
                    df = pd.read_csv(io.StringIO(csv_data))
                    if max_rows > 0:
                        df = df.head(max_rows)
                except Exception:
                    return pd.DataFrame()
        elif ext == ".txt":
            for sep in ["\t", ",", " ", "|"]:
                try:
                    df = pd.read_csv(path, sep=sep,
                                     nrows=max_rows if max_rows > 0 else None)
                    if len(df.columns) >= 3:
                        break
                except Exception:
                    pass
            else:
                return pd.DataFrame()
        elif ext == ".xml":
            df = pd.read_xml(path)
        else:
            return pd.DataFrame()

        if numeric:
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()

def _load_file_preview(path: Path, n: int = 20) -> pd.DataFrame:
    """Load first N rows only — never freezes."""
    return _load_file(path, max_rows=n, numeric=False)

def _save_df(df: pd.DataFrame, path: Path):
    """Save DataFrame to CSV, padding n-columns to consistent width.

    Operates on a copy so the caller's DataFrame is never mutated. (Previously
    this added missing n-columns in place, silently altering the object the
    caller still held.)
    """
    if df.empty:
        return
    out = df.copy()
    nc = sorted([c for c in out.columns if re.match(r'^n\d+$', c)],
                key=lambda x: int(x[1:]))
    if nc:
        max_n = int(nc[-1][1:])
        for i in range(1, max_n+1):
            if f"n{i}" not in out.columns:
                out[f"n{i}"] = None
    out.to_csv(path, index=False)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_styled_excel(df: pd.DataFrame, cvi_set: set,
                    n_cols: list, sheet_name: str = "Data") -> bytes:
    """
    Export DataFrame to Excel with highlighted cells.
    Cells whose value is in cvi_set get a salmon/pink background.
    Count column gets a yellow background.
    Returns bytes suitable for st.download_button.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    PINK   = PatternFill("solid", fgColor="FFB3B3")
    YELLOW = PatternFill("solid", fgColor="FFFF99")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    EXCEL_ROW_LIMIT = 1_048_575  # Excel max rows minus header
    truncated = False
    if len(df) > EXCEL_ROW_LIMIT:
        df = df.iloc[:EXCEL_ROW_LIMIT].copy()
        truncated = True

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    cols = list(df.columns)
    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = GREY
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = 8

    if truncated:
        # Write a notice row beneath the header
        ws.insert_rows(2)
        notice_cell = ws.cell(row=2, column=1,
                              value=f"⚠ Truncated to {EXCEL_ROW_LIMIT:,} rows "
                                    f"(Excel limit). Full data available as CSV.")
        notice_cell.font = Font(bold=True, color="FF0000")

    # Data rows (start at row 3 if truncated notice inserted, else row 2)
    row_offset = 3 if truncated else 2
    for ri, (_, row) in enumerate(df.iterrows(), row_offset):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci,
                           value=None if pd.isna(val) else val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            if col == "Count":
                cell.fill = YELLOW
            elif col in n_cols:
                try:
                    if int(round(float(val))) in cvi_set:
                        cell.fill = PINK
                except (ValueError, TypeError):
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# 5. HIGH-PERFORMANCE MATCHING ENGINE (numpy vectorised, chunked, parallel)
# ═══════════════════════════════════════════════════════════════════════════════

# ── DuckDB engine — for files >50MB / 63M+ rows ───────────────────────────────
DUCKDB_THRESHOLD = 50 * 1024 * 1024   # 50 MB CSV → use DuckDB

def _duckdb_available() -> bool:
    try:
        import duckdb
        return True
    except ImportError:
        return False


def _match_duckdb(csv_path: Path, cvi_nums: list[int],
                  n_col_count: int) -> np.ndarray:
    """
    DuckDB-based matching for very large CSV files.
    Reads CSV directly from disk — never loads full file into RAM.
    Works on 63M rows in ~30 seconds vs pandas which would freeze.

    csv_path:     path to the main data CSV file
    cvi_nums:     list of CVI numbers to match against
    n_col_count:  number of n-columns (n1..nN)

    Returns np.ndarray of match counts per row (length = row count in file).
    """
    import duckdb

    n_cols_sql = " + ".join(
        f"(CASE WHEN n{i} IN ({','.join(str(x) for x in cvi_nums)}) "
        f"THEN 1 ELSE 0 END)"
        for i in range(1, n_col_count + 1)
    )

    query = f"""
        SELECT {n_cols_sql} AS match_count
        FROM read_csv_auto('{csv_path}',
            ignore_errors=true,
            sample_size=10000)
        ORDER BY rowid
    """

    try:
        con = duckdb.connect(database=":memory:")
        result = con.execute(query).fetchnumpy()
        con.close()
        counts = result["match_count"].astype(np.int32)
        return counts
    except Exception:
        return np.array([], dtype=np.int32)


def _smart_match(main_df_or_path, n_cols: list,
                 cvi_nums: np.ndarray) -> np.ndarray:
    """
    Route to DuckDB (large files) or numpy (in-memory).
    main_df_or_path: pd.DataFrame already loaded, OR Path to CSV file.
    """
    if isinstance(main_df_or_path, Path):
        size = main_df_or_path.stat().st_size
        if size > DUCKDB_THRESHOLD and _duckdb_available():
            return _match_duckdb(
                main_df_or_path, cvi_nums.tolist(), len(n_cols))
        else:
            # Load in chunks if large
            chunks = pd.read_csv(main_df_or_path, chunksize=500_000)
            df = pd.concat(chunks, ignore_index=True)
            return _count_matches(df, n_cols, cvi_nums)
    else:
        return _count_matches(main_df_or_path, n_cols, cvi_nums)


def _match_chunk(chunk_arr: np.ndarray, cvi_arr: np.ndarray) -> np.ndarray:
    """
    Vectorised intersection count.
    chunk_arr: (N, K) float32 — main data chunk
    cvi_arr:   (M,)   float32 — CVI numbers
    returns:   (N,)   int32   — match count per row
    Handles 500K rows in ~0.05s on M3 Pro Max.

    NOTE (restored): the `def` line for this function had been lost in a
    previous edit. That left this body as unreachable code *inside*
    `_smart_match`, so `_match_chunk` was never actually defined — and every
    call to `_count_matches` / `run_matching` raised `NameError`. It is now a
    proper module-level function again. Logic is unchanged.
    """
    counts = np.zeros(chunk_arr.shape[0], dtype=np.int32)
    for j in range(chunk_arr.shape[1]):
        counts += np.isin(chunk_arr[:, j], cvi_arr).astype(np.int32)
    return counts


DISPLAY_THRESHOLD = 500_000   # rows — above this, don't store per-row DataFrames


def _parse_cvi_col(series: pd.Series) -> list[int]:
    """
    Extract valid integers from a CVI column.
    Handles floats (1.0→1), strings ("1"), NaN safely.
    No upper range cap — CVI values are not raw lottery balls
    and can legitimately exceed 45 depending on the collation formula.
    """
    result = []
    for v in series.dropna():
        try:
            n = float(str(v).strip())
            if not pd.isna(n) and n >= 1:          # only require positive integer
                result.append(int(round(n)))
        except (ValueError, TypeError):
            pass
    return result


def _count_matches(df: pd.DataFrame, n_cols: list,
                   cvi_nums: np.ndarray) -> np.ndarray:
    """Vectorised match count — chunks handled internally.

    Accumulates each chunk's counts in a list and concatenates once at the end.
    (Previously this re-allocated and copied a growing array on every chunk —
    O(n²) total copying for many chunks. Output is identical.)
    """
    parts = []
    for start in range(0, len(df), CHUNK_SIZE):
        chunk  = df.iloc[start:start+CHUNK_SIZE]
        arr    = chunk[n_cols].to_numpy(dtype=np.float32)
        arr    = np.nan_to_num(arr, nan=-1.0)
        parts.append(_match_chunk(arr, cvi_nums))
    return np.concatenate(parts) if parts else np.empty(0, dtype=np.int32)


def _count_str(counts: np.ndarray) -> str:
    return ",".join(str(int(k)) for k in sorted(np.unique(counts)))


def _breakdown_str(counts: np.ndarray, sc_set: set,
                   prefix_sel="S", prefix_unsel="U") -> tuple[str, str, str, str]:
    """
    Returns (sel_bd_str, unsel_bd_str, unsel_count_str, count_dist_dict_str).
    """
    unique = sorted(int(k) for k in np.unique(counts))
    sel_parts, unsel_parts = [], []
    for k in unique:
        n_k = int((counts == k).sum())
        if k in sc_set:
            sel_parts.append(f"{prefix_sel}{k}:{n_k}")
        else:
            unsel_parts.append(f"{prefix_unsel}{k}:{n_k}")
    unsel_counts = [k for k in unique if k not in sc_set]
    return (
        "  ".join(sel_parts)   or "No breakdown",
        "  ".join(unsel_parts) or "No breakdown",
        ",".join(str(k) for k in unsel_counts) or "No count",
        {f"S{k}": int((counts==k).sum()) for k in unique},
    )


def run_matching(main_df: pd.DataFrame,
                 cvi_df:  pd.DataFrame,
                 sc_dict: dict,
                 carry_fwd: dict,
                 main_path=None) -> dict:
    """
    Sequential matching engine — corrected carry-forward logic.

    carry_fwd[w] = "U" or "S"  (default "U" for any missing key)
    Meaning of toggle at Row N:
        "U" → Row N's present = Row N-1's UNSELECTED
        "S" → Row N's present = Row N-1's SELECTED

    Main Count: pre-computed once against ORIGINAL main_df for every w-column.
                Never recalculated regardless of carry-forward state.

    Memory: if M > DISPLAY_THRESHOLD, per-row DataFrames are NOT stored —
            only count distributions are kept. Final stage data always stored.

    main_path (optional): path to the CSV that `main_df` was loaded from. When
        supplied AND it is a large CSV AND DuckDB is installed AND it is safe
        (file row-count == M and the n-columns are a contiguous n1…nN block),
        the heavy *main-count pre-pass* (one full-M scan per w-column) is run in
        DuckDB instead of pandas. If DuckDB returns an unexpected length for any
        column, that column silently falls back to the pandas path, so results
        are always identical to the pure-pandas engine. Passing None (the
        default) keeps the original all-pandas behavior.

        NOTE on scope: only the pre-pass uses DuckDB. The staged carry-forward
        matching still runs in memory on `main_df`, because each stage filters a
        shrinking pool of rows (selected vs unselected) — that is row-level and
        stateful, not a single SQL aggregate. So this lowers the cost of the
        most expensive *uniform* passes but does NOT remove the need to hold
        `main_df` in RAM. Fully streaming the staged engine is a larger,
        separate change.
    """
    if main_df.empty or cvi_df.empty:
        return {"selected": pd.DataFrame(), "unselected": pd.DataFrame(),
                "fig9_table": pd.DataFrame(), "breakdown": pd.DataFrame(),
                "debug_rows": []}

    # ── Detect number columns ──────────────────────────────────────────────
    # Prefer explicit n1…nN columns when present. Only fall back to the
    # "≥90% numeric" heuristic when there are NO explicit n-columns — otherwise
    # numeric metadata columns (Postcode, Length, Draw Number, …) get treated
    # as lottery numbers and corrupt every match count.
    _explicit_n = [c for c in main_df.columns if re.match(r'^n\d+$', c, re.I)]
    if _explicit_n:
        n_cols = sorted(
            _explicit_n,
            key=lambda x: int(re.sub(r'\D', '', x) or 0)
        )[:20]
    else:
        n_cols = sorted(
            [c for c in main_df.columns
             if pd.to_numeric(main_df[c], errors="coerce").notna().mean() > 0.9],
            key=lambda x: int(re.sub(r'\D', '', x) or 0)
        )[:20]

    w_cols = sorted(
        [c for c in cvi_df.columns if re.match(r'^w\d+$', c)],
        key=lambda x: int(x[1:])
    )

    M = len(main_df)
    small_enough = (M <= DISPLAY_THRESHOLD)

    # ── PRE-COMPUTE Main Count AND Main Breakdown for every w-column ──────
    # Computed once against original M rows. Never recalculated.
    main_count_map: dict[str, str] = {}
    main_bd_map:    dict[str, str] = {}   # full distribution e.g. "S0:1 S1:3 S2:9..."
    main_counts_map: dict[str, np.ndarray] = {}  # raw counts for filter UI

    # ── Decide whether the pre-pass may use DuckDB (safe, opt-in) ──────────
    # Only when: a path was given, it's a large CSV, DuckDB is installed, the
    # file's row count matches M (so DuckDB row order lines up with main_df),
    # and the n-columns form a contiguous n1…nN block (so _match_duckdb's
    # generated SQL references columns that actually exist).
    _duck_path = None
    try:
        if main_path is not None:
            _p = Path(main_path)
            _contiguous_n = (
                bool(n_cols)
                and all(c.lower() == f"n{i}" for i, c in enumerate(n_cols, 1))
            )
            if (_p.exists() and _p.suffix.lower() == ".csv"
                    and _duckdb_available()
                    and _p.stat().st_size > DUCKDB_THRESHOLD
                    and _contiguous_n
                    and _count_csv_rows(_p) == M):
                _duck_path = _p
    except Exception:
        _duck_path = None

    for w in w_cols:
        raw = _parse_cvi_col(cvi_df[w])
        if raw:
            mc = None
            if _duck_path is not None:
                # DuckDB streams the file from disk for this one w-column.
                duck_mc = _match_duckdb(_duck_path, raw, len(n_cols))
                # Trust it only if it returned exactly M counts; otherwise the
                # pandas path below guarantees a correct result.
                if duck_mc.shape[0] == M:
                    mc = duck_mc
            if mc is None:
                mc = _count_matches(
                    main_df, n_cols, np.array(raw, dtype=np.float32))
            main_counts_map[w] = mc
            main_count_map[w]  = _count_str(mc)
            # Full breakdown (all M rows, no SC filter — just distribution)
            unique_k = sorted(int(k) for k in np.unique(mc))
            main_bd_map[w] = "  ".join(
                f"S{k}:{int((mc==k).sum())}" for k in unique_k
            )
        else:
            main_count_map[w]   = "—"
            main_bd_map[w]      = "—"
            main_counts_map[w]  = np.array([], dtype=np.int32)

    # ── Stage state ────────────────────────────────────────────────────────
    # prev_sel / prev_unsel are the two pools produced by the PREVIOUS stage.
    # The toggle at the CURRENT stage picks which one becomes "present".
    prev_sel   = pd.DataFrame()        # Row 0 has no previous selected
    prev_unsel = main_df.copy().reset_index(drop=True)  # Row 1 default = all M

    fig9_rows      = []
    breakdown_rows = []
    debug_rows     = []

    final_sel   = pd.DataFrame()
    final_unsel = prev_unsel.copy()

    for idx, w in enumerate(w_cols):
        w_num     = int(w[1:])
        direction = carry_fwd.get(w, carry_fwd.get(f"w{w_num}", "U")).upper()

        # ── Determine present_df for this stage ───────────────────────
        if idx == 0:
            # First row always receives full original main data
            present_df    = main_df.copy().reset_index(drop=True)
            present_label = f"M:{M}"
        elif direction == "S":
            present_df    = prev_sel.reset_index(drop=True)
            present_label = f"S:{len(prev_sel)}"
        else:  # "U"
            present_df    = prev_unsel.reset_index(drop=True)
            present_label = f"U:{len(prev_unsel)}"

        present_count = len(present_df)
        # Defensive: sync label with actual count (catches any reference drift)
        if present_label.startswith("U:"):
            present_label = f"U:{present_count}"
        elif present_label.startswith("S:") and not present_label.startswith("S:0"):
            present_label = f"S:{present_count}"
        raw_cvi       = _parse_cvi_col(cvi_df[w])
        cvi_nums      = np.array(raw_cvi, dtype=np.float32)
        sc_this       = sc_dict.get(w, sc_dict.get(str(w_num), []))
        if isinstance(sc_this, (int, float)):
            sc_this = [int(sc_this)]
        elif isinstance(sc_this, str):
            sc_this = [int(x.strip()) for x in sc_this.split(",")
                       if x.strip().lstrip('-').isdigit()]
        sc_set  = set(sc_this)
        sc_str  = ",".join(str(s) for s in sorted(sc_set)) if sc_set else "—"

        # ── Empty CVI column ──────────────────────────────────────────
        if not len(cvi_nums):
            fig9_rows.append({
                "Row":            f"Row {w_num}",
                "Main\nData":     f"M:{M}",
                "CVI":            w,
                "Dir":            direction,
                "Main\nCount":    main_count_map[w],
                "Main\nBreakdown":main_bd_map[w],
                "Present\nData":  present_label,
                "Present\nCount": "— No CVI",
                "SC":             sc_str,
                "Selected":       "S:0",
                "Sel\nBreakdown": "—",
                "Unsel\nCount":   "—",
                "Unselected":     f"U:{present_count} (fwd)",
                "Unsel\nBreakdown":"—",
            })
            debug_rows.append({
                "w": w, "direction": direction,
                "cvi_numbers": [], "cvi_set": set(),
                "present_in": present_count, "present_label": present_label,
                "sc": list(sc_set), "selected_n": 0,
                "unselected_n": present_count,
                "main_count_str": main_count_map[w],
                "main_bd_str":    main_bd_map[w],
                "main_counts":    main_counts_map[w],
                "pres_count_str": "—", "count_dist": {},
                "note": "No CVI data — all carry forward",
                "sel_df": None,
                "unsel_df": present_df.copy() if small_enough else None,
                "n_cols": n_cols,
            })
            prev_sel   = pd.DataFrame()
            prev_unsel = present_df.copy()
            final_sel   = pd.DataFrame()
            final_unsel = present_df.copy()
            continue

        # ── Match present_df against this w-column ────────────────────
        # Via the _smart_match dispatcher: with an in-memory DataFrame this is
        # exactly _count_matches(...). present_df is a shrinking pool (it cannot
        # be re-read from disk), so this stage stays in RAM by design.
        pres_counts = _smart_match(present_df, n_cols, cvi_nums)
        sel_bd_str, unsel_bd_str, unsel_count_str, count_dist = \
            _breakdown_str(pres_counts, sc_set)
        pres_count_str = _count_str(pres_counts)

        sel_mask   = np.isin(pres_counts, list(sc_set))
        sel_df     = present_df[sel_mask].reset_index(drop=True)
        unsel_df   = present_df[~sel_mask].reset_index(drop=True)

        # Add count column to sel/unsel for display
        if small_enough and not sel_df.empty:
            sel_counts_arr = pres_counts[sel_mask]
            sel_df = sel_df.copy()
            sel_df["Count"] = sel_counts_arr
        if small_enough and not unsel_df.empty:
            unsel_df = unsel_df.copy()
            unsel_df["Count"] = pres_counts[~sel_mask]

        fig9_rows.append({
            "Row":            f"Row {w_num}",
            "Main\nData":     f"M:{M}",
            "CVI":            w,
            "Dir":            direction,
            "Main\nCount":    main_count_map[w],
            "Main\nBreakdown":main_bd_map[w],
            "Present\nData":  present_label,
            "Present\nCount": pres_count_str,
            "SC":             sc_str,
            "Selected":       f"S:{len(sel_df)}" if not sel_df.empty else "S:0",
            "Sel\nBreakdown": sel_bd_str,
            "Unsel\nCount":   unsel_count_str,
            "Unselected":     f"U:{len(unsel_df)}" if not unsel_df.empty else "U:0",
            "Unsel\nBreakdown":unsel_bd_str,
        })

        bd = {**count_dist, "w_col": w}
        breakdown_rows.append(bd)

        # Also add count to main_df for display purposes
        main_df_with_count = None
        if small_enough:
            mc_arr = main_counts_map.get(w, np.array([], dtype=np.int32))
            if len(mc_arr) == M:
                main_df_with_count = main_df.copy()
                main_df_with_count["Count"] = mc_arr

        debug_rows.append({
            "w":              w,
            "direction":      direction,
            "cvi_numbers":    raw_cvi,
            "cvi_set":        set(raw_cvi),
            "present_in":     present_count,
            "present_label":  present_label,
            "sc":             sorted(sc_set),
            "selected_n":     len(sel_df),
            "unselected_n":   len(unsel_df),
            "main_count_str": main_count_map[w],
            "main_bd_str":    main_bd_map[w],
            "main_counts":    main_counts_map[w],
            "pres_count_str": pres_count_str,
            "count_dist":     count_dist,
            "note":           "",
            "sel_df":         sel_df   if small_enough else None,
            "unsel_df":       unsel_df if small_enough else None,
            "main_df_wc":     main_df_with_count,
            "n_cols":         n_cols,
        })

        # ── Carry forward: store both pools for next stage ────────────
        prev_sel   = sel_df.drop(columns=["Count"], errors="ignore")
        prev_unsel = unsel_df.drop(columns=["Count"], errors="ignore")
        final_sel   = sel_df
        final_unsel = unsel_df

        # ── If present exhausted, fill remaining rows ─────────────────
        if present_df.empty or (sel_df.empty and unsel_df.empty):
            for rw in w_cols[idx+1:]:
                rw_num = int(rw[1:])
                fig9_rows.append({
                    "Row": f"Row {rw_num}", "Main\nData": f"M:{M}",
                    "CVI": rw, "Dir": carry_fwd.get(rw,"U"),
                    "Main\nCount": main_count_map.get(rw,"—"),
                    "Main\nBreakdown": main_bd_map.get(rw,"—"),
                    "Present\nData": "U:0", "Present\nCount": "—",
                    "SC": sc_dict.get(rw,"—"), "Selected": "S:0",
                    "Sel\nBreakdown": "—", "Unsel\nCount": "—",
                    "Unselected": "U:0", "Unsel\nBreakdown": "—",
                })
                debug_rows.append({
                    "w": rw, "direction": carry_fwd.get(rw,"U"),
                    "cvi_numbers": [], "cvi_set": set(),
                    "present_in": 0, "present_label": "U:0",
                    "sc": [], "selected_n": 0, "unselected_n": 0,
                    "main_count_str": main_count_map.get(rw,"—"),
                    "main_bd_str": main_bd_map.get(rw,"—"),
                    "main_counts": np.array([], dtype=np.int32),
                    "pres_count_str": "—", "count_dist": {},
                    "note": "No present data — exhausted",
                    "sel_df": None, "unsel_df": None,
                    "main_df_wc": None, "n_cols": n_cols,
                })
            break

    return {
        "selected":    final_sel,
        "unselected":  final_unsel,
        "fig9_table":  pd.DataFrame(fig9_rows),
        "breakdown":   pd.DataFrame(breakdown_rows).fillna(0),
        "debug_rows":  debug_rows,
        "n_cols":      n_cols,
        "small_enough": small_enough,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# 6. COLLATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
def sort_d_longest_first(D: pd.DataFrame) -> pd.DataFrame:
    """Return D with rows ordered LONGEST entry (most numbers) → SHORTEST.

    'Length' = how many w-columns are filled in that row (a System 20 entry = 20,
    a standard game = 6). Stable sort, so equal-length rows keep their order.
    """
    if D is None or D.empty:
        return D
    wcols = [c for c in D.columns if re.match(r'^w\d+$', str(c), re.I)]
    if not wcols:
        return D
    lengths = D[wcols].notna().sum(axis=1)
    order = lengths.sort_values(ascending=False, kind="stable").index
    return D.loc[order].reset_index(drop=True)


def prepare_d_input_sets(D: pd.DataFrame, n: int) -> pd.DataFrame:
    """Peel off the N LONGEST entries of D as the generator input columns.

    Per the pipeline brief: the longest D rows become the w1, w2, w3, w4 (… w8) input
    SETS for the generators — Sp/So take the first 4 longest, Ep takes the first 8.
    Each chosen entry (a row of numbers in D) is laid DOWN one column, so the result
    is a DataFrame whose columns are w1, w2, w3, w4, … each holding one long entry's
    numbers.

    Example: D's three System-20 rows + one System-19 row (3+1=4) → columns w1,w2,w3,w4.
    Note: In ALL D, each syndicate IS a row labeled w1,w2… (number positions across).
    Here we peel those rows into COLUMNS w1,w2,w3,w4 (one syndicate per column) so the
    generators can work with them as independent sets.
    """
    if D is None or D.empty or n < 1:
        return pd.DataFrame()
    wcols = [c for c in D.columns if re.match(r'^w\d+$', str(c), re.I)]
    if not wcols:
        return pd.DataFrame()
    ordered = sort_d_longest_first(D)
    top = ordered[wcols].head(n)
    labels = [f"w{i+1}" for i in range(n)]  # w1, w2, w3, w4, …
    cols = {}
    for label, (_, row) in zip(labels, top.iterrows()):
        nums = [int(x) for x in row.dropna()]
        cols[label] = pd.Series(nums, dtype="Int64")
    return pd.DataFrame(cols)


# ═════════════════════════════════════════════════════════════════════════════
# VARIABLE GENERATORS (all inlined here — one file)
#   R  ← generate_rainbow        (input: Since-Last table)
#   Sp ← generate_splits         (input: D's 4 longest rows; auto half-splits)
#   So ← generate_splits_combi   (input: D's 4 longest rows; auto half-splits)
#   Ep ← generate_excelpro       (ExcelPro Java→Python; input: D's 8 longest + R wt)  [pending]
# Each returns COLUMN-oriented sets; the row-collation transposes them into w-rows.
# Parameters are auto-decided (half-splits, safe-max); the UI offers a timed window
# to override before the automatic choice is committed.
# ═════════════════════════════════════════════════════════════════════════════
def auto_half_splits(sets_dict: dict) -> list:
    """Split each set in halves, rounding UP: len 20→10, 17→9 (= (len+1)//2)."""
    return [(len(v) + 1) // 2 for v in sets_dict.values()]


def _d_input_to_sets(d_input_df: pd.DataFrame) -> dict:
    """Columns w1,w2,w3,w4… (each a long D entry down the rows) → {label: [numbers]}."""
    out = {}
    for col in d_input_df.columns:
        out[str(col)] = [int(x) for x in d_input_df[col].dropna().tolist()]
    return out


# ── R : Rainbow (ported from task2.py) ──────────────────────────────────────
def _bounded_combos(keys, max_comb):
    """Combinations of `keys` sized 1..max_comb, generated directly (no 2**K tail)."""
    from itertools import chain, combinations
    keys = list(keys)
    hi = min(max_comb, len(keys))
    return chain.from_iterable(combinations(keys, r) for r in range(1, hi + 1))


def _combo_count(n_keys, max_comb):
    from math import comb
    hi = min(max_comb, n_keys)
    return sum(comb(n_keys, r) for r in range(1, hi + 1))


def _normalise_sl(sl_df: pd.DataFrame) -> pd.DataFrame:
    df = sl_df.copy()
    lower = {str(c).strip().lower(): c for c in df.columns}
    def pick(*cands):
        for cand in cands:
            if cand in lower:
                return lower[cand]
        return None
    c_num = pick("numbers", "number", "n", "ball")
    c_sl = pick("since last", "since_last", "sincelast", "since", "sl")
    c_keep = pick("to_keep", "to keep", "tokeep", "keep")
    if c_num is None or c_sl is None:
        raise ValueError("Since-Last table needs a number column and a 'Since Last' "
                         "column (got: %s)" % list(df.columns))
    out = pd.DataFrame({
        "numbers": pd.to_numeric(df[c_num], errors="coerce"),
        "since_last": pd.to_numeric(df[c_sl], errors="coerce"),
    })
    out["to_keep"] = (pd.to_numeric(df[c_keep], errors="coerce")
                      if c_keep is not None else out["numbers"])
    out = out.dropna(subset=["since_last", "numbers"], how="any")
    out["since_last"] = out["since_last"].astype(int) + 1
    out["numbers"] = out["numbers"].astype(int)
    return out


def _wt_writer(df: pd.DataFrame) -> pd.DataFrame:
    """task2.py's 'wt' sheet (SL/wt/…/wt_abcd) — Ep reads this as input2 'All wt'."""
    listed = df.groupby("since_last")["numbers"].apply(list).to_dict()
    keep = set(df["to_keep"].dropna().tolist())
    i0, w0, i1, w1, iF, wF = [], [], [], [], [], []
    for i in sorted(listed.keys()):
        i0 += [i] + [None] * (len(listed[i]) - 1)
        i1 += [i] + [None] * len(listed[i])
        w0 += listed[i]
        w1 += listed[i] + [None]
        kref = [e for e in listed[i] if e in keep]
        iF += [i] + [None] * (len(kref) - 1)
        wF += kref
    gen = [i0, w0, i1, w1, iF, wF]
    out = pd.DataFrame({n: pd.Series(gen[n]) for n in range(len(gen))})
    out.columns = ["SL", "wt", "SL_", "wt_", "_SL_", "wt_abcd"]
    return out


def generate_rainbow(sl_df: pd.DataFrame, max_comb=None, combo_guard: int = 200_000):
    """R from the Since-Last table. Returns (result_df, wt_df, info). Safe-max guard
    auto-lowers max_comb until the combo-count fits — no 2**K explosion."""
    df = _normalise_sl(sl_df)
    grouped = df.groupby("since_last")["numbers"].apply(list).to_dict()
    to_keep = set(df["to_keep"].dropna().tolist())
    keys = sorted(grouped.keys())
    n = len(keys)
    requested = n if max_comb is None else max(1, min(int(max_comb), n))
    capped = False
    while requested > 1 and _combo_count(n, requested) > combo_guard:
        requested -= 1
        capped = True
    result = {}
    for comb in _bounded_combos(keys, requested):
        ref = []
        for g in comb:
            ref += grouped[g]
        result[str(comb)] = [e for e in ref if e in to_keep]
    result_df = pd.DataFrame({k: pd.Series(v, dtype="Int64") for k, v in result.items()})
    if not result_df.empty:
        result_df = result_df[result_df.isna().sum().sort_values(kind="stable").index]
    info = {"n_groups": n, "max_comb": requested,
            "n_combos": result_df.shape[1], "capped": capped, "combo_guard": combo_guard}
    return result_df, _wt_writer(df), info


# ── Sp : Splits (ported from task1b.py) ──────────────────────────────────────
def generate_splits(d_input_df: pd.DataFrame, splitter=None) -> pd.DataFrame:
    """Sp from D's 4 longest rows (columns w1,w2,w3,w4). Auto half-splits unless given."""
    from itertools import combinations
    sets_ready = _d_input_to_sets(d_input_df)
    keys = list(sets_ready.keys())
    if splitter is None:
        splitter = auto_half_splits(sets_ready)
    split_sets = {}
    for idx, k in enumerate(keys):
        split_sets[k + "0"] = sets_ready[k][:splitter[idx]]
        split_sets[k + "1"] = sets_ready[k][splitter[idx]:]
    comb3 = list(combinations(keys, 3))
    comb2 = list(combinations(keys, 2))
    let_3, fin_3 = ["e", "f", "g", "h"], ["o", "p", "q", "r"]
    let_2, fin_2 = ["i", "j", "k", "l", "m", "n"], ["s", "t", "u", "v", "w", "x"]
    comb3dict, comb2dict, result = {}, {}, {}
    universe = {e for k in keys for e in sets_ready[k]}
    for j, comb in enumerate(comb3):
        iter_set = {e for L in comb for e in sets_ready[L]}
        i = 0
        for suffix in ("0", "1"):
            for letter in comb:
                comb3dict[let_3[j] + str(i)] = iter_set - set(split_sets[letter + suffix]); i += 1
    for j, comb in enumerate(comb2):
        iter_set = {e for L in comb for e in sets_ready[L]}
        i = 0
        for suffix in ("0", "1"):
            for letter in comb:
                comb2dict[let_2[j] + str(i)] = iter_set - set(split_sets[letter + suffix]); i += 1
    for i in range(len(let_3)):
        for j in range(len(comb3dict) // len(let_3)):
            result[fin_3[i] + str(j)] = universe - comb3dict[let_3[i] + str(j)]
    for i in range(len(let_2)):
        for j in range(len(comb2dict) // len(let_2)):
            result[fin_2[i] + str(j)] = universe - comb2dict[let_2[i] + str(j)]
    skeys = sorted(split_sets.keys())
    for i, key in enumerate(split_sets):
        result[key] = split_sets[key]
        result["y" + str(i)] = universe - set(split_sets[skeys[i]])
    result.update(comb3dict)
    result.update(comb2dict)
    return pd.DataFrame({k: pd.Series(list(v), dtype="Int64") for k, v in result.items()})


# ── So : Splits-combi (ported from automation_vba.py, len==4 branch) ─────────
def generate_splits_combi(d_input_df: pd.DataFrame, splitter=None) -> pd.DataFrame:
    """So from D's 4 longest rows (columns w1,w2,w3,w4). Auto half-splits unless given.
    Ports the set algebra of automation_vba.sets_sampling; skips the cosmetic
    column-relabeling (not needed for the pipeline variable).

    Key naming: split keys are base_key + "0"/"1" (e.g. 'w10','w11' for base 'w1').
    Membership check strips the trailing suffix char to recover the base key.
    """
    from itertools import combinations
    sets_ready = _d_input_to_sets(d_input_df)
    keys = list(sets_ready.keys())
    if splitter is None:
        splitter = auto_half_splits(sets_ready)
    split_sets = {}
    for idx, k in enumerate(keys):
        split_sets[k + "0"] = sets_ready[k][:splitter[idx]]
        split_sets[k + "1"] = sets_ready[k][splitter[idx]:]
    universe = {e for s in sets_ready.values() for e in s}
    result = {"U": universe}
    comb_pairs = {p: set(split_sets[p[0]]) | set(split_sets[p[1]])
                  for p in combinations(split_sets.keys(), 2)}
    comb_three = {t: set().union(*[set(sets_ready[x]) for x in t])
                  for t in combinations(keys, 3)}
    for ti, tset in comb_three.items():
        for pj, pset in comb_pairs.items():
            # Strip trailing "0"/"1" suffix to get the base key, then check membership
            # in the triple. Works for any key length (w1, w2, ... or a, b, ...).
            base0 = pj[0][:-1]  # e.g. 'w10' → 'w1', 'a0' → 'a'
            base1 = pj[1][:-1]  # e.g. 'w21' → 'w2', 'b1' → 'b'
            if pset.issubset(tset) and base0 in ti and base1 in ti:
                result["U-" + str(ti) + "-" + str(pj)] = universe - (tset - pset)
                result[str(ti) + "-" + str(pj)] = tset - pset
        result[str(ti)] = tset
    for pj, pset in comb_pairs.items():
        result["U-" + str(pj)] = universe - pset
        result[str(pj)] = pset
    return pd.DataFrame({k: pd.Series(sorted(v), dtype="Int64") for k, v in result.items()})


# ── Ep : ExcelPro (ported from Java Main.java — the substantive 'All' output) ─
def prepare_ep_objects(D: pd.DataFrame, mode: str = "pairs") -> dict:
    """Build the 4 ExcelPro objects a, b, c, d — each = (arrayOne, arrayTwo) — from D.

    mode='pairs' (default): the 8 LONGEST D rows paired into 4 objects:
      a = (D-row#1, D-row#2)  →  i.e. the w1 and w2 longest rows
      b = (D-row#3, D-row#4)  →  w3, w4
      c = (D-row#5, D-row#6)  →  w5, w6
      d = (D-row#7, D-row#8)  →  w7, w8
    Each pair's first member becomes arrayOne, second becomes arrayTwo.
    Ep's wt filter list is supplied by R (rainbow).

    mode='halves': the 4 LONGEST D rows, each split in half (arrayOne=first half,
      arrayTwo=second half) — kept as an alternative.
    """
    if mode == "pairs":
        inp = prepare_d_input_sets(D, 8)   # columns: w1, w2, … w8 (8 longest rows)
        cols = list(inp.columns)
        objects = {}
        for i, lab in enumerate(["a", "b", "c", "d"]):
            # pair i: cols[2i] = arrayOne (e.g. w1), cols[2i+1] = arrayTwo (e.g. w2)
            one = [int(x) for x in inp[cols[2 * i]].dropna()] if 2 * i < len(cols) else []
            two = [int(x) for x in inp[cols[2 * i + 1]].dropna()] if 2 * i + 1 < len(cols) else []
            objects[lab] = (one, two)
        return objects
    # mode='halves'
    inp = prepare_d_input_sets(D, 4)       # columns: w1, w2, w3, w4 (4 longest rows)
    objects = {}
    for lab, col in zip(["a", "b", "c", "d"], inp.columns):
        nums = [int(x) for x in inp[col].dropna()]
        half = (len(nums) + 1) // 2
        objects[lab] = (nums[:half], nums[half:])
    return objects


def generate_excelpro(objects: dict, wt_list) -> pd.DataFrame:
    """Ep — the ExcelPro 'All' result (faithful to Main.java comboList → 'All').

    For each pair (x, y) of the 4 objects a, b, c, d, emit 4 columns:
      a_<xy>  — wt numbers found in x.arrayOne
      b_<xy>  — wt numbers found in x.arrayTwo
      c_<xy>  — wt numbers found in y.arrayOne
      d_<xy>  — wt numbers found in y.arrayTwo

    Column naming uses the concatenated object labels (single chars), so pairs produce:
      a_ab, b_ab, c_ab, d_ab,
      a_ac, b_ac, c_ac, d_ac,
      a_ad, b_ad, c_ad, d_ad,
      a_bc, b_bc, c_bc, d_bc,
      a_bd, b_bd, c_bd, d_bd,
      a_cd, b_cd, c_cd, d_cd
    """
    from itertools import combinations as _ep_combos
    keys = list(objects.keys())   # ['a', 'b', 'c', 'd']
    wt = [int(w) for w in wt_list]
    result = {}
    for x, y in _ep_combos(keys, 2):
        h = x + y                 # e.g. 'ab', 'ac', 'ad', 'bc', 'bd', 'cd'
        xo, xt = set(objects[x][0]), set(objects[x][1])
        yo, yt = set(objects[y][0]), set(objects[y][1])
        result["a_" + h] = [w for w in wt if w in xo]
        result["b_" + h] = [w for w in wt if w in xt]
        result["c_" + h] = [w for w in wt if w in yo]
        result["d_" + h] = [w for w in wt if w in yt]
    return pd.DataFrame({k: pd.Series(v, dtype="Int64") for k, v in result.items()})


def _to_w_rows(df: pd.DataFrame, is_direct: bool = False) -> pd.DataFrame:
    """Return a variable as ROW-oriented w-sets — each ROW is one combination.

    Spreadsheet reality: Excel allows ~1,048,576 ROWS but only 16,384 COLUMNS, so
    the big dimension (hundreds of thousands of syndicates) must live in ROWS.

      • D (Direct): rows are ALREADY combinations → keep rows, w-columns only.
        Identified by the caller (is_direct=True) OR by syndicate metadata — so a
        numbers-only D (no Syndicate_ID/draw columns) is still kept as rows and NOT
        transposed back into the column wall.
      • B / R / Ep / Sp / So are stored COLUMN-wise (each column is a combination)
        → TRANSPOSE so each column becomes a row.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    wcols = [c for c in df.columns if re.match(r'^w\d+$', str(c), re.I)]
    D_META = {"syndicate_id", "syndicate_name", "game", "games", "pb",
              "draw_number", "draw_numbers", "outlet_id", "outlet_name",
              "postcode", "state", "share_cost", "available_shares",
              "total_shares", "address", "suburb"}
    has_d_meta = any(str(c).strip().lower() in D_META for c in df.columns)

    if is_direct or has_d_meta:
        # D: rows already ARE combinations — keep them, w-columns only.
        sub = df[wcols] if wcols else df
        out = sub.reset_index(drop=True)
    else:
        # B/R/Ep/Sp/So: columns ARE combinations → transpose to rows.
        sub = df[wcols] if wcols else df
        out = sub.T.reset_index(drop=True)
    # Drop fully-empty rows, coerce to numbers.
    out = out.apply(pd.to_numeric, errors="coerce")
    out = out.dropna(how="all").reset_index(drop=True)
    return out


def _sets_df_to_rows(df: pd.DataFrame, set_col: str = "set") -> pd.DataFrame:
    """Transpose a column-oriented set DataFrame to row-oriented for display/export.

    Input:  columns = set names (e.g. 'ab', 'U', 'w10', 'a_ab'),
            rows    = values padded with NaN to equal length.
    Output: one row per set; first column = set_col (the set label);
            remaining columns = pos_1, pos_2, … (the actual numbers).
    Empty/all-NaN value columns are dropped.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    t = df.T.reset_index()
    t.columns = [set_col] + [f"pos_{i+1}" for i in range(t.shape[1] - 1)]
    # Drop value columns that are entirely NaN
    val_cols = [c for c in t.columns if c != set_col]
    t = t.dropna(subset=val_cols, how="all")
    # Drop trailing all-NaN position columns
    non_empty_val = [c for c in val_cols if t[c].notna().any()]
    t = t[[set_col] + non_empty_val].reset_index(drop=True)
    return t


# ── Universal file reader ──────────────────────────────────────────────────
_UPLOAD_TYPES = [
    "csv", "txt", "tsv",
    "xlsx", "xls", "xlsb", "xlsm", "ods",
    "html", "htm",
    "json", "xml",
    "accdb", "mdb",
]

def _read_uploaded(f, **kw) -> "pd.DataFrame":
    """Read any Streamlit UploadedFile into a DataFrame regardless of extension.

    Supports: CSV/TSV/TXT, Excel (xlsx/xls/xlsb/xlsm/ods),
              HTML tables, JSON, XML, and Access (accdb/mdb via mdbtools).
    Extra kwargs are forwarded to the underlying pandas reader.
    """
    import io as _io
    name = f.name.lower()
    raw = f.getvalue()

    if name.endswith((".xlsx", ".xlsm", ".xlsb")):
        return pd.read_excel(_io.BytesIO(raw), engine="openpyxl", **kw)
    if name.endswith(".xls"):
        try:
            return pd.read_excel(_io.BytesIO(raw), engine="xlrd", **kw)
        except Exception:
            return pd.read_excel(_io.BytesIO(raw), engine="openpyxl", **kw)
    if name.endswith(".ods"):
        return pd.read_excel(_io.BytesIO(raw), engine="odf", **kw)
    if name.endswith((".html", ".htm")):
        tables = pd.read_html(_io.BytesIO(raw), **kw)
        return tables[0] if tables else pd.DataFrame()
    if name.endswith(".json"):
        return pd.read_json(_io.BytesIO(raw), **kw)
    if name.endswith(".xml"):
        return pd.read_xml(_io.BytesIO(raw), **kw)
    if name.endswith(".tsv"):
        return pd.read_csv(_io.BytesIO(raw), sep="\t", **kw)
    if name.endswith((".accdb", ".mdb")):
        try:
            import tempfile, subprocess, os
            with tempfile.NamedTemporaryFile(suffix=os.path.splitext(name)[1],
                                            delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            tables_out = subprocess.run(
                ["mdb-tables", "-1", tmp_path],
                capture_output=True, text=True, timeout=30)
            tbl = tables_out.stdout.strip().splitlines()[0] if tables_out.returncode == 0 else ""
            if tbl:
                result = subprocess.run(
                    ["mdb-export", tmp_path, tbl],
                    capture_output=True, text=True, timeout=30)
                if result.returncode == 0:
                    return pd.read_csv(_io.StringIO(result.stdout), **kw)
            raise RuntimeError(f"mdb-export failed. stderr: {tables_out.stderr}")
        except FileNotFoundError:
            raise RuntimeError(
                "Access (.accdb/.mdb) files need mdbtools installed on the server.\n"
                "  macOS: brew install mdbtools\n"
                "  Linux: sudo apt install mdbtools")
    # Default → CSV (handles .txt and plain .csv)
    for sep in (",", ";", "\t", "|", None):
        try:
            kwargs = dict(**kw)
            if sep is not None:
                kwargs["sep"] = sep
            else:
                kwargs.update(sep=None, engine="python")
            return pd.read_csv(_io.BytesIO(raw), **kwargs)
        except Exception:
            continue
    raise ValueError(f"Could not parse uploaded file: {f.name}")


def execute_collation(components: list[str]) -> pd.DataFrame:
    """
    Build a formula's CVI by STACKING each variable's w-sets as ROWS (vertically)
    and numbering the position columns w1, w2, … across the widest combination.

    Data model (confirmed — ROW orientation, to respect the spreadsheet column
    ceiling of ~16k columns vs ~1M rows):
      • Each w-set (a syndicate pick, a base set, a rainbow set, …) is one ROW.
      • D (Direct) is already row-oriented (one syndicate per row) → kept as rows.
      • B, R, Ep, Sp, So are stored column-wise → TRANSPOSED so each becomes a row.
      • A numbered/pure formula (e.g. D1D2D3) = "all of that variable joined" → a
        SINGLE block (strip trailing digits + de-duplicate): D1D2D3 == one D block.
    Result is TALL. e.g. sat BRD = B rows + R rows + 354,682 D rows, columns
    w1…w(longest combination). Matching is row-vs-row against Main_Data.
    """
    # Resolve tokens → base variables: strip trailing digits (D1→D, Sp2→Sp,
    # B3→B) and de-duplicate while preserving left-to-right order.
    seen, base_vars = set(), []
    for comp in components:
        base = re.sub(r"\d+$", "", str(comp)).strip()
        if base and base not in seen:
            seen.add(base)
            base_vars.append(base)

    pieces = []
    for var in base_vars:
        df = S.get(var) if S else None
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        block = _to_w_rows(df, is_direct=(var == "D"))   # ROW-oriented w-sets
        if block is None or block.empty:
            continue
        block.columns = [f"w{i+1}" for i in range(len(block.columns))]
        block.insert(0, "Source", var)  # remember which variable each row came from
        pieces.append(block)

    if not pieces:
        return pd.DataFrame()

    # Stack vertically; align to the widest combination (pad missing positions).
    combined = pd.concat(pieces, axis=0, ignore_index=True)
    wcols = [c for c in combined.columns if str(c).startswith("w")]
    combined = combined[["Source"] + wcols]
    return combined

# ═══════════════════════════════════════════════════════════════════════════════
# 7. SCRAPER ENGINE (Playwright + requests)
# ═══════════════════════════════════════════════════════════════════════════════
def _extract_nums(text) -> list[int]:
    """Extract positive integers from any text. No upper cap — supports any jurisdiction."""
    return sorted(set(int(n) for n in re.findall(r'\b(\d{1,3})\b', str(text))
                  if int(n) >= 1))

def _valid_combo(nums: list) -> bool:
    if not (6 <= len(nums) <= 25): return False
    if max(nums) - min(nums) < 10: return False
    return True

def _parse_json(body: str, pc: int, state: str) -> list[dict]:
    """Walk any JSON shape to extract syndicate number combinations."""
    try:
        data = json.loads(body)
    except Exception:
        return []
    records, seen = [], set()

    def walk(obj, depth=0):
        if depth > 8: return
        if isinstance(obj, list):
            for item in obj: walk(item, depth+1)
        elif isinstance(obj, dict):
            nums = []
            for f in ["numbers","combination","selections","picks","balls",
                      "gameNumbers","game_numbers","selectedNumbers","ticketNumbers"]:
                if f in obj and isinstance(obj[f], list):
                    nums = sorted(set(int(x) for x in obj[f]
                                  if str(x).isdigit() and int(x) >= 1))
                    if nums: break
            if _valid_combo(nums):
                key = tuple(nums)
                if key not in seen:
                    seen.add(key)
                    def g(*keys):
                        for k in keys:
                            v = obj.get(k)
                            if v is not None and str(v).strip(): return str(v).strip()
                        return ""
                    row = {
                        "Postcode":        g("postcode","post_code") or str(pc),
                        "State":           g("state","jurisdiction") or state,
                        "Syndicate Number": g("id","syndicateId","syndicateNumber"),
                        "Agent":           g("agentName","retailer","outlet"),
                        "Title":           g("title","name","description"),
                        "Price per Share": g("price","sharePrice"),
                        "Shares":          g("shares","sharesAvailable"),
                        "Draw Number":     g("drawNumber","draw_number","drawId","drawNo"),
                        "Draw Date":       g("drawDate","draw_date","scheduledDate"),
                        "Game Type":       g("gameType","type","product","gameName") or
                                          ("Normal" if len(nums)<=7 else f"System {len(nums)}"),
                        "Length":          len(nums),
                    }
                    for j,n in enumerate(nums,1): row[f"n{j}"] = n
                    records.append(row)
            for v in obj.values(): walk(v, depth+1)

    walk(data)
    return records


def _parse_html(html: str, pc: int, state: str) -> list[dict]:
    """HTML fallback — extracts number clusters from rendered page."""
    if not REQ_OK or not html: return []
    soup = BeautifulSoup(html, "html.parser")
    for t in soup(["script","style","noscript","head"]): t.decompose()
    records, seen = [], set()
    for el in soup.find_all(True):
        if len(list(el.children)) > 15: continue
        nums = _extract_nums(el.get_text(" ",strip=True))
        if _valid_combo(nums):
            key = tuple(nums)
            if key in seen: continue
            seen.add(key)
            row = {"Postcode":pc,"State":state,
                   "Syndicate Number":"","Agent":"","Title":"",
                   "Price per Share":"","Shares":"",
                   "Draw Number":"","Draw Date":"",
                   "Game Type":"Normal" if len(nums)<=7 else f"System {len(nums)}",
                   "Length":len(nums)}
            for j,n in enumerate(nums,1): row[f"n{j}"] = n
            records.append(row)
    return records


async def _pw_scrape_state(state: str, postcodes: list,
                            api_url: str, cookie_str: str,
                            pb, stx) -> list[dict]:
    """Playwright scrape — intercepts JSON API responses per postcode."""
    all_recs = []
    total = len(postcodes)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            user_agent=SCRAPE_HEADERS["User-Agent"],
            locale="en-AU", timezone_id="Australia/Sydney",
            ignore_https_errors=True,
        )
        if cookie_str:
            cookies = [{"name":k.strip(),"value":v.strip(),
                        "domain":".thelott.com","path":"/"}
                       for kv in cookie_str.split(";")
                       for k,_,v in [kv.partition("=")]]
            await ctx.add_cookies(cookies)
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        page = await ctx.new_page()

        for i, pc in enumerate(postcodes):
            pb.progress((i+1)/total,
                text=f"[{state}] {pc} ({i+1}/{total}) — {len(all_recs)} records")
            stx.caption(f"Scraping {state} · postcode {pc}")

            captured = []
            async def on_resp(resp):
                ct = resp.headers.get("content-type","")
                if resp.status == 200 and \
                   ("json" in ct or any(k in resp.url.lower()
                    for k in ["syndic","api","play","search"])):
                    try:
                        body = (await resp.body()).decode("utf-8","ignore")
                        if body.strip()[:1] in ["{","["]:
                            captured.append({"url":resp.url,"body":body})
                    except Exception: pass

            page.on("response", on_resp)
            try:
                url = api_url.format(pc=pc) if "{pc}" in api_url \
                      else f"{api_url}?postcode={pc}"
                await page.goto(url, timeout=25000, wait_until="load")
                await asyncio.sleep(3)
                for sel in ["input[placeholder*='postcode' i]",
                            "input[type='search']"]:
                    try:
                        el = page.locator(sel).first
                        if await el.is_visible(timeout=1000):
                            await el.fill(str(pc))
                            await el.press("Enter")
                            await asyncio.sleep(2)
                            break
                    except Exception: pass
                await asyncio.sleep(2)

                recs = []
                for cap in captured:
                    recs.extend(_parse_json(cap["body"], pc, state))
                if not recs:
                    html = await page.content()
                    recs = _parse_html(html, pc, state)
                all_recs.extend(recs)

            except Exception as ex:
                st.session_state.S["scrape_log"].append(f"{state} {pc}: {ex}")
            finally:
                page.remove_listener("response", on_resp)

            await asyncio.sleep(random.uniform(1.0, 2.5))

        await browser.close()

    return all_recs


def _run_sync(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try: return loop.run_until_complete(coro)
    finally: loop.close()

# ═══════════════════════════════════════════════════════════════════════════════
# 7b. THELOTT CONFIRMED API — two-step outlet → syndicate scraper
# ═══════════════════════════════════════════════════════════════════════════════
import ssl as _ssl
import urllib.request as _urlreq

# Company IDs confirmed by testing (NT=no syndicates, WA=Lotterywest no API)
_STATE_COMPANY = {"NSW": 3, "ACT": 3, "VIC": 1, "TAS": 1, "QLD": 2, "SA": 6}

_TLOTT_HEADERS = {
    "Accept": "application/json",
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Origin":  "https://www.thelott.com",
    "Referer": "https://www.thelott.com/play/syndicates/syndicate-shares",
}

def _ssl_ctx():
    ctx = _ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl.CERT_NONE
    return ctx

def _api_get(url: str):
    """Raw GET with SSL bypass (required on Mac). Returns parsed JSON or None."""
    req = _urlreq.Request(url, headers=_TLOTT_HEADERS)
    try:
        with _urlreq.urlopen(req, timeout=15, context=_ssl_ctx()) as r:
            return json.loads(r.read().decode("utf-8"))
    except Exception:
        return None

def _is_online() -> bool:
    """Quick connectivity check to thelott API."""
    try:
        req = _urlreq.Request(
            "https://api.thelott.com/outlet/outlets?state=NSW&postcode_or_locality=2000",
            headers=_TLOTT_HEADERS)
        with _urlreq.urlopen(req, timeout=8, context=_ssl_ctx()):
            return True
    except Exception:
        return False

def _fetch_outlets(state: str, postcode: int) -> list:
    """Step 1: Get all outlet IDs for a postcode."""
    url = (f"https://api.thelott.com/outlet/outlets"
           f"?state={state}&postcode_or_locality={postcode}")
    data = _api_get(url)
    if not data:
        return []
    outlets = data if isinstance(data, list) else data.get("outlets", [])
    ids = []
    for o in (outlets if isinstance(outlets, list) else []):
        oid = str(o.get("id") or o.get("outletId") or "").strip()
        if oid:
            ids.append(oid)
    return ids

def _fetch_syndicates_batch(company: int, outlet_ids: list) -> list:
    """Step 2: Get syndicates for a comma-separated batch of outlet IDs."""
    if not outlet_ids:
        return []
    url = (f"https://api.thelott.com/syndicates/api/search"
           f"?company={company}&outlets={','.join(outlet_ids)}&limit=100")
    data = _api_get(url)
    if not data:
        return []
    items = (data if isinstance(data, list)
             else data.get("syndicates", data.get("items", data.get("data", []))))
    return items if isinstance(items, list) else []

def _parse_syndicate_row(syn: dict, postcode: int, state: str) -> dict:
    """Map a thelott syndicate API object to the standard CSV schema."""
    def g(*keys):
        for k in keys:
            v = syn.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""
    outlet = syn.get("outlet") or syn.get("store") or {}
    if not isinstance(outlet, dict):
        outlet = {}
    # Collect draw number strings from games/combinations arrays
    draw_numbers = ""
    for gkey in ("games", "combinations", "entries"):
        games = syn.get(gkey, [])
        if isinstance(games, list) and games:
            parts = []
            for gm in games:
                nums = (gm.get("numbers") or gm.get("selections") or []
                        if isinstance(gm, dict) else gm if isinstance(gm, list) else [])
                if nums:
                    parts.append(" ".join(str(n) for n in sorted(nums)))
            if parts:
                draw_numbers = " | ".join(parts)
            break
    return {
        "Postcode":         postcode,
        "State":            state,
        "Syndicate_ID":     g("syndicateId", "id", "syndicateNumber"),
        "Syndicate_Name":   g("syndicateName", "title", "name", "description"),
        "Draw_Date":        g("drawDate", "draw_date", "scheduledDate", "closeDate"),
        "Share_Cost":       g("sharePrice", "price", "costPerShare", "shareCost"),
        "Available_Shares": g("availableShares", "sharesAvailable", "sharesRemaining"),
        "Total_Shares":     g("totalShares", "shareCount", "shares"),
        "Games":            g("gameType", "gameName", "product", "type"),
        "Draw_Numbers":     draw_numbers,
        "Outlet_ID":        g("outletId") or str(outlet.get("id", "")),
        "Outlet_Name":      g("agentName", "storeName") or str(outlet.get("name", "")),
        "Address":          str(outlet.get("address", g("address"))),
        "Suburb":           str(outlet.get("suburb", g("suburb"))),
    }

def _sweep_state_thelott(state: str, postcodes: list,
                          pb, stx) -> pd.DataFrame:
    """
    Full sweep for one state using confirmed two-step thelott API.
    Deduplicates by Syndicate_ID. Returns DataFrame.
    """
    company = _STATE_COMPANY.get(state)
    if company is None:
        return pd.DataFrame()

    seen_ids = {}   # syndicate_id → row dict
    total = len(postcodes)

    for i, pc in enumerate(postcodes):
        pb.progress(
            (i + 1) / total,
            text=f"[{state}] {pc}  ({i+1}/{total}) — {len(seen_ids)} syndicates so far"
        )
        stx.caption(f"🔍 {state} · postcode {pc}")

        outlet_ids = _fetch_outlets(state, pc)
        if not outlet_ids:
            time.sleep(random.uniform(0.1, 0.25))
            continue

        # Batch outlets in groups of 20 (comma-separated — critical format)
        for b in range(0, len(outlet_ids), 20):
            batch = outlet_ids[b:b + 20]
            syns  = _fetch_syndicates_batch(company, batch)
            for syn in syns:
                row = _parse_syndicate_row(syn, pc, state)
                sid = row["Syndicate_ID"] or f"{pc}_{b}_{syns.index(syn)}"
                if sid not in seen_ids:
                    seen_ids[sid] = row
            time.sleep(random.uniform(0.08, 0.20))

        time.sleep(random.uniform(0.20, 0.55))

    stx.empty()
    return pd.DataFrame(list(seen_ids.values())) if seen_ids else pd.DataFrame()

def _data_status() -> list:
    """Return status info for each D_*.csv file in Main_Data/."""
    rows = []
    for state in ["NSW", "VIC", "QLD", "SA", "TAS"]:
        fp = DIRS["Global_Scraper"] / f"D_{state}.csv"
        if fp.exists():
            stat = fp.stat()
            age_h = (datetime.now().timestamp() - stat.st_mtime) / 3600
            try:
                n = sum(1 for _ in open(fp)) - 1
            except Exception:
                n = 0
            rows.append({
                "state": state, "file": fp.name, "rows": n,
                "size_kb": round(stat.st_size / 1024, 1),
                "age_h": round(age_h, 1),
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%d %b %H:%M"),
                "exists": True,
            })
        else:
            rows.append({"state": state, "file": f"D_{state}.csv",
                         "rows": 0, "size_kb": 0, "age_h": 9999,
                         "modified": "—", "exists": False})
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# 7c. PICKS SCRAPER — full syndicate picks via thelott details API
#     (previously thelott_picks_scraper.py — now inlined, no external file needed)
#     Output: w1..wN columns per game line, saved to Global_Scraper/D_{STATE}.csv
# ═══════════════════════════════════════════════════════════════════════════════
import threading as _threading
from concurrent.futures import ThreadPoolExecutor as _ThreadPoolExecutor, as_completed as _as_completed
import csv as _csv

_PICKS_MAX_WORKERS    = 6
_PICKS_PER_REQ_PAUSE  = 0.12
_PICKS_RETRY_TIMES    = 4
_PICKS_RETRY_BACKOFF  = 1.5
_PICKS_THROTTLE_CD    = 20
_picks_print_lock     = _threading.Lock()
_picks_throttle_lock  = _threading.Lock()
_picks_throttle_hits  = 0

_GAME_BY_COMPANY_PRODUCT = {
    (3, 6): "sat", (3, 22): "sat", (3, 7): "mwf", (3, 25): "mwf",
    (3, 8): "oz",  (3, 23): "oz",  (3, 9): "pb",  (3, 24): "pb",
    (3, 10): "sfl", (3, 27): "sfl", (3, 37): "sat", (3, 1): "sat",
    (3, 2): "oz", (3, 3): "pb", (3, 4): "mwf", (3, 5): "sfl",
    (1, 1): "sat", (1, 2): "oz", (1, 3): "pb", (1, 4): "mwf", (1, 5): "sfl",
    (2, 14): "sat", (2, 19): "sat", (2, 15): "oz", (2, 20): "oz",
    (2, 16): "pb",  (2, 21): "pb",  (2, 17): "mwf", (2, 18): "sfl", (2, 3): "pb",
    (6, 22): "sat", (6, 14): "sat", (6, 15): "oz", (6, 16): "pb",
    (6, 17): "mwf", (6, 18): "sfl", (6, 19): "sat",
}

_GAME_KEY_TO_NAME = {
    "pb": "Powerball", "oz": "Oz Lotto", "sat": "Saturday Lotto",
    "mwf": "Monday & Wednesday Lotto", "sfl": "Set for Life",
}

_PICKS_POOL_MAX = {gk: cfg["pool"] for gk, cfg in GAMES_CFG.items()}

_PICKS_DRAW_RANGES = {
    "pb":  (1400, 1900), "oz":  (1500, 1900),
    "sat": (4400, 4900), "mwf": (4400, 4900), "sfl": (1, 99999),
}

_PICKS_STATE_COMPANY = {"NSW": 3, "ACT": 3, "VIC": 1, "TAS": 1, "QLD": 2, "SA": 6}


def _picks_api_get(url: str) -> dict:
    import urllib.request as _ur, ssl as _ssl2
    ctx = _ssl2.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl2.CERT_NONE
    req = _ur.Request(url, headers={
        "Accept": "application/json",
        "Origin": "https://www.thelott.com",
        "Referer": "https://www.thelott.com/",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    })
    with _ur.urlopen(req, timeout=15, context=ctx) as r:
        return json.loads(r.read().decode("utf-8-sig"))


def _picks_outlets(postcode: str, state: str) -> list:
    data = _picks_api_get(
        f"https://api.thelott.com/outlet/outlets"
        f"?state={state}&postcode_or_locality={postcode}")
    if data.get("is_out_of_jurisdiction"):
        return []
    seen, ids = set(), []
    for loc in data.get("locality_outlets", []):
        for o in loc.get("outlets", []):
            oid = o["outlet_id"]
            if oid not in seen:
                seen.add(oid); ids.append(oid)
    return ids


def _picks_syndicate_ids(outlet_ids: list, company: int, batch: int = 20) -> list:
    seen = set()
    for i in range(0, len(outlet_ids), batch):
        csv_ids = ",".join(str(o) for o in outlet_ids[i:i + batch])
        try:
            data = _picks_api_get(
                f"https://api.thelott.com/syndicates/api/search"
                f"?company={company}&outlets={csv_ids}&limit=100")
            for s in data.get("data", []):
                seen.add(s["syndicateId"])
        except Exception as ex:
            print(f"    [warn] search batch: {ex}")
        time.sleep(0.15)
    return sorted(seen)


def _game_for(company, product, draw_number):
    g = _GAME_BY_COMPANY_PRODUCT.get((company, product))
    if g and draw_number:
        lo, hi = _PICKS_DRAW_RANGES.get(g, (1, 99999))
        if not (lo <= int(draw_number) <= hi):
            return g, f"CHECK(draw {draw_number} out of {g} range)"
    return (g or f"unknown_c{company}_p{product}"), ""


def _resolve_game(company, product, draw_number, selections):
    g, note = _game_for(company, product, draw_number)
    mx = max(int(n) for n in selections) if selections else 0
    pool = _PICKS_POOL_MAX.get(g)
    if pool is not None and mx <= pool and g in _PICKS_POOL_MAX:
        return g, note
    dn = int(draw_number) if str(draw_number).isdigit() else 0
    inferred = None
    if 4400 <= dn <= 4900:
        inferred = "mwf" if g == "mwf" else "sat"
    elif 1400 <= dn <= 1900:
        inferred = "pb" if mx <= 35 else "oz"
    elif mx >= 46:
        inferred = "oz"
    elif mx == 45:
        inferred = "mwf" if g == "mwf" else "sat"
    elif mx <= 44 and g == "sfl":
        inferred = "sfl"
    else:
        for gk in ("pb", "sfl", "sat", "mwf", "oz"):
            if _PICKS_POOL_MAX.get(gk, 0) >= mx:
                inferred = gk; break
    if inferred and _PICKS_POOL_MAX.get(inferred, 0) >= mx:
        tag = "retagged" if g in _PICKS_POOL_MAX else "inferred"
        return inferred, f"{tag} {g}->{inferred} (max {mx}, draw {dn})"
    return (g or "unknown"), f"UNRESOLVED (max {mx})"


def _picks_fetch_details(syndicate_id: int, company: int) -> list:
    try:
        b = _picks_api_get(
            f"https://api.thelott.com/syndicates/api/details"
            f"?syndicateId={syndicate_id}&companyId={company}")
    except Exception as ex:
        print(f"    [warn] details {syndicate_id}: {ex}"); return []
    name  = (b.get("syndicateName") or "").strip()
    cost  = b.get("shareCost", "")
    avail = b.get("availableShares", "")
    total = b.get("totalShares", "")
    outlets = b.get("outlets", [])
    outlet_id = ""
    if isinstance(outlets, list) and outlets:
        outlet_id = (outlets[0].get("outletId") if isinstance(outlets[0], dict) else "") or ""
    rows = []
    for bet in b.get("syndicateBets", []):
        product    = bet.get("product")
        draws      = bet.get("draws", [])
        draw_no    = draws[0].get("drawNumber") if draws else ""
        draw_dt    = (draws[0].get("drawDate", "")[:10]) if draws else ""
        entries    = bet.get("entries", [])
        entry_type = entries[0].get("entryType", "") if entries else ""
        for g in bet.get("games", []):
            sels = g.get("selections", []) or []
            if not sels:
                continue
            game, warn = _resolve_game(company, product, draw_no, sels)
            mx = max(int(n) for n in sels)
            if mx > max(_PICKS_POOL_MAX.values()) or mx < 1:
                continue
            pb_raw = g.get("powerball", "")
            if game != "pb":
                pb_val = ""
            elif g.get("powerHit") and (pb_raw in (0, "0", None, "")):
                pb_val = "PH"
            elif pb_raw in (0, "0", None, ""):
                pb_val = ""
            else:
                pb_val = pb_raw
            row = {
                "Syndicate_ID": syndicate_id, "Syndicate_Name": name,
                "Game": game, "Games": _GAME_KEY_TO_NAME.get(game, game),
                "Game_Check": warn, "Product": product, "CompanyId": company,
                "Draw_Number": draw_no, "Draw_Date": draw_dt,
                "Entry_Type": entry_type, "System_Number": g.get("systemNumber", ""),
                "PowerHit": g.get("powerHit", ""), "Share_Cost": cost,
                "Available_Shares": avail, "Total_Shares": total,
                "Outlet_ID": outlet_id, "PB": pb_val,
            }
            for i, n in enumerate(sels, 1):
                row[f"w{i}"] = n
            rows.append(row)
    return rows


def _picks_fetch_retry(syndicate_id: int, company: int) -> list:
    global _picks_throttle_hits
    import urllib.error as _urlerr2
    for attempt in range(1, _PICKS_RETRY_TIMES + 1):
        try:
            rows = _picks_fetch_details(syndicate_id, company)
            time.sleep(_PICKS_PER_REQ_PAUSE)
            return rows
        except _urlerr2.HTTPError as ex:
            if ex.code == 403:
                with _picks_throttle_lock:
                    _picks_throttle_hits += 1
                    first = (_picks_throttle_hits == 1)
                if first:
                    with _picks_print_lock:
                        print("    [throttle] 403 rate limit — cooling down and retrying")
                if attempt == _PICKS_RETRY_TIMES:
                    return []
                time.sleep(_PICKS_THROTTLE_CD * attempt)
            else:
                if attempt == _PICKS_RETRY_TIMES:
                    return []
                time.sleep(_PICKS_RETRY_BACKOFF * attempt)
        except Exception:
            if attempt == _PICKS_RETRY_TIMES:
                return []
            time.sleep(_PICKS_RETRY_BACKOFF * attempt)
    return []


def _picks_collect_ids(state: str, company: int) -> dict:
    """Phase 1 (sequential): collect unique syndicate IDs across all postcodes."""
    pcs = [str(p) for p in STATE_POSTCODES.get(state, [])]
    seen: dict = {}
    for i, pc in enumerate(pcs):
        try:
            outlets = _picks_outlets(pc, state)
        except Exception as ex:
            print(f"  [{i+1}/{len(pcs)}] {pc} outlet error: {ex}"); continue
        if not outlets:
            continue
        ids = _picks_syndicate_ids(outlets, company)
        new = sum(1 for sid in ids if sid not in seen)
        for sid in ids:
            if sid not in seen:
                seen[sid] = pc
        print(f"  [{i+1}/{len(pcs)}] {pc}: {len(ids)} here, +{new} new (unique: {len(seen)})")
        time.sleep(0.1)
    return seen


def _picks_dedup(rows: list) -> list:
    seen: set = set()
    out = []
    for r in rows:
        wkey = tuple(r.get(f"w{i}") for i in range(1, 40) if r.get(f"w{i}") is not None)
        key = (r.get("Syndicate_ID"), r.get("Draw_Number"), r.get("Game"), wkey, r.get("PB"))
        if key not in seen:
            seen.add(key); out.append(r)
    return out


def _picks_columns(rows: list) -> list:
    preferred = ["Syndicate_ID", "Syndicate_Name", "Game", "Games", "Game_Check",
                 "Product", "CompanyId", "Draw_Number", "Draw_Date", "Entry_Type",
                 "System_Number", "PowerHit", "Share_Cost", "Available_Shares",
                 "Total_Shares", "Outlet_ID", "Postcode", "State", "PB"]
    keys: set = set()
    for r in rows:
        keys.update(r.keys())
    wcols = sorted((k for k in keys if len(k) > 1 and k[0] == "w" and k[1:].isdigit()),
                   key=lambda x: int(x[1:]))
    ordered = [c for c in preferred if c in keys]
    rest = sorted(k for k in keys if k not in ordered
                  and not (len(k) > 1 and k[0] == "w" and k[1:].isdigit()))
    return ordered + rest + wcols


def sweep_state_picks(state: str, workers: int = _PICKS_MAX_WORKERS,
                      save_path: "Path | None" = None) -> list:
    """Two-phase dedup-first sweep — returns rows with w1..wN actual picks.
    Saves to Global_Scraper/D_{state}.csv (or override with save_path).

    Replaces: python3 thelott_picks_scraper.py sweep {state}
    Now call:  from masterapp import sweep_state_picks; sweep_state_picks('NSW')
    """
    global _picks_throttle_hits
    _picks_throttle_hits = 0
    company = _PICKS_STATE_COMPANY.get(state)
    if company is None:
        print(f"  no company id for {state}"); return []

    print(f"--- Phase 1: collecting unique syndicate IDs for {state} ---")
    t0 = time.time()
    id_map = _picks_collect_ids(state, company)
    ids = list(id_map.keys())
    print(f"--- Phase 1 done: {len(ids)} unique syndicates in {time.time()-t0:.0f}s ---")
    if not ids:
        print("  no syndicates found."); return []

    print(f"--- Phase 2: fetching details ({workers} workers) ---")
    t1 = time.time()
    all_rows: list = []
    done = 0
    with _ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_picks_fetch_retry, sid, company): sid for sid in ids}
        for fut in _as_completed(futures):
            sid = futures[fut]
            recs = fut.result() or []
            pc = id_map.get(sid, "")
            for r in recs:
                r["Postcode"] = pc; r["State"] = state
            all_rows.extend(recs)
            done += 1
            if done % 50 == 0 or done == len(ids):
                with _picks_print_lock:
                    print(f"    details {done}/{len(ids)} ({len(all_rows)} rows so far)")
    print(f"--- Phase 2 done in {time.time()-t1:.0f}s ---")

    before = len(all_rows)
    all_rows = _picks_dedup(all_rows)
    if not all_rows:
        print("  (no rows)"); return []
    out_path = save_path or (DIRS["Global_Scraper"] / f"D_{state}.csv")
    cols = _picks_columns(all_rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = _csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader(); w.writerows(all_rows)
    print(f"  saved {len(all_rows)} unique rows (deduped from {before}) -> {out_path}")
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# 8. CVI MATRIX — transposition + variable slicing
# ═══════════════════════════════════════════════════════════════════════════════
def build_w_matrix(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Sort rows longest→shortest, transpose each into a w-column, pad with None.

    The source 'number' columns of a D file are the syndicate picks, labelled
    **w1…wN** (project convention: syndicate = w, main data = n). Older/main-data
    files may instead use n1…nN. Accept EITHER prefix as the input columns so D
    (w-columns) loads correctly. (Previously this only looked for n1…nN, so D —
    which is w-columns — produced an empty matrix and the "No n-columns found"
    error.) The OUTPUT is always w-columns regardless of input naming.
    """
    src_cols = sorted(
        [c for c in df_raw.columns if re.match(r'^[wn]\d+$', c, re.I)],
        key=lambda x: int(x[1:]))
    if not src_cols:
        return pd.DataFrame()
    rows_as_lists = []
    for _, row in df_raw.iterrows():
        nums = [int(row[c]) for c in src_cols
                if pd.notna(row[c]) and str(row[c]).strip().lstrip('-').isdigit()]
        if nums: rows_as_lists.append(sorted(nums))
    if not rows_as_lists: return pd.DataFrame()
    rows_as_lists.sort(key=len, reverse=True)
    max_len = max(len(r) for r in rows_as_lists)
    w_dict = {f"w{i+1}": (r + [None]*(max_len-len(r)))
              for i, r in enumerate(rows_as_lists)}
    return pd.DataFrame(w_dict)


def d_to_w_only(df: pd.DataFrame) -> pd.DataFrame:
    """Return a variable as CLEAN w-columns, transposing ONLY raw syndicate data.

    Two very different shapes arrive here:
      • RAW D  — one ROW per syndicate, picks in w1…wN, plus syndicate metadata
                 (Syndicate_ID, Game, Games, PB, Draw_Number, Outlet_ID, …).
                 This must be TRANSPOSED → one w-column per syndicate.
      • ALREADY-WIDE (B, R, Ep, Sp, So, or a built W-matrix) — already one COLUMN
                 per w-set. These must be returned AS-IS (never transposed); we only
                 drop any stray non-w label column.

    The previous heuristic transposed whenever ANY non-w column was present, which
    flattened B (720 w-columns) into one long column. The real signal that a frame
    is raw D is the presence of SYNDICATE METADATA — so we key on that instead.
    """
    if df is None or df.empty:
        return df
    wcols = [c for c in df.columns if re.match(r'^w\d+$', str(c), re.I)]
    ncols = [c for c in df.columns if re.match(r'^n\d+$', str(c), re.I)]
    D_META = {"syndicate_id", "syndicate_name", "game", "games", "pb",
              "draw_number", "draw_numbers", "outlet_id", "outlet_name",
              "postcode", "state", "share_cost", "available_shares",
              "total_shares", "address", "suburb"}
    has_d_meta = any(str(c).strip().lower() in D_META for c in df.columns)

    # RAW D (rows = syndicates): transpose to one w-column per syndicate.
    if has_d_meta:
        return build_w_matrix(df)
    # ALREADY-WIDE variable (B/R/Ep/Sp/So or built matrix): keep w-columns as-is.
    if wcols:
        return df[wcols]
    # No w-columns, no D metadata, but n-columns present (main-data style): transpose.
    if ncols:
        return build_w_matrix(df)
    # Nothing recognisable — return unchanged rather than risk a bad transpose.
    return df


def slice_variables(w_mat: pd.DataFrame) -> dict:
    """Auto-slice Ep (top 8), Sp (top 4 lanes a-d), So (union of Sp lanes)."""
    w_keys = sorted([c for c in w_mat.columns if re.match(r'^w\d+$',c)],
                    key=lambda x: int(x[1:]))
    def col_ints(k):
        result = []
        for v in w_mat[k].dropna():
            try:
                n = float(str(v))
                if not pd.isna(n) and n >= 1:
                    result.append(int(round(n)))
            except (ValueError, TypeError):
                pass
        return result

    ep_keys = w_keys[:8]
    sp_keys = w_keys[:4]
    ep_raw = {k: col_ints(k) for k in ep_keys}
    sp_raw = {"lane_a": col_ints(sp_keys[0]) if len(sp_keys)>0 else [],
              "lane_b": col_ints(sp_keys[1]) if len(sp_keys)>1 else [],
              "lane_c": col_ints(sp_keys[2]) if len(sp_keys)>2 else [],
              "lane_d": col_ints(sp_keys[3]) if len(sp_keys)>3 else []}

    def to_df(d):
        ml = max((len(v) for v in d.values()), default=0)
        return pd.DataFrame({k: v+[None]*(ml-len(v)) for k,v in d.items()})

    so_union = sorted(set(sum(sp_raw.values(),[])))
    return {"Ep": to_df(ep_raw), "Sp": to_df(sp_raw),
            "So": pd.DataFrame({"So_union": so_union})}


def fetch_since_last(url: str, pool: int) -> dict | None:
    """Best-effort scrape of a lottolyzer number-frequencies page.

    Returns {number: since_last} for numbers 1..pool, or None if the page
    can't be parsed (in which case the caller falls back to manual upload).
    Runs on the user's machine (this app has network); kept defensive because
    lottolyzer's markup can change.
    """
    try:
        import urllib.request
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                         "AppleWebKit/537.36 (KHTML, like Gecko) "
                                         "Chrome/124.0 Safari/537.36"})
        html = urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "ignore")
        tables = pd.read_html(html)            # needs lxml/bs4 (usually present)
    except Exception:
        return None
    # Find the table that has a Number column and a "Since Last"-type column.
    for t in tables:
        cols = {str(c).strip().lower(): c for c in t.columns}
        num_c = next((cols[k] for k in cols
                      if k in ("number", "no", "no.", "ball") or "number" in k), None)
        sl_c = next((cols[k] for k in cols
                     if "since" in k or "games since" in k or "last seen" in k), None)
        if not (num_c and sl_c):
            continue
        d = {}
        for _, row in t.iterrows():
            try:
                n = int(float(str(row[num_c]).strip()))
                s = int(float(str(row[sl_c]).strip()))
                if 1 <= n <= pool:
                    d[n] = s
            except (ValueError, TypeError):
                pass
        if d:
            return d
    return None


def save_since_last(sl_dict: dict, game_key: str, label: str, pool: int,
                    url: str, sl_file) -> dict:
    """Persist a since_last dict to since_last.json (shared by fetch + upload)."""
    all_wt = sorted(sl_dict.keys(), key=lambda n: (sl_dict[n], n))
    data = {
        "since_last_dict": {str(k): v for k, v in sl_dict.items()},
        "all_wt": all_wt,
        "to_keep": list(all_wt),
        "game": game_key,
        "game_name": label,
        "pool_size": pool,
        "scraped_at": datetime.now().isoformat(),
        "url": url,
    }
    sl_file.parent.mkdir(parents=True, exist_ok=True)
    sl_file.write_text(json.dumps(data, indent=2))
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# 8b. AUTO-WIRE GENERATORS — runs Sp / So (and Ep if B+R ready) from D
# ═══════════════════════════════════════════════════════════════════════════════
def _auto_wire_generators(gdirs: dict, gkey: str):
    """Auto-run Sp, So (and Ep when B+R are available) immediately after D loads.

    Uses prepare_d_input_sets to peel the 4/8 longest D rows (w1,w2,w3,w4…) as
    column-oriented sets, then feeds them to the generators. Results are stored
    in S and written to disk so subsequent tabs show them instantly.
    """
    d_df = S.get("D", pd.DataFrame())
    if d_df is None or d_df.empty:
        return

    auto_status = st.empty()
    msgs = []

    # ── Sp (Splits / task1b) ────────────────────────────────────────────────
    try:
        sp_input = prepare_d_input_sets(d_df, 4)  # columns: w1,w2,w3,w4
        if not sp_input.empty:
            sp_df = generate_splits(sp_input)
            if not sp_df.empty:
                S["Sp"] = sp_df
                sp_path = gdirs["Splits"] / f"Sp_{gkey}.csv"
                sp_df.to_csv(sp_path, index=False)
                msgs.append(f"Sp ({sp_df.shape[1]} cols)")
    except Exception as _sp_ex:
        msgs.append(f"Sp error: {_sp_ex}")

    # ── So (SplitsCombi / automation_vba) ──────────────────────────────────
    try:
        so_input = prepare_d_input_sets(d_df, 4)  # columns: w1,w2,w3,w4
        if not so_input.empty:
            so_df = generate_splits_combi(so_input)
            if not so_df.empty:
                S["So"] = so_df
                so_path = gdirs["Splits_Combi"] / f"So_{gkey}.csv"
                so_df.to_csv(so_path, index=False)
                msgs.append(f"So ({so_df.shape[1]} cols)")
    except Exception as _so_ex:
        msgs.append(f"So error: {_so_ex}")

    # ── Ep (ExcelPro) — requires R's wt list ───────────────────────────────
    b_df = S.get("B", pd.DataFrame())
    r_wt_df = S.get("_R_wt", pd.DataFrame())
    if not b_df.empty and not r_wt_df.empty:
        try:
            ep_objs = prepare_ep_objects(d_df, mode="pairs")  # 8 longest rows → 4 pairs
            wt_list = r_wt_df["wt"].dropna().tolist() if "wt" in r_wt_df.columns else []
            if not wt_list:
                # Fall back to all_wt column or any numeric column
                wt_list = r_wt_df.iloc[:, 0].dropna().tolist()
            if ep_objs and wt_list:
                ep_df = generate_excelpro(ep_objs, wt_list)
                if not ep_df.empty:
                    S["Ep"] = ep_df
                    ep_path = gdirs["ExcelPro"] / f"Ep_{gkey}.csv"
                    ep_df.to_csv(ep_path, index=False)
                    msgs.append(f"Ep ({ep_df.shape[1]} cols)")
        except Exception as _ep_ex:
            msgs.append(f"Ep error: {_ep_ex}")

    if msgs:
        auto_status.markdown(
            f'<div class="ok">⚡ Auto-generated: {" · ".join(msgs)}</div>',
            unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# 9. UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def am_toggle(section: str) -> str:
    """Global Auto/Manual toggle — ONE state shared across ALL pages.
    Clicking AUTO or MANUAL anywhere sets it site-wide and reruns immediately,
    so no page can silently break the auto chain.
    """
    # ── Read / initialise global state ────────────────────────────────────
    mode = S["auto"].get("_global", "Manual")

    c1, c2, _ = st.columns([1, 1, 8])
    with c1:
        if st.button("▶ AUTO", key=f"_a_{section}",
                     type="primary" if mode == "Auto" else "secondary",
                     use_container_width=True):
            S["auto"]["_global"] = "Auto"
            st.rerun()
    with c2:
        if st.button("✋ MANUAL", key=f"_m_{section}",
                     type="primary" if mode == "Manual" else "secondary",
                     use_container_width=True):
            S["auto"]["_global"] = "Manual"
            st.rerun()
    return mode


def edit_var(var_key: str, folder: Path, filename: str, label: str, note=""):
    """Editable variable input table — reads from and saves to real folder."""
    st.markdown(f"**{label}**")
    if note: st.caption(note)
    p = folder / filename
    if var_key not in S or (isinstance(S.get(var_key), pd.DataFrame) and S[var_key].empty and p.exists()):
        S[var_key] = _load_file(p)
    df = S.get(var_key, pd.DataFrame())
    edited = st.data_editor(df, key=f"ed_{var_key}",
                            use_container_width=True, num_rows="dynamic", height=280)
    for col in edited.columns:
        edited[col] = pd.to_numeric(edited[col], errors="coerce")
    S[var_key] = edited
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {label} CSV", to_csv_bytes(edited),
                           f"{label}.csv","text/csv",key=f"dl_{var_key}")
    with c2:
        if st.button(f"💾 Save to {filename}", key=f"sv_{var_key}"):
            edited.to_csv(p, index=False); st.success(f"Saved to {p}")


def set_container_status(db: str, val: str):
    for i in range(len(S["container_status"][db])):
        S["container_status"][db].at[i,"Status"] = val

# ═══════════════════════════════════════════════════════════════════════════════
# 10. PAGE CONFIG & CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Syndicate System", layout="wide",
                   initial_sidebar_state="collapsed")
st.markdown("""
<style>
  .block-container{padding-top:.4rem}
  .sec-hdr{font-size:.9rem;font-weight:900;color:#fff;padding:5px 14px;
            border-radius:5px;margin:4px 0 6px 0;display:inline-block}
  .hdr-blue{background:#1d3557} .hdr-green{background:#2d6a4f}
  .hdr-orange{background:#e07c00} .hdr-purple{background:#6d3b8e}
  .hdr-red{background:#c1121f} .hdr-teal{background:#0077b6}
  .hdr-brown{background:#6b4226}
  /* Status boxes — fixed foreground colour so they show on dark AND light themes */
  .note{background:#fff3cd;border-left:4px solid #e07c00;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c3d00 !important}
  .info{background:#cce5ff;border-left:4px solid #0077b6;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#003a5c !important}
  .warn{background:#ffd6d6;border-left:4px solid #c1121f;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#5c0000 !important}
  .ok  {background:#c3e6cb;border-left:4px solid #28a745;
         padding:8px 12px;border-radius:4px;font-size:.79rem;margin:4px 0 8px 0;
         color:#0a3d1f !important}
  /* Make all custom div text always visible */
  .note *, .info *, .warn *, .ok * { color: inherit !important; }

  /* ── Main navigation tabs (st.radio) — high visibility ───────────────── */
  /* Each tab rendered as a clear bordered pill so faint tabs are easy to read */
  div[role="radiogroup"] > label{
      background:#1b2436;
      border:1.5px solid #3a4a66;
      border-radius:8px;
      padding:7px 14px !important;
      margin:3px 6px 3px 0 !important;
      transition:all .12s ease;
  }
  div[role="radiogroup"] > label:hover{
      border-color:#5b8def;
      background:#243049;
  }
  /* Tab label text — brighter and bolder than default */
  div[role="radiogroup"] > label p{
      color:#dfe7f5 !important;
      font-weight:700 !important;
      font-size:.92rem !important;
  }
  /* The SELECTED tab — strong amber highlight so the active page is obvious */
  div[role="radiogroup"] > label:has(input:checked){
      background:#f4a000 !important;
      border-color:#ffd166 !important;
      box-shadow:0 0 0 2px rgba(244,160,0,.35);
  }
  div[role="radiogroup"] > label:has(input:checked) p{
      color:#1a1300 !important;
      font-weight:900 !important;
  }

  /* ── Game selector PRIMARY buttons — make selected game very obvious ── */
  /* Secondary (inactive) game buttons — cool dark background, clear text */
  div[data-testid="stHorizontalBlock"] button[kind="secondary"],
  div[data-testid="column"] button[kind="secondary"]{
      background:#1f2d45 !important;
      border:1.5px solid #3a4e6e !important;
      color:#b8cce4 !important;
      font-weight:600 !important;
  }
  /* Primary (active) game button — bright amber, impossible to miss */
  div[data-testid="stHorizontalBlock"] button[kind="primary"],
  div[data-testid="column"] button[kind="primary"]{
      background:linear-gradient(135deg,#f4a000,#ffcc44) !important;
      border:2px solid #ffd166 !important;
      color:#1a1300 !important;
      font-weight:900 !important;
      box-shadow:0 0 10px rgba(244,160,0,.5) !important;
  }

  /* ── Variable Input sub-tabs — make Streamlit tab bar highly visible ── */
  div[data-testid="stTabs"] button[role="tab"]{
      font-weight:700 !important;
      font-size:.88rem !important;
      color:#c5d6f0 !important;
      border-bottom:2px solid transparent !important;
      padding:8px 16px !important;
  }
  div[data-testid="stTabs"] button[role="tab"][aria-selected="true"]{
      color:#ffd166 !important;
      border-bottom:2px solid #f4a000 !important;
      font-weight:900 !important;
  }
  div[data-testid="stTabs"] button[role="tab"]:hover{
      color:#ffffff !important;
      background:#243049 !important;
  }
</style>
""", unsafe_allow_html=True)

_init_state()
S = st.session_state.S

# ═══════════════════════════════════════════════════════════════════════════════
# 11. NAVIGATION
# ═══════════════════════════════════════════════════════════════════════════════
st.title("🎰 Syndicate System")
st.caption(f"Root: `{ROOT}`")

# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL SCRAPER — collapsible, above game selector, standalone
# ═══════════════════════════════════════════════════════════════════════════════
with st.expander("🕷️ Global Scraper — sweep all states, all games", expanded=False):
        st.markdown('<span class="sec-hdr hdr-blue">🕷️ Scraper — Global D Variable Sweep (all games)</span>',
                    unsafe_allow_html=True)
        st.markdown("""
        <div class="info">
        <b>This page is standalone — not game-specific.</b>
        It sweeps thelott.com for ALL syndicate games at once and saves raw state
        files to <code>Global_Scraper/</code>. After sweeping, click
        <b>🔀 Promote All + Split by Game</b> to route each row to its correct
        game folder. The game selector above does not affect this page.
        </div>
        """, unsafe_allow_html=True)
        am_toggle("scraper")

        # ── Connectivity notice (informational only — buttons always active) ───
        online = _is_online()
        if not online:
            st.markdown(
                '<div class="warn">⚠️ <b>Streamlit cannot reach thelott API directly</b> '
                '(Mac SSL restriction in browser process). '
                'Use the <b>terminal commands</b> in Section 7 below to run sweeps — '
                'they work perfectly from terminal. Cached data is shown here.</div>',
                unsafe_allow_html=True)

        # ── Info ───────────────────────────────────────────────────────────────
        st.markdown("""
        <div class="info">
        <b>Confirmed API</b> — two-step: outlets per postcode → syndicates per outlet batch.<br>
        No auth, no cookies, no Playwright required. Saves to <code>Global_Scraper/D_{STATE}.csv</code>.<br>
        Coverage: <b>NSW · VIC · QLD · SA · TAS</b> (ACT inside NSW data · NT=0 syndicates · WA=Lotterywest, in-store only).
        </div>
        """, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — DATA STATUS DASHBOARD
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📊 Data Status")

        status_rows = _data_status()
        col_labels = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
        for label in ["State", "File", "Rows", "Size", "Age", "Last Updated", "Freshness"]:
            col_labels[["State","File","Rows","Size","Age","Last Updated","Freshness"]
                        .index(label)].markdown(f"**{label}**")

        for row in status_rows:
            cols = st.columns([1.2, 2.5, 1.2, 1.2, 1.5, 2.0, 1.5])
            cols[0].write(row["state"])
            cols[1].write(f"`{row['file']}`")
            cols[2].write(f"{row['rows']:,}" if row["exists"] else "—")
            cols[3].write(f"{row['size_kb']} KB" if row["exists"] else "—")
            cols[4].write(f"{row['age_h']}h" if row["exists"] else "—")
            cols[5].write(row["modified"])
            if not row["exists"]:
                cols[6].markdown('<span style="color:#c1121f">⛔ Missing</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 6:
                cols[6].markdown('<span style="color:#28a745">🟢 Fresh</span>',
                                 unsafe_allow_html=True)
            elif row["age_h"] < 48:
                cols[6].markdown('<span style="color:#e07c00">🟡 Stale</span>',
                                 unsafe_allow_html=True)
            else:
                cols[6].markdown('<span style="color:#c1121f">🔴 Old</span>',
                                 unsafe_allow_html=True)

        total_rows = sum(r["rows"] for r in status_rows if r["exists"])
        freshest   = min((r["age_h"] for r in status_rows if r["exists"]), default=9999)
        oldest     = max((r["age_h"] for r in status_rows if r["exists"]), default=0)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total syndicates cached", f"{total_rows:,}")
        m2.metric("States with data", sum(1 for r in status_rows if r["exists"]))
        m3.metric("Freshest file", f"{freshest}h ago" if freshest < 9999 else "—")
        m4.metric("Oldest file",   f"{oldest}h ago"   if oldest > 0     else "—")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — SWEEP CONTROLS
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 🔄 Refresh Data")

        SWEEP_STATES  = ["NSW", "VIC", "QLD", "SA", "TAS"]
        SWEEP_POSTCODES = {
            "NSW": list(range(2000, 3000)),
            "VIC": list(range(3000, 4000)),
            "QLD": list(range(4000, 5000)),
            "SA":  list(range(5000, 5600)),
            "TAS": list(range(7000, 7800)),
        }

        c1, c2, c3 = st.columns([2, 1, 2])
        with c1:
            max_pc = st.number_input(
                "Max postcodes per state (0 = all)",
                min_value=0, max_value=9999, value=0,
                help="Set to a small number (e.g. 50) for a quick test run. 0 = full sweep."
            )
        with c2:
            st.write("")
            smart_skip = st.checkbox("Smart skip", value=True,
                                     help="Skip states where data is under 6 hours old")
        with c3:
            est_pcs = sum(min(max_pc or len(v), len(v))
                          for v in SWEEP_POSTCODES.values())
            est_min = max(round(est_pcs * 0.8 / 60), 1)
            st.metric("Est. full sweep time", f"~{est_min} min")

        # ── Per-state buttons ─────────────────────────────────────────────────
        st.markdown("**Sweep individual state:**")
        btn_cols = st.columns(len(SWEEP_STATES))
        triggered_state = None
        for i, s in enumerate(SWEEP_STATES):
            with btn_cols[i]:
                row = next(r for r in status_rows if r["state"] == s)
                age_label = f"{row['age_h']}h" if row["exists"] else "no data"
                if st.button(f"{s}\n({age_label})", key=f"sw_{s}",
                             use_container_width=True):
                    triggered_state = s

        # ── Sweep all button ──────────────────────────────────────────────────
        sweep_all = st.button(
            "🚀 SWEEP ALL STATES",
            type="primary", use_container_width=True,
        )

        def _run_sweep(states_to_sweep: list):
            """Execute sweep for given list of states, saving each immediately."""
            for state in states_to_sweep:
                row = next(r for r in status_rows if r["state"] == state)
                if smart_skip and row["exists"] and row["age_h"] < 6:
                    st.markdown(
                        f'<div class="info">⏭️ <b>{state}</b> skipped — data is only '
                        f'{row["age_h"]}h old (smart skip on).</div>',
                        unsafe_allow_html=True)
                    continue

                pcs = list(SWEEP_POSTCODES[state])
                if max_pc:
                    pcs = pcs[:max_pc]

                out_path = DIRS["Global_Scraper"] / f"D_{state}.csv"
                st.markdown(f"**▶ Sweeping {state} — {len(pcs):,} postcodes…**")
                pb  = st.progress(0)
                stx = st.empty()

                df_result = _sweep_state_thelott(state, pcs, pb, stx)
                pb.empty()

                if not df_result.empty:
                    df_result.to_csv(out_path, index=False)
                    st.markdown(
                        f'<div class="ok">✅ <b>{state}</b>: {len(df_result):,} syndicates '
                        f'→ <code>{out_path.name}</code></div>',
                        unsafe_allow_html=True)
                    show_paginated_df(df_result, key=f"scrape_result_{state}", use_container_width=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  {state}  {len(df_result):,} rows")
                else:
                    st.markdown(
                        f'<div class="warn">⚠️ <b>{state}</b>: 0 syndicates found. '
                        f'Check connectivity or try again later.</div>',
                        unsafe_allow_html=True)
                    S["scrape_log"].append(
                        f"{datetime.now():%Y-%m-%d %H:%M}  {state}  0 rows — possible block")

        if triggered_state:
            _run_sweep([triggered_state])
        elif sweep_all:
            _run_sweep(SWEEP_STATES)

        # ── Health warning ────────────────────────────────────────────────────
        zero_states = [r["state"] for r in status_rows if r["exists"] and r["rows"] < 10]
        if zero_states:
            st.markdown(
                f'<div class="warn">⚠️ <b>Health check:</b> {", ".join(zero_states)} '
                f'returned very few records. The API may be throttling — '
                f'wait and retry, or check logs below.</div>',
                unsafe_allow_html=True)

        # ── Scrape log ────────────────────────────────────────────────────────
        if S["scrape_log"]:
            with st.expander("🪵 Sweep log"):
                for line in S["scrape_log"][-60:]:
                    st.text(line)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — QUICK API TEST (single postcode)
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("🔍 Quick API test — single postcode"):
            tc1, tc2, tc3 = st.columns([1.5, 1.5, 1])
            with tc1:
                test_state = st.selectbox("State", SWEEP_STATES, key="test_state")
            with tc2:
                defaults = {"NSW": 2000, "VIC": 3000, "QLD": 4000, "SA": 5000, "TAS": 7000}
                test_pc_val = st.number_input("Postcode", value=defaults.get(test_state, 2000),
                                              key="test_pc")
            with tc3:
                st.write(""); st.write("")
                do_test = st.button("▶ Test", key="do_test")

            if do_test:
                with st.spinner(f"Fetching outlets for {test_state} / {test_pc_val}…"):
                    outlets = _fetch_outlets(test_state, int(test_pc_val))
                st.write(f"**Outlets found:** {len(outlets)}  —  IDs: `{', '.join(outlets[:8])}`{'…' if len(outlets)>8 else ''}")
                if outlets:
                    with st.spinner("Fetching syndicates…"):
                        company = _STATE_COMPANY[test_state]
                        syns = _fetch_syndicates_batch(company, outlets[:20])
                    if syns:
                        rows = [_parse_syndicate_row(s, int(test_pc_val), test_state)
                                for s in syns]
                        st.markdown(
                            f'<div class="ok">✅ {len(rows)} syndicates returned.</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(pd.DataFrame(rows), key="quick_api_rows", use_container_width=True)
                    else:
                        st.markdown(
                            '<div class="warn">0 syndicates for these outlets.</div>',
                            unsafe_allow_html=True)
                else:
                    st.markdown(
                        '<div class="warn">No outlets found for this postcode.</div>',
                        unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — PROMOTE TO DIRECT/
        # ══════════════════════════════════════════════════════════════════════


with st.expander("🗂️ Game Breakdown — promote & split by game", expanded=False):
        # SECTION 4 — PROMOTE ALL + SPLIT BY GAME
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### ✅ Promote All → Split by Game")

        # ── Re-split warning if any game folders are empty ────────────────────
        empty_games = []
        for gk in GAME_KEYS:
            gd = game_dirs(gk)["Direct"]
            if not list(gd.glob("D_*.csv")):
                empty_games.append(f"{GAMES_CFG[gk]['emoji']} {GAMES_CFG[gk]['label']}")
        if empty_games:
            st.markdown(
                f'<div class="warn">⚠️ These games have no D files yet: '
                f'<b>{", ".join(empty_games)}</b><br>'
                f'Run <b>🔀 Promote All + Split by Game</b> to populate them. '
                f'Key fix: "Monday &amp; Wednesday Lotto" is now correctly mapped to MWF.</div>',
                unsafe_allow_html=True)

        st.markdown("""
        <div class="info">
        One click: reads all D_*.csv files from <b>Main_Data/</b>, splits every row
        by its <b>Games</b> column, saves game-specific files into each game's
        <code>Games_Breakdown/</code> folder (one per game).<br>
        Multi-game syndicates (e.g. "Oz Lotto | Powerball") → a copy in <b>each</b>
        matching game folder.<br>
        Mapped: TattsLotto/Gold Lotto/X Lotto → SAT &nbsp;·&nbsp;
        Monday &amp; Wednesday Lotto → MWF &nbsp;·&nbsp;
        Super 66/Lucky Lotteries → skipped.
        </div>
        """, unsafe_allow_html=True)

        raw_d_files = sorted(DIRS["Global_Scraper"].glob("D_*.csv"))

        if raw_d_files:
            # Show quick game breakdown from first available file
            with st.expander("📊 Preview game counts before splitting"):
                for src in raw_d_files[:2]:  # show first 2 to avoid slowness
                    try:
                        df_prev = pd.read_csv(src, usecols=["Games"])
                        gc = df_prev["Games"].value_counts().reset_index()
                        gc.columns = ["Games value", "Rows"]
                        st.write(f"**{src.name}:**")
                        show_paginated_df(gc, key=f"game_breakdown_prev_{src.name}", use_container_width=True, height=200)
                    except Exception:
                        pass

            st.write(f"**Files ready to split:** {[f.name for f in raw_d_files]}")

            if st.button("🔀 PROMOTE ALL + SPLIT BY GAME",
                         type="primary", use_container_width=True,
                         key="promote_all_split"):
                all_results = {}
                progress = st.progress(0)
                for fi, src in enumerate(raw_d_files):
                    progress.progress((fi + 1) / len(raw_d_files),
                                      text=f"Splitting {src.name}…")
                    result = split_d_by_game(src, ROOT)
                    all_results[src.name] = result
                progress.empty()

                # ── Combine all states into ONE national file per game ────────
                # Each state's syndicates are distinct; concatenate them so each
                # game's CVI sees the whole country. Writes D_ALL_<game>.csv.
                games_touched = set()
                for res in all_results.values():
                    for gkey, count in res.items():
                        if gkey.startswith("_"):
                            continue
                        if count and count > 0:
                            games_touched.add(gkey)
                combine_rows = []
                for gkey in sorted(games_touched):
                    info = combine_states_for_game(gkey)
                    if info.get("files"):
                        combine_rows.append({
                            "Game":   GAMES_CFG[gkey]["label"],
                            "States combined": ", ".join(info["states"]),
                            "National rows": f"{info['rows']:,}",
                            "File": f"D_ALL_{gkey}.csv",
                        })
                if combine_rows:
                    st.markdown('<div class="ok">✅ National combine complete — '
                                'each game now has a D_ALL_&lt;game&gt;.csv across all '
                                'states (this is the default CVI source).</div>',
                                unsafe_allow_html=True)
                    show_paginated_df(pd.DataFrame(combine_rows), key="combine_rows_tbl", use_container_width=True)

                # Summary table
                st.markdown("**Split results:**")
                summary_rows = []
                all_unknowns = set()
                all_skipped  = set()
                for fname, res in all_results.items():
                    unknowns = res.pop("_unknown_games", [])
                    skipped  = res.pop("_skipped_games", [])
                    all_unknowns.update(unknowns)
                    all_skipped.update(skipped)
                    for gkey, count in res.items():
                        if count > 0:
                            summary_rows.append({
                                "Source file":   fname,
                                "Game":          GAMES_CFG[gkey]["label"],
                                "Rows written":  f"{count:,}",
                                "Saved to":      f"Games/{gkey.upper()}/Games_Breakdown/",
                            })

                if summary_rows:
                    show_paginated_df(pd.DataFrame(summary_rows), key="split_summary_tbl", use_container_width=True)
                    st.markdown('<div class="ok">✅ Split complete — '
                                'each game folder now has its own D variable files.</div>',
                                unsafe_allow_html=True)
                else:
                    st.warning("No rows were written. Check the Games column values.")

                if all_unknowns:
                    st.markdown(
                        f'<div class="warn">⚠️ Unrecognised game names (skipped): '
                        f'<b>{sorted(all_unknowns)}</b><br>'
                        f'Add these to GAME_NAME_MAP in masterapp.py if needed.</div>',
                        unsafe_allow_html=True)
                if all_skipped:
                    st.markdown(
                        f'<div class="info">ℹ️ Intentionally skipped '
                        f'(supplementary games): {sorted(all_skipped)}</div>',
                        unsafe_allow_html=True)

        else:
            st.info("No D_*.csv files in Main_Data/ yet. Run a sweep first.")

        # ── Promote files to Direct/ ──────────────────────────────────────────
        with st.expander("📤 Promote files to Direct/"):
            promote_candidates = (sorted(DIRS["Global_Scraper"].glob("D_*.csv")) +
                                  sorted(active_game_dirs()["Scraper"].glob("D_*.csv")))
            if promote_candidates:
                import shutil as _shutil_promote

                # ── Button row: ALL states at once OR one at a time ───────────
                st.markdown(
                    '<div class="info">Choose <b>Promote ALL</b> to copy every '
                    'state/territory file in one click, or use the selector below '
                    'to promote a single file.</div>',
                    unsafe_allow_html=True)

                btn_all, btn_single_col = st.columns(2)

                with btn_all:
                    if st.button("📤 Promote ALL States/Territories to Direct/",
                                 use_container_width=True, key="do_promote_all",
                                 type="primary"):
                        _direct = active_game_dirs()["Direct"]
                        _done, _errs = [], []
                        _prog = st.progress(0)
                        for _fi, _fp in enumerate(promote_candidates):
                            try:
                                _dst = _direct / _fp.name
                                _shutil_promote.copy2(_fp, _dst)
                                _done.append(_fp.name)
                            except Exception as _e:
                                _errs.append(f"{_fp.name}: {_e}")
                            _prog.progress((_fi + 1) / len(promote_candidates))
                        _prog.empty()
                        if _done:
                            st.success(
                                f"✅ Promoted {len(_done)} file(s) → "
                                f"{_direct.parent.name}/{_direct.name}/\n\n" +
                                "\n".join(f"• {n}" for n in _done))
                        if _errs:
                            st.error("Errors:\n" + "\n".join(_errs))

                st.markdown("---")

                # ── Single-file promote ───────────────────────────────────────
                st.markdown("**Promote a single file:**")
                chosen = st.selectbox("File to promote:",
                                      [f"{fp.parent.name}/{fp.name}"
                                       for fp in promote_candidates],
                                      key="promote_sel")
                chosen_fp = next(fp for fp in promote_candidates
                                 if f"{fp.parent.name}/{fp.name}" == chosen)
                c1p, c2p = st.columns([2, 1])
                with c1p:
                    df_prev = pd.read_csv(chosen_fp)
                    st.write(f"{len(df_prev):,} rows · {list(df_prev.columns)}")
                with c2p:
                    st.write("")
                    if st.button("📤 Promote to Direct/",
                                 use_container_width=True, key="do_promote"):
                        dst = active_game_dirs()["Direct"] / chosen_fp.name
                        _shutil_promote.copy2(chosen_fp, dst)
                        S["D"] = _load_file(dst)
                        st.success(f"Promoted → {dst.parent.name}/{chosen_fp.name} "
                                   f"({len(S['D']):,} rows loaded)")
            else:
                st.info("No D_*.csv files found.")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 5 — DATA BROWSER
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### 📂 Cached Data Browser")
        for row in status_rows:
            if not row["exists"]:
                continue
            fp = DIRS["Global_Scraper"] / row["file"]
            with st.expander(
                f"📄 {row['file']}  —  {row['rows']:,} rows  ·  {row['size_kb']} KB  "
                f"·  updated {row['modified']}"
            ):
                df_view = pd.read_csv(fp)
                # Show game breakdown
                if "Games" in df_view.columns:
                    st.markdown("**Games breakdown:**")
                    gc = df_view["Games"].value_counts().reset_index()
                    gc.columns = ["Game", "Rows"]
                    show_paginated_df(gc, key=f"cached_browser_gc_{row['state']}", use_container_width=True, height=200)

                # ── Current draw vs future draws ───────────────────────────
                if "Draw_Date" in df_view.columns:
                    try:
                        _dv_dd = pd.to_datetime(df_view["Draw_Date"], errors="coerce")
                        _dv_sorted = sorted(_dv_dd.dropna().unique())
                        if _dv_sorted:
                            _dv_cur  = _dv_sorted[0]
                            _dv_fut  = [d for d in _dv_sorted if d > _dv_cur]
                            _dv_c1, _dv_c2 = st.columns(2)
                            _n_cur_dv = int((_dv_dd == _dv_cur).sum())
                            _n_fut_dv = int((_dv_dd.isin(_dv_fut)).sum())
                            _dv_c1.metric(
                                "📅 Current draw syndicates", f"{_n_cur_dv:,}",
                                help=f"Draw date: {str(_dv_cur.date())}")
                            _dv_c2.metric(
                                "🔮 Future draw syndicates", f"{_n_fut_dv:,}",
                                help=f"{len(_dv_fut)} future date(s)")
                    except Exception:
                        pass

                show_paginated_df(df_view, key=f"cached_browser_view_{row['state']}", use_container_width=True)
                st.download_button(
                    f"⬇ Download {row['file']} — all {len(df_view):,} rows",
                    to_csv_bytes(df_view),
                    row["file"], "text/csv",
                    key=f"dl_view_{row['state']}"
                )

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 6 — UPLOAD PRE-SCRAPED FILE
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("📤 Upload a pre-scraped CSV"):
            up = st.file_uploader("Upload data file (CSV, Excel, HTML, JSON, XML, …)",
                                  type=_UPLOAD_TYPES, key="scraper_upload")
            if up:
                df_up = _read_uploaded(up)
                stem = up.name.rsplit(".", 1)[0]
                sp = active_game_dirs()["Main_Data"] / f"{stem}.csv"
                df_up.to_csv(sp, index=False)
                st.markdown(f'<div class="ok">✅ {sp.name} — {len(df_up):,} rows saved '
                            f'to Main_Data/</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 7 — TERMINAL COMMANDS + SCHEDULER
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("⏰ Run sweeps from terminal + Schedule nightly (Mac)"):
            script_path = Path(__file__).resolve()
            log_path    = ROOT / "logs" / "scraper.log"

            st.markdown(
                '<div class="note">💡 <b>Tip:</b> Use the <b>SWEEP STATES</b> buttons in '
                'Section 2 above for in-app scraping. The scraper is fully inlined in '
                'masterapp.py — <b>no separate thelott_picks_scraper.py needed.</b></div>',
                unsafe_allow_html=True)

            st.markdown("**Copy a command → paste directly into your terminal:**")

            _cmds_state = {
                "NSW": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('NSW')\"",
                "VIC": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('VIC')\"",
                "QLD": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('QLD')\"",
                "SA":  f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('SA')\"",
                "TAS": f"cd {script_path.parent} && python3 -c \"from masterapp import sweep_state_picks; sweep_state_picks('TAS')\"",
            }
            for _state, _cmd in _cmds_state.items():
                st.markdown(f"**Sweep {_state}:**")
                st.code(_cmd, language="bash")

            st.markdown("**Sweep ALL states at once (recommended):**")
            st.code(
                f"cd {script_path.parent} && python3 -c \"\n"
                f"from masterapp import sweep_state_picks\n"
                f"for s in ['NSW','VIC','QLD','SA','TAS']:\n"
                f"    print(f'=== sweeping {{s}} ===')\n"
                f"    sweep_state_picks(s)\n"
                f"\"",
                language="bash")

            st.markdown("**After sweeping — split and combine by game:**")
            st.code(
                f"cd {script_path.parent} && python3 -c \"\n"
                f"import sys; sys.path.insert(0, '.')\n"
                f"from masterapp import split_d_by_game, combine_states_for_game, ROOT, DIRS\n"
                f"from pathlib import Path\n"
                f"for f in sorted(DIRS['Global_Scraper'].glob('D_*.csv')):\n"
                f"    r = split_d_by_game(f, ROOT)\n"
                f"    print(f.name, r)\n"
                f"\"",
                language="bash")

            st.markdown("**Schedule nightly at 2 am (Mac crontab):**")
            st.code(
                f"# 1. Open crontab editor:\n"
                f"crontab -e\n\n"
                f"# 2. Add this line, then save:\n"
                f"0 2 * * * cd {script_path.parent} && python3 -c \""
                f"from masterapp import sweep_state_picks; "
                f"[sweep_state_picks(s) for s in ['NSW','VIC','QLD','SA','TAS']]"
                f"\" >> {log_path} 2>&1",
                language="bash")


st.markdown("---")

# ── GAME SELECTOR ─────────────────────────────────────────────────────────────
_game_cols = st.columns(len(GAME_KEYS) + 1)
with _game_cols[0]:
    st.markdown("**Select Game:**")
for _i, _gk in enumerate(GAME_KEYS):
    with _game_cols[_i + 1]:
        _cfg = GAMES_CFG[_gk]
        _active = (active_game() == _gk)
        _btn_type = "primary" if _active else "secondary"
        if st.button(
            f"{_cfg['emoji']} {_cfg['label']}\n{_cfg['draw_day']}",
            key=f"game_btn_{_gk}",
            type=_btn_type,
            use_container_width=True,
        ):
            st.session_state["active_game"] = _gk
            # Auto-reload B for the newly selected game so it's available
            # immediately (before the user clicks the B tab).
            _b_new = _auto_load_b(_gk)
            if not _b_new.empty:
                S["B"] = _b_new
            st.rerun()

_gcfg  = active_game_cfg()
_gdirs = active_game_dirs()
st.markdown(
    f'<div class="info">🎮 Active game: <b>{_gcfg["emoji"]} {_gcfg["label"]}</b> '
    f'— Pool: 1–{_gcfg["pool"]} · Pick {_gcfg["pick"]} · Draws: {_gcfg["draw_day"]} '
    f'· Data folder: <code>Games/{active_game().upper()}/</code></div>',
    unsafe_allow_html=True,
)
st.markdown("---")

page = st.radio("", [
    "📥 Main Data",
    "🔄 CVI Matrix",
    "🧩 Variable Inputs",
    "📦 Container Formula",
    "🖥️ Container Dashboards",
    "📤 Master Outputs",
    "🗂️ Cluster Manager",
], horizontal=True, label_visibility="collapsed")
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MAIN DATA
# ═══════════════════════════════════════════════════════════════════════════════
if page == "📥 Main Data":
    st.markdown('<span class="sec-hdr hdr-teal">📥 Main Data — Manually uploaded each run</span>',
                unsafe_allow_html=True)

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    Main Data for <b>{_gcfg['emoji']} {_gcfg['label']}</b> —
    your historical combination dataset matched against variable inputs.<br>
    Saved to disk immediately — <b>survives browser refresh</b>.<br>
    <b>Large files (63M rows):</b> save CSV to the Main_Data folder directly —
    no upload needed. The app reads from disk.<br>
    <b>accdb files:</b> place them directly in
    <code>Games/{_gkey.upper()}/Main_Data/</code> — Streamlit cannot upload
    Access databases via browser (browser security restriction).
    Install <code>mdb-tools</code> via Homebrew to read them.
    </div>
    """, unsafe_allow_html=True)

    # ── Folder path hint ───────────────────────────────────────────────────
    with st.expander("📁 Drop large files here directly (no upload needed)"):
        st.code(str(_gdirs["Main_Data"]), language=None)
        st.markdown("""
**For 63M row files:**
```bash
# Copy your main data file directly — no size limit
cp /path/to/your/maindata.csv ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/
```
Then click **🔄 Scan folder** below — the app reads it from disk.

**For accdb files:**
```bash
# 1. Install mdb-tools if not done
brew install mdb-tools

# 2. Convert accdb to CSV (no size limit)
mdb-export /path/to/file.accdb TableName > ~/Desktop/Sika/o_Automation_Suite/Games/SAT/Main_Data/maindata.csv
```
        """)

    # ── Upload (for smaller files under config limit) ──────────────────────
    up = st.file_uploader(
        "Upload Main Data file (CSV, Excel, HTML, JSON, XML, TSV, …)",
        type=_UPLOAD_TYPES,
        key="up_main_data")
    if up:
        path = _gdirs["Main_Data"] / up.name
        path.write_bytes(up.getvalue())
        st.markdown(
            f'<div class="ok">✅ Saved to disk: <code>{path.name}</code> '
            f'({path.stat().st_size / 1024 / 1024:.1f} MB) — '
            f'persists across refresh.</div>',
            unsafe_allow_html=True)

    # ── Scan folder for ALL files (disk-first approach) ────────────────────
    if st.button("🔄 Scan Main_Data folder", key="scan_main_data",
                 use_container_width=True):
        st.rerun()

    existing = sorted(
        list(_gdirs["Main_Data"].glob("*.csv")) +
        list(_gdirs["Main_Data"].glob("*.xlsx")) +
        list(_gdirs["Main_Data"].glob("*.accdb")) +
        list(_gdirs["Main_Data"].glob("*.mdb"))
    )
    # Also check legacy 1n subfolder
    existing += sorted(
        list(_gdirs["Main_Data"].glob("1n")) +
        [_gdirs["Main_Data"] / "1n" / f
         for f in ["Main_Data.xlsx", "Main_Data.csv"]
         if (_gdirs["Main_Data"] / "1n" / f).exists()]
    )
    existing = list(dict.fromkeys(existing))  # deduplicate

    if existing:
        st.markdown(f"**{len(existing)} file(s) found in Main_Data folder:**")
        for fp in existing:
            if not fp.is_file():
                continue
            size_mb = fp.stat().st_size / 1024 / 1024
            # Fast row count for CSV (no load)
            if fp.suffix == ".csv":
                rows = _count_csv_rows(fp)
                row_label = f"{rows:,} rows"
            else:
                row_label = f"{size_mb:.1f} MB"

            c1, c2, c3, c4 = st.columns([3, 1.2, 1, 0.8])
            c1.write(f"`{fp.name}`")
            c2.write(row_label)
            c3.write(f"{size_mb:.1f} MB")
            with c4:
                if st.button("Load", key=f"load_md_{fp.name}"):
                    with st.spinner(f"Loading {fp.name}…"):
                        df_ex = _load_file(fp, numeric=False)
                    if not df_ex.empty:
                        S["main_data"]      = df_ex
                        S["main_data_path"] = str(fp)
                        st.success(f"Loaded {fp.name}: {len(df_ex):,} rows")
                    else:
                        st.error(f"Could not read {fp.name}")

    # ── Auto-reload from disk if session state lost ────────────────────────
    if S.get("main_data", pd.DataFrame()).empty:
        saved_path = S.get("main_data_path", "")
        if saved_path and Path(saved_path).exists():
            with st.spinner("Reloading main data from disk…"):
                S["main_data"] = _load_file(Path(saved_path), numeric=False)
        else:
            # Auto-load most recent file in folder
            auto_files = sorted(
                _gdirs["Main_Data"].glob("*.csv"),
                key=lambda f: f.stat().st_mtime, reverse=True)
            if auto_files:
                newest = auto_files[0]
                st.markdown(
                    f'<div class="info">ℹ️ Auto-loading most recent file: '
                    f'<b>{newest.name}</b></div>', unsafe_allow_html=True)
                with st.spinner(f"Loading {newest.name}…"):
                    S["main_data"] = _load_file(newest, numeric=False)
                S["main_data_path"] = str(newest)

    md = S.get("main_data", pd.DataFrame())
    if not md.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Rows", f"{len(md):,}")
        c2.metric("Columns", len(md.columns))
        c3.metric("Est. match time (17 formulas)",
                  f"~{max(len(md)*7*17//16//1_000_000, 1)} sec")
        st.markdown("**Preview (first 20 rows):**")
        show_paginated_df(md, key="main_data_preview", use_container_width=True)
        st.download_button("⬇ Export Main Data CSV",
                           to_csv_bytes(md), "main_data.csv", "text/csv",
                           key="dl_main_data")
    else:
        st.info("No Main Data loaded. Upload a file, drop it in the folder, "
                "or click Load on an existing file above.")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CVI MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 CVI Matrix":
    st.markdown('<span class="sec-hdr hdr-teal">🔄 CVI Matrix — Transposition & Variable Slicing</span>',
                unsafe_allow_html=True)
    am_toggle("cvi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b><br>
    Reads D from <code>Games/{_gkey.upper()}/games_breakdown_{_gkey}/</code>
    (run <b>Promote All + Split by Game</b> on the Scraper page first).<br>
    Defaults to <b>D_ALL_{_gkey}.csv</b> — all states combined (national view).
    Each syndicate row → one w-column. Sorted longest→shortest.
    Ep = top 8 w-columns · Sp = top 4 lanes (a,b,c,d) · So = union combis.
    </div>
    """, unsafe_allow_html=True)

    # Game-specific Games_Breakdown/ folder — national D_ALL first (default)
    direct_dir = _gdirs["Games_Breakdown"]
    _all = sorted(direct_dir.glob(f"D_ALL_{_gkey}.csv"))
    _per_state = sorted(p for p in direct_dir.glob("D_*.csv")
                        if not p.name.startswith("D_ALL_"))
    raw_files  = _all + _per_state          # national file listed first

    # Fallback: also show global raw scrapes if game-specific not yet split
    global_files = sorted(DIRS["Global_Scraper"].glob("D_*.csv"))

    if not raw_files and not global_files:
        st.warning("No D files found. Run the Scraper, then Promote All + Split by Game.")
    else:
        if raw_files:
            st.markdown(f'<div class="ok">✅ {len(raw_files)} game-specific D file(s) '
                        f'found in Games/{_gkey.upper()}/Games_Breakdown/</div>',
                        unsafe_allow_html=True)
            all_files = raw_files
            file_labels = [f.name for f in all_files]
        else:
            st.markdown(
                f'<div class="warn">⚠️ No game-specific D files yet for '
                f'{_gcfg["label"]}. Showing global Main_Data files — '
                f'run Promote All + Split by Game first.</div>',
                unsafe_allow_html=True)
            all_files = global_files
            file_labels = [f"Main_Data/{f.name}" for f in all_files]

        chosen = st.selectbox("Source D file:", file_labels, key="cvi_d_sel")
        chosen_fp = all_files[file_labels.index(chosen)]
        df_raw = pd.read_csv(chosen_fp)
        _raw_wcols = [c for c in df_raw.columns if re.match(r'^w\d+$', str(c), re.I)]
        _raw_meta  = [c for c in df_raw.columns if c not in _raw_wcols]
        st.write(f"**{len(df_raw):,} syndicates · {len(_raw_wcols)} number positions**")
        # Restore original source view: all metadata columns, then 'w' label, then numbers.
        # Each ROW = one syndicate. The 'w' column labels it (w1, w2, w3…).
        # The original w1..w27 column headers are renamed to 1, 2, 3… (position numbers).
        _prev_n  = min(50, len(df_raw))
        _raw_disp = df_raw.head(_prev_n).copy()
        # Insert 'w' label column right after the last metadata column
        _insert_at = len(_raw_meta)
        _raw_disp.insert(_insert_at, "w", [f"w{i+1}" for i in range(len(_raw_disp))])
        # Rename w1..w27 → 1, 2, 3…
        _raw_disp = _raw_disp.rename(columns={wc: str(i+1) for i, wc in enumerate(_raw_wcols)})
        st.caption(f"Showing first {_prev_n} of {len(df_raw):,} syndicates — 'w' column labels each row:")
        show_paginated_df(_raw_disp, key="cvi_raw_disp", use_container_width=True)

        if st.button("🔄 BUILD W-MATRIX & SLICE Ep / Sp / So",
                     type="primary", use_container_width=True):
            # ROW-ORIENTATION NOTE: D now stays as ROWS (each syndicate = a row),
            # because transposing hundreds of thousands of syndicates into columns
            # exceeds the spreadsheet ceiling (~16k cols). Only transpose for small
            # files; for large D we keep rows and just slice the placeholder sets.
            EXCEL_COL_LIMIT = 16384
            if len(df_raw) > EXCEL_COL_LIMIT:
                st.info(f"Row orientation: {len(df_raw):,} syndicates exceed the "
                        f"{EXCEL_COL_LIMIT:,}-column spreadsheet limit, so D is kept "
                        f"as rows (no transpose). Collate directly in Container "
                        f"Formula — the CVI stacks w-sets as rows.")
                w_mat = df_raw  # keep as rows; do not transpose
            else:
                with st.spinner("Building…"):
                    w_mat = build_w_matrix(df_raw)
            if w_mat.empty:
                st.error("No number columns found. A D (syndicate) file needs "
                         "w1…wN columns (main-data n1…nN is also accepted).")
            else:
                mfname = chosen_fp.name.replace("D_", "CVI_Matrix_")
                cvi_out = _gdirs["CVI"] / mfname
                w_mat.to_csv(cvi_out, index=False)
                st.markdown(f'<div class="ok">✅ W-Matrix: {len(w_mat.columns):,} '
                            f'w-sets · longest pick = {len(w_mat)} numbers '
                            f'→ saved to Games/{_gkey.upper()}/</div>',
                            unsafe_allow_html=True)

                S["D"] = w_mat

                # ── ROW VIEW: w label on the LEFT, numbers across each row ──
                # The w-matrix as built has each w-set as a COLUMN (w1, w2, …).
                # We transpose so that each row = one w-set and the leftmost
                # column is the w label — longest pick at the top.
                st.markdown("---")
                st.markdown("### 📋 W-Matrix — row view (w label · numbers across · longest first)")
                st.markdown(
                    '<div class="info">Each row is one syndicate w-set. '
                    'The <b>w</b> column on the left is the set name. '
                    'Numbers run across to the right. '
                    'Rows sorted <b>longest → shortest</b> (most numbers at top).</div>',
                    unsafe_allow_html=True)

                # Transpose: columns of w_mat become rows
                _wcols_mat = [c for c in w_mat.columns if re.match(r'^w\d+$', str(c), re.I)]
                _w_sub = w_mat[_wcols_mat]
                # Each column is a w-set; T gives us: index=old col names, cols=row positions
                _w_T = _w_sub.T.reset_index()
                _w_T.columns = ["w"] + [f"pos_{i+1}" for i in range(_w_T.shape[1] - 1)]
                # Sort by count of non-null values (longest first)
                _w_T["_len"] = _w_T[[c for c in _w_T.columns if c.startswith("pos_")]].notna().sum(axis=1)
                _w_T = _w_T.sort_values("_len", ascending=False).drop(columns=["_len"])
                _w_T = _w_T.reset_index(drop=True)

                # ── RENUMBERING TABLE: old_w → new_w after sort ───────────
                st.markdown("#### 🔢 Renumbering — old position → new position after sort")
                st.markdown(
                    '<div class="note">After sorting longest → shortest, each w-set gets '
                    'a new sequential number. This table shows the before/after mapping. '
                    'The <b>old_w</b> column is dropped once you are satisfied with the '
                    'new order.</div>',
                    unsafe_allow_html=True)
                _renum_rows = []
                for _new_i, _old_w in enumerate(_w_T["w"].tolist()):
                    _n_nums = int(w_mat[_old_w].notna().sum()) if _old_w in w_mat.columns else 0
                    _renum_rows.append({
                        "new_w":    f"w{_new_i + 1}",
                        "old_w":    _old_w,
                        "numbers":  _n_nums,
                    })
                _renum_df = pd.DataFrame(_renum_rows)
                show_paginated_df(_renum_df, key="wmat_renum_df", use_container_width=True)

                # Apply new w labels to the transposed view
                _w_T.insert(0, "new_w", [f"w{i+1}" for i in range(len(_w_T))])
                # Show preview (cap at 200 rows for display speed)
                _show_n = min(200, len(_w_T))
                if len(_w_T) > _show_n:
                    st.caption(f"Preview — first {_show_n} of {len(_w_T):,} rows:")
                show_paginated_df(_w_T, key="wmat_rows_view", use_container_width=True)
                st.download_button(
                    f"⬇ Download W-Matrix (row view) — {len(_w_T):,} rows",
                    to_csv_bytes(_w_T),
                    mfname.replace("CVI_Matrix_", "W_Rows_"),
                    "text/csv",
                    key="dl_wmat_rows"
                )

                st.markdown("---")
                st.markdown("#### Variable slices (Ep · Sp · So)")
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown("**Ep** — top 8 w-sets")
                    show_paginated_df(slices["Ep"], key="cvi_slice_ep", use_container_width=True, height=200)
                with c2:
                    st.markdown("**Sp** — top 4 lanes (a–d)")
                    show_paginated_df(slices["Sp"], key="cvi_slice_sp", use_container_width=True, height=200)
                with c3:
                    st.markdown("**So** — union of Sp lanes")
                    show_paginated_df(slices["So"], key="cvi_slice_so", use_container_width=True, height=200)

        # Browse existing CVI matrices for this game
        cvi_files = sorted(_gdirs["CVI"].glob("CVI_Matrix_*.csv"))
        if cvi_files:
            st.markdown("---")
            cm = st.selectbox("Inspect saved matrix:", [f.name for f in cvi_files],
                              key="cvi_inspect_sel")
            df_m = pd.read_csv(_gdirs["CVI"] / cm)
            _wcols_m = [c for c in df_m.columns if re.match(r'^w\d+$', str(c), re.I)]
            _meta_m  = [c for c in df_m.columns if c not in _wcols_m]
            st.write(f"**{len(df_m):,} syndicates · {len(_wcols_m)} number positions**")
            # Each ROW = one syndicate = one w-set.
            # Add 'w' label column, rename w1..wN headers to 1, 2, 3… (position numbers).
            # No transposing needed — just relabel.
            _cap_m = min(50, len(df_m))
            _df_m_disp = df_m[_wcols_m].head(_cap_m).copy()
            _df_m_disp.insert(0, "w", [f"w{i+1}" for i in range(len(_df_m_disp))])
            _df_m_disp = _df_m_disp.rename(columns={wc: str(i+1) for i, wc in enumerate(_wcols_m)})
            st.caption(f"Showing first {_cap_m} of {len(df_m):,} syndicates — each row is one w-set:")
            show_paginated_df(_df_m_disp, key="cvi_inspect_disp", use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: VARIABLE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧩 Variable Inputs":
    st.markdown('<span class="sec-hdr hdr-green">🧩 Variable Inputs — B · R · D · Ep · Sp · So</span>',
                unsafe_allow_html=True)
    am_toggle("vi")

    _gcfg  = active_game_cfg()
    _gdirs = active_game_dirs()
    _gkey  = active_game()

    st.markdown(f"""
    <div class="info">
    <b>Active game: {_gcfg['emoji']} {_gcfg['label']}</b> — Pool 1–{_gcfg['pool']},
    Pick {_gcfg['pick']}, draws {_gcfg['draw_day']}.<br>
    <b>B</b> = pre-loaded w-columns from <b>Base_&lt;game&gt;.xlsx</b> (e.g. Base_sat.xlsx, sheet B_sat;
    auto-loaded from your project folder — rarely changes).<br>
    <b>Ep</b> = ExcelPro output: pairs the 8 longest rows of D (w1–w8) into 4 objects, then
    filters R's wt list through each object's two halves. <b>Auto-runs when D is loaded.</b><br>
    <b>Sp</b> = Splits (task1b): uses the 4 longest rows of ALL D (w1,w2,w3,w4) as input sets —
    splits each in half and computes set-algebra combinations. <b>Auto-runs when D is loaded.</b><br>
    <b>So</b> = SplitsCombi (auto_vba): same 4 longest rows of D → union combis of split halves.
    <b>Auto-runs when D is loaded.</b><br>
    <b>R</b>  = Rainbow (task2): Since Last from lottolyzer → powerset combos.<br>
    <b>D</b>  = Syndicate w-columns (standalone + feeds formula row 11).<br>
    <b>Note:</b> In ALL D, each syndicate IS a row; w1,w2,… label the number positions across.
    The generators peel the N longest rows and re-orient them as columns w1,w2,… (one set per column)
    so the set-algebra code can work on them as independent pools.
    </div>
    """, unsafe_allow_html=True)

    vtabs = st.tabs(["B (Base)", "R (Rainbow)", "D (Direct)",
                     "Ep (ExcelPro)", "Sp (Splits)", "So (SplitsCombi)",
                     "Since Last"])

    # ── TAB: B (Base) ──────────────────────────────────────────────────────
    with vtabs[0]:
        st.markdown(f"**B — Base variable for {_gcfg['label']}** "
                    f"(sheet: `{_gcfg['b_sheet']}` in {_gcfg.get('b_file','Base.xlsx')})")

        # Load B: prefer this game's own Base_<game>.xlsx; then a shared Base.xlsx;
        # then the legacy f_rules_Gclaude.xlsx. Non-breaking migration.
        b_file = _gcfg.get("b_file", "Base.xlsx")
        b_rules_candidates = (list(ROOT.rglob(b_file))
                              or list(ROOT.rglob("Base.xlsx"))
                              or list(ROOT.rglob("f_rules_Gclaude.xlsx")))
        if b_rules_candidates:
            b_rules_path = b_rules_candidates[0]
            try:
                xl_b = pd.ExcelFile(b_rules_path, engine="openpyxl")
                # Prefer this game's clean sheet (B_pb/B_sat/…); then the uppercase
                # key (PB/SAT/…); then the legacy cryptic sheet name; finally, if the
                # workbook is a single-game file, just use its only sheet.
                sheet = None
                gk = _gkey.upper()
                for cand in (_gcfg.get("b_sheet"), gk, _gcfg.get("b_sheet_legacy")):
                    if cand and cand in xl_b.sheet_names:
                        sheet = cand
                        break
                if sheet is None and len(xl_b.sheet_names) == 1:
                    sheet = xl_b.sheet_names[0]   # per-game file with one sheet
                if sheet is not None:
                    df_b_raw = xl_b.parse(sheet, header=None)

                    def _nums_from_series(s):
                        """Extract positive integers from a pandas Series."""
                        out = []
                        for v in s.dropna():
                            try:
                                fv = float(v)
                                if fv >= 1:
                                    out.append(int(fv))
                            except (ValueError, TypeError):
                                pass
                        return out

                    b_data = {}

                    # ── Strategy 1: row 0 contains "w"-prefixed headers ────────
                    w_cols_b = [str(df_b_raw.iloc[0, c]).strip()
                                for c in range(df_b_raw.shape[1])
                                if str(df_b_raw.iloc[0, c]).strip().lower().startswith("w")]
                    if w_cols_b:
                        for i, wc in enumerate(w_cols_b):
                            nums = _nums_from_series(df_b_raw.iloc[1:, i])
                            if nums:
                                b_data[wc] = pd.Series(nums)

                    # ── Strategy 2: parsed with header=0, columns named w* ────
                    if not b_data:
                        df_b_hdr = xl_b.parse(sheet)
                        w_hdr_cols = [c for c in df_b_hdr.columns
                                      if str(c).strip().lower().startswith("w")]
                        for wc in w_hdr_cols:
                            nums = _nums_from_series(df_b_hdr[wc])
                            if nums:
                                b_data[str(wc).strip()] = pd.Series(nums)

                    # ── Strategy 3: auto-detect — every numeric column = a w-set
                    if not b_data:
                        w_idx = 1
                        for ci in range(df_b_raw.shape[1]):
                            nums = _nums_from_series(df_b_raw.iloc[:, ci])
                            if nums:
                                b_data[f"w{w_idx}"] = pd.Series(nums)
                                w_idx += 1
                        if b_data:
                            st.info(
                                f"ℹ️ Sheet '{sheet}' had no 'w'-prefixed headers — "
                                f"auto-named {len(b_data)} numeric column(s) as "
                                f"w1 … w{len(b_data)}. "
                                f"Rename your sheet's header row to w1, w2, … to "
                                f"control the labels.")
                    if b_data:
                        df_b = pd.DataFrame(b_data)
                        S["B"] = df_b
                        st.markdown(
                            f'<div class="ok">✅ B loaded: {len(b_data)} w-columns '
                            f'from sheet <b>{sheet}</b> in {b_rules_path.name}</div>',
                            unsafe_allow_html=True)
                        # Show w-column lengths
                        len_info = {wc: len(b_data[wc].dropna()) for wc in b_data}
                        st.write("Column lengths:", len_info)

                        # ── View toggle: columns (as stored) vs rows ──────────
                        _b_view = st.radio(
                            "Display B as:",
                            ["Columns (as stored — each w is a column)",
                             "Rows (after transpose — each w becomes a row)"],
                            horizontal=True,
                            key="b_view_mode"
                        )
                        if _b_view.startswith("Rows"):
                            # Transpose so each w-column becomes a row with w label on left
                            df_b_T = df_b.T.reset_index()
                            df_b_T.columns = ["w"] + [f"pos_{i+1}"
                                                       for i in range(df_b_T.shape[1]-1)]
                            # Sort longest → shortest
                            _pos_cols_b = [c for c in df_b_T.columns if c.startswith("pos_")]
                            df_b_T["_len"] = df_b_T[_pos_cols_b].notna().sum(axis=1)
                            df_b_T = df_b_T.sort_values("_len", ascending=False).drop(columns=["_len"]).reset_index(drop=True)
                            # Add new_w as first column (renumbering after sort)
                            df_b_T.insert(0, "new_w", [f"w{i+1}" for i in range(len(df_b_T))])
                            st.markdown(
                                '<div class="info">ℹ️ B transposed — each row is one '
                                'w-set. <b>new_w</b> = renumbered after longest→shortest sort. '
                                '<b>w</b> = original name. Numbers run across to the right.</div>',
                                unsafe_allow_html=True)
                            show_paginated_df(df_b_T, key="b_transposed_view", use_container_width=True)
                            st.download_button("⬇ Download B_as_rows.csv",
                                               to_csv_bytes(df_b_T), "B_as_rows.csv",
                                               "text/csv", key="dl_b_rows")
                        else:
                            show_paginated_df(df_b, key="b_cols_view", use_container_width=True)
                            st.download_button("⬇ Download B.csv",
                                               to_csv_bytes(df_b), "B.csv",
                                               "text/csv", key="dl_b_rules")

                        # ── Renumbering table (always shown below) ────────────
                        st.markdown("---")
                        st.markdown("**🔢 B Renumbering — original → new position (longest → shortest)**")
                        _b_renum = []
                        _b_wcols_sorted = sorted(b_data.keys(),
                                                  key=lambda k: len(b_data[k].dropna()),
                                                  reverse=True)
                        for _bi, _bwc in enumerate(_b_wcols_sorted):
                            _b_renum.append({
                                "new_w": f"w{_bi+1}",
                                "old_w": _bwc,
                                "numbers": int(len(b_data[_bwc].dropna())),
                            })
                        show_paginated_df(pd.DataFrame(_b_renum), key="b_renum_tbl", use_container_width=True)
                    else:
                        st.warning(f"Sheet '{sheet}' found but no numeric w-columns parsed.")
                else:
                    st.warning(f"No B sheet for this game. Looked for "
                               f"'{_gcfg.get('b_sheet')}' (or legacy "
                               f"'{_gcfg.get('b_sheet_legacy')}'). "
                               f"Available: {xl_b.sheet_names}")
            except Exception as ex:
                st.error(f"Error reading {b_rules_path.name}: {ex}")
        else:
            st.markdown(f'<div class="warn">{b_file} not found in project folder '
                        '(shared Base.xlsx or legacy f_rules_Gclaude.xlsx also accepted). '
                        'Upload it below — the app will load it immediately.</div>',
                        unsafe_allow_html=True)
            up_b = st.file_uploader(
                f"Upload {b_file} (or Base.xlsx / f_rules_Gclaude.xlsx)",
                type=_UPLOAD_TYPES, key="up_b_rules")
            if up_b:
                # Save under both the game-specific name AND the legacy name
                dest_game   = ROOT / b_file
                dest_legacy = ROOT / "f_rules_Gclaude.xlsx"
                _raw = up_b.read()
                dest_game.write_bytes(_raw)
                if dest_game != dest_legacy:
                    dest_legacy.write_bytes(_raw)
                st.success(f"Saved to {dest_game.name}. Loading now…")
                # Immediately parse and load B so user doesn't need to refresh
                try:
                    xl_up = pd.ExcelFile(dest_game, engine="openpyxl")
                    _sheet_up = None
                    for _cand in (_gcfg.get("b_sheet"), _gkey.upper(),
                                  _gcfg.get("b_sheet_legacy")):
                        if _cand and _cand in xl_up.sheet_names:
                            _sheet_up = _cand
                            break
                    if _sheet_up is None and len(xl_up.sheet_names) == 1:
                        _sheet_up = xl_up.sheet_names[0]
                    if _sheet_up:
                        _df_b_up = xl_up.parse(_sheet_up, header=None)
                        _w_cols_up = [str(_df_b_up.iloc[0, c])
                                      for c in range(_df_b_up.shape[1])
                                      if str(_df_b_up.iloc[0, c]).startswith("w")]
                        _b_data_up = {}
                        for _i, _wc in enumerate(_w_cols_up):
                            _col_vals = _df_b_up.iloc[1:, _i].dropna()
                            _nums = [int(float(v)) for v in _col_vals
                                     if str(v).replace(".", "").replace("-", "").isdigit()
                                     and float(v) >= 1]
                            if _nums:
                                _b_data_up[_wc] = pd.Series(_nums)
                        if _b_data_up:
                            S["B"] = pd.DataFrame(_b_data_up)
                            st.markdown(
                                f'<div class="ok">✅ B loaded immediately: '
                                f'{len(_b_data_up)} w-columns from sheet '
                                f'<b>{_sheet_up}</b>.</div>',
                                unsafe_allow_html=True)
                            show_paginated_df(S["B"], key="b_uploaded_view", use_container_width=True)
                        else:
                            st.warning("File saved but no w-columns found. "
                                       "Check sheet layout — row 0 must have w1, w2, …")
                    else:
                        st.warning(f"File saved but sheet '{_gcfg.get('b_sheet')}' not found. "
                                   f"Available sheets: {xl_up.sheet_names}")
                except Exception as _ex_up:
                    st.error(f"Saved OK but could not parse: {_ex_up}")

    # ── TAB: R (Rainbow) ───────────────────────────────────────────────────
    with vtabs[1]:
        st.markdown("**R — Rainbow (task2.py): Since Last → powerset combos**")

        sl_file = _gdirs["SinceLast"] / "since_last.json"
        # Auto-fetch from lottolyzer if we don't have it cached yet.
        # Use the overridden URL from the Since Last tab if the user changed it.
        _r_sl_url = (st.session_state.get("sl_url_override")
                     or _gcfg["lottolyzer"])
        if not sl_file.exists():
            with st.spinner(f"Fetching Since Last from lottolyzer for {_gcfg['label']}…"):
                sl_dict_auto = fetch_since_last(_r_sl_url, _gcfg["pool"])
            if sl_dict_auto:
                save_since_last(sl_dict_auto, _gkey, _gcfg["label"],
                                _gcfg["pool"], _r_sl_url, sl_file)
                st.markdown(
                    f'<div class="ok">✅ Auto-fetched Since Last — '
                    f'{len(sl_dict_auto)} numbers from lottolyzer.</div>',
                    unsafe_allow_html=True)
        if sl_file.exists():
            try:
                sl_data = json.loads(sl_file.read_text())
                since_last_dict = {int(k): int(v)
                                   for k, v in sl_data.get("since_last_dict", {}).items()}
                all_wt = sl_data.get("all_wt", [])
                to_keep_list = sl_data.get("to_keep", [])
                scraped_at = sl_data.get("scraped_at", "unknown")

                st.markdown(
                    f'<div class="ok">✅ Since Last loaded — {len(since_last_dict)} numbers '
                    f'| scraped: {scraped_at[:16]}</div>',
                    unsafe_allow_html=True)

                st.write(f"**all_wt** (first 15 / most recent → oldest): "
                         f"`{all_wt[:15]}...`")

                # AUTO: pick the highest SAFE max_comb (no malfunction). Tick the
                # box to intervene with a manual value instead.
                _n_groups = len(set(int(v) + 1 for v in since_last_dict.values()))
                manual_max = None
                if st.checkbox("Set max groups manually (else auto safe-max)",
                               key="r_manual"):
                    manual_max = st.slider("Max Since Last groups to combine:",
                                           1, max(2, _n_groups), min(3, _n_groups),
                                           key="r_max_comb")
                if st.button("▶ Generate Rainbow (R)", key="gen_R",
                             type="primary", use_container_width=True):
                    try:
                        sl_df = pd.DataFrame({
                            "numbers": list(since_last_dict.keys()),
                            "Since Last": list(since_last_dict.values()),
                        })
                        sl_df["to_keep"] = pd.Series(all_wt)
                        r_df, r_wt, r_info = generate_rainbow(sl_df, max_comb=manual_max)
                        S["R"] = r_df
                        S["_R_wt"] = r_wt    # R's to_keep/wt -> Ep input2
                        r_path = _gdirs["Rainbow"] / f"R_{_gkey}.csv"
                        r_df.to_csv(r_path, index=False)
                        _cap = " (auto-capped to stay safe)" if r_info["capped"] else ""
                        st.markdown(
                            f'<div class="ok">✅ R generated: {r_info["n_combos"]} combos '
                            f'from {r_info["n_groups"]} groups · max_comb={r_info["max_comb"]}'
                            f'{_cap} → {r_path.name}</div>', unsafe_allow_html=True)
                        show_paginated_df(r_df, key="r_generated_view", use_container_width=True)
                    except Exception as ex:
                        st.error(f"Rainbow error: {ex}")

            except Exception as ex:
                st.error(f"Error loading Since Last: {ex}")
        else:
            st.markdown(
                '<div class="warn">⚠️ Couldn\'t auto-fetch Since Last from lottolyzer '
                '(page may be unreachable or its layout changed). Use the '
                '<b>Since Last</b> tab to fetch again or upload it manually.</div>',
                unsafe_allow_html=True)

        if not S.get("R", pd.DataFrame()).empty:
            st.markdown("**Current R in memory:**")
            show_paginated_df(S["R"], key="r_current_memory", use_container_width=True, height=200)

    # ── TAB: D (Direct) ────────────────────────────────────────────────────
    with vtabs[2]:
        st.markdown(f"**D — Syndicate data for {_gcfg['label']}**")

        # ── Build GAME-SPECIFIC file lists (Q1 fix) ────────────────────────
        # Per-state files for THIS game only  e.g. D_NSW_sat.csv, D_VIC_sat.csv…
        _d_state_files = sorted(
            p for p in _gdirs["Games_Breakdown"].glob("D_*.csv")
            if not p.name.startswith("D_ALL_"))
        # Pre-combined national file for this game (D_ALL_sat.csv)
        _d_all_combined = sorted(_gdirs["Games_Breakdown"].glob(f"D_ALL_{_gkey}.csv"))
        # Additional files already in Direct/
        _d_direct_files = (sorted(_gdirs["Direct"].glob("D*.csv"))
                           + sorted(_gdirs["Direct"].glob("D*.xlsx")))

        # Extract state abbreviations from per-state file names for display
        _state_abbrs = []
        for _sf in _d_state_files:
            _parts = _sf.stem.upper().split("_")   # D_NSW_SAT → ['D','NSW','SAT']
            if len(_parts) >= 2:
                _state_abbrs.append(_parts[1])      # 'NSW', 'VIC' …

        # Selectbox list: pre-combined first (already merged), then per-state, then Direct
        _d_select_files = _d_all_combined + _d_state_files + _d_direct_files
        _seen_fp, _d_select_dedup = set(), []
        for _f in _d_select_files:
            if _f not in _seen_fp:
                _seen_fp.add(_f); _d_select_dedup.append(_f)
        _d_select_files = _d_select_dedup

        # Helper: derive full-game key for storage
        _d_full_key = f"D_full_{_gkey}"   # stores unfiltered version
        _d_draw_key = f"active_draw_{_gkey}"  # stores active draw number

        st.markdown(
            f'<div class="info">'
            f'<b>D scope — {_gcfg["label"]} only.</b> '
            f'Per-state files: <code>'
            + (", ".join(f"D_{s}_{_gkey}.csv" for s in _state_abbrs) if _state_abbrs
               else f"none found yet in Games_Breakdown/")
            + f'</code><br>'
            f'<b>Load ALL</b> combines only the per-state files for {_gcfg["label"]} '
            f'({len(_d_state_files)} file{"s" if len(_d_state_files)!=1 else ""}). '
            f'Then use <b>Set Active Draw ▼</b> to lock one draw into memory before '
            f'CVI Matrix and all downstream steps.</div>',
            unsafe_allow_html=True)

        if _d_select_files:
            _d_labels = [f.name for f in _d_select_files]
            ch_d = st.selectbox("Select file to load (or use Load ALL below):",
                                _d_labels, key="sel_d_file")

            c_load1, c_loadall = st.columns(2)

            # ── Load single file ────────────────────────────────────────────
            with c_load1:
                if st.button("▶ Load selected file", key="btn_d_load_one",
                             use_container_width=True):
                    _fp = _d_select_files[_d_labels.index(ch_d)]
                    with st.spinner(f"Loading {_fp.name}…"):
                        df_d_loaded = _load_file(_fp)
                        if not df_d_loaded.empty:
                            df_d_loaded = sort_d_longest_first(df_d_loaded)
                    S["D"] = df_d_loaded
                    S[_d_full_key] = df_d_loaded.copy()
                    S.pop(_d_draw_key, None)           # clear any prior draw filter
                    st.success(f"✅ Loaded {len(df_d_loaded):,} rows from {_fp.name} "
                               f"(sorted longest → shortest)")
                    _auto_wire_generators(_gdirs, _gkey)

            # ── Load ALL per-state files for this game ──────────────────────
            with c_loadall:
                _load_all_lbl = (
                    f"📦 Load ALL {_gcfg['label']} state files "
                    f"({len(_d_state_files)} file{'s' if len(_d_state_files)!=1 else ''}"
                    + (f": {' + '.join(_state_abbrs)}" if _state_abbrs else "")
                    + ")"
                )
                if _d_state_files:
                    if st.button(_load_all_lbl, key="btn_d_load_all",
                                 type="primary", use_container_width=True):
                        _dfs_states = []
                        _prog_d = st.progress(0,
                            text=f"Loading {_gcfg['label']} state files…")
                        for _fi, _fp in enumerate(_d_state_files):
                            _prog_d.progress((_fi + 1) / len(_d_state_files),
                                             text=f"Loading {_fp.name}…")
                            try:
                                _tmp = _load_file(_fp)
                                if not _tmp.empty:
                                    _dfs_states.append(_tmp)
                            except Exception as _ex:
                                st.warning(f"Skipped {_fp.name}: {_ex}")
                        _prog_d.empty()
                        if _dfs_states:
                            _d_combined = pd.concat(_dfs_states, ignore_index=True)
                            _d_combined = sort_d_longest_first(_d_combined)
                            # De-duplicate by Syndicate_ID + Draw_Number
                            _dedup_cols = [c for c in
                                           ["Syndicate_ID", "Draw_Number", "Game"]
                                           if c in _d_combined.columns]
                            if _dedup_cols:
                                _d_combined = _d_combined.drop_duplicates(
                                    subset=_dedup_cols, keep="first")
                            S["D"] = _d_combined
                            S[_d_full_key] = _d_combined.copy()
                            S.pop(_d_draw_key, None)       # clear prior draw filter
                            st.success(
                                f"✅ Combined {len(_dfs_states)} state file(s) for "
                                f"{_gcfg['label']} "
                                f"({' + '.join(_state_abbrs) if _state_abbrs else ''}) → "
                                f"**{len(_d_combined):,} rows** · "
                                f"{len(_d_combined.columns)} cols "
                                f"(duplicates removed, sorted longest → shortest)")
                            _auto_wire_generators(_gdirs, _gkey)
                        else:
                            st.error("No state D files could be loaded.")
                else:
                    st.warning(
                        f"No per-state D files found for {_gcfg['label']} in "
                        f"Games_Breakdown/. Run Promote All + Split by Game first, "
                        f"then return here.")

        # ── Active Draw Selector (Q2 fix) ───────────────────────────────────
        _df_full = S.get(_d_full_key, S.get("D", pd.DataFrame()))
        if not _df_full.empty:
            st.markdown("---")
            st.markdown("#### 🎯 Active Draw — Lock one draw into memory")
            st.markdown(
                '<div class="info">'
                'Selecting a draw and clicking <b>Set Active Draw</b> filters '
                '<code>S["D"]</code> to <b>only that draw\'s syndicates</b>. '
                'CVI Matrix, Container Formula, and all downstream steps will then '
                'work exclusively on that draw\'s data. '
                'Click <b>Restore all draws</b> to go back to the full dataset.'
                '</div>',
                unsafe_allow_html=True)

            _dn_col = "Draw_Number" if "Draw_Number" in _df_full.columns else None
            _dd_col = "Draw_Date"   if "Draw_Date"   in _df_full.columns else None
            _cur_active_draw = S.get(_d_draw_key)

            # Build draw options from Draw_Number column
            if _dn_col:
                try:
                    _draw_num_vals = (
                        _df_full[_dn_col]
                        .dropna()
                        .astype(str)
                        .str.extract(r'(\d+)')[0]
                        .dropna()
                        .astype(int)
                        .sort_values()
                        .unique()
                        .tolist()
                    )
                except Exception:
                    _draw_num_vals = []
            else:
                _draw_num_vals = []

            _draw_opts = ["All draws"] + [str(d) for d in _draw_num_vals]
            _def_draw_idx = 0
            if _cur_active_draw and str(_cur_active_draw) in _draw_opts:
                _def_draw_idx = _draw_opts.index(str(_cur_active_draw))

            _adc1, _adc2, _adc3 = st.columns([2, 1, 2])
            with _adc1:
                _draw_sel = st.selectbox(
                    f"Draw # for {_gcfg['label']}:",
                    _draw_opts,
                    index=_def_draw_idx,
                    key=f"d_draw_sel_{_gkey}")
            with _adc2:
                st.write(""); st.write("")
                if st.button("🎯 Set Active Draw",
                             key=f"btn_set_draw_{_gkey}",
                             type="primary",
                             use_container_width=True):
                    if _draw_sel == "All draws":
                        S["D"] = S[_d_full_key].copy()
                        S.pop(_d_draw_key, None)
                        st.success(
                            f"✅ Restored all draws — "
                            f"{len(S['D']):,} rows in S['D'].")
                    else:
                        S[_d_draw_key] = int(_draw_sel)
                        if _dn_col:
                            try:
                                _dn_s = (_df_full[_dn_col]
                                         .astype(str)
                                         .str.extract(r'(\d+)')[0]
                                         .fillna("-1")
                                         .astype(int))
                                _filt = _df_full[_dn_s == int(_draw_sel)].copy()
                            except Exception:
                                _filt = _df_full.copy()
                        elif _dd_col:
                            # fallback: filter by Draw_Date
                            try:
                                _dds = pd.to_datetime(_df_full[_dd_col], errors="coerce")
                                _uniq_dt = sorted(_dds.dropna().unique())
                                _di = min(int(_draw_sel), len(_uniq_dt)) - 1
                                _tgt = _uniq_dt[max(0, _di)]
                                _filt = _df_full[
                                    _dds.dt.normalize() ==
                                    pd.Timestamp(_tgt).normalize()
                                ].copy()
                            except Exception:
                                _filt = _df_full.copy()
                        else:
                            _filt = _df_full.copy()
                        S["D"] = _filt
                        st.success(
                            f"✅ Active draw set to **Draw {_draw_sel}** → "
                            f"**{len(_filt):,} rows** now in S['D']. "
                            f"CVI Matrix and all downstream steps use this draw only.")
                    st.rerun()

            # ── Active draw status banner ───────────────────────────────────
            with _adc3:
                if _cur_active_draw:
                    _filt_n = len(S.get("D", pd.DataFrame()))
                    st.success(
                        f"🎯 **Active: Draw {_cur_active_draw}** "
                        f"({_filt_n:,} rows in memory)")
                else:
                    _full_n = len(_df_full)
                    st.info(f"📋 No draw filter — all {_full_n:,} rows loaded")

            # ── Per-draw breakdown metrics ──────────────────────────────────
            if _dn_col and _draw_num_vals:
                try:
                    _dn_s2 = (_df_full[_dn_col]
                              .astype(str)
                              .str.extract(r'(\d+)')[0]
                              .dropna()
                              .astype(int))
                    _draw_counts = _dn_s2.value_counts().sort_index()
                    _mc = st.columns(min(len(_draw_counts), 6))
                    for _ci, (_dv, _dc) in enumerate(_draw_counts.items()):
                        _flag = "🎯 " if str(_dv) == str(_cur_active_draw) else ""
                        _mc[_ci % len(_mc)].metric(
                            f"{_flag}Draw {_dv}", f"{_dc:,}",
                            help=f"{_dc:,} syndicates for draw {_dv}")
                except Exception:
                    pass
            elif _dd_col:
                try:
                    _dda = pd.to_datetime(_df_full[_dd_col], errors="coerce")
                    _udates = sorted(_dda.dropna().unique())
                    if _udates:
                        _mc = st.columns(min(len(_udates), 6))
                        for _ci, _ud in enumerate(_udates):
                            _cnt = int((_dda.dt.normalize() ==
                                        pd.Timestamp(_ud).normalize()).sum())
                            _mc[_ci % len(_mc)].metric(
                                str(pd.Timestamp(_ud).date()), f"{_cnt:,}")
                except Exception:
                    pass

        # ── Display current S["D"] (filtered or full) ──────────────────────
        df_d = S.get("D", pd.DataFrame())
        if not df_d.empty:
            _d_wcols_tab = [c for c in df_d.columns
                            if re.match(r'^w\d+$', str(c), re.I)]
            _d_col_orient = (len(_d_wcols_tab) > 0 and
                             len(df_d) < len(_d_wcols_tab) and
                             len(_d_wcols_tab) > 20)

            if _d_col_orient:
                _d_norm = df_d[_d_wcols_tab].T.reset_index(drop=True)
            else:
                _d_norm = (df_d[_d_wcols_tab].reset_index(drop=True)
                           if _d_wcols_tab else df_d.reset_index(drop=True))

            _d_norm = _d_norm.apply(pd.to_numeric, errors="coerce")
            _d_sort_key = _d_norm.notna().sum(axis=1)
            _d_norm = _d_norm.loc[
                _d_sort_key.sort_values(ascending=False, kind="stable").index
            ].reset_index(drop=True)
            _d_norm.columns = list(range(1, len(_d_norm.columns) + 1))
            _d_norm.insert(0, "w", [f"w{i+1}" for i in range(len(_d_norm))])

            n_syn = len(_d_norm)
            n_pos = len(_d_norm.columns) - 1
            _draw_lbl = (f" · Draw {S.get(_d_draw_key)}"
                         if S.get(_d_draw_key) else " · all draws")
            st.write(f"**{n_syn:,} syndicates · up to {n_pos} number "
                     f"positions{_draw_lbl}**")
            show_paginated_df(_d_norm, key="d_tab_main_view",
                              use_container_width=True)
            st.download_button(
                f"⬇ Download D (current view) as CSV — {n_syn:,} rows",
                to_csv_bytes(_d_norm),
                f"D_{_gkey}_export.csv",
                "text/csv",
                key="dl_d_tab"
            )
        elif _d_select_files:
            st.info("Load a file above to preview syndicates.")
        else:
            st.info(f"No D files in Games/{_gkey.upper()}/. Run the Scraper page "
                    f"→ Promote All + Split by Game.")

    # ── TAB: Ep (ExcelPro) ─────────────────────────────────────────────────
    with vtabs[3]:
        st.markdown("**Ep — ExcelPro** · objects a,b,c,d from D's 8 longest rows "
                    "→ 6 pair blocks filtered through R's wt list")
        st.markdown(
            "Objects: **a** = D rows 1+2 (w1,w2) · **b** = rows 3+4 (w3,w4) · "
            "**c** = rows 5+6 (w5,w6) · **d** = rows 7+8 (w7,w8). "
            "wt list from R; falls back to unique numbers in D's top-8 rows if R not loaded. "
            "Auto-runs when D loads — use button to re-run manually.")

        d_df = S.get("D", pd.DataFrame())
        r_df = S.get("R", pd.DataFrame())

        if d_df.empty:
            st.warning("Load D first (Direct tab).")
        else:
            wt_list_ep: list = []
            if not r_df.empty:
                _wt_col = next((c for c in ("wt", "all_wt") if c in r_df.columns), None)
                if _wt_col:
                    wt_list_ep = r_df[_wt_col].dropna().astype(int).tolist()
            if not wt_list_ep:
                _d_top8 = prepare_d_input_sets(d_df, 8)
                wt_list_ep = sorted({int(v) for col in _d_top8.columns
                                     for v in _d_top8[col].dropna()})

            st.write(f"wt list: **{len(wt_list_ep)} numbers** "
                     f"({'from R' if wt_list_ep and not r_df.empty else 'from D fallback'})")

            if st.button("▶ Run Ep", type="primary",
                         key="run_ep_btn", use_container_width=True):
                try:
                    _ep_objs = prepare_ep_objects(d_df, mode="pairs")
                    _ep_df   = generate_excelpro(_ep_objs, wt_list_ep)
                    S["Ep"]  = _ep_df
                    _ep_path = _gdirs["ExcelPro"] / f"Ep_{_gkey}.csv"
                    _ep_df.to_csv(_ep_path, index=False)
                    st.markdown(
                        f'<div class="ok">✅ Ep: {_ep_df.shape[1]} cols → '
                        f'{_ep_path.name}</div>',
                        unsafe_allow_html=True)
                except Exception as _ep_ex:
                    st.error(f"Ep error: {_ep_ex}")

        ep_df_view = S.get("Ep", pd.DataFrame())
        if not ep_df_view.empty:
            st.markdown("**Ep output — row-oriented (one row per set):**")
            show_paginated_df(_sets_df_to_rows(ep_df_view, set_col="set"), key="ep_rows_view", use_container_width=True)

            try:
                import io as _ep_io
                _ep_buf = _ep_io.BytesIO()
                with pd.ExcelWriter(_ep_buf, engine="openpyxl") as _ep_xl:
                    _sets_df_to_rows(ep_df_view, set_col="set").to_excel(
                        _ep_xl, sheet_name="All", index=False)
                    from itertools import combinations as _ep_combos2
                    for _p in ["".join(p) for p in _ep_combos2(["a","b","c","d"], 2)]:
                        _pcols = [c for c in ep_df_view.columns if c.endswith("_" + _p)]
                        if _pcols:
                            _pb = ep_df_view[_pcols].copy()
                            _pb.columns = [c.split("_")[0] for c in _pcols]
                            _sets_df_to_rows(_pb, set_col="obj").to_excel(
                                _ep_xl, sheet_name=_p, index=False)
                _ep_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download Ep_{_gkey}.xlsx  (All + ab ac ad bc bd cd)",
                    data=_ep_buf.getvalue(),
                    file_name=f"Ep_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _ep_dl_ex:
                st.warning(f"Excel export unavailable: {_ep_dl_ex}")

    # ── TAB: Sp (Splits) ───────────────────────────────────────────────────
    with vtabs[4]:
        st.markdown("**Sp — Splits (task1b.py): top 4 w-columns of D + 4 split points**")

        d_df = S.get("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)

            top4 = w_cols_d[:4]
            st.write(f"Top 4 w-columns → Sp input: `{top4}`")

            sets_ready = {}
            for col in top4:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready[col] = vals

            if sets_ready:
                st.markdown("**Set split points** (0 = midpoint auto):")
                sp_cols = st.columns(4)
                splitters = []
                for i, (col, vals) in enumerate(sets_ready.items()):
                    with sp_cols[i]:
                        default_split = len(vals) // 2
                        sp = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_split,
                            key=f"sp_split_{col}_{i}")
                        splitters.append(sp if sp > 0 else default_split)

                if st.button("▶ Run Splits (Sp)", type="primary",
                             key="run_sp", use_container_width=True):
                    try:
                        from itertools import combinations as _sp_combos

                        keys = list(sets_ready.keys())
                        split_sets = {}
                        for j, key in enumerate(keys):
                            vals = sets_ready[key]
                            sp_val = splitters[j]
                            split_sets[key + "0"] = vals[:sp_val]
                            split_sets[key + "1"] = vals[sp_val:]

                        universe = set(n for v in sets_ready.values() for n in v)
                        comb3 = [c for r in range(3, 4)
                                 for c in _sp_combos(keys, r)]
                        comb2 = [c for c in _sp_combos(keys, 2)]

                        result_sp = {}
                        # Split sets
                        for i, key in enumerate(sorted(split_sets.keys())):
                            result_sp[key] = split_sets[key]
                            result_sp[f"y{i}"] = list(
                                universe - set(split_sets[key]))

                        # 3-combinations
                        let_3 = ["e", "f", "g", "h"]
                        comb3dict = {}
                        for j, combo in enumerate(comb3):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb3dict[let_3[j % 4] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb3dict.items():
                            result_sp[key] = list(val)

                        # 2-combinations
                        let_2 = ["i", "j", "k", "l", "m", "n"]
                        comb2dict = {}
                        for j, combo in enumerate(comb2):
                            iter_set = set(n for col in combo
                                           for n in sets_ready[col])
                            i = 0
                            for suffix in ["0", "1"]:
                                for col in combo:
                                    comb2dict[let_2[j % 6] + str(i)] = \
                                        iter_set - set(split_sets[col + suffix])
                                    i += 1
                        for key, val in comb2dict.items():
                            result_sp[key] = list(val)

                        sp_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_sp.items()})
                        order_sp = sp_df.isna().sum().sort_values().index
                        sp_df = sp_df[order_sp]
                        S["Sp"] = sp_df
                        sp_path = _gdirs["Splits"] / f"Sp_{_gkey}.csv"
                        sp_df.to_csv(sp_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ Sp generated: {sp_df.shape[1]} columns '
                            f'→ {sp_path.name}</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(sp_df, key="sp_generated_view", use_container_width=True)
                    except Exception as ex:
                        st.error(f"Splits error: {ex}")
            else:
                st.warning("Could not extract numeric data from D columns.")

        sp_view = S.get("Sp", pd.DataFrame())
        if not sp_view.empty:
            st.markdown("**Current Sp in memory (row-oriented — one row per set):**")
            show_paginated_df(_sets_df_to_rows(sp_view, set_col="set"), key="sp_current_memory", use_container_width=True)

            # ── Excel export: row-oriented, one sheet per group ──────────
            st.markdown("#### Download as Excel (one sheet per set group, row-oriented)")
            try:
                import io as _sp_io
                _sp_buf = _sp_io.BytesIO()
                with pd.ExcelWriter(_sp_buf, engine="openpyxl") as _sp_writer:
                    # Sheet "All" — full row-oriented table
                    _sets_df_to_rows(sp_view, set_col="set").to_excel(
                        _sp_writer, sheet_name="All", index=False)
                    # Group columns by prefix, then export each group row-oriented
                    _sp_groups = {
                        "Splits_w":      [c for c in sp_view.columns
                                          if re.match(r'^w\d+[01]$', str(c))],
                        "Complements_y": [c for c in sp_view.columns
                                          if re.match(r'^y\d+$', str(c))],
                        "Combo3_efgh":   [c for c in sp_view.columns
                                          if re.match(r'^[efgh]\d+$', str(c))],
                        "Combo2_ijklmn": [c for c in sp_view.columns
                                          if re.match(r'^[ijklmn]\d+$', str(c))],
                    }
                    for _sheet, _cols in _sp_groups.items():
                        if _cols:
                            _sets_df_to_rows(sp_view[_cols], set_col="set").to_excel(
                                _sp_writer, sheet_name=_sheet, index=False)
                _sp_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download Sp_{_gkey}.xlsx (All + 4 group sheets)",
                    data=_sp_buf.getvalue(),
                    file_name=f"Sp_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _sp_dl_ex:
                st.warning(f"Excel export unavailable: {_sp_dl_ex}")

    # ── TAB: So (SplitsCombi) ──────────────────────────────────────────────
    with vtabs[5]:
        st.markdown("**So — SplitsCombi (automation_vba.py): top 4 w-columns → union combis**")

        d_df = S.get("D", pd.DataFrame())
        if d_df.empty:
            st.warning("Load D variable first (Direct tab).")
        else:
            w_cols_d_so = sorted(
                [c for c in d_df.columns
                 if re.match(r'^w\d+$', c) or re.match(r'^n\d+$', c, re.I)],
                key=lambda x: d_df[x].dropna().shape[0], reverse=True)
            top4_so = w_cols_d_so[:4]
            st.write(f"Top 4 w-columns → So input: `{top4_so}`")

            sets_ready_so = {}
            for col in top4_so:
                vals = [int(float(v)) for v in d_df[col].dropna()
                        if str(v).replace(".","").isdigit() and float(v) >= 1]
                if vals:
                    sets_ready_so[col] = vals

            if sets_ready_so and len(sets_ready_so) == 4:
                st.markdown("**Set split points for So** (0 = midpoint auto):")
                so_cols = st.columns(4)
                splitters_so = []
                for i, (col, vals) in enumerate(sets_ready_so.items()):
                    with so_cols[i]:
                        default_s = len(vals) // 2
                        sv = st.number_input(
                            f"{col} (len={len(vals)})",
                            min_value=0, max_value=len(vals),
                            value=default_s,
                            key=f"so_split_{col}_{i}")
                        splitters_so.append(sv if sv > 0 else default_s)

                if st.button("▶ Run SplitsCombi (So)", type="primary",
                             key="run_so", use_container_width=True):
                    try:
                        from itertools import combinations as _so_combos
                        from collections import OrderedDict

                        keys_so = list(sets_ready_so.keys())
                        split_sets_so = {}
                        for j, key in enumerate(keys_so):
                            vals = sets_ready_so[key]
                            sv = splitters_so[j]
                            split_sets_so[key + "0"] = vals[:sv]
                            split_sets_so[key + "1"] = vals[sv:]

                        universe_so = set(
                            n for v in sets_ready_so.values() for n in v)

                        # Build pair/triple dicts; store actual tuples separately so
                        # membership check can use proper key extraction instead of
                        # fragile string-index tricks (which broke with multi-char
                        # keys like w1,w2,w3,w4).
                        _pair_tuples = list(_so_combos(list(split_sets_so.keys()), 2))
                        _triple_tuples = list(_so_combos(keys_so, 3))

                        comb_pairs = {}
                        _pair_map: dict = {}   # str(tuple) → actual tuple
                        for _pt in _pair_tuples:
                            _k = str(_pt)
                            comb_pairs[_k] = set(split_sets_so[_pt[0]]).union(
                                set(split_sets_so[_pt[1]]))
                            _pair_map[_k] = _pt

                        comb_three = {}
                        _triple_map: dict = {}
                        for _tt in _triple_tuples:
                            _k = str(_tt)
                            comb_three[_k] = set(sets_ready_so[_tt[0]]).union(
                                set(sets_ready_so[_tt[1]]),
                                set(sets_ready_so[_tt[2]]))
                            _triple_map[_k] = _tt

                        result_so: dict = {}
                        result_so["U"] = universe_so

                        for set_i, tset in comb_three.items():
                            triple_base_keys = set(_triple_map[set_i])
                            for set_j, pset in comb_pairs.items():
                                # Extract base key by stripping trailing "0"/"1" suffix.
                                # Works for any key length: 'w10'→'w1', 'a0'→'a'.
                                _pair_t = _pair_map[set_j]
                                base0 = _pair_t[0][:-1]
                                base1 = _pair_t[1][:-1]
                                if (pset.issubset(tset)
                                        and base0 in triple_base_keys
                                        and base1 in triple_base_keys):
                                    result_so[f"U-{set_i}-{set_j}"] = \
                                        universe_so - (tset - pset)
                                    result_so[f"{set_i}-{set_j}"] = tset - pset
                            result_so[str(set_i)] = tset

                        for set_j, pset in comb_pairs.items():
                            result_so[f"U-{set_j}"] = universe_so - pset
                            result_so[str(set_j)] = pset

                        so_df = pd.DataFrame(
                            {k: pd.Series(list(v)) for k, v in result_so.items()})
                        order_so = so_df.isna().sum().sort_values().index
                        so_df = so_df[order_so]
                        S["So"] = so_df
                        so_path = _gdirs["Splits_Combi"] / f"So_{_gkey}.csv"
                        so_df.to_csv(so_path, index=False)
                        st.markdown(
                            f'<div class="ok">✅ So generated: {so_df.shape[1]} columns '
                            f'→ {so_path.name}</div>',
                            unsafe_allow_html=True)
                        show_paginated_df(so_df, key="so_generated_view", use_container_width=True)
                    except Exception as ex:
                        st.error(f"SplitsCombi error: {ex}")
            elif sets_ready_so:
                st.info(f"So engine needs exactly 4 w-columns. Found {len(sets_ready_so)}.")
            else:
                st.warning("Could not extract numeric data from D columns.")

        so_view = S.get("So", pd.DataFrame())
        if not so_view.empty:
            st.markdown("**Current So in memory (row-oriented — one row per set):**")
            show_paginated_df(_sets_df_to_rows(so_view, set_col="set"), key="so_current_memory", use_container_width=True)

            # ── Excel export: row-oriented, one sheet per combination type
            st.markdown("#### Download as Excel (one sheet per combination type, row-oriented)")
            try:
                import io as _so_io
                _so_buf = _so_io.BytesIO()

                def _so_group(col: str) -> str:
                    """Classify a So column name into its combination category."""
                    c = str(col)
                    if c == "U":
                        return "Universe"
                    if c.startswith("U-") and c.count("(") >= 2:
                        return "U_minus_triple_pair"
                    if not c.startswith("U-") and c.count("(") >= 2:
                        return "Triple_minus_pair"
                    if c.count("(") == 1 and "," in c and c.count(",") >= 2:
                        return "Triples"
                    if c.startswith("U-"):
                        return "U_minus_pair"
                    return "Pairs"

                from collections import defaultdict as _sodd
                _so_groups: dict = _sodd(list)
                for _c in so_view.columns:
                    _so_groups[_so_group(_c)].append(_c)

                with pd.ExcelWriter(_so_buf, engine="openpyxl") as _so_writer:
                    # Sheet "All" — full row-oriented table
                    _sets_df_to_rows(so_view, set_col="set").to_excel(
                        _so_writer, sheet_name="All", index=False)
                    _sheet_order = ["Universe", "Triples", "Triple_minus_pair",
                                    "U_minus_triple_pair", "Pairs", "U_minus_pair"]
                    for _sh in _sheet_order:
                        _cols = _so_groups.get(_sh, [])
                        if _cols:
                            _sets_df_to_rows(so_view[_cols], set_col="set").to_excel(
                                _so_writer, sheet_name=_sh[:31], index=False)

                _so_buf.seek(0)
                st.download_button(
                    label=f"⬇ Download So_{_gkey}.xlsx (All + 6 type sheets)",
                    data=_so_buf.getvalue(),
                    file_name=f"So_{_gkey}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _so_dl_ex:
                st.warning(f"Excel export unavailable: {_so_dl_ex}")

    # ── TAB: Since Last ────────────────────────────────────────────────────
    with vtabs[6]:
        st.markdown(f"**Since Last — fetch from lottolyzer for {_gcfg['label']}**")
        _default_sl_url = _gcfg["lottolyzer"]
        st.markdown(
            '<div class="note">⚠️ If the link below opens the wrong game on lottolyzer '
            '(e.g. shows Set for Life instead of Saturday Lotto), paste the correct URL '
            'in the field below and click <b>Fetch now</b>.</div>',
            unsafe_allow_html=True)
        sl_url = st.text_input(
            "Lottolyzer URL (editable — fix if it points to the wrong game):",
            value=_default_sl_url,
            key="sl_url_override"
        )
        st.markdown(f"→ [Open this link in browser]({sl_url})")

        sl_file = _gdirs["SinceLast"] / "since_last.json"

        if st.button("⤓ Fetch now from lottolyzer", type="primary",
                     use_container_width=True, key="fetch_sl_now"):
            with st.spinner("Fetching from lottolyzer…"):
                d = fetch_since_last(sl_url, _gcfg["pool"])
            if d:
                save_since_last(d, _gkey, _gcfg["label"], _gcfg["pool"],
                                sl_url, sl_file)
                st.success(f"Fetched and cached {len(d)} numbers. Refresh to see the table.")
            else:
                st.error("Couldn't parse lottolyzer (unreachable or layout changed). "
                         "Use the manual upload below.")
        if sl_file.exists():
            try:
                sl_data = json.loads(sl_file.read_text())
                scraped_at = sl_data.get("scraped_at", "unknown")[:16]
                n_nums = len(sl_data.get("since_last_dict", {}))
                st.markdown(
                    f'<div class="ok">✅ Since Last cached — {n_nums} numbers '
                    f'| scraped: {scraped_at}</div>',
                    unsafe_allow_html=True)
                # Show table
                sl_dict = {int(k): int(v)
                           for k, v in sl_data.get("since_last_dict", {}).items()}
                all_wt = sl_data.get("all_wt", [])
                sl_display = pd.DataFrame([
                    {"Rank": i+1, "Number": num,
                     "Since_Last": sl_dict.get(num, "?"),
                     "Group": ("🔴 Last Draw" if sl_dict.get(num,99)==0
                               else "🟠 Very Recent" if sl_dict.get(num,99)<=3
                               else "🟡 Recent" if sl_dict.get(num,99)<=9
                               else "🟢 Moderate" if sl_dict.get(num,99)<=19
                               else "🔵 Old" if sl_dict.get(num,99)<=29
                               else "⚫ Very Old")}
                    for i, num in enumerate(all_wt)
                ])
                show_paginated_df(sl_display, key="sl_display_tbl", use_container_width=True)
            except Exception as ex:
                st.error(f"Error loading Since Last: {ex}")

        st.markdown("---")
        st.markdown("**Manual upload** (paste from lottolyzer CSV export):")
        up_sl = st.file_uploader(
            "Upload Since Last file (columns: Number, Since Last) — CSV, Excel, …",
            type=_UPLOAD_TYPES, key="up_sl")
        if up_sl:
            try:
                df_sl = _read_uploaded(up_sl)
                # Find Since Last column
                sl_col = next((c for c in df_sl.columns
                               if "since" in c.lower()), None)
                num_col = next((c for c in df_sl.columns
                                if "number" in c.lower() or c == "Number"), None)
                if sl_col and num_col:
                    sl_dict_up = {int(row[num_col]): int(row[sl_col])
                                  for _, row in df_sl.iterrows()
                                  if pd.notna(row[sl_col]) and pd.notna(row[num_col])}
                    pool = _gcfg["pool"]
                    all_wt_up = sorted(sl_dict_up.keys(),
                                       key=lambda n: (sl_dict_up[n], n))
                    # Same ordering as all_wt; kept as a separate list so the two
                    # keys can diverge later without surprises.
                    to_keep_up = list(all_wt_up)
                    save_data = {
                        "since_last_dict": {str(k): v
                                             for k, v in sl_dict_up.items()},
                        "all_wt": all_wt_up,
                        "to_keep": to_keep_up,
                        "game": _gkey,
                        "game_name": _gcfg["label"],
                        "pool_size": pool,
                        "scraped_at": datetime.now().isoformat(),
                        "url": sl_url,
                    }
                    _gdirs["SinceLast"].mkdir(parents=True, exist_ok=True)
                    sl_file.write_text(json.dumps(save_data, indent=2))
                    st.success(f"Since Last saved: {len(sl_dict_up)} numbers. "
                               f"Refresh page to see.")
                else:
                    st.error("Could not find 'Number' and 'Since Last' columns.")
            except Exception as ex:
                st.error(f"Upload error: {ex}")

        st.markdown("---")
        st.markdown("""
        **To get Since Last manually from lottolyzer:**
        1. Go to the link above
        2. Export or copy the frequency table
        3. Save as CSV with columns: `Number`, `Since Last`
        4. Upload above
        """)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER FORMULA
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📦 Container Formula":
    st.markdown('<span class="sec-hdr hdr-purple">📦 Container Formula — Live Collation</span>',
                unsafe_allow_html=True)
    am_toggle("cf")

    st.markdown("""
    <div class="info">
    Each row = 1 container. 17 rows = 17 containers running in parallel (Python multiprocessing).
    No Docker needed — Python processes handle parallel execution natively.
    Sp = S throughout. Collation: drop col-0 of each component, concatenate, relabel w1…wN.
    </div>
    """, unsafe_allow_html=True)

    # Variable status
    st.markdown("**Variables in memory:**")
    vc = st.columns(6)
    for i, k in enumerate(["B","R","D","Sp","So","Ep"]):
        with vc[i]:
            df = S.get(k, pd.DataFrame())
            if isinstance(df, pd.DataFrame) and not df.empty:
                st.markdown(f'<div class="ok">✅ {k} {len(df)}r×{len(df.columns)}c</div>',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="warn">⚠️ {k} empty</div>',
                            unsafe_allow_html=True)

    st.markdown("---")
    cf_df = pd.DataFrame({
        "#": [r[0] for r in CF_ROWS],
        "Name": [r[1] for r in CF_ROWS],
        "Type": [r[2] for r in CF_ROWS],
        "Components": [" + ".join(r[3]) for r in CF_ROWS],
        "Active": [S["cf_active"].get(r[1],True) for r in CF_ROWS],
    })
    edited_cf = st.data_editor(cf_df, key="cf_tbl",
        use_container_width=True, num_rows="fixed", height=530,
        column_config={
            "Active": st.column_config.CheckboxColumn("Active"),
            "Type":   st.column_config.SelectboxColumn("Type",
                          options=["Formed","Ready-made"]),
        })
    for _, row in edited_cf.iterrows():
        S["cf_active"][row["Name"]] = row["Active"]

    st.markdown("---")
    active_names = [r[1] for r in CF_ROWS if S["cf_active"].get(r[1],True)]
    chosen_f = st.selectbox("Collate formula:", active_names)
    comps = COMP_MAP.get(chosen_f, [])
    st.write(f"**Components:** {' + '.join(comps)}")
    # Resolve to base variables (strip trailing digits, de-dup) — must match
    # execute_collation, so D1D2D3 checks for D (not D1/D2/D3), B1B2B3 → B, etc.
    _seen, base_needed = set(), []
    for c in comps:
        b = re.sub(r"\d+$", "", str(c)).strip()
        if b and b not in _seen:
            _seen.add(b); base_needed.append(b)
    missing = [b for b in base_needed if b not in S or
               (isinstance(S.get(b), pd.DataFrame) and S[b].empty)]
    if missing:
        st.markdown(f'<div class="warn">⚠️ Missing/empty: {missing} '
                    f'(load these in Variable Inputs, or Build W-Matrix for D)</div>',
                    unsafe_allow_html=True)

    # ── JOIN DEMO: show how variables stack BEFORE running collation ─────
    with st.expander("🔍 Collation Join Preview — see exactly how variables stack", expanded=False):
        st.markdown(
            '<div class="info"><b>Rule:</b> B rows come first (all of them), then R rows '
            '(if loaded), then D rows underneath. Each block is separated by a divider. '
            'Numbers fill across columns w1, w2, … matching the widest row.</div>',
            unsafe_allow_html=True)

        _demo_pieces = []
        _demo_summary = []
        for _dv in base_needed:
            _df_dv = S.get(_dv)
            if _df_dv is None or (isinstance(_df_dv, pd.DataFrame) and _df_dv.empty):
                continue
            _block = _to_w_rows(_df_dv, is_direct=(_dv == "D"))
            if _block is None or _block.empty:
                continue
            _block = _block.copy()
            _block.columns = [f"w{i+1}" for i in range(len(_block.columns))]
            _block.insert(0, "Source", _dv)
            _demo_pieces.append(_block)
            _demo_summary.append({"Variable": _dv, "Rows": f"{len(_block):,}",
                                   "Widest pick": len(_block.columns) - 1})

        if _demo_summary:
            st.markdown("**Block sizes — in stacking order:**")
            _sum_df = pd.DataFrame(_demo_summary)
            _sum_df["Cumulative rows"] = _sum_df["Rows"].str.replace(",","").astype(int).cumsum().apply(lambda x: f"{x:,}")
            show_paginated_df(_sum_df, key="cf_demo_sum", use_container_width=True, hide_index=True)

            for _pi, _piece in enumerate(_demo_pieces):
                _var_name = _piece["Source"].iloc[0]
                _n_rows = len(_piece)
                _preview_n = min(8, _n_rows)
                st.markdown(
                    f'<div class="ok" style="margin-top:10px">▶ <b>{_var_name}</b> block — '
                    f'{_n_rows:,} row{"s" if _n_rows!=1 else ""} '
                    f'(showing first {_preview_n})</div>',
                    unsafe_allow_html=True)
                show_paginated_df(_piece, key=f"cf_piece_preview_{_pi}", use_container_width=True)
                if _pi < len(_demo_pieces) - 1:
                    _next_var = _demo_pieces[_pi+1]["Source"].iloc[0]
                    st.markdown(
                        f'<div class="note" style="text-align:center;font-size:.85rem">'
                        f'⬇ {_next_var} rows continue below this point</div>',
                        unsafe_allow_html=True)
        else:
            st.info("Load variables in Variable Inputs first to preview the join.")

    st.markdown("---")

    if st.button(f"▶ Collate {chosen_f}", type="primary", use_container_width=True):
        with st.spinner("Collating…"):
            result = execute_collation(comps)
        if result.empty:
            st.error("Result empty — load components in Variable Inputs first.")
        else:
            out = _gdirs["CVI"] / f"CVI_{chosen_f}.csv"
            result.to_csv(out, index=False)
            S["cvi"][chosen_f] = result
            n_wcols = sum(1 for c in result.columns if str(c).startswith("w"))
            st.markdown(f'<div class="ok">✅ {chosen_f}: {len(result):,} rows '
                        f'(w-sets) × {n_wcols} number columns (w1…w{n_wcols}) '
                        f'→ <code>{out.name}</code></div>',
                        unsafe_allow_html=True)

            # ── Stacked result — show each variable's block clearly ───────
            _res_sources = result["Source"].unique().tolist() if "Source" in result.columns else []
            if _res_sources:
                st.markdown("**Result breakdown by variable block:**")
                _blk_rows = []
                for _src in _res_sources:
                    _blk = result[result["Source"] == _src]
                    _blk_rows.append({"Variable": _src,
                                      "Rows": f"{len(_blk):,}",
                                      "Starts at row": f"{result[result['Source']==_src].index[0]+1:,}"})
                show_paginated_df(pd.DataFrame(_blk_rows), key="cf_blk_rows", use_container_width=True, hide_index=True)

            n_show_c = min(60, result.shape[1])
            st.caption(f"Result — {result.shape[0]:,} rows (full matrix saved to disk):")
            show_paginated_df(result.iloc[:, :n_show_c], key="cf_result_view", use_container_width=True)
            st.download_button(f"⬇ CVI_{chosen_f}.csv", to_csv_bytes(result),
                               f"CVI_{chosen_f}.csv","text/csv")

    if S["cvi"]:
        with st.expander("📋 All collated CVIs in memory"):
            for fname, df in S["cvi"].items():
                st.markdown(f"**{fname}** — {len(df)} rows × {len(df.columns)} cols")
                show_paginated_df(df, key=f"cvi_mem_expander_{fname}", use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: CONTAINER DASHBOARDS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🖥️ Container Dashboards":
    st.markdown('<span class="sec-hdr hdr-red">🖥️ Container Dashboards</span>',
                unsafe_allow_html=True)
    am_toggle("cd")

    # ── PARALLEL RUN — runs all active containers simultaneously ──────────
    st.markdown("### ▶ Run All Active Containers in Parallel")
    st.markdown("""
    <div class="info">
    Fires all CVI files found in <code>Container_Variable_Inputs/</code>
    simultaneously using Python multiprocessing.
    Each container gets its own worker process.
    Carry-forward defaults to all-U for parallel run.
    </div>
    """, unsafe_allow_html=True)

    # Cluster config
    clusters = load_clusters()
    next_id  = next_cluster_id(clusters)
    all_main_files = scan_main_data_files()
    all_cvi_files  = scan_cvi_files()

    p1,p2,p3,p4 = st.columns(4)
    with p1:
        st.metric("Next Cluster ID", next_id)
    with p2:
        cluster_label = st.text_input("Cluster label:", f"1n",
                                      key="par_cluster_label",
                                      help="e.g. 1n, Draw1567, May28")
    with p3:
        lotto_sel = st.selectbox("Lotto type:",
                                 list(LOTTO_TYPES.keys()),
                                 format_func=lambda x: f"{x} — {LOTTO_TYPES[x]}",
                                 key="par_lotto")
    with p4:
        draw_date_par = st.text_input("Draw date (YYYY_MM_DD):",
                                      datetime.now().strftime("%Y_%m_%d"),
                                      key="par_date")

    # Auto-scan Main Data
    matching_main = [f for f in all_main_files
                     if f.get("lotto") == lotto_sel or not f.get("lotto")]
    if matching_main:
        main_choice = st.selectbox(
            "Main Data file (auto-scanned from Main_Data/):",
            [f["raw"] for f in matching_main],
            format_func=lambda x: f"{x}  ({next(f['rows'] for f in matching_main if f['raw']==x):,} rows)",
            key="par_main_choice"
        )
        chosen_main_info = next(f for f in matching_main if f["raw"] == main_choice)
        draw_no_par = chosen_main_info.get("draw","D?")
    else:
        st.warning("No Main Data files found in Main_Data/. "
                   f"Expected pattern: {{cluster}}_{lotto_sel}_D{{draw_no}}.csv")
        main_choice  = None
        draw_no_par  = "D?"

    # Show available CVI files
    matching_cvi = [f for f in all_cvi_files
                    if f.get("lotto") == lotto_sel or not f.get("lotto")]
    st.markdown(f"**CVI files found for `{lotto_sel}`: {len(matching_cvi)}**")
    if matching_cvi:
        cvi_preview_df = pd.DataFrame([{
            "Formula": f["formula"], "Date": f["date"], "File": f["raw"]
        } for f in matching_cvi])
        show_paginated_df(cvi_preview_df, key="cd_cvi_preview", use_container_width=True, hide_index=True)
    else:
        st.warning("No CVI files found. Collate formulas in Container Formula first.")

    run_all_btn = st.button(
        f"▶ RUN ALL {len(matching_cvi)} CONTAINERS IN PARALLEL",
        type="primary", use_container_width=True,
        key="run_all_par",
        disabled=(not main_choice or not matching_cvi)
    )

    if run_all_btn and main_choice and matching_cvi:
        main_path  = _gdirs["Main_Data"] / main_choice
        output_dir = _gdirs["Outputs"] / f"Cluster_{next_id}_{cluster_label}"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Build worker args
        worker_args = []
        for cvi_info in matching_cvi:
            formula = cvi_info["formula"]
            sc_path = _gdirs["Selected_Counts"] / f"SC_{formula}.csv"
            worker_args.append((
                formula,
                cvi_info["path"],
                str(main_path),
                str(sc_path) if sc_path.exists() else "",
                next_id,
                cluster_label,
                lotto_sel,
                draw_no_par,
                draw_date_par,
                str(output_dir),
            ))

        st.info(f"Launching {len(worker_args)} workers…")
        prog = st.progress(0, text="Starting workers…")

        # Pool and cpu_count are imported at module top.
        n_workers = min(len(worker_args), cpu_count())
        st.caption(f"Using {n_workers} CPU cores of {cpu_count()} available")

        with Pool(processes=n_workers) as pool:
            results = []
            for i, res in enumerate(
                pool.imap_unordered(_parallel_worker, worker_args)
            ):
                results.append(res)
                prog.progress(
                    (i+1)/len(worker_args),
                    text=f"Completed {i+1}/{len(worker_args)}: "
                         f"{res['formula']} — "
                         f"{'✅' if res['status']=='complete' else '❌'} "
                         f"S:{res['selected_n']} U:{res['unselected_n']}"
                )

        prog.empty()

        # Save cluster to registry
        new_cluster = {
            "id":          next_id,
            "label":       cluster_label,
            "lotto":       lotto_sel,
            "draw_no":     draw_no_par,
            "draw_date":   draw_date_par,
            "main_file":   main_choice,
            "output_dir":  str(output_dir),
            "containers":  len(worker_args),
            "status":      "complete",
            "results":     results,
            "timestamp":   datetime.now().isoformat(),
        }
        clusters.append(new_cluster)
        save_clusters(clusters)

        # Show summary
        ok  = [r for r in results if r["status"] == "complete"]
        err = [r for r in results if r["status"] != "complete"]
        st.success(f"✅ Cluster {next_id} '{cluster_label}' complete — "
                   f"{len(ok)} succeeded, {len(err)} errors")
        res_df = pd.DataFrame([{
            "Formula":    r["formula"],
            "Status":     r["status"],
            "Selected":   r["selected_n"],
            "Unselected": r["unselected_n"],
            "Error":      r.get("error",""),
        } for r in results])
        show_paginated_df(res_df, key="cd_parallel_results", use_container_width=True, hide_index=True)
        if err:
            for r in err:
                st.error(f"❌ {r['formula']}: {r['error']}")

    st.markdown("---")

    # ── Dashboard stack ────────────────────────────────────────────────────
    st.markdown("**Individual dashboard (open one at a time):**")
    # Build list from all CVI files + hardcoded defaults
    cvi_found = [parse_cvi_filename(f.name)["formula"]
                 for f in sorted(_gdirs["CVI"].glob("CVI_*.csv"))]
    all_db_names = list(dict.fromkeys(
        [f"1n & {f}" for f in cvi_found] + DASHBOARDS
    ))
    for db_name in all_db_names:
        st.markdown(
            f'<div style="background:#1a1a2e;border:1px solid #444;color:#ccc;'
            f'font-size:.71rem;padding:2px 12px;margin-bottom:2px;'
            f'display:inline-block;min-width:260px;">'
            f'{db_name}…  Container Dashboard</div><br>',
            unsafe_allow_html=True)

    st.markdown("---")
    db = st.selectbox("Open Dashboard:", all_db_names,
                      format_func=lambda x: f"{x}… Container Dashboard")
    formula_name = db.replace("1n & ","")

    st.markdown(f'<div style="font-size:1rem;font-weight:900;border:2px solid #555;'
                f'padding:5px 14px;display:inline-block;margin-bottom:8px;">'
                f'{db}…  Container Dashboard</div>', unsafe_allow_html=True)

    # ── Auto-load CVI for this formula ────────────────────────────────────
    cvi_df = S["cvi"].get(formula_name)
    if cvi_df is None:
        # Try all CVI files for this formula (any lotto type, any date)
        for cvi_fp in sorted(_gdirs["CVI"].glob(f"CVI_*{formula_name}*.csv")):
            cvi_df = _load_file(cvi_fp)
            if not cvi_df.empty:
                S["cvi"][formula_name] = cvi_df
                break
        if cvi_df is None:
            for search_dir in [_gdirs["CVI"], _gdirs["Rainbow"]]:
                cvi_path = search_dir / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_df = _load_file(cvi_path)
                    break
    if cvi_df is None:
        cvi_df = pd.DataFrame()

    # ── Auto-load SC ───────────────────────────────────────────────────────
    sc_auto = {}
    sc_folder = _gdirs["Selected_Counts"]
    for sc_file in (sorted(sc_folder.glob(f"SC_{formula_name}*.csv")) +
                    sorted(sc_folder.glob(f"SC_{formula_name}*.xlsx"))):
        try:
            df_sc = (pd.read_excel(sc_file, engine="openpyxl")
                     if sc_file.suffix.lower() == ".xlsx"
                     else pd.read_csv(sc_file))
            if "w" in df_sc.columns and "Selected Count" in df_sc.columns:
                for _, row in df_sc.iterrows():
                    sc_auto[str(row["w"])] = str(row["Selected Count"])
        except Exception: pass

    # ── Auto-scan and load Main Data ──────────────────────────────────────
    main_df = S.get("main_data", pd.DataFrame())
    avail_main = scan_main_data_files()

    # ── Preview panel ──────────────────────────────────────────────────────
    st.markdown("#### Preview — Loaded Data")
    pv1, pv2, pv3 = st.columns(3)
    with pv1:
        if not cvi_df.empty:
            w_cols_prev = [c for c in cvi_df.columns
                           if re.match(r'^w\d+$', c)]
            st.success(f"✅ **CVI** — {len(w_cols_prev)} w-columns × "
                       f"{len(cvi_df)} rows")
            with st.expander("Preview CVI"):
                show_paginated_df(cvi_df, key="cd_cvi_preview_auto", use_container_width=True,
                             hide_index=True, height=220)
        else:
            st.warning("⚠️ CVI not loaded")

    with pv2:
        if sc_auto:
            st.success(f"✅ **SC** — {len(sc_auto)} w-columns")
            with st.expander("Preview SC"):
                show_paginated_df(pd.DataFrame([
                    {"w": k, "SC": v} for k,v in sc_auto.items()
                ]), key="cd_sc_preview", use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ SC not loaded")
            up_sc = st.file_uploader(
                "Upload SC (cols: w, Selected Count) — CSV, Excel, …",
                type=_UPLOAD_TYPES,
                key="dash_sc_upload", label_visibility="collapsed")
            if up_sc is not None:
                try:
                    df_up = _read_uploaded(up_sc)
                    if "w" in df_up.columns and "Selected Count" in df_up.columns:
                        sc_folder.mkdir(parents=True, exist_ok=True)
                        dest = sc_folder / f"SC_{formula_name}.csv"
                        df_up.to_csv(dest, index=False)
                        st.success(f"Saved → {dest.name}. Refresh to load.")
                    else:
                        st.error("Need columns: 'w' and 'Selected Count'.")
                except Exception as ex:
                    st.error(f"Upload error: {ex}")

    with pv3:
        if not main_df.empty:
            # Validate it's actually main data (has n-columns, not w-columns)
            n_check = [c for c in main_df.columns if re.match(r'^n\d+$', c, re.I)]
            w_check = [c for c in main_df.columns if re.match(r'^w\d+$', c)]
            if w_check and not n_check:
                st.error("⚠️ Loaded file looks like a CVI file (w-columns). "
                         "Please load correct Main Data below.")
                S["main_data"] = pd.DataFrame()
                main_df = pd.DataFrame()
            else:
                st.success(f"✅ **Main Data** — {len(main_df):,} rows × "
                           f"{len(main_df.columns)} cols")
            with st.expander("Preview Main Data"):
                if avail_main:
                    sel_main = st.selectbox(
                        "Switch Main Data file:",
                        [f["raw"] for f in avail_main],
                        key=f"switch_main_{db}"
                    )
                    if st.button("Load selected", key=f"load_main_{db}"):
                        chosen = next(
                            f for f in avail_main if f["raw"] == sel_main)
                        main_df = _load_file(Path(chosen["path"]))
                        S["main_data"] = main_df
                        S["main_data_path"] = chosen["path"]
                        st.success(f"Loaded {sel_main}")
                        st.rerun()
                show_paginated_df(main_df, key="cd_main_df_preview", use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ Main Data not loaded")
            if avail_main:
                sel_main2 = st.selectbox(
                    "Available Main Data files:",
                    [f["raw"] for f in avail_main],
                    key=f"sel_main2_{db}"
                )
                if st.button("Load", key=f"load_main2_{db}", type="primary"):
                    chosen2 = next(
                        f for f in avail_main if f["raw"] == sel_main2)
                    main_df = _load_file(Path(chosen2["path"]))
                    S["main_data"] = main_df
                    S["main_data_path"] = chosen2["path"]
                    st.rerun()

    st.markdown("---")

    # ── Data Status + Re-upload (always visible, always replaceable) ──────
    st.markdown("#### Data Status")
    col_s1, col_s2 = st.columns(2)

    # ── CVI ───────────────────────────────────────────────────────────────
    with col_s1:
        if not cvi_df.empty:
            st.success(f"✅ CVI loaded: {len(cvi_df.columns)} columns, "
                       f"{len(cvi_df)} rows")
            # Always show replace button even when loaded
            if st.button("🔄 Replace CVI", key=f"clr_cvi_{db}",
                         use_container_width=True):
                S["cvi"].pop(formula_name, None)
                # Remove from disk so it won't auto-reload
                cvi_path = _gdirs["CVI"] / f"CVI_{formula_name}.csv"
                if cvi_path.exists():
                    cvi_path.unlink()
                cvi_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning(f"⚠️ No CVI loaded for {formula_name}")

        # Upload always visible — key uses a counter so it resets after clear
        upload_key_cvi = f"up_cvi_{db}_{S.get('cvi_upload_v', {}).get(db, 0)}"
        up_cvi = st.file_uploader(
            "Upload CVI file (CSV, Excel, HTML, JSON, …):",
            type=_UPLOAD_TYPES, key=upload_key_cvi
        )
        if up_cvi:
            df_cvi_up = _read_uploaded(up_cvi)
            cvi_save = _gdirs["CVI"] / f"CVI_{formula_name}.csv"
            df_cvi_up.to_csv(cvi_save, index=False)
            S["cvi"][formula_name] = df_cvi_up
            # Bump version so key resets next time Replace is clicked
            if "cvi_upload_v" not in S: S["cvi_upload_v"] = {}
            S["cvi_upload_v"][db] = S["cvi_upload_v"].get(db, 0) + 1
            cvi_df = df_cvi_up
            st.success(f"✅ CVI saved → Container_Variable_Inputs/CVI_{formula_name}.csv")
            st.rerun()

    # ── Main Data ─────────────────────────────────────────────────────────
    with col_s2:
        if not main_df.empty:
            st.success(f"✅ Main Data: {len(main_df):,} rows × "
                       f"{len(main_df.columns)} cols")
            if st.button("🔄 Replace Main Data", key=f"clr_md_{db}",
                         use_container_width=True):
                S["main_data"] = pd.DataFrame()
                S["main_data_path"] = ""   # drop stale path so DuckDB can't misfire
                main_df = pd.DataFrame()
                st.rerun()
        else:
            st.warning("⚠️ No Main Data loaded")

        upload_key_md = f"up_md_{db}_{S.get('md_upload_v', {}).get(db, 0)}"
        up_md = st.file_uploader(
            "Upload Main Data file (CSV, Excel, HTML, JSON, …):",
            type=_UPLOAD_TYPES, key=upload_key_md
        )
        if up_md:
            df_md_up = _read_uploaded(up_md)
            for col in df_md_up.columns:
                df_md_up[col] = pd.to_numeric(df_md_up[col], errors="coerce")
            S["main_data"] = df_md_up
            # Uploaded in-memory data has no stable on-disk source; clear the
            # path so the DuckDB pre-pass guard won't match it to another file.
            S["main_data_path"] = ""
            if "md_upload_v" not in S: S["md_upload_v"] = {}
            S["md_upload_v"][db] = S["md_upload_v"].get(db, 0) + 1
            main_df = df_md_up
            st.success(f"✅ Main Data: {len(df_md_up):,} rows × "
                       f"{len(df_md_up.columns)} cols")
            st.rerun()

    st.markdown("---")
    # Selected Count
    st.markdown("**Selected Count input**")

    # Show if auto-loaded from file
    if sc_auto:
        st.success(f"✅ Selected Counts loaded from SC_{formula_name}*.csv")
        st.caption("  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))

    # SC upload — always replaceable
    with st.expander("📂 Upload / Replace Selected Count file"):
        st.caption("File must have columns: w, Selected Count")
        upload_key_sc = f"up_sc_{db}_{S.get('sc_upload_v', {}).get(db, 0)}"
        up_sc = st.file_uploader("Upload SC file (CSV, Excel, HTML, JSON, …)",
                                  type=_UPLOAD_TYPES, key=upload_key_sc)
        if up_sc:
            df_sc_up = _read_uploaded(up_sc)
            sp_sc = _gdirs["Selected_Counts"] / f"SC_{formula_name}.csv"
            df_sc_up.to_csv(sp_sc, index=False)
            if "sc_upload_v" not in S: S["sc_upload_v"] = {}
            S["sc_upload_v"][db] = S["sc_upload_v"].get(db, 0) + 1
            st.success(f"✅ Saved → SC_{formula_name}.csv")
            st.rerun()

    c1,c2 = st.columns(2)
    with c1:
        sc_method = st.radio("Method:", ["Custom","Same for all","Range"],
                             horizontal=True, key=f"scm_{db}")
    with c2:
        # Show per-column summary — do NOT merge into one string
        if sc_auto:
            st.caption("Per-column SC (from file): " +
                       "  |  ".join(f"{k}={v}" for k,v in sc_auto.items()))
        fallback_sc = st.text_input(
            "Fallback Count (used if no SC file loaded):", "7",
            key=f"scr_{db}")

    # Build sc_dict: per w-column, specific threshold
    # sc_auto keys look like "w1","w2",… or "1","2",… normalise to "w1","w2"
    sc_dict = {}
    for k, v in sc_auto.items():
        key = k if k.startswith("w") else f"w{k}"
        sc_dict[key] = [int(x.strip()) for x in str(v).split(",")
                        if x.strip().lstrip('-').isdigit()]

    # If no SC file loaded, use fallback for all w-columns
    if not sc_dict and fallback_sc:
        fallback_list = [int(x.strip()) for x in fallback_sc.split(",")
                         if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = fallback_list

    if sc_method == "Same for all" and fallback_sc:
        same_list = [int(x.strip()) for x in fallback_sc.split(",")
                     if x.strip().lstrip('-').isdigit()]
        w_keys = [c for c in (cvi_df.columns if not cvi_df.empty else [])
                  if re.match(r'^w\d+$', c)]
        for wk in w_keys:
            sc_dict[wk] = same_list

    # ── Carry-forward direction toggles ───────────────────────────────────
    st.markdown("---")
    st.markdown("**Carry-forward direction — which pool feeds the next row**")
    st.caption("U = Unselected carries forward (default)  |  "
               "S = Selected carries forward  |  "
               "Toggle at Row N controls what Row N receives from Row N-1.")

    # Initialise carry_fwd in session state per dashboard
    cf_key = f"carry_fwd_{db}"
    if cf_key not in S:
        S[cf_key] = {}

    # Get w-columns from loaded CVI or fallback
    w_keys_cf = sorted(
        [c for c in (cvi_df.columns if not cvi_df.empty else [])
         if re.match(r'^w\d+$', c)],
        key=lambda x: int(x[1:])
    )

    # ── Compact carry-forward control ─────────────────────────────────────
    # Global default + per-row overrides via a single editable table.
    # No per-column button grid — works cleanly with any number of w-columns.
    gd1, gd2, gd3 = st.columns([1.5, 1.5, 5])
    with gd1:
        global_dir = st.selectbox(
            "Default direction:",
            ["U — Unselected", "S — Selected"],
            key=f"gdir_{db}",
            help="Sets carry-forward direction for ALL rows at once."
        )
        _gd = "U" if global_dir.startswith("U") else "S"
    with gd2:
        if st.button("Apply to all", key=f"gdir_apply_{db}",
                     use_container_width=True):
            for wk in w_keys_cf:
                S[cf_key][wk] = _gd
            st.rerun()
    with gd3:
        st.caption(
            f"{len(w_keys_cf)} w-columns loaded. "
            "Edit the Direction column in the table below to override individual rows. "
            "U = Unselected carries forward · S = Selected carries forward."
        )

    if w_keys_cf:
        # Build override table: one row per w-column
        _cf_rows = [{"Row": f"Row {int(wk[1:])}  ·  {wk}",
                     "Direction": S[cf_key].get(wk, "U")}
                    for wk in w_keys_cf]
        _cf_edit_df = pd.DataFrame(_cf_rows)
        _cf_edited = st.data_editor(
            _cf_edit_df,
            key=f"cf_tbl_{db}",
            use_container_width=True,
            num_rows="fixed",
            height=min(40 * len(w_keys_cf) + 50, 340),
            column_config={
                "Row":       st.column_config.TextColumn("Row", disabled=True,
                                                          width="medium"),
                "Direction": st.column_config.SelectboxColumn(
                                 "Direction", options=["U", "S"], width="small"),
            },
            hide_index=True,
        )
        # Persist edits back into session state
        for _idx, _row in _cf_edited.iterrows():
            _wk = w_keys_cf[_idx]
            S[cf_key][_wk] = _row["Direction"]

    # Build final carry_fwd dict (defaults to U)
    carry_fwd = {wk: S[cf_key].get(wk, "U") for wk in w_keys_cf}

    # ── Is SC Available toggle ────────────────────────────────────────────
    sc_avail_key = f"sc_avail_{db}"
    if sc_avail_key not in S:
        S[sc_avail_key] = "YES"

    # ── Individual run cluster label ───────────────────────────────────────
    ind_cl1, ind_cl2, ind_cl3 = st.columns([2, 2, 4])
    with ind_cl1:
        ind_cluster = st.text_input(
            "Cluster label for this run:",
            value="1n",
            key=f"ind_cluster_{db}",
            help="Used in output filenames: e.g. 1n, Draw1567, Test01"
        )
    with ind_cl2:
        ind_lotto = st.selectbox(
            "Lotto type:",
            list(LOTTO_TYPES.keys()),
            format_func=lambda x: f"{x}",
            key=f"ind_lotto_{db}"
        )
    with ind_cl3:
        ind_draw = st.text_input(
            "Draw number:",
            value="D????",
            key=f"ind_draw_{db}",
            help="e.g. D1567"
        )

    # File prefix for individual run outputs
    ind_prefix = f"{ind_cluster}_{ind_lotto}_{ind_draw}_{formula_name}"
    st.markdown("---")
    sa1, sa2, _ = st.columns([1,1,6])
    with sa1:
        if st.button("🟢 SC: YES (Auto)", key=f"sca_y_{db}",
                     type="primary" if S[sc_avail_key]=="YES" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "YES"
    with sa2:
        if st.button("🔴 SC: NO (Manual)", key=f"sca_n_{db}",
                     type="primary" if S[sc_avail_key]=="NO" else "secondary",
                     use_container_width=True):
            S[sc_avail_key] = "NO"
    if S[sc_avail_key] == "NO":
        st.info("Manual mode: SC not provided upfront. "
                "Process pauses after each stage for researcher to provide SC.")
    else:
        st.caption("Automated mode: all stages run sequentially without stopping.")

    # RUN MATCHING
    st.markdown("---")
    can_run = not main_df.empty and not cvi_df.empty
    if not can_run:
        st.info("📤 Upload CVI and Main Data above to enable matching.")

    if st.button(f"▶ RUN MATCHING — {db}", type="primary",
                 use_container_width=True, key=f"run_{db}",
                 disabled=not can_run):
        n = len(main_df)
        st.info(f"Matching M={n:,} rows × {len(w_keys_cf)} w-columns…")
        with st.spinner("Matching engine running…"):
            # Pass the source path (if known). run_matching only uses it when
            # it is safe — large CSV, DuckDB present, row-count == M, contiguous
            # n-columns — and falls back to pandas otherwise.
            res = run_matching(main_df, cvi_df, sc_dict, carry_fwd,
                               main_path=S.get("main_data_path"))
        S["results"][db] = res

    # ── Display results (persists after run) ──────────────────────────────
    if db in S["results"]:
        res   = S["results"][db]
        fig9  = res["fig9_table"]
        sel   = res["selected"]
        unsel = res["unselected"]
        bd    = res["breakdown"]
        dbg   = res.get("debug_rows", [])
        small = res.get("small_enough", True)
        n_cols_res = res.get("n_cols", [])

        st.success(f"✅ Final stage — Selected: {len(sel):,}  ·  "
                   f"Unselected: {len(unsel):,}")

        # ── Helper: show df with count filter + highlights inline ──────
        def show_filtered_highlighted(df_in, cvi_set_in, n_cols_in,
                                      section_key, default_counts=None):
            if df_in is None or df_in.empty:
                st.info("No data.")
                return
            count_col = "Count" if "Count" in df_in.columns else None
            avail_counts = []
            if count_col:
                avail_counts = sorted(df_in[count_col].dropna()
                                      .astype(int).unique().tolist())
            if avail_counts:
                filt = st.multiselect(
                    "Filter by count:", avail_counts,
                    default=default_counts or avail_counts,
                    key=f"flt_{db}_{section_key}"
                )
                df_show = df_in[df_in[count_col].isin(filt)] if filt else df_in
            else:
                df_show = df_in
            if df_show.empty:
                st.info("No rows match the selected count filter.")
                return
            # Highlight matching cells
            nc = [c for c in n_cols_in if c in df_show.columns]
            if nc and cvi_set_in:
                def _style(val):
                    try:
                        if int(round(float(val))) in cvi_set_in:
                            return ("background-color:#d4a0a0;"
                                    "color:#000;font-weight:bold")
                    except (ValueError, TypeError):
                        pass
                    return ""
                def _cnt_style(val):
                    return "background-color:#ffffaa;color:#000;font-weight:bold"
                try:
                    styled = df_show.style.applymap(_style, subset=nc)
                    if count_col:
                        styled = styled.applymap(_cnt_style,
                                                 subset=[count_col])
                    show_paginated_df(styled, key="cd_formula_result_styled", use_container_width=True, hide_index=True)
                except Exception:
                    show_paginated_df(df_show, key="cd_formula_result_plain", use_container_width=True, hide_index=True)
            else:
                show_paginated_df(df_show, key="cd_formula_result_unstyled", use_container_width=True, hide_index=True)

        # ── Interactive Matching Table — per-cell popover detail ──────
        st.markdown("#### Matching Table")
        if not fig9.empty:
            _dbg_by_w = {d["w"]: d for d in dbg}

            # Column config: (header label, relative width)
            _MT = [
                ("Row",        1.5), ("M",          1.2), ("CVI",       1.0),
                ("Dir",        0.6), ("Main Ct",    2.5), ("Main Bkdn", 3.5),
                ("Present",    1.2), ("Pres Ct",    2.5), ("SC",        0.8),
                ("Selected",   1.2), ("Sel Bkdn",   3.5),
                ("Unsel Ct",   1.2), ("Unselected", 1.2), ("Unsel Bkdn",3.5),
            ]
            _widths = [c[1] for c in _MT]

            # Header row
            _hdrs = st.columns(_widths)
            for _hc, (_hn, _) in zip(_hdrs, _MT):
                _hc.markdown(f"<small><b>{_hn}</b></small>",
                             unsafe_allow_html=True)
            st.divider()

            def _trunc(s, n=16):
                s = str(s)
                return (s[:n] + "…") if len(s) > n else s

            for _ri, _fr in fig9.reset_index(drop=True).iterrows():
                _w       = _fr["CVI"]
                _d       = _dbg_by_w.get(_w, {})
                _cvi_set = _d.get("cvi_set", set())
                _sc_set  = set(_d.get("sc", []))
                _rc      = st.columns(_widths)

                # 0 · Row — plain
                _rc[0].write(_fr["Row"])

                # 1 · M:29 — popover → main data table
                _fr_main_data = _fr["Main\nData"]
                with _rc[1].popover(_fr_main_data,
                                    use_container_width=True):
                    st.markdown(f"**Main Data — {_fr_main_data}**")
                    _md_wc = _d.get("main_df_wc")
                    if _md_wc is not None and not _md_wc.empty:
                        show_paginated_df(_md_wc, key=f"pop_md_wc_{_w}_{_ri}", use_container_width=True)
                    else:
                        st.info("Full data not stored "
                                "(M > threshold or no CVI).")

                # 2 · CVI (w1…) — popover → CVI numbers list
                with _rc[2].popover(_fr["CVI"],
                                    use_container_width=True):
                    st.markdown(f"**CVI numbers for {_w}** "
                                f"({len(_cvi_set)} values):")
                    st.write(sorted(_cvi_set) if _cvi_set else "— None —")

                # 3 · Dir — plain
                _rc[3].write(_fr["Dir"])

                # 4 · Main Count — popover → full count string
                with _rc[4].popover(_trunc(_fr["Main\nCount"]),
                                    use_container_width=True):
                    st.markdown("**Main Count distribution:**")
                    st.write(str(_fr["Main\nCount"]))

                # 5 · Main Breakdown — popover → breakdown + dist table
                with _rc[5].popover(_trunc(_fr["Main\nBreakdown"]),
                                    use_container_width=True):
                    st.markdown("**Main Breakdown:**")
                    st.write(str(_fr["Main\nBreakdown"]))
                    _cd = _d.get("count_dist", {})
                    if _cd:
                        show_paginated_df(pd.DataFrame([
                            {"Count": int(k[1:]), "Rows": v,
                             "Status": "✅ Sel" if int(k[1:]) in _sc_set
                                       else "— Unsel"}
                            for k, v in sorted(_cd.items())
                        ]), key=f"pop_cd_{_w}_{_ri}", hide_index=True, use_container_width=True)

                # 6 · Present Data — popover → row count entering stage
                _fr_present_data = _fr["Present\nData"]
                with _rc[6].popover(_fr_present_data,
                                    use_container_width=True):
                    st.markdown(f"**Present at this stage:** "
                                f"{_fr_present_data}")
                    st.write(f"Rows entering: "
                             f"{_d.get('present_in', '—')}")

                # 7 · Present Count — popover → count breakdown string
                with _rc[7].popover(_trunc(_fr["Present\nCount"]),
                                    use_container_width=True):
                    st.markdown("**Present Count distribution:**")
                    st.write(_d.get("pres_count_str",
                                    str(_fr["Present\nCount"])))

                # 8 · SC — plain
                _rc[8].write(_fr["SC"])

                # 9 · Selected — popover → sel_df table + CSV download
                with _rc[9].popover(_fr["Selected"],
                                    use_container_width=True):
                    st.markdown(f"**Selected rows — {_fr['Selected']}**")
                    _sel_df_p = _d.get("sel_df")
                    if _sel_df_p is not None and not _sel_df_p.empty:
                        show_paginated_df(_sel_df_p, key=f"pop_sel_{db}_{_ri}_df", use_container_width=True)
                        st.download_button(
                            f"⬇ {_w} Selected CSV",
                            to_csv_bytes(_sel_df_p),
                            f"{ind_prefix}_{_w}_sel.csv",
                            "text/csv",
                            key=f"pop_sel_{db}_{_ri}",
                        )
                    elif _d.get("selected_n", 0) == 0:
                        st.info("No rows selected at this stage.")
                    else:
                        st.info(f"S:{_d['selected_n']} — not stored "
                                f"(M > {DISPLAY_THRESHOLD:,})")

                # 10 · Sel Breakdown — popover → breakdown text
                with _rc[10].popover(_trunc(_fr["Sel\nBreakdown"]),
                                     use_container_width=True):
                    st.markdown("**Selected Breakdown:**")
                    st.write(str(_fr["Sel\nBreakdown"]))

                # 11 · Unsel Count — popover → unsel count string
                with _rc[11].popover(_trunc(_fr["Unsel\nCount"]),
                                     use_container_width=True):
                    st.markdown("**Unselected Count distribution:**")
                    st.write(str(_fr["Unsel\nCount"]))

                # 12 · Unselected — popover → unsel_df table + CSV download
                with _rc[12].popover(_fr["Unselected"],
                                     use_container_width=True):
                    st.markdown(f"**Unselected rows — {_fr['Unselected']}**")
                    _unsel_df_p = _d.get("unsel_df")
                    if _unsel_df_p is not None and not _unsel_df_p.empty:
                        show_paginated_df(_unsel_df_p, key=f"pop_unsel_{db}_{_ri}_df", use_container_width=True)
                        st.download_button(
                            f"⬇ {_w} Unselected CSV",
                            to_csv_bytes(_unsel_df_p),
                            f"{ind_prefix}_{_w}_unsel.csv",
                            "text/csv",
                            key=f"pop_unsel_{db}_{_ri}",
                        )
                    elif _d.get("unselected_n", 0) == 0:
                        st.info("No unselected rows remaining.")
                    else:
                        st.info(f"U:{_d['unselected_n']} — not stored "
                                f"(M > {DISPLAY_THRESHOLD:,})")

                # 13 · Unsel Breakdown — popover → breakdown text
                with _rc[13].popover(_trunc(_fr["Unsel\nBreakdown"]),
                                     use_container_width=True):
                    st.markdown("**Unselected Breakdown:**")
                    st.write(str(_fr["Unsel\nBreakdown"]))

            # Download the full table as CSV
            st.download_button("⬇ Matching Table CSV",
                               to_csv_bytes(fig9),
                               f"{ind_prefix}_matching_table.csv",
                               "text/csv", key=f"dl_fig9_{db}")

        # ── Per-row summary table + single-row inspector ──────────────
        # Show all rows as a compact summary table first, then let the
        # user pick ONE row to drill into — avoids rendering hundreds of
        # expanders at once.
        st.markdown("#### Row-by-Row Summary")
        _summary_rows = []
        for _d in dbg:
            _icon = "✅" if _d["selected_n"] > 0 else (
                    "⚪" if _d["present_in"] == 0 else "🔵")
            _note = f"  ⚠️{_d.get('note','')}" if _d.get("note") else ""
            _summary_rows.append({
                "":        _icon,
                "Row":     int(_d["w"][1:]),
                "w":       _d["w"],
                "Dir":     _d.get("direction", "U"),
                "Present": _d["present_in"],
                "SC":      str(list(_d["sc"])),
                "S":       _d["selected_n"],
                "U":       _d["unselected_n"],
                "Note":    _note.strip(),
            })
        if _summary_rows:
            _sum_df = pd.DataFrame(_summary_rows)
            show_paginated_df(_sum_df, key=f"cd_row_summary_{db}", use_container_width=True, hide_index=True)

            # ── One-click export of the full summary ───────────────────
            _bdl1, _bdl2 = st.columns(2)
            with _bdl1:
                st.download_button(
                    "⬇ Download Row Summary CSV",
                    to_csv_bytes(_sum_df.drop(columns=[""])),
                    f"{ind_prefix}_row_summary.csv",
                    "text/csv",
                    key=f"dl_sum_csv_{db}",
                    use_container_width=True,
                )
            with _bdl2:
                try:
                    import io as _sum_io
                    _sum_buf = _sum_io.BytesIO()
                    with pd.ExcelWriter(_sum_buf, engine="openpyxl") as _sum_xl:
                        _export_df = _sum_df.drop(columns=[""]).copy()
                        _export_df.to_excel(_sum_xl, sheet_name="Row_Summary",
                                            index=False)
                        # Conditional formatting: green S>0, grey S=0
                        from openpyxl.styles import PatternFill as _PF, Font as _Fnt
                        _ws = _sum_xl.sheets["Row_Summary"]
                        _green = _PF("solid", fgColor="C6EFCE")
                        _red   = _PF("solid", fgColor="FFC7CE")
                        _s_col = list(_export_df.columns).index("S") + 1
                        _u_col = list(_export_df.columns).index("U") + 1
                        for _r in range(2, len(_export_df) + 2):
                            _sc = _ws.cell(row=_r, column=_s_col)
                            _uc = _ws.cell(row=_r, column=_u_col)
                            try:
                                if int(_sc.value or 0) > 0:
                                    _sc.fill = _green
                                else:
                                    _sc.fill = _red
                            except (ValueError, TypeError):
                                pass
                    _sum_buf.seek(0)
                    st.download_button(
                        "⬇ Download Row Summary Excel (S highlighted)",
                        _sum_buf.getvalue(),
                        f"{ind_prefix}_row_summary.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_sum_xl_{db}",
                        use_container_width=True,
                    )
                except Exception as _sum_xl_ex:
                    st.warning(f"Excel export unavailable: {_sum_xl_ex}")

        # ── Single-row inspector ───────────────────────────────────────
        st.markdown("#### Inspect a Row")
        _row_labels = [
            f"Row {int(d['w'][1:])}  ·  {d['w']}  ·  "
            f"Dir:{d.get('direction','U')}  ·  Present:{d['present_in']}  ·  "
            f"S:{d['selected_n']}  ·  U:{d['unselected_n']}"
            for d in dbg
        ]
        if _row_labels:
            _sel_row_label = st.selectbox(
                "Select row to inspect:",
                _row_labels,
                key=f"row_inspect_{db}"
            )
            _sel_i = _row_labels.index(_sel_row_label)
            d = dbg[_sel_i]
            i = _sel_i
            w_lbl     = d["w"]
            cvi_set_d = d.get("cvi_set", set(d.get("cvi_numbers", [])))
            n_cols_d  = d.get("n_cols", [])

            if not cvi_set_d:
                st.warning(f"{w_lbl}: No CVI numbers — column is empty. "
                           "All present rows carry forward.")
            else:
                # CVI numbers + distribution
                cca, ccb = st.columns(2)
                with cca:
                    st.markdown(f"**{w_lbl} numbers ({len(cvi_set_d)}):**")
                    st.write(sorted(cvi_set_d))
                with ccb:
                    st.markdown("**Count distribution (present rows):**")
                    if d["count_dist"]:
                        dist_df = pd.DataFrame([
                            {"Count": int(k[1:]),
                             "Rows": v,
                             "": "✅ Sel" if int(k[1:]) in d["sc"]
                                 else "— Unsel"}
                            for k, v in sorted(d["count_dist"].items())
                        ])
                        show_paginated_df(dist_df, key=f"cd_dist_df_{_w}_{_ri}", hide_index=True, use_container_width=True)
                    else:
                        st.info("—")

                st.markdown("---")

                # Main Data breakdown (collapsible)
                with st.expander(
                    f"📊 Main Data Breakdown  ·  M:{len(main_df)}"
                    f"  ·  {d.get('main_bd_str','—')}"
                ):
                    md_wc = d.get("main_df_wc")
                    if md_wc is not None and not md_wc.empty:
                        show_filtered_highlighted(
                            md_wc, cvi_set_d, n_cols_d, f"main_{i}")
                        c1e, c2e = st.columns(2)
                        with c1e:
                            st.download_button(
                                "⬇ Main Breakdown CSV",
                                to_csv_bytes(md_wc),
                                f"{ind_prefix}_{w_lbl}_main.csv",
                                "text/csv", key=f"dl_main_{db}_{i}")
                        with c2e:
                            st.download_button(
                                "⬇ Main Breakdown Excel (highlighted)",
                                to_styled_excel(md_wc, cvi_set_d, n_cols_d,
                                                f"{w_lbl}_Main"),
                                f"{ind_prefix}_{w_lbl}_main.xlsx",
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                                key=f"dl_main_xl_{db}_{i}")
                    else:
                        st.info("Main data breakdown not available "
                                f"(M>{DISPLAY_THRESHOLD:,} or no CVI).")

                st.markdown("---")

                # Selected + Unselected side by side (collapsible)
                ts, tu = st.columns(2)
                with ts:
                    d_sel = d.get("sel_df")
                    with st.expander(
                        f"✅ Selected  S:{d['selected_n']}", expanded=True
                    ):
                        if d_sel is not None and not d_sel.empty:
                            show_filtered_highlighted(
                                d_sel, cvi_set_d, n_cols_d,
                                f"sel_{i}",
                                default_counts=list(d["sc"]))
                            c1s, c2s = st.columns(2)
                            with c1s:
                                st.download_button(
                                    f"⬇ {w_lbl} Selected CSV",
                                    to_csv_bytes(d_sel),
                                    f"{ind_prefix}_{w_lbl}_sel.csv",
                                    "text/csv", key=f"dl_sel_{db}_{i}")
                            with c2s:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(d_sel, cvi_set_d,
                                                    n_cols_d, f"{w_lbl}_Sel"),
                                    f"{ind_prefix}_{w_lbl}_sel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_sel_xl_{db}_{i}")
                        elif d["selected_n"] == 0:
                            st.info("No rows selected at this stage.")
                        else:
                            st.info(f"S:{d['selected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

                with tu:
                    d_unsel = d.get("unsel_df")
                    with st.expander(
                        f"🔵 Unselected  U:{d['unselected_n']} → fwd",
                        expanded=True
                    ):
                        if d_unsel is not None and not d_unsel.empty:
                            show_filtered_highlighted(
                                d_unsel, cvi_set_d, n_cols_d, f"unsel_{i}")
                            c1u, c2u = st.columns(2)
                            with c1u:
                                st.download_button(
                                    f"⬇ {w_lbl} Unselected CSV",
                                    to_csv_bytes(d_unsel),
                                    f"{ind_prefix}_{w_lbl}_unsel.csv",
                                    "text/csv", key=f"dl_unsel_{db}_{i}")
                            with c2u:
                                st.download_button(
                                    "⬇ Excel (highlighted)",
                                    to_styled_excel(d_unsel, cvi_set_d,
                                                    n_cols_d, f"{w_lbl}_Unsel"),
                                    f"{ind_prefix}_{w_lbl}_unsel.xlsx",
                                    "application/vnd.openxmlformats-"
                                    "officedocument.spreadsheetml.sheet",
                                    key=f"dl_unsel_xl_{db}_{i}")
                        elif d["unselected_n"] == 0:
                            st.info("No rows remaining.")
                        else:
                            st.info(f"U:{d['unselected_n']} — not stored "
                                    f"(M>{DISPLAY_THRESHOLD:,})")

        # ── Final stage permanent tabs ─────────────────────────────────
        st.markdown("---")
        st.markdown("#### Final Stage Output")

        # Get final w CVI set
        final_cvi_set = set()
        if dbg:
            last_d = next((d for d in reversed(dbg)
                           if d.get("cvi_set")), None)
            if last_d:
                final_cvi_set = last_d.get("cvi_set", set())

        ft1, ft2, ft3 = st.tabs([
            f"Selected (final)  S:{len(sel)}",
            f"Unselected (final)  U:{len(unsel)}",
            "Breakdown S0/S1/S2… (all stages)",
        ])

        with ft1:
            if sel.empty:
                st.info("No rows selected at final stage.")
            else:
                n_f = [c for c in n_cols_res if c in sel.columns]
                show_filtered_highlighted(
                    sel, final_cvi_set, n_f, "final_sel",
                    default_counts=None)
                fpath = _gdirs["Outputs"] / f"{ind_prefix}_selected.csv"
                sel.to_csv(fpath, index=False)
                c1f, c2f = st.columns(2)
                with c1f:
                    st.download_button(
                        "⬇ Selected CSV", to_csv_bytes(sel),
                        f"{ind_prefix}_selected.csv","text/csv",
                        key=f"dl_s_{db}")
                with c2f:
                    st.download_button(
                        "⬇ Selected Excel (highlighted)",
                        to_styled_excel(sel, final_cvi_set, n_f, "Selected"),
                        f"{ind_prefix}_selected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_s_xl_{db}")

        with ft2:
            if unsel.empty:
                st.info("No rows remaining at final stage.")
            else:
                n_u = [c for c in n_cols_res if c in unsel.columns]
                show_filtered_highlighted(
                    unsel, final_cvi_set, n_u, "final_unsel")
                c1u2, c2u2 = st.columns(2)
                with c1u2:
                    st.download_button(
                        "⬇ Unselected CSV", to_csv_bytes(unsel),
                        f"{ind_prefix}_unselected.csv","text/csv",
                        key=f"dl_u_{db}")
                with c2u2:
                    st.download_button(
                        "⬇ Unselected Excel (highlighted)",
                        to_styled_excel(unsel, final_cvi_set, n_u,
                                        "Unselected"),
                        f"{ind_prefix}_unselected.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        key=f"dl_u_xl_{db}")

        with ft3:
            st.caption("S0=0 matches · S1=1 match … per w-column")
            if bd.empty:
                st.info("No breakdown data.")
            else:
                show_paginated_df(bd, key=f"cd_breakdown_{db}", use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇ Breakdown CSV", to_csv_bytes(bd),
                    f"{ind_prefix}_breakdown.csv","text/csv",
                    key=f"dl_bd_{db}")

    # Controls
    st.markdown("---")
    if db not in S["container_status"]:
        S["container_status"][db] = pd.DataFrame({
            "Name":[db],"Status":["Stopped"],"Memory Usage":[""],
            "CPU Usage":[""],"Percent progress %":[0],
            "Time Guestimate Min":[""],"Time Guestimate Max":[""],
        })

    b1,b2,b3,b4,b5,b6,b7 = st.columns(7)
    with b1:
        if st.button("▶ Start", type="primary", use_container_width=True, key=f"s1_{db}"):
            set_container_status(db,"Running"); st.toast("Started!")
    with b2:
        if st.button("■ Stop", use_container_width=True, key=f"s2_{db}"):
            set_container_status(db,"Stopped")
    with b3:
        if st.button("⚡ Kill", use_container_width=True, key=f"s3_{db}"):
            set_container_status(db,"Killed")
    with b4:
        if st.button("↺ Restart", use_container_width=True, key=f"s4_{db}"):
            st.toast("Restarting…")
    with b5:
        if st.button("⏸ Pause", use_container_width=True, key=f"s5_{db}"):
            set_container_status(db,"Paused")
    with b6:
        if st.button("▶ Resume", use_container_width=True, key=f"s6_{db}"):
            set_container_status(db,"Running")
    with b7:
        if st.button("🗑 Remove", use_container_width=True, key=f"s7_{db}"):
            st.toast("Removed.")

    S["container_status"][db] = st.data_editor(
        S["container_status"][db], key=f"stbl_{db}",
        use_container_width=True, num_rows="dynamic",
        column_config={
            "Status": st.column_config.SelectboxColumn(
                "Status", options=["Stopped","Running","Paused","Killed","Error"]),
            "Percent progress %": st.column_config.ProgressColumn(
                "Percent progress %", min_value=0, max_value=100),
        })

    c1,c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇ {db} Status CSV",
            to_csv_bytes(S["container_status"][db]),
            f"{db.replace(' ','_')}_status.csv","text/csv",key=f"dls_{db}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MASTER OUTPUTS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Master Outputs":
    st.markdown('<span class="sec-hdr hdr-brown">📤 Master Outputs — Aggregated Results</span>',
                unsafe_allow_html=True)
    am_toggle("out")

    # Collect from all dashboards
    all_sel, all_unsel, all_bd = [], [], []
    for db in DASHBOARDS:
        res = S["results"].get(db)
        if res:
            if not res["selected"].empty:
                s = res["selected"].copy(); s.insert(0,"Dashboard",db)
                all_sel.append(s)
            if not res["unselected"].empty:
                u = res["unselected"].copy(); u.insert(0,"Dashboard",db)
                all_unsel.append(u)
            if not res["breakdown"].empty:
                b = res["breakdown"].copy(); b.insert(0,"Dashboard",db)
                all_bd.append(b)

    n_done = len(all_sel)
    if n_done:
        st.markdown(f'<div class="ok">✅ Results from {n_done} dashboard(s).</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="warn">⚠️ No results yet. Run matching in Container Dashboards.</div>',
                    unsafe_allow_html=True)

    def cc(dfs):
        if not dfs: return pd.DataFrame()
        return pd.concat(dfs, ignore_index=True).drop_duplicates()

    def show_output_tab(df: pd.DataFrame, label: str, fname: str):
        if df.empty:
            st.info("No data yet.")
            return
        st.write(f"**{len(df):,} unique rows**")
        show_paginated_df(df, key=f"master_out_{fname}", use_container_width=True)
        df.to_csv(_gdirs["Outputs"]/fname, index=False)
        st.download_button(f"⬇ Export", to_csv_bytes(df), fname,"text/csv",
                           key=f"dl_out_{fname}")

    t1,t2,t3,t4,t5 = st.tabs([
        "Cluster List Combined Selected",
        "Cluster Combined Selected",
        "Cluster List Combined Unselected",
        "Cluster Combined Unselected",
        "Selected Breakdown S0/S1/S2…",
    ])

    with t1:
        show_output_tab(cc(all_sel), "Cluster List Combined Selected",
                        "cluster_list_combined_selected.csv")
    with t2:
        df_t2 = cc(all_sel)
        nc = [c for c in df_t2.columns if re.match(r'^n\d+$',c)] if not df_t2.empty else []
        show_output_tab(df_t2[nc].drop_duplicates() if nc else pd.DataFrame(),
                        "Cluster Combined Selected", "cluster_combined_selected.csv")
    with t3:
        show_output_tab(cc(all_unsel), "Cluster List Combined Unselected",
                        "cluster_list_combined_unselected.csv")
    with t4:
        df_t4 = cc(all_unsel)
        nc4 = [c for c in df_t4.columns if re.match(r'^n\d+$',c)] if not df_t4.empty else []
        show_output_tab(df_t4[nc4].drop_duplicates() if nc4 else pd.DataFrame(),
                        "Cluster Combined Unselected", "cluster_combined_unselected.csv")
    with t5:
        show_output_tab(cc(all_bd), "Selected Breakdown",
                        "selected_breakdown.csv")

# ════════════════════════════════════════════════════════════════════════════
# PAGE: CLUSTER MANAGER
# ════════════════════════════════════════════════════════════════════════════
elif page == "🗂️ Cluster Manager":
    st.markdown('<span class="sec-hdr hdr-purple">🗂️ Cluster Manager</span>',
                unsafe_allow_html=True)

    st.markdown("""
    <div class="info">
    Each cluster = one complete run of all active containers against one Main Data file.
    <b>1n</b> = current run · <b>1n+1</b> = next run · <b>1n-1</b> = previous run.
    Labels are user-assigned. System auto-assigns numeric ID (001, 002…).
    </div>
    """, unsafe_allow_html=True)

    clusters = load_clusters()

    if not clusters:
        st.info("No clusters run yet. Use the parallel run button on Container Dashboards.")
    else:
        # ── Cluster registry table ─────────────────────────────────────
        st.markdown("### All Clusters")
        reg_df = pd.DataFrame([{
            "ID":         c["id"],
            "Label":      c["label"],
            "Lotto":      c.get("lotto",""),
            "Draw":       c.get("draw_no",""),
            "Date":       c.get("draw_date",""),
            "Main Data":  c.get("main_file",""),
            "Containers": c.get("containers",0),
            "Status":     c.get("status",""),
            "Run at":     c.get("timestamp","")[:16].replace("T"," "),
        } for c in clusters])
        reg_df.index = range(1, len(reg_df)+1)
        show_paginated_df(reg_df, key="cm_reg_df", use_container_width=True)

        # ── Select cluster to inspect ──────────────────────────────────
        st.markdown("---")
        st.markdown("### Inspect a Cluster")
        cluster_ids = [f"{c['id']} — {c['label']}" for c in clusters]
        chosen_c = st.selectbox("Select cluster:", cluster_ids,
                                key="cm_chosen")
        chosen_idx = cluster_ids.index(chosen_c)
        c_data = clusters[chosen_idx]

        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Cluster ID",   c_data["id"])
        c2.metric("Label",        c_data["label"])
        c3.metric("Lotto",        c_data.get("lotto",""))
        c4.metric("Containers",   c_data.get("containers",0))

        # ── Per-container results for this cluster ─────────────────────
        results = c_data.get("results", [])
        if results:
            st.markdown("**Container results:**")
            res_df = pd.DataFrame([{
                "Formula":    r["formula"],
                "Status":     "✅" if r["status"]=="complete" else "❌",
                "Selected":   r["selected_n"],
                "Unselected": r["unselected_n"],
                "Error":      r.get("error",""),
            } for r in results])
            res_df.index = range(1, len(res_df)+1)
            show_paginated_df(res_df, key=f"cm_res_df_{chosen_c}", use_container_width=True, hide_index=False)

        # ── Load and browse cluster output files ───────────────────────
        out_dir = Path(c_data.get("output_dir", ""))
        if out_dir.exists():
            st.markdown("---")
            st.markdown("### Browse Cluster Output Files")
            out_files = sorted(out_dir.glob("*.csv"))
            if out_files:
                chosen_file = st.selectbox(
                    "File:", [f.name for f in out_files],
                    key="cm_file"
                )
                fp = out_dir / chosen_file
                df_view = pd.read_csv(fp)
                st.write(f"**{len(df_view):,} rows · {len(df_view.columns)} cols**")
                show_paginated_df(df_view, key=f"cm_file_view_{chosen_file}", use_container_width=True, hide_index=True)
                st.download_button(f"⬇ {chosen_file}",
                                   df_view.to_csv(index=False).encode(),
                                   chosen_file,"text/csv",
                                   key=f"dl_cm_{chosen_file}")
            else:
                st.info("No output files found in this cluster's output folder.")

        # ── Cross-cluster comparison ───────────────────────────────────
        if len(clusters) >= 2:
            st.markdown("---")
            st.markdown("### Cross-cluster Comparison")
            st.caption("Compare selected/unselected counts across clusters "
                       "for the same formula.")

            all_formulas = sorted(set(
                r["formula"]
                for c in clusters for r in c.get("results",[])
            ))
            if all_formulas:
                comp_formula = st.selectbox("Formula to compare:",
                                            all_formulas, key="cm_comp")
                comp_rows = []
                for c in clusters:
                    res = next((r for r in c.get("results",[])
                                if r["formula"] == comp_formula), None)
                    if res:
                        comp_rows.append({
                            "Cluster ID":  c["id"],
                            "Label":       c["label"],
                            "Lotto":       c.get("lotto",""),
                            "Draw":        c.get("draw_no",""),
                            "Date":        c.get("draw_date",""),
                            "Selected":    res["selected_n"],
                            "Unselected":  res["unselected_n"],
                            "Status":      res["status"],
                        })
                if comp_rows:
                    comp_df = pd.DataFrame(comp_rows)
                    comp_df.index = range(1, len(comp_df)+1)
                    show_paginated_df(comp_df, key="cm_comp_df", use_container_width=True)
                    st.download_button(
                        f"⬇ Cross-cluster comparison CSV",
                        comp_df.to_csv(index=False).encode(),
                        f"cross_cluster_{comp_formula}.csv","text/csv",
                        key="dl_cm_comp"
                    )

        # ── Rename / delete cluster ────────────────────────────────────
        st.markdown("---")
        st.markdown("### Manage This Cluster")
        new_label = st.text_input("Rename cluster label:",
                                  value=c_data["label"],
                                  key="cm_rename")
        rc1,rc2,_ = st.columns([1,1,4])
        with rc1:
            if st.button("💾 Save label", key="cm_save_label"):
                clusters[chosen_idx]["label"] = new_label
                save_clusters(clusters)
                st.success("Label updated.")
                st.rerun()
        with rc2:
            if st.button("🗑 Delete cluster", key="cm_delete"):
                clusters.pop(chosen_idx)
                save_clusters(clusters)
                st.success("Cluster deleted from registry.")
                st.rerun()