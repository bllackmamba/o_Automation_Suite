#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reshape pre-computed grouping output into the 5-section "All" layout (xlsx).

Reshape ONLY — never recompute the algorithm. Reads:
  {prefix}_grouped.csv     (Section 2 n1-n5 incl. blank separators + Section 3 N1-N6 header)
  {prefix}_regenerate.csv  (Section 4)
  {prefix}_remnant.csv     (Section 5)

Sections sit side by side, separated by one blank column, all aligned by a
single global row index. When that index passes ROWS_PER_SHEET the layout
continues on a new sheet with the same headers, picking up where it left off.
"""
import csv
import sys

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

ROWS_PER_SHEET = 1_000_000
TITLE_FONT = Font(bold=True, color="FFFFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")  # dark blue
HDR_FONT = Font(bold=True)

# (title, start_col_0based, [column headers])  -- one blank col between sections
SECTIONS = [
    ("Input",      0,  ["n1", "n2", "n3", "n4", "n5"]),
    ("Grouped",    6,  ["n1", "n2", "n3", "n4", "n5"]),
    ("NList",      12, ["N1", "N2", "N3", "N4", "N5", "N6"]),
    ("Regenerate", 19, ["n1", "n2", "n3", "n4", "n5", "n6"]),
    ("Remnant",    26, ["n1", "n2", "n3", "n4", "n5"]),
]
TOTAL_COLS = 31  # last section spans cols 26..30


def _to_int(s):
    s = s.strip()
    return int(s) if s else None


def grouped_all(path):
    """Yield (n1-5 list, N1-6 list) for every grouped row incl. blank separators."""
    with open(path, newline="") as f:
        r = csv.reader(f)
        next(r, None)  # header
        for row in r:
            row = row + [""] * (12 - len(row))
            yield [_to_int(x) for x in row[0:5]], [_to_int(x) for x in row[6:12]]


def grouped_nonblank(path):
    """Yield n1-5 list for grouped rows that actually carry data (no separators)."""
    with open(path, newline="") as f:
        r = csv.reader(f)
        next(r, None)
        for row in r:
            row = row + [""] * (5 - len(row))
            vals = [_to_int(x) for x in row[0:5]]
            if any(v is not None for v in vals):
                yield vals


def plain_csv(path):
    """Yield list-of-int rows from a simple header+data csv."""
    with open(path, newline="") as f:
        r = csv.reader(f)
        next(r, None)
        for row in r:
            if not any(x.strip() for x in row):
                continue  # skip stray blank lines
            yield [_to_int(x) for x in row]


def _styled(ws, value, font=None, fill=None):
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    return c


def _new_sheet(wb, title):
    ws = wb.create_sheet(title=title)
    ws.sheet_format.defaultColWidth = 5
    # Row 1: section titles (5 styled cells, rest blank)
    title_row = [None] * TOTAL_COLS
    for name, start, _hdr in SECTIONS:
        title_row[start] = _styled(ws, name, TITLE_FONT, TITLE_FILL)
    ws.append(title_row)
    # Row 2: column headers
    hdr_row = [None] * TOTAL_COLS
    for _name, start, hdr in SECTIONS:
        for i, h in enumerate(hdr):
            hdr_row[start + i] = _styled(ws, h, HDR_FONT)
    ws.append(hdr_row)
    return ws


def build(prefix, out_path):
    g_path = f"{prefix}_grouped.csv"
    rg_path = f"{prefix}_regenerate.csv"
    rm_path = f"{prefix}_remnant.csv"

    sec_all = grouped_all(g_path)        # Sections 2 + 3 (every row)
    sec_input = grouped_nonblank(g_path)  # Section 1 (no separators)
    sec_regen = plain_csv(rg_path)       # Section 4
    remnant = list(plain_csv(rm_path))   # Section 5 (tiny)

    wb = Workbook(write_only=True)
    sheet_idx = 1
    ws = _new_sheet(wb, "All")
    sheets = ["All"]
    rows_on_sheet = 0
    total = 0

    input_done = regen_done = all_done = False
    while not all_done:
        # advance to a new sheet at the boundary
        if rows_on_sheet >= ROWS_PER_SHEET:
            sheet_idx += 1
            name = f"All (cont. {sheet_idx})"
            ws = _new_sheet(wb, name)
            sheets.append(name)
            rows_on_sheet = 0

        try:
            n15, N16 = next(sec_all)  # drives the global index (longest section)
        except StopIteration:
            all_done = True
            break

        row = [None] * TOTAL_COLS
        # Section 1 Input (cols 0-4)
        if not input_done:
            try:
                v = next(sec_input)
                row[0:5] = v
            except StopIteration:
                input_done = True
        # Section 2 Grouped (cols 6-10)
        row[6:11] = n15
        # Section 3 NList (cols 12-17)
        row[12:18] = N16
        # Section 4 Regenerate (cols 19-24)
        if not regen_done:
            try:
                v = next(sec_regen)
                row[19:19 + len(v)] = v
            except StopIteration:
                regen_done = True
        # Section 5 Remnant (cols 26-30) -- only the first global rows
        if total < len(remnant):
            v = remnant[total]
            row[26:26 + len(v)] = v

        ws.append(row)
        rows_on_sheet += 1
        total += 1

    wb.save(out_path)
    return sheets, total


if __name__ == "__main__":
    prefix = sys.argv[1] if len(sys.argv) > 1 else "core/c545"
    out = sys.argv[2] if len(sys.argv) > 2 else "core/c545_combined.xlsx"
    sheets, total = build(prefix, out)
    print(f"sheets={len(sheets)} {sheets}")
    print(f"data_rows={total}")
